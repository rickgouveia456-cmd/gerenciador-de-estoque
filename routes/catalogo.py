"""Catalogo Centralizado de Insumos."""
from flask import Blueprint, render_template, request, redirect, url_for, flash, jsonify, send_file
from extensions import db
from models import CatalogoInsumo, Item, Almoxarifado, agora
from core import login_required, admin_required, almoxarife_required, usuario_atual, is_admin_ou_ggo, admin_ou_ggo_required, almoxarifados_do_ggo
import io, logging
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

logger = logging.getLogger(__name__)

catalogo_bp = Blueprint('catalogo_bp', __name__)

CATEGORIAS = ['geral', 'epi', 'maquinario', 'eletrica', 'hidraulica', 'gas']
CAT_LABEL  = {
    'epi': 'EPI', 'maquinario': 'Maquinario', 'eletrica': 'Eletrica',
    'hidraulica': 'Hidraulica', 'gas': 'Gas', 'geral': 'Geral',
}


@catalogo_bp.route('/catalogo/insumos')
@login_required
def catalogo_insumos():
    u     = usuario_atual()
    q     = request.args.get('q', '').strip()
    cat   = request.args.get('categoria', '')
    query = CatalogoInsumo.query.filter_by(ativo=True)
    if not is_admin_ou_ggo(u):
        ids_alm = u.almoxarifados_permitidos()
        if ids_alm:
            query = query.filter(
                db.or_(
                    CatalogoInsumo.almoxarifado_id == None,
                    CatalogoInsumo.almoxarifado_id.in_(ids_alm)
                )
            )
    if q:
        query = query.filter(CatalogoInsumo.nome.ilike(f'%{q}%'))
    if cat:
        query = query.filter_by(categoria=cat)
    alm_id_filtro = request.args.get('alm', type=int)
    if alm_id_filtro:
        query = query.filter(CatalogoInsumo.almoxarifado_id == alm_id_filtro)

    insumos = query.order_by(CatalogoInsumo.nome).all()

    # Montar mapa nome_lower -> item do almoxarifado do usuário
    # para mostrar quantidade e valor em estoque na coluna extra
    mapa_estoque = {}  # nome_lower -> Item
    alm_nome = None

    if u.perfil == 'admin':
        # Admin vê todos os almoxarifados — não filtra por um específico
        pass
    else:
        ids_alm = u.almoxarifados_permitidos()
        if ids_alm:
            # Pega o almoxarifado principal do usuário como referência
            alm_id_ref = u.almoxarifado_id or next(iter(ids_alm))
            alm_ref = db.session.get(Almoxarifado, alm_id_ref)
            if alm_ref:
                alm_nome = alm_ref.nome
                itens_alm = Item.query.filter_by(
                    almoxarifado_id=alm_id_ref, ativo=True
                ).all()
                for it in itens_alm:
                    mapa_estoque[it.nome.lower().strip()] = it

    if is_admin_ou_ggo(u):
        almoxarifados_lista = Almoxarifado.query.order_by(Almoxarifado.nome).all()
    else:
        ids_alm2 = u.almoxarifados_permitidos()
        almoxarifados_lista = Almoxarifado.query.filter(Almoxarifado.id.in_(ids_alm2)).order_by(Almoxarifado.nome).all() if ids_alm2 else []

    return render_template('catalogo_insumos.html',
                           insumos=insumos, q=q, cat=cat,
                           categorias=CATEGORIAS, cat_label=CAT_LABEL,
                           mapa_estoque=mapa_estoque, alm_nome=alm_nome,
                           almoxarifados_lista=almoxarifados_lista,
                           alm_id_filtro=alm_id_filtro)


@catalogo_bp.route('/catalogo/insumos/novo', methods=['GET', 'POST'])
@almoxarife_required
def catalogo_novo():
    u = usuario_atual()
    if is_admin_ou_ggo(u):
        almoxarifados = Almoxarifado.query.order_by(Almoxarifado.nome).all()
    else:
        ids = u.almoxarifados_permitidos()
        almoxarifados = Almoxarifado.query.filter(Almoxarifado.id.in_(ids)).order_by(Almoxarifado.nome).all() if ids else []

    if request.method == 'POST':
        nome = request.form.get('nome', '').strip()
        if not nome:
            flash('Nome e obrigatorio.', 'danger')
            return render_template('catalogo_form.html', insumo=None,
                                   categorias=CATEGORIAS, cat_label=CAT_LABEL,
                                   almoxarifados=almoxarifados, form_data=request.form)
        alm_id = request.form.get('almoxarifado_id', type=int) or None
        existente = CatalogoInsumo.query.filter(
            CatalogoInsumo.nome.ilike(nome),
            CatalogoInsumo.ativo == True,
            CatalogoInsumo.almoxarifado_id == alm_id
        ).first()
        if existente:
            flash(f'Ja existe um insumo com o nome "{existente.nome}" neste catalogo.', 'warning')
            return render_template('catalogo_form.html', insumo=None,
                                   categorias=CATEGORIAS, cat_label=CAT_LABEL,
                                   almoxarifados=almoxarifados, form_data=request.form)
        try:
            valor = float(request.form.get('valor_unitario', '').replace(',', '.') or 0) or None
        except (ValueError, AttributeError):
            valor = None
        ins = CatalogoInsumo(
            nome=nome,
            codigo_ref=request.form.get('codigo_ref', '').strip() or None,
            unidade=request.form.get('unidade', 'un').strip(),
            categoria=request.form.get('categoria', 'geral'),
            ca=request.form.get('ca', '').strip() or None,
            descricao=request.form.get('descricao', '').strip() or None,
            valor_unitario=valor,
            almoxarifado_id=alm_id,
            criado_por=u.nome if u else None,
        )
        db.session.add(ins)
        db.session.commit()
        _sincronizar_valor_itens(ins)
        flash(f'Insumo "{ins.nome}" adicionado ao catalogo!', 'success')
        return redirect(url_for('catalogo_bp.catalogo_insumos'))
    return render_template('catalogo_form.html', insumo=None,
                           categorias=CATEGORIAS, cat_label=CAT_LABEL,
                           almoxarifados=almoxarifados, form_data={})


@catalogo_bp.route('/catalogo/insumos/<int:id>/editar', methods=['GET', 'POST'])
@almoxarife_required
def catalogo_editar(id):
    ins = CatalogoInsumo.query.get_or_404(id)
    if request.method == 'POST':
        ins.nome       = request.form.get('nome', ins.nome).strip()
        ins.codigo_ref = request.form.get('codigo_ref', '').strip() or None
        ins.unidade    = request.form.get('unidade', ins.unidade).strip()
        ins.categoria  = request.form.get('categoria', ins.categoria)
        ins.ca         = request.form.get('ca', '').strip() or None
        ins.descricao  = request.form.get('descricao', '').strip() or None
        try:
            ins.valor_unitario = float(request.form.get('valor_unitario', '').replace(',', '.') or 0) or None
        except (ValueError, AttributeError):
            ins.valor_unitario = None
        db.session.commit()
        # Sincronizar valor com itens cadastrados
        _sincronizar_valor_itens(ins)
        flash('Insumo atualizado!', 'success')
        return redirect(url_for('catalogo_bp.catalogo_insumos'))
    return render_template('catalogo_form.html', insumo=ins,
                           categorias=CATEGORIAS, cat_label=CAT_LABEL, form_data={})


@catalogo_bp.route('/catalogo/insumos/<int:id>/deletar', methods=['POST'])
@admin_ou_ggo_required
def catalogo_deletar(id):
    ins = CatalogoInsumo.query.get_or_404(id)
    ins.ativo = False
    db.session.commit()
    flash(f'Insumo "{ins.nome}" removido do catalogo.', 'warning')
    return redirect(url_for('catalogo_bp.catalogo_insumos'))


@catalogo_bp.route('/catalogo/insumos/importar', methods=['GET', 'POST'])
@almoxarife_required
def catalogo_importar():
    """Importa insumos do Excel e sincroniza valores com o estoque."""
    u = usuario_atual()
    if request.method == 'POST':
        arquivo = request.files.get('arquivo')
        if not arquivo or not arquivo.filename.endswith(('.xlsx', '.xls')):
            flash('Envie um arquivo Excel (.xlsx ou .xls).', 'danger')
            return redirect(url_for('catalogo_bp.catalogo_importar'))
        try:
            wb = openpyxl.load_workbook(arquivo, data_only=True)
            ws = wb.active
            inseridos = atualizados = erros = 0
            for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                if not any(row):
                    continue
                try:
                    nome      = str(row[0]).strip() if row[0] else None
                    codigo_ref= str(row[1]).strip() if row[1] else None
                    unidade   = str(row[2]).strip() if row[2] else 'un'
                    try:
                        valor = float(str(row[3]).replace(',', '.')) if row[3] else None
                    except (ValueError, TypeError):
                        valor = None
                    categoria = str(row[4]).strip().lower() if len(row) > 4 and row[4] else 'geral'
                    if categoria not in CATEGORIAS:
                        categoria = 'geral'
                except Exception:
                    erros += 1
                    continue
                if not nome:
                    erros += 1
                    continue
                existente = CatalogoInsumo.query.filter(
                    CatalogoInsumo.nome.ilike(nome), CatalogoInsumo.ativo == True
                ).first()
                if existente:
                    # Atualizar valor se fornecido
                    if codigo_ref:
                        existente.codigo_ref = codigo_ref
                    if unidade:
                        existente.unidade = unidade
                    if valor is not None:
                        existente.valor_unitario = valor
                    existente.categoria = categoria
                    _sincronizar_valor_itens(existente)
                    atualizados += 1
                else:
                    alm_id_cat = request.form.get('almoxarifado_id', type=int) or None
                    novo = CatalogoInsumo(
                        nome=nome, codigo_ref=codigo_ref, unidade=unidade,
                        categoria=categoria, valor_unitario=valor,
                        almoxarifado_id=alm_id_cat,
                        criado_por=u.nome if u else None
                    )
                    db.session.add(novo)
                    db.session.flush()
                    _sincronizar_valor_itens(novo)
                    inseridos += 1
            db.session.commit()
            flash(f'Importacao concluida: {inseridos} inseridos, {atualizados} atualizados'
                  + (f', {erros} com erro.' if erros else '.'), 'success' if not erros else 'warning')
        except Exception as e:
            flash(f'Erro ao processar arquivo: {e}', 'danger')
        return redirect(url_for('catalogo_bp.catalogo_insumos'))
    if u.perfil == 'admin':
        almoxarifados = Almoxarifado.query.order_by(Almoxarifado.nome).all()
    else:
        ids = u.almoxarifados_permitidos()
        almoxarifados = Almoxarifado.query.filter(Almoxarifado.id.in_(ids)).order_by(Almoxarifado.nome).all() if ids else []
    return render_template('catalogo_importar.html', almoxarifados=almoxarifados)


@catalogo_bp.route('/catalogo/insumos/modelo-excel')
@login_required
def catalogo_modelo_excel():
    """Gera modelo Excel para importacao do catalogo."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Catalogo'
    h_fill = PatternFill('solid', fgColor='1A3A5C')
    h_font = Font(bold=True, color='FFFFFF', size=11)
    borda  = Border(left=Side(style='thin'), right=Side(style='thin'),
                    top=Side(style='thin'), bottom=Side(style='thin'))
    headers = ['Nome', 'Codigo Referencia', 'Unidade', 'Valor Unitario (R$)', 'Categoria']
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=col, value=h)
        c.font = h_font; c.fill = h_fill
        c.alignment = Alignment(horizontal='center'); c.border = borda
    exemplos = [
        ('Capacete de Seguranca Classe A', 'CAP-001', 'un', 45.90, 'epi'),
        ('Cimento CP-II', 'CIM-001', 'sc', 38.50, 'geral'),
        ('Cabo Eletrico 2.5mm', 'CAB-025', 'm', 4.20, 'eletrica'),
    ]
    for r, ex in enumerate(exemplos, 2):
        for c, v in enumerate(ex, 1):
            ws.cell(row=r, column=c, value=v).border = borda
    for i, w in enumerate([40, 20, 10, 20, 15], 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    buf = io.BytesIO()
    wb.save(buf); buf.seek(0)
    return send_file(buf, as_attachment=True, download_name='modelo_catalogo.xlsx',
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')


@catalogo_bp.route('/catalogo/valor-estoque')
@login_required
def catalogo_valor_estoque():
    """Mostra valor total do estoque por almoxarifado (quantidade x valor_unitario)."""
    u = usuario_atual()
    if is_admin_ou_ggo(u):
        alms = Almoxarifado.query.all()
    elif u.perfil == 'analista':
        alms = [db.session.get(Almoxarifado, u.almoxarifado_id)] if u.almoxarifado_id else []
    else:
        ids = u.almoxarifados_permitidos()
        alms = Almoxarifado.query.filter(Almoxarifado.id.in_(ids)).all() if ids else []

    resumo = []
    total_geral = 0.0
    for alm in alms:
        itens = Item.query.filter_by(almoxarifado_id=alm.id, ativo=True).all()
        valor_alm = 0.0
        itens_com_valor = []
        itens_sem_valor = 0
        for it in itens:
            if it.valor_unitario and it.valor_unitario > 0:
                vt = it.quantidade * it.valor_unitario
                valor_alm += vt
                itens_com_valor.append({
                    'item': it,
                    'valor_total': vt,
                    'valor_unitario': it.valor_unitario
                })
            else:
                itens_sem_valor += 1
        itens_com_valor.sort(key=lambda x: x['valor_total'], reverse=True)
        total_geral += valor_alm
        resumo.append({
            'almoxarifado': alm,
            'valor_total': valor_alm,
            'itens': itens_com_valor,
            'itens_sem_valor': itens_sem_valor,
            'total_itens': len(itens),
        })
    resumo.sort(key=lambda x: x['valor_total'], reverse=True)
    return render_template('catalogo_valor_estoque.html',
                           resumo=resumo, total_geral=total_geral)


@catalogo_bp.route('/api/catalogo/buscar')
@login_required
def api_catalogo_buscar():
    q   = request.args.get('q', '').strip()
    cat = request.args.get('categoria', '')
    query = CatalogoInsumo.query.filter_by(ativo=True)
    if q:
        query = query.filter(CatalogoInsumo.nome.ilike(f'%{q}%'))
    if cat:
        query = query.filter_by(categoria=cat)
    insumos = query.order_by(CatalogoInsumo.nome).limit(15).all()
    return jsonify([{
        'id': i.id, 'nome': i.nome, 'codigo_ref': i.codigo_ref or '',
        'unidade': i.unidade, 'categoria': i.categoria, 'ca': i.ca or '',
        'descricao': i.descricao or '',
        'valor_unitario': i.valor_unitario or 0,
    } for i in insumos])


def _sincronizar_valor_itens(ins):
    """Propaga valor_unitario do catalogo para todos os itens com mesmo nome."""
    if not ins.valor_unitario:
        return
    itens = Item.query.filter(Item.nome.ilike(ins.nome), Item.ativo == True).all()
    for it in itens:
        it.valor_unitario = ins.valor_unitario
    if itens:
        try:
            db.session.commit()
            logger.info(f'Valor sincronizado: {len(itens)} itens com nome "{ins.nome}" -> R${ins.valor_unitario}')
        except Exception as e:
            logger.error(f'Erro ao sincronizar valor: {e}')
            db.session.rollback()


@catalogo_bp.route('/admin/seed-catalogo-estrutura', methods=['POST'])
@admin_required
def seed_catalogo_estrutura():
    """Seed dos itens do catálogo da Estrutura Ventura Patamares."""
    from models import Almoxarifado
    # Busca o almoxarifado de estrutura
    alm = Almoxarifado.query.filter(Almoxarifado.nome.ilike('%estrutura%')).first()
    alm_id = alm.id if alm else None

    ITENS = [
        ("Abraçadeira Nylon 20 cm 4.8 mm","9488","un",0.13,"geral"),
        ("Distanciador Parede Concreto D.100mm Aço 4,2-6,3mm","8448","un",0.18,"geral"),
        ("Distanciador Parede Concreto Cob.120mm Aço 4,2-8,0mm","14315","un",0.30,"geral"),
        ("Espaçador Centopeia Multiapoio 20mm","14377","un",0.34,"geral"),
        ("Joelho 90° soldável 25mm","10473","UND",0.39,"geral"),
        ("Gabarito Redondo D.100mm Parede 10cm","8456","un",0.52,"geral"),
        ("Mascara Descartável PFF2 S/ Filtro Azul","9207","un",0.78,"epi"),
        ("Esponja Dupla Face","14375","UND",0.78,"geral"),
        ("Gabarito Redondo D.120mm Parede 12cm","14316","un",0.90,"geral"),
        ("Distanciador Parede Concreto Cob.150mm Aço 4,2-8,0mm","8449","un",0.91,"geral"),
        ("Espatula 8","7113","un",0.92,"geral"),
        ("Protetor auricular plug","9049","UND",0.95,"epi"),
        ("lapis carpinteiro","6982","un",1.10,"geral"),
        ("Luva Eletroduto Reforçado Laranja 32mm","8431","UND",1.10,"eletrica"),
        ("Luva Eletroduto Reforçado Laranja 25mm","8430","un",1.21,"eletrica"),
        ("Conector Elétrico Wago 3 Saídas","9268","UND",1.25,"eletrica"),
        ("Eletroduto Reforçado Laranja 20mm","8426","m",1.40,"eletrica"),
        ("Conector Elétrico Wago 5 Saídas","9269","UND",1.92,"eletrica"),
        ("Luva Flextactil","7794","un",2.43,"epi"),
        ("SACO DE RAFIA","18002","un",2.50,"geral"),
        ("Corda Poliamida 12 mm","9743","M",2.68,"geral"),
        ("Eletroduto Reforçado Laranja 25mm","8427","m",2.80,"eletrica"),
        ("CAMURÇA","9065","un",2.90,"geral"),
        ("Joelho 90° PVC Marrom Soldável LR 25mm","12850","UND",2.93,"geral"),
        ("Óculos de Proteção Escuro","7795","UND",3.36,"epi"),
        ("Óculos de Proteção incolor","9047","UND",3.36,"epi"),
        ("Eletroduto Reforçado Laranja 32mm","8428","M",3.90,"eletrica"),
        ("Óculos de proteção Sobrepor Escuro","14441","un",4.16,"epi"),
        ("Plug Elétrico Macho 2P+T 20A","8199","un",4.80,"eletrica"),
        ("Fita Zebrada","7790","un",5.13,"geral"),
        ("Óculos de Sobrepor Incolor","9205","UND",5.28,"epi"),
        ("Plug Elétrico Fêmea 2P+T 20A","8200","un",5.50,"eletrica"),
        ("Fita Veda Rosca 18mm x 50m","8486","UND",5.90,"hidraulica"),
        ("Linha de Nylon para Pedreiro","6972","UND",6.00,"geral"),
        ("FITA ISOLANTE 18X20M","7484","UND",6.20,"eletrica"),
        ("Aço CA-50 Ø 12,5mm (1/2\")","5402","kg",6.27,"geral"),
        ("Aço CA-50 Ø 10mm","7658","kg",6.27,"geral"),
        ("Treliça Ptg 8L 12 metros","1018","un",6.62,"geral"),
        ("ESPATULA 4 POL","7112","un",7.00,"geral"),
        ("Aço CA-50 Ø 8mm (5/16\")","5387","KG",7.15,"geral"),
        ("Aço CA-50 Ø 6,3mm (1/4\") Cortado e Dobrado","6056","KG",7.18,"geral"),
        ("Mangueira Cristal Trançada 3/4'' PT 250","8704","MT",7.35,"hidraulica"),
        ("Lamina de Serra","7040","un",7.83,"geral"),
        ("Aço CA-60 Ø 5mm (3/16\")","5401","KG",7.87,"geral"),
        ("Pasta Lubrificante 400g","8324","un",8.68,"geral"),
        ("Trinchão","6971","UND",8.70,"geral"),
        ("FITA CREPE 50MMX50M","4828","un",9.20,"geral"),
        ("Garfo reforçado gaiola Rolo 23cm","9364","UND",9.50,"geral"),
        ("Cabo PP Flexível 3X2,5mm2 Preto","7005","m",9.80,"eletrica"),
        ("Balde plástico","11367","un",9.85,"geral"),
        ("Protetor auricular concha","6958","un",11.62,"epi"),
        ("Óculos de Segurança Ampla Visão","9048","UND",11.81,"epi"),
        ("Agente Desmoldante de Alto Desempenho Ecológico: Vegetal","8461","L",11.98,"geral"),
        ("Luva Vaqueta Cano Curto com Elástico","9127","PAR",12.23,"epi"),
        ("Desempenadeira Aço Galo Lisa 12x29cm","15143","UND",13.00,"geral"),
        ("Vassoura Piaçava","5733","un",13.50,"geral"),
        ("Rolo de Lã antigota 23cm","17367","UND",14.00,"geral"),
        ("Disco Diamantado Makita Continuo 105mm","6983","un",15.00,"maquinario"),
        ("Desempenadeira Dentada 10mm","6969","UND",15.43,"geral"),
        ("Mascara Respirador c/ 2 Filtro","9208","UND",15.63,"epi"),
        ("CAPA DE CHUVA","7788","UNI",15.69,"epi"),
        ("VASSOURÃO DE PIAÇACAVA 40CM COM CABO","7527","UND",16.00,"geral"),
        ("Papel toalha interfolhas 1000 unid","7524","pc",17.50,"geral"),
        ("Macacão Tyvek","6962","un",17.90,"epi"),
        ("Pigmento Pó xadrez Vermelho","10089","UND",18.00,"geral"),
        ("Espátula","11227","UND",18.00,"geral"),
        ("Luva Vaqueta Cano Longo","16565","un",18.84,"epi"),
        ("Arame Recozido Nº 18","4719","un",19.30,"geral"),
        ("Misturador de Tinta e Gesso 100x600mm SDS Plus","11002","un",19.35,"geral"),
        ("Balde de Ferro","7550","un",19.90,"geral"),
        ("Arame Revestido com PVC Verde BWG 16","16357","RL",20.00,"geral"),
        ("Sarrafo em Alumínio","4777","UND",21.00,"geral"),
        ("Fitilho para Amarração 950g","18323","un",21.00,"geral"),
        ("Colher de Pedreiro","12004","un",21.87,"geral"),
        ("Pá Quadrada","7543","UND",21.90,"geral"),
        ("Pá de Bico","7544","UND",21.90,"geral"),
        ("Fita Guia Passa Fio 10m","16717","un",22.00,"eletrica"),
        ("CARNEIRA PARA CAPACETE","12863","UND",22.70,"epi"),
        ("PREGO 18X27","5330","KG",24.99,"geral"),
        ("Martelo tipo Unha 27mm C/ Cabo Fibra","9725","un",25.00,"geral"),
        ("Mosquetão Oval Aço Trava Roscada 25kN","14362","UND",25.41,"epi"),
        ("Chave de Dobrar ferro Reforçada 12mm","9480","PC",26.24,"geral"),
        ("Torqueza 12\"","9169","un",31.00,"geral"),
        ("Perneira de segurança","17549","UND",33.20,"epi"),
        ("Tarkoprimer","15515","un",33.58,"geral"),
        ("Capacete Amarelo","8707","UND",35.50,"epi"),
        ("CAPACETE AZUL","8708","UND",35.50,"epi"),
        ("Capacete Cinza","8710","UND",35.50,"epi"),
        ("Capacete Marron","8712","UND",35.50,"epi"),
        ("Capacete Verde","8713","UND",35.50,"epi"),
        ("Capacete Vermelho","8714","UND",35.50,"epi"),
        ("Plug industrial 2P+T IP44 16A 220V","16334","un",37.54,"eletrica"),
        ("PNEU PARA CARRINHO PLATAFORMA COMPLETO","9561","PC",38.00,"geral"),
        ("Kit Giz de Linha Abs 30 Metros Vermelho","10088","un",40.00,"geral"),
        ("Bota de Couro N 43","7786","PAR",40.12,"epi"),
        ("Bota de Couro N 41","8665","PAR",40.12,"epi"),
        ("Bota de Couro N 44","8667","PAR",40.12,"epi"),
        ("Bota de Couro N 38","9128","un",40.12,"epi"),
        ("Bota de Couro Nº 39","9449","un",40.12,"epi"),
        ("BOTA PVC 42","11371a","par",40.25,"epi"),
        ("BOTA PVC 44","11371b","par",40.25,"epi"),
        ("BOTA PVC 45","11371c","un",40.25,"epi"),
        ("BOTA PVC 39","11371d","un",40.25,"epi"),
        ("BOTA PVC 40","11371e","un",40.25,"epi"),
        ("BOTA PVC 43","11371f","par",40.25,"epi"),
        ("BOTA PVC 41","11371g","par",40.25,"epi"),
        ("LIMA CHATA COM CABO FINO","10673","un",41.71,"geral"),
        ("Camisa Brim Leve Manga Longa M","7950","un",45.00,"epi"),
        ("CAMISA BRIM LEVE MANGA LONGA G","7951","UND",45.00,"epi"),
        ("CAMISA BRIM LEVE MANGA LONGA GG","7952","UND",45.00,"epi"),
        ("CALÇA BRIM CORDÃO ELASTICO M","7954","un",45.00,"epi"),
        ("CALÇA BRIM CORDÃO ELASTICO G","7955","un",45.00,"epi"),
        ("CALÇA BRIM CORDÃO ELASTICO EX GG","7956","un",45.00,"epi"),
        ("CAMISA BRIM LEVE MANGA LONGA EX GG","9730","UND",45.00,"epi"),
        ("Garrafa termica 5L","10477","un",46.81,"geral"),
        ("Desempenadeira","11622","un",49.99,"geral"),
        ("Caixa para massa 60x43x17cm","9087","un",54.00,"geral"),
        ("Alavanca de Aço","7547","UND",75.00,"geral"),
        ("Tarkomassa Preparação (20 Kg)","17096","pc",96.00,"geral"),
        ("Bota Segurança Metatarso NR10 Composite N 39","17815","PAR",169.00,"epi"),
        ("Bota Segurança Metatarso NR10 Composite N 40","17816","PAR",169.00,"epi"),
        ("Bota Segurança Metatarso NR10 Composite N 41","17817","PAR",169.00,"epi"),
        ("Bota Segurança Metatarso NR10 Composite N 42","17818","PAR",169.00,"epi"),
        ("Bota Segurança Metatarso NR10 Composite N 43","17819","PAR",169.00,"epi"),
        ("Bota de Segurança Metatarso NR10 Composite N 44","17820","par",169.00,"epi"),
        ("Bota Segurança Metatarso NR10 Composite N 45","17821","PAR",169.00,"epi"),
        ("Cinto de Segurança tipo Paraquedista STF","9463","UND",128.00,"epi"),
        ("Maçarico Portátil","9652","UND",130.00,"maquinario"),
        ("Tesoura Corta Vergalhão de 24","8275","UND",140.00,"geral"),
        ("Bota Segurança Metatarso NR10 Composite N 46","17822","un",169.00,"epi"),
        ("Disco de desbaste diamantado DG-CW 125/5","16154","un",163.00,"maquinario"),
        ("Tela Soldada Q-92 Malha 15x15cm Fio 4,2mm","8618","UND",164.80,"eletrica"),
        ("Talabarte segurança duplo fita poliéster HL032YEA","16672","UND",168.00,"epi"),
        ("Tela Soldada EQ-92 Malha POP 15x15cm Fio 4,2mm","8621","UND",178.20,"eletrica"),
        ("CHAVE DE GRIFO 36","9658","un",180.00,"geral"),
        ("Peça de Isopor 1,44mx1,205mx0,22m","16152","un",184.86,"geral"),
        ("Colar Cervical Regulável","17577","un",186.75,"epi"),
        ("Tela Soldada Q-138 Malha 10x10cm Fio 4,2mm 2,45x6,00m","8620","UND",234.99,"eletrica"),
        ("Tela Soldada Q-166 Malha 10x10cm Fio 4,6mm 2,45x6,0m","15331","UND",236.72,"eletrica"),
        ("Broca diamantada 42mm x 320mm Perfuratriz","19491","un",238.00,"maquinario"),
        ("Broca Diamantada 63mm x 370mm Perfuratriz","15219","UND",266.31,"maquinario"),
        ("Broca Diamantada 160mm x 450mm Perfuratriz","18916","un",388.08,"maquinario"),
        ("Fita intumescente","11175","UND",498.92,"geral"),
        ("Prego para betão X-C 30 B3 MX 30mm","14320","un",572.83,"geral"),
        ("Prego para betão X-C 20 B3 MX","15271","un",584.01,"geral"),
        ("Trava-quedas retrátil 3,5m STF CQTQ0173","10964","un",680.00,"epi"),
        ("ESQUADRO DE ALUMÍNIO","9468","un",980.00,"geral"),
        ("Broca diamantada 112mm x 430mm Perfuratriz","19492","UND",770.00,"maquinario"),
    ]

    inseridos = atualizados = 0
    for nome, cod, unid, valor, cat in ITENS:
        existente = CatalogoInsumo.query.filter(
            CatalogoInsumo.nome.ilike(nome), CatalogoInsumo.ativo == True
        ).first()
        if existente:
            existente.valor_unitario = valor
            existente.almoxarifado_id = alm_id
            _sincronizar_valor_itens(existente)
            atualizados += 1
        else:
            novo = CatalogoInsumo(
                nome=nome, codigo_ref=cod, unidade=unid,
                categoria=cat, valor_unitario=valor,
                almoxarifado_id=alm_id,
                criado_por='seed-estrutura'
            )
            db.session.add(novo)
            db.session.flush()
            _sincronizar_valor_itens(novo)
            inseridos += 1
    db.session.commit()
    flash(f'✅ Catálogo da Estrutura: {inseridos} inseridos, {atualizados} atualizados!', 'success')
    return redirect(url_for('catalogo_bp.catalogo_insumos'))
