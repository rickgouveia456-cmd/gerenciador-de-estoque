#!/usr/bin/env python3
"""
Script de backup automático executado pelo GitHub Actions.
Conecta no banco PostgreSQL, gera Excel e envia por email.
"""
import os
import sys
import smtplib
from email.mime.multipart import MIMEMultipart
from email.mime.base import MIMEBase
from email.mime.text import MIMEText
from email import encoders
from datetime import datetime, date
import io
import psycopg2
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

def conectar_banco():
    """Conecta no banco PostgreSQL usando a URL do Railway."""
    database_url = os.environ.get('DATABASE_URL')
    if not database_url:
        print('❌ DATABASE_URL não configurada')
        sys.exit(1)
    
    # Converte postgres:// para postgresql://
    if database_url.startswith('postgres://'):
        database_url = database_url.replace('postgres://', 'postgresql://', 1)
    
    try:
        conn = psycopg2.connect(database_url)
        print('✅ Conectado ao banco PostgreSQL')
        return conn
    except Exception as e:
        print(f'❌ Erro ao conectar no banco: {e}')
        sys.exit(1)

def buscar_dados(conn):
    """Busca todos os dados necessários do banco."""
    cursor = conn.cursor()
    
    # Busca almoxarifados
    cursor.execute('SELECT id, nome FROM almoxarifado ORDER BY nome')
    almoxarifados = [{'id': row[0], 'nome': row[1]} for row in cursor.fetchall()]
    
    # Busca itens de cada almoxarifado
    for alm in almoxarifados:
        cursor.execute('''
            SELECT codigo, descricao, unidade, quantidade, minimo, maximo, localizacao
            FROM item
            WHERE almoxarifado_id = %s AND ativo = true
            ORDER BY codigo
        ''', (alm['id'],))
        alm['itens'] = cursor.fetchall()
    
    # Busca usuários com email
    cursor.execute('''
        SELECT id, nome, email, perfil, almoxarifado_id
        FROM usuario
        WHERE ativo = true AND email IS NOT NULL AND email != ''
    ''')
    usuarios = [{'id': row[0], 'nome': row[1], 'email': row[2], 'perfil': row[3], 'almoxarifado_id': row[4]} 
                for row in cursor.fetchall()]
    
    cursor.close()
    return almoxarifados, usuarios

def gerar_excel_almoxarifado(alm):
    """Gera Excel com itens de um almoxarifado específico."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = alm['nome'][:31]  # Limite de 31 caracteres
    
    # Cabeçalho
    ws['A1'] = f"BACKUP DO ALMOXARIFADO: {alm['nome']}"
    ws['A1'].font = Font(bold=True, size=14)
    ws['A2'] = f"Data: {datetime.now().strftime('%d/%m/%Y %H:%M')}"
    
    # Títulos das colunas
    colunas = ['Código', 'Descrição', 'Unidade', 'Quantidade', 'Mínimo', 'Máximo', 'Localização']
    for col, titulo in enumerate(colunas, 1):
        cell = ws.cell(4, col, titulo)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color='CCCCCC', end_color='CCCCCC', fill_type='solid')
    
    # Dados
    for row_idx, item in enumerate(alm['itens'], 5):
        for col_idx, valor in enumerate(item, 1):
            ws.cell(row_idx, col_idx, valor)
    
    # Ajusta largura das colunas
    for col in range(1, 8):
        ws.column_dimensions[get_column_letter(col)].width = 15
    
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf

def gerar_excel_completo(almoxarifados):
    """Gera Excel com todos os almoxarifados."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Backup Completo'
    
    # Cabeçalho
    ws['A1'] = 'BACKUP COMPLETO DO ESTOQUE'
    ws['A1'].font = Font(bold=True, size=14)
    ws['A2'] = f"Data: {datetime.now().strftime('%d/%m/%Y %H:%M')}"
    
    # Títulos das colunas
    colunas = ['Almoxarifado', 'Código', 'Descrição', 'Unidade', 'Quantidade', 'Mínimo', 'Máximo', 'Localização']
    for col, titulo in enumerate(colunas, 1):
        cell = ws.cell(4, col, titulo)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color='CCCCCC', end_color='CCCCCC', fill_type='solid')
    
    # Dados
    row_idx = 5
    for alm in almoxarifados:
        for item in alm['itens']:
            ws.cell(row_idx, 1, alm['nome'])
            for col_idx, valor in enumerate(item, 2):
                ws.cell(row_idx, col_idx, valor)
            row_idx += 1
    
    # Ajusta largura das colunas
    for col in range(1, 9):
        ws.column_dimensions[get_column_letter(col)].width = 15
    
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf

def enviar_email(destinatario, assunto, corpo, anexo_buf, nome_arquivo):
    """Envia email com anexo."""
    remetente = os.environ.get('BACKUP_EMAIL_FROM')
    senha = os.environ.get('BACKUP_EMAIL_PASS')
    
    if not remetente or not senha:
        print('❌ Credenciais de email não configuradas')
        return False
    
    msg = MIMEMultipart()
    msg['From'] = remetente
    msg['To'] = destinatario
    msg['Subject'] = assunto
    msg.attach(MIMEText(corpo, 'plain'))
    
    part = MIMEBase('application', 'octet-stream')
    part.set_payload(anexo_buf.read())
    encoders.encode_base64(part)
    part.add_header('Content-Disposition', f'attachment; filename="{nome_arquivo}"')
    msg.attach(part)
    
    try:
        with smtplib.SMTP_SSL('smtp.gmail.com', 465) as smtp:
            smtp.login(remetente, senha.replace(' ', ''))
            smtp.send_message(msg)
        return True
    except Exception as e:
        print(f'❌ Erro ao enviar email para {destinatario}: {e}')
        return False

def main():
    """Função principal."""
    print('=' * 60)
    print('BACKUP AUTOMÁTICO - GitHub Actions')
    print(f'Iniciado em: {datetime.now().strftime("%d/%m/%Y %H:%M")}')
    print('=' * 60)
    
    # Conecta no banco
    conn = conectar_banco()
    
    # Busca dados
    print('📊 Buscando dados do banco...')
    almoxarifados, usuarios = buscar_dados(conn)
    conn.close()
    
    print(f'✅ {len(almoxarifados)} almoxarifados encontrados')
    print(f'✅ {len(usuarios)} usuários com email encontrados')
    
    hoje = date.today().strftime('%d/%m/%Y')
    enviados = 0
    erros = 0
    
    # Envia backup individual para cada almoxarife
    for alm in almoxarifados:
        almoxarifes = [u for u in usuarios if u['perfil'] == 'almoxarife' and u['almoxarifado_id'] == alm['id']]
        
        for user in almoxarifes:
            print(f'📧 Enviando backup de "{alm["nome"]}" para {user["email"]}...')
            buf = gerar_excel_almoxarifado(alm)
            nome_arquivo = f"backup_{alm['nome'].replace(' ', '_')}_{date.today()}.xlsx"
            corpo = (
                f'Backup automático do almoxarifado: {alm["nome"]}\n'
                f'Data: {datetime.now().strftime("%d/%m/%Y %H:%M")}\n\n'
                f'Este arquivo contém o estoque atual do seu almoxarifado.\n'
                f'Guarde em local seguro.'
            )
            
            if enviar_email(user['email'], f'Backup {alm["nome"]} — {hoje}', corpo, buf, nome_arquivo):
                print(f'  ✅ Enviado para {user["email"]}')
                enviados += 1
            else:
                erros += 1
    
    # Envia backup completo para admins
    admins = [u for u in usuarios if u['perfil'] == 'admin']
    email_fixo = os.environ.get('BACKUP_EMAIL_TO', 'rickgouveia157@gmail.com')
    
    # Adiciona email fixo se não estiver na lista
    emails_admins = list(set([u['email'] for u in admins] + [email_fixo]))
    
    for email in emails_admins:
        print(f'📧 Enviando backup completo para {email}...')
        buf = gerar_excel_completo(almoxarifados)
        nome_arquivo = f"backup_completo_{date.today()}.xlsx"
        corpo = (
            f'Backup automático completo do sistema de estoque.\n'
            f'Data: {datetime.now().strftime("%d/%m/%Y %H:%M")}\n\n'
            f'Este arquivo contém todos os almoxarifados.\n'
            f'Guarde em local seguro.'
        )
        
        if enviar_email(email, f'Backup Completo Estoque — {hoje}', corpo, buf, nome_arquivo):
            print(f'  ✅ Enviado para {email}')
            enviados += 1
        else:
            erros += 1
    
    print('=' * 60)
    print(f'✅ Backup concluído!')
    print(f'   Enviados: {enviados}')
    print(f'   Erros: {erros}')
    print('=' * 60)
    
    sys.exit(0 if erros == 0 else 1)

if __name__ == '__main__':
    main()
