import io
import json
import os
import re
from collections import defaultdict
from datetime import datetime, date
from flask import render_template, request, redirect, url_for, jsonify, flash, send_file, session, current_app
from flask_wtf.csrf import CSRFError
from markupsafe import escape

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from . import db
from .models import (
    Almoxarifado, Item, Movimentacao, Requisicao,
    RequisicaoMestre, RequisicaoMestreItem, Usuario,
    Colaborador, AcessoExtra
)
from .utils import (
    flash_html, login_required, admin_required, almoxarife_required,
    usuario_atual, extrair_colaborador, agora,
    registrar_tentativa_login, valida_limite_login
)
from .backup import gerar_excel_backup, gerar_excel_backup_almoxarifado, enviar_backup_por_almoxarifado


def init_routes(app):
    @app.errorhandler(CSRFError)
    def handle_csrf_error(e):
        flash('Sessão expirada ou requisição inválida. Tente novamente.', 'warning')
        return redirect(request.referrer or url_for('index'))

    @app.route('/login', methods=['GET', 'POST'])
    def login():
        if request.method == 'POST':
            ip = request.remote_addr or '0.0.0.0'
            if not valida_limite_login(ip):
                flash('Muitas tentativas. Aguarde 5 minutos.', 'danger')
                return render_template('login.html'), 429
            login_val = request.form.get('login', '').strip()
            senha_val = request.form.get('senha', '')
            if not login_val or not senha_val:
                flash('Preencha login e senha.', 'warning')
                return render_template('login.html')
            u = Usuario.query.filter_by(login=login_val, ativo=True).first()
            if u and u.check_senha(senha_val):
                session.clear()
                session['usuario_id'] = u.id
                flash(f'Bem-vindo, {u.nome}!', 'success')
                return redirect(url_for('index'))
            registrar_tentativa_login(ip)
            flash('Login ou senha incorretos.', 'danger')
        return render_template('login.html')

    @app.route('/logout')
    def logout():
        session.clear()
        return redirect(url_for('login'))

    @app.route('/')
    @login_required
    def index():
        u = usuario_atual()
        if u.perfil in ('mestre', 'tecnico_seguranca'):
            return redirect(url_for('mestre_requisicoes'))
        if u.perfil == 'admin':
            almoxarifados = Almoxarifado.query.all()
            alertas = Item.query.filter(Item.quantidade <= Item.estoque_minimo, Item.ativo == True).all()
        else:
            ids = u.almoxarifados_permitidos()
            almoxarifados = Almoxarifado.query.filter(Almoxarifado.id.in_(ids)).all() if ids else []
            alertas = Item.query.filter(
                Item.quantidade <= Item.estoque_minimo,
                Item.almoxarifado_id.in_(ids),
                Item.ativo == True
            ).all() if ids else []
        stats = {
            'total_almoxarifados': len(almoxarifados),
            'total_itens': sum(len(a.itens) for a in almoxarifados),
            'itens_alerta': len([a for a in alertas if a.quantidade > 0]),
            'itens_criticos': len([a for a in alertas if a.quantidade <= 0]),
        }
        return render_template('index.html', almoxarifados=almoxarifados, alertas=alertas, stats=stats)

    @app.route('/almoxarifado/<int:id>')
    @login_required
    def almoxarifado(id):
        u = usuario_atual()
        if u.perfil in ('mestre', 'tecnico_seguranca'):
            flash('Acesso restrito. Use a tela de requisições.', 'warning')
            return redirect(url_for('mestre_requisicoes'))
        alm = Almoxarifado.query.get_or_404(id)
        if u.perfil != 'admin' and id not in u.almoxarifados_permitidos():
            flash('Acesso negado.', 'danger')
            return redirect(url_for('index'))
        itens = Item.query.filter_by(almoxarifado_id=id).order_by(Item.ativo.desc(), Item.nome).all()
        return render_template('almoxarifado.html', almoxarifado=alm, itens=itens)

    @app.route('/item/<int:id>')
    @login_required
    def item(id):
        it = Item.query.get_or_404(id)
        u = usuario_atual()
        if u.perfil != 'admin' and it.almoxarifado_id not in u.almoxarifados_permitidos():
            flash('Acesso negado.', 'danger')
            return redirect(url_for('index'))
        movs = Movimentacao.query.filter_by(item_id=id).order_by(Movimentacao.data.desc()).limit(50).all()
        return render_template('item.html', item=it, movimentacoes=movs)

    @app.route('/almoxarifado/novo', methods=['GET', 'POST'])
    @admin_required
    def novo_almoxarifado():
        if request.method == 'POST':
            alm = Almoxarifado(nome=request.form['nome'], descricao=request.form.get('descricao', ''))
            db.session.add(alm)
            db.session.commit()
            flash(f'Almoxarifado "{alm.nome}" criado!', 'success')
            return redirect(url_for('index'))
        return render_template('form_almoxarifado.html', almoxarifado=None)

    @app.route('/almoxarifado/<int:id>/editar', methods=['GET', 'POST'])
    @login_required
    def editar_almoxarifado(id):
        alm = Almoxarifado.query.get_or_404(id)
        u = usuario_atual()
        if u.perfil != 'admin' and id not in u.almoxarifados_permitidos():
            flash('Acesso negado.', 'danger')
            return redirect(url_for('index'))
        if request.method == 'POST':
            alm.nome = request.form['nome']
            alm.descricao = request.form.get('descricao', '')
            db.session.commit()
            flash('Almoxarifado atualizado!', 'success')
            return redirect(url_for('index'))
        return render_template('form_almoxarifado.html', almoxarifado=alm)

    @app.route('/almoxarifado/<int:id>/deletar', methods=['POST'])
    @admin_required
    def deletar_almoxarifado(id):
        alm = Almoxarifado.query.get_or_404(id)
        db.session.delete(alm)
        db.session.commit()
        flash('Almoxarifado removido!', 'warning')
        return redirect(url_for('index'))

    @app.route('/item/novo', methods=['GET', 'POST'])
    @almoxarife_required
    def novo_item():
        almoxarifados = Almoxarifado.query.all()
        if request.method == 'POST':
            it = Item(
                nome=request.form['nome'],
                codigo=request.form['codigo'],
                unidade=request.form['unidade'],
                quantidade=float(request.form.get('quantidade', 0)),
                estoque_minimo=float(request.form.get('estoque_minimo', 0)),
                almoxarifado_id=int(request.form['almoxarifado_id']),
                categoria=request.form.get('categoria', 'geral'),
                ca=request.form.get('ca', '').strip() or None
            )
            db.session.add(it)
            db.session.commit()
            flash(f'Item "{it.nome}" cadastrado!', 'success')
            return redirect(url_for('almoxarifado', id=it.almoxarifado_id))
        return render_template('form_item.html', item=None, almoxarifados=almoxarifados)

    @app.route('/item/<int:id>/editar', methods=['GET', 'POST'])
    @login_required
    def editar_item(id):
        it = Item.query.get_or_404(id)
        u = usuario_atual()
        if u.perfil in ('mestre', 'tecnico_seguranca') or (u.perfil != 'admin' and it.almoxarifado_id not in u.almoxarifados_permitidos()):
            flash('Acesso negado.', 'danger')
            return redirect(url_for('item', id=id))
        almoxarifados = Almoxarifado.query.all()
        if request.method == 'POST':
            it.nome = request.form['nome']
            it.codigo = request.form['codigo']
            it.unidade = request.form['unidade']
            it.estoque_minimo = float(request.form.get('estoque_minimo', 0))
            it.almoxarifado_id = int(request.form['almoxarifado_id'])
            it.categoria = request.form.get('categoria', 'geral')
            it.ca = request.form.get('ca', '').strip() or None
            qtd_str = request.form.get('quantidade')
            if qtd_str is not None and qtd_str != '':
                try:
                    nova_qtd = float(qtd_str)
                    if nova_qtd != it.quantidade:
                        diff = nova_qtd - it.quantidade
                        tipo = 'entrada' if diff > 0 else 'saida'
                        db.session.add(Movimentacao(
                            tipo=tipo,
                            quantidade=abs(diff),
                            responsavel=u.nome if u else 'Sistema',
                            observacao=f'Ajuste manual: {it.quantidade} → {nova_qtd} {it.unidade}',
                            item_id=it.id
                        ))
                        it.quantidade = nova_qtd
                except ValueError:
                    pass
            db.session.commit()
            flash('Item atualizado!', 'success')
            return redirect(url_for('almoxarifado', id=it.almoxarifado_id))
        return render_template('form_item.html', item=it, almoxarifados=almoxarifados)

    @app.route('/item/<int:id>/deletar', methods=['POST'])
    @login_required
    def deletar_item(id):
        it = Item.query.get_or_404(id)
        u = usuario_atual()
        if u.perfil != 'admin':
            flash('Apenas administradores podem deletar itens.', 'danger')
            return redirect(url_for('item', id=id))
        alm_id = it.almoxarifado_id
        db.session.delete(it)
        db.session.commit()
        flash('Item removido!', 'warning')
        return redirect(url_for('almoxarifado', id=alm_id))

    @app.route('/movimentacao/lote', methods=['GET', 'POST'])
    @login_required
    def movimentacao_lote():
        u = usuario_atual()
        almoxarifados = Almoxarifado.query.all() if u.perfil == 'admin' else \
            Almoxarifado.query.filter(Almoxarifado.id.in_(u.almoxarifados_permitidos())).all()
        if u.perfil == 'admin':
            historico = Movimentacao.query.order_by(Movimentacao.data.desc()).limit(20).all()
            requisicoes_hist = Requisicao.query.order_by(Requisicao.data_retirada.desc()).limit(20).all()
        else:
            ids = u.almoxarifados_permitidos()
            historico = (Movimentacao.query.join(Item)
                         .filter(Item.almoxarifado_id.in_(ids))
                         .order_by(Movimentacao.data.desc()).limit(20).all())
            requisicoes_hist = (Requisicao.query.join(Item)
                                .filter(Item.almoxarifado_id.in_(ids))
                                .order_by(Requisicao.data_retirada.desc()).limit(20).all())
        itens_json = {}
        for alm in almoxarifados:
            itens_json[str(alm.id)] = [
                {'id': it.id, 'nome': it.nome, 'quantidade': it.quantidade,
                 'unidade': it.unidade, 'categoria': it.categoria or 'geral',
                 'ca': it.ca or ''}
                for it in alm.itens
            ]
        if request.method == 'POST':
            alm_id = int(request.form['almoxarifado_id'])
            if u.perfil != 'admin' and alm_id not in u.almoxarifados_permitidos():
                flash('Acesso negado.', 'danger')
                return redirect(url_for('movimentacao_lote'))
            tipo = request.form['tipo']
            responsavel = request.form.get('responsavel', '')
            observacao = request.form.get('observacao', '')
            erros = []
            movs = []
            indices = set()
            for key in request.form.keys():
                if key.startswith('item_id_'):
                    try:
                        indices.add(int(key.split('_')[-1]))
                    except ValueError:
                        pass
            for i in sorted(indices):
                item_id = request.form.get(f'item_id_{i}')
                qtd_str = request.form.get(f'quantidade_{i}')
                colab = request.form.get(f'colaborador_{i}', '').strip()
                resp_linha = request.form.get(f'responsavel_{i}', '').strip() or responsavel
                if not item_id or not qtd_str:
                    continue
                it = db.session.get(Item, item_id)
                try:
                    qtd = float(qtd_str)
                except ValueError:
                    continue
                if not it or qtd <= 0:
                    continue
                if tipo == 'saida' and qtd > it.quantidade:
                    erros.append(f'"{it.nome}": estoque insuficiente ({it.quantidade} {it.unidade})')
                    continue
                it.quantidade = round(it.quantidade + qtd if tipo == 'entrada' else it.quantidade - qtd, 4)
                if tipo == 'saida' and colab:
                    obs_linha = f'liberado P/ {colab}'
                    if observacao:
                        obs_linha += f' | {observacao}'
                else:
                    obs_linha = observacao
                movs.append(Movimentacao(
                    tipo=tipo, quantidade=qtd,
                    responsavel=resp_linha,
                    observacao=obs_linha,
                    item_id=it.id
                ))
            if movs:
                db.session.add_all(movs)
                db.session.commit()
                tipo_label = '📥 Entrada' if request.form['tipo'] == 'entrada' else '📤 Saída'
                alm = db.session.get(Almoxarifado, alm_id)
                flash_html(
                    f'<strong>{escape(tipo_label)} registrada!</strong> '
                    f'{len(movs)} item(ns) movimentado(s) em <strong>{escape(alm.nome if alm else "")}</strong>. '
                    f'<a href="/almoxarifado/{alm_id}" class="alert-link">Ver Almoxarifado</a>',
                    'success'
                )
            elif not erros:
                flash('Adicione pelo menos um item antes de confirmar.', 'warning')
            for e in erros:
                flash_html(
                    f'<strong>Estoque insuficiente:</strong> {escape(e)} '
                    f'<a href="/almoxarifado/{alm_id}" class="alert-link">Consultar Estoque</a>',
                    'danger'
                )
            return redirect(url_for('movimentacao_lote'))
        return render_template('movimentacao_lote.html',
                               almoxarifados=almoxarifados,
                               itens_json=json.dumps(itens_json),
                               historico=historico,
                               requisicoes=requisicoes_hist)

    @app.route('/item/<int:id>/movimentar', methods=['POST'])
    @login_required
    def movimentar(id):
        it = Item.query.get_or_404(id)
        u = usuario_atual()
        if u.perfil != 'admin' and it.almoxarifado_id not in u.almoxarifados_permitidos():
            flash('Acesso negado.', 'danger')
            return redirect(url_for('item', id=id))
        tipo = request.form['tipo']
        qtd = float(request.form['quantidade'])
        responsavel = request.form.get('responsavel', '').strip()
        observacao = request.form.get('observacao', '').strip()
        if tipo == 'saida' and qtd > it.quantidade:
            flash('Quantidade insuficiente em estoque!', 'danger')
            return redirect(url_for('item', id=id))
        it.quantidade = round(it.quantidade + qtd if tipo == 'entrada' else it.quantidade - qtd, 4)
        mov = Movimentacao(
            tipo=tipo, quantidade=qtd,
            responsavel=responsavel,
            observacao=observacao,
            item_id=id
        )
        db.session.add(mov)
        db.session.commit()
        flash(f'{'Entrada' if tipo == 'entrada' else 'Saída'} de {qtd} {it.unidade} registrada!', 'success')
        return redirect(url_for('item', id=id))

    @app.route('/requisicoes')
    @login_required
    def requisicoes():
        u = usuario_atual()
        colaborador = request.args.get('colaborador', '')
        status = request.args.get('status', '')
        data_ini = request.args.get('data_ini', '')
        data_fim = request.args.get('data_fim', '')
        q = Requisicao.query
        if u.perfil != 'admin':
            ids = u.almoxarifados_permitidos()
            q = q.join(Item).filter(Item.almoxarifado_id.in_(ids)) if ids else q.filter(False)
        if colaborador:
            q = q.filter(Requisicao.colaborador.ilike(f'%{colaborador}%'))
        if status:
            q = q.filter(Requisicao.status == status)
        if data_ini:
            q = q.filter(Requisicao.data_retirada >= data_ini)
        if data_fim:
            q = q.filter(Requisicao.data_retirada <= data_fim + ' 23:59:59')
        reqs = q.order_by(Requisicao.data_retirada.desc()).all()
        return render_template('requisicoes.html',
            requisicoes=reqs,
            total=len(reqs),
            em_uso=sum(1 for r in reqs if r.status == 'aberta'),
            devolvidos=sum(1 for r in reqs if r.status == 'devolvida'),
            filtro_colaborador=colaborador,
            filtro_status=status,
            filtro_data_ini=data_ini,
            filtro_data_fim=data_fim
        )

    @app.route('/requisicoes/nova', methods=['GET', 'POST'])
    @login_required
    def requisicao_nova():
        u = usuario_atual()
        almoxarifados = Almoxarifado.query.all() if u.perfil == 'admin' else \
            Almoxarifado.query.filter(Almoxarifado.id.in_(u.almoxarifados_permitidos())).all()
        itens_json = {}
        for alm in almoxarifados:
            itens_json[str(alm.id)] = [
                {'id': it.id, 'nome': it.nome, 'quantidade': it.quantidade, 'unidade': it.unidade}
                for it in alm.itens
            ]
        if request.method == 'POST':
            colaborador = request.form.get('colaborador', '')
            observacao = request.form.get('observacao', '')
            criados = 0
            indices = set()
            for key in request.form.keys():
                if key.startswith('item_id_'):
                    try:
                        indices.add(int(key.split('_')[-1]))
                    except ValueError:
                        pass
            for i in sorted(indices):
                item_id = request.form.get(f'item_id_{i}')
                qtd_str = request.form.get(f'quantidade_{i}')
                if not item_id or not qtd_str:
                    continue
                it = db.session.get(Item, item_id)
                try:
                    qtd = float(qtd_str)
                except ValueError:
                    continue
                if not it or qtd <= 0:
                    continue
                if u.perfil != 'admin' and it.almoxarifado_id not in u.almoxarifados_permitidos():
                    continue
                if qtd > it.quantidade:
                    flash_html(
                        f'<strong>Estoque insuficiente:</strong> "{escape(it.nome)}" tem apenas '
                        f'<strong>{it.quantidade} {escape(it.unidade)}</strong> disponível. '
                        f'<a href="/almoxarifado/{it.almoxarifado_id}" class="alert-link">Consultar Estoque</a>',
                        'danger'
                    )
                    continue
                it.quantidade -= qtd
                db.session.add(Requisicao(
                    colaborador=colaborador,
                    observacao=observacao,
                    quantidade=qtd,
                    item_id=it.id
                ))
                db.session.add(Movimentacao(
                    tipo='saida', quantidade=qtd,
                    responsavel=colaborador,
                    observacao=f'Requisicao — {observacao}',
                    item_id=it.id
                ))
                criados += 1
            if criados:
                db.session.commit()
                flash_html(
                    f'<strong>✅ Requisição registrada!</strong> '
                    f'{criados} item(ns) retirado(s) com sucesso para <strong>{escape(colaborador)}</strong>. '
                    f'<a href="/requisicoes" class="alert-link">Ver Requisições</a>',
                    'success'
                )
            elif not any(True for key in request.form.keys() if key.startswith('item_id_')):
                flash('Adicione pelo menos um item antes de registrar.', 'warning')
            return redirect(url_for('requisicoes'))
        return render_template('requisicao_nova.html',
                               almoxarifados=almoxarifados,
                               itens_json=json.dumps(itens_json))

    @app.route('/requisicoes/<int:id>/devolver', methods=['POST'])
    @login_required
    def devolver_requisicao(id):
        req = Requisicao.query.get_or_404(id)
        u = usuario_atual()
        if u.perfil != 'admin' and req.item.almoxarifado_id not in u.almoxarifados_permitidos():
            flash('Acesso negado.', 'danger')
            return redirect(url_for('requisicoes'))
        if req.status == 'aberta':
            req.status = 'devolvida'
            req.data_devolucao = agora()
            req.item.quantidade += req.quantidade
            db.session.add(Movimentacao(
                tipo='entrada', quantidade=req.quantidade,
                responsavel=req.colaborador,
                observacao=f'Devolução de requisição #{req.id}',
                item_id=req.item_id
            ))
            db.session.commit()
            flash(f'Devolução de "{req.item.nome}" registrada!', 'success')
        return redirect(url_for('requisicoes'))

    @app.route('/almoxarifado/<int:id>/importar', methods=['GET', 'POST'])
    @login_required
    def importar_itens(id):
        u = usuario_atual()
        alm = Almoxarifado.query.get_or_404(id)
        if u.perfil != 'admin' and id not in u.almoxarifados_permitidos():
            flash('Acesso negado.', 'danger')
            return redirect(url_for('index'))
        if request.method == 'POST':
            arquivo = request.files.get('arquivo')
            modo = request.form.get('modo', 'adicionar')
            if not arquivo or not arquivo.filename.endswith(('.xlsx', '.xls')):
                flash('Envie um arquivo Excel (.xlsx ou .xls).', 'danger')
                return redirect(url_for('importar_itens', id=id))
            try:
                wb = openpyxl.load_workbook(arquivo, data_only=True)
                ws = wb.active
                inseridos = 0
                atualizados = 0
                erros = []
                for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                    if not any(row):
                        continue
                    try:
                        codigo = str(row[0]).strip() if row[0] else None
                        nome = str(row[1]).strip() if row[1] else None
                        unidade = str(row[2]).strip() if row[2] else 'un'
                        quantidade = float(row[3]) if row[3] is not None else 0
                        est_min = float(row[4]) if row[4] is not None else 0
                    except Exception:
                        erros.append(f'Linha {row_num}: formato inválido')
                        continue
                    if not codigo or not nome:
                        erros.append(f'Linha {row_num}: código ou nome vazio')
                        continue
                    item_existente = Item.query.filter_by(codigo=codigo).first()
                    if item_existente:
                        if modo in ('atualizar', 'substituir'):
                            item_existente.nome = nome
                            item_existente.unidade = unidade
                            item_existente.estoque_minimo = est_min
                            item_existente.almoxarifado_id = id
                            if modo == 'substituir':
                                tem_movimentacoes = Movimentacao.query.filter_by(item_id=item_existente.id).count() > 0
                                if not tem_movimentacoes:
                                    item_existente.quantidade = quantidade
                            else:
                                item_existente.quantidade += quantidade
                            atualizados += 1
                    else:
                        novo = Item(
                            codigo=codigo, nome=nome, unidade=unidade,
                            quantidade=quantidade, estoque_minimo=est_min,
                            almoxarifado_id=id
                        )
                        db.session.add(novo)
                        inseridos += 1
                db.session.commit()
                msg = f'Importação concluída: {inseridos} inseridos, {atualizados} atualizados.'
                if erros:
                    msg += f' {len(erros)} linha(s) com erro.'
                flash(msg, 'success' if not erros else 'warning')
                for e in erros[:10]:
                    flash(e, 'danger')
            except Exception as e:
                flash(f'Erro ao processar arquivo: {str(e)}', 'danger')
            return redirect(url_for('almoxarifado', id=id))
        return render_template('importar_itens.html', almoxarifado=alm)

    @app.route('/almoxarifado/<int:id>/modelo_excel')
    @login_required
    def modelo_excel(id):
        alm = Almoxarifado.query.get_or_404(id)
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Itens'
        h_fill = PatternFill('solid', fgColor='1A3A5C')
        h_font = Font(bold=True, color='FFFFFF', size=11)
        borda = Border(left=Side(style='thin'), right=Side(style='thin'),
                        top=Side(style='thin'), bottom=Side(style='thin'))
        for col, h in enumerate(['Codigo', 'Nome', 'Unidade', 'Quantidade', 'Estoque Minimo'], 1):
            c = ws.cell(row=1, column=col, value=h)
            c.font = h_font; c.fill = h_fill
            c.alignment = Alignment(horizontal='center'); c.border = borda
        for r, ex in enumerate([('CIM-001','Cimento CP-II','sc',500,100),
                                 ('ARG-002','Areia Grossa','m3',30,5),
                                 ('EPI-003','Capacete','un',20,5)], 2):
            for c, v in enumerate(ex, 1):
                ws.cell(row=r, column=c, value=v).border = borda
        for i, w in enumerate([14,40,10,14,16], 1):
            ws.column_dimensions[get_column_letter(i)].width = w
        buf = io.BytesIO()
        wb.save(buf); buf.seek(0)
        return send_file(buf, as_attachment=True,
                         download_name=f'modelo_{alm.nome.replace(" ", "_")}.xlsx',
                         mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

    @app.route('/relatorios/consumo')
    @login_required
    def relatorio_consumo():
        u = usuario_atual()
        alm_id = request.args.get('almoxarifado_id', type=int)
        data_ini = request.args.get('data_ini', str(date.today().replace(day=1)))
        data_fim = request.args.get('data_fim', str(date.today()))
        if alm_id and u.perfil != 'admin' and alm_id not in u.almoxarifados_permitidos():
            flash('Acesso negado.', 'danger')
            return redirect(url_for('index'))
        query = Movimentacao.query.filter(
            Movimentacao.tipo == 'saida',
            Movimentacao.data >= data_ini,
            Movimentacao.data <= data_fim + ' 23:59:59'
        )
        if alm_id:
            query = query.join(Item).filter(Item.almoxarifado_id == alm_id)
        elif u.perfil != 'admin':
            query = query.join(Item).filter(Item.almoxarifado_id.in_(u.almoxarifados_permitidos()))
        movs = query.order_by(Movimentacao.data.desc()).all()
        almoxarifados = Almoxarifado.query.all()
        return render_template('relatorio_consumo.html', movimentacoes=movs,
                               almoxarifados=almoxarifados, data_ini=data_ini,
                               data_fim=data_fim, alm_id=alm_id)

    @app.route('/relatorios/consumo/exportar')
    @login_required
    def exportar_consumo():
        u = usuario_atual()
        alm_id = request.args.get('almoxarifado_id', type=int)
        data_ini = request.args.get('data_ini', str(date.today().replace(day=1)))
        data_fim = request.args.get('data_fim', str(date.today()))
        if alm_id and u.perfil != 'admin' and alm_id not in u.almoxarifados_permitidos():
            flash('Acesso negado.', 'danger')
            return redirect(url_for('index'))
        query = Movimentacao.query.filter(
            Movimentacao.tipo == 'saida',
            Movimentacao.data >= data_ini,
            Movimentacao.data <= data_fim + ' 23:59:59'
        )
        if alm_id:
            query = query.join(Item).filter(Item.almoxarifado_id == alm_id)
        elif u.perfil != 'admin':
            query = query.join(Item).filter(Item.almoxarifado_id.in_(u.almoxarifados_permitidos()))
        movs = query.order_by(Movimentacao.data.desc()).all()
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Consumo'
        h_fill = PatternFill('solid', fgColor='1A3A5C')
        h_font = Font(bold=True, color='FFFFFF', size=11)
        borda = Border(left=Side(style='thin'), right=Side(style='thin'),
                       top=Side(style='thin'), bottom=Side(style='thin'))
        alm_nome = db.session.get(Almoxarifado, alm_id).nome if alm_id else 'Todos os Almoxarifados'
        ws.merge_cells('A1:G1')
        ws['A1'] = f'Relatório de Consumo — {alm_nome}'
        ws['A1'].font = Font(bold=True, size=13, color='1A3A5C')
        ws['A1'].alignment = Alignment(horizontal='center')
        ws.merge_cells('A2:G2')
        ws['A2'] = f'Período: {data_ini} a {data_fim}   |   Gerado em: {agora().strftime("%d/%m/%Y %H:%M")}'
        ws['A2'].font = Font(italic=True, size=10, color='666666')
        ws['A2'].alignment = Alignment(horizontal='center')
        headers = ['Data', 'Código', 'Item', 'Almoxarifado', 'Quantidade', 'Responsável', 'Observação']
        for col, h in enumerate(headers, 1):
            c = ws.cell(row=4, column=col, value=h)
            c.font = h_font; c.fill = h_fill
            c.alignment = Alignment(horizontal='center'); c.border = borda
        for row_num, mov in enumerate(movs, 5):
            dados = [
                mov.data.strftime('%d/%m/%Y %H:%M'),
                mov.item.codigo, mov.item.nome,
                mov.item.almoxarifado.nome,
                f'{mov.quantidade} {mov.item.unidade}',
                mov.responsavel or '', mov.observacao or ''
            ]
            for col, val in enumerate(dados, 1):
                c = ws.cell(row=row_num, column=col, value=val)
                c.border = borda
                if row_num % 2 == 0:
                    c.fill = PatternFill('solid', fgColor='F0F4F8')
        for i, w in enumerate([18, 14, 45, 35, 14, 20, 40], 1):
            ws.column_dimensions[get_column_letter(i)].width = w
        buf = io.BytesIO()
        wb.save(buf); buf.seek(0)
        return send_file(buf, as_attachment=True,
                         download_name=f'consumo_{data_ini}_a_{data_fim}.xlsx',
                         mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

    @app.route('/relatorios/consumo-por-pessoa')
    @login_required
    def relatorio_consumo_pessoa():
        u = usuario_atual()
        alm_id = request.args.get('almoxarifado_id', type=int)
        data_ini = request.args.get('data_ini', str(date.today().replace(day=1)))
        data_fim = request.args.get('data_fim', str(date.today()))
        responsavel_filtro = request.args.get('responsavel', '').strip()
        if alm_id and u.perfil != 'admin' and alm_id not in u.almoxarifados_permitidos():
            flash('Acesso negado.', 'danger')
            return redirect(url_for('index'))
        query = Movimentacao.query.filter(
            Movimentacao.tipo == 'saida',
            Movimentacao.data >= data_ini,
            Movimentacao.data <= data_fim + ' 23:59:59'
        )
        if alm_id:
            query = query.join(Item).filter(Item.almoxarifado_id == alm_id)
        elif u.perfil != 'admin':
            query = query.join(Item).filter(Item.almoxarifado_id.in_(u.almoxarifados_permitidos()))
        movs = query.order_by(Movimentacao.data.desc()).all()
        por_pessoa = defaultdict(list)
        for mov in movs:
            colaborador = extrair_colaborador(mov)
            if responsavel_filtro and responsavel_filtro.lower() not in colaborador.lower():
                continue
            por_pessoa[colaborador].append(mov)
        resumo = []
        for nome, lista in sorted(por_pessoa.items()):
            resumo.append({
                'nome': nome,
                'movimentacoes': lista,
                'total_movs': len(lista),
                'itens_distintos': len(set(m.item_id for m in lista)),
            })
        almoxarifados = Almoxarifado.query.all()
        return render_template('relatorio_consumo_pessoa.html',
                               resumo=resumo,
                               almoxarifados=almoxarifados,
                               data_ini=data_ini,
                               data_fim=data_fim,
                               alm_id=alm_id,
                               responsavel_filtro=responsavel_filtro,
                               total_geral=sum(p['total_movs'] for p in resumo))

    @app.route('/relatorios/consumo-por-pessoa/exportar')
    @login_required
    def exportar_consumo_pessoa():
        u = usuario_atual()
        alm_id = request.args.get('almoxarifado_id', type=int)
        data_ini = request.args.get('data_ini', str(date.today().replace(day=1)))
        data_fim = request.args.get('data_fim', str(date.today()))
        responsavel_filtro = request.args.get('responsavel', '').strip()
        if alm_id and u.perfil != 'admin' and alm_id not in u.almoxarifados_permitidos():
            flash('Acesso negado.', 'danger')
            return redirect(url_for('index'))
        query = Movimentacao.query.filter(
            Movimentacao.tipo == 'saida',
            Movimentacao.data >= data_ini,
            Movimentacao.data <= data_fim + ' 23:59:59'
        )
        if alm_id:
            query = query.join(Item).filter(Item.almoxarifado_id == alm_id)
        elif u.perfil != 'admin':
            query = query.join(Item).filter(Item.almoxarifado_id.in_(u.almoxarifados_permitidos()))
        movs = query.order_by(Movimentacao.data.asc()).all()
        por_pessoa = defaultdict(list)
        for mov in movs:
            colab = extrair_colaborador(mov)
            if responsavel_filtro and responsavel_filtro.lower() not in colab.lower():
                continue
            por_pessoa[colab].append(mov)
        h_fill = PatternFill('solid', fgColor='1A3A5C')
        h_font = Font(bold=True, color='FFFFFF', size=11)
        z_fill = PatternFill('solid', fgColor='F0F4F8')
        borda = Border(left=Side(style='thin'), right=Side(style='thin'),
                       top=Side(style='thin'), bottom=Side(style='thin'))
        centro = Alignment(horizontal='center', vertical='center')
        esq = Alignment(horizontal='left', vertical='center')
        alm_nome = db.session.get(Almoxarifado, alm_id).nome if alm_id else 'Todos'
        total_geral = sum(len(v) for v in por_pessoa.values())
        wb = openpyxl.Workbook()
        ws1 = wb.active
        ws1.title = 'Resumo'
        for col, w in enumerate([6, 40, 20, 18, 18], 1):
            ws1.column_dimensions[get_column_letter(col)].width = w
        ws1.merge_cells('A1:E1')
        ws1['A1'].value = f'Consumo por Pessoa — {alm_nome}'
        ws1['A1'].font = Font(bold=True, size=13, color='1A3A5C')
        ws1['A1'].fill = PatternFill('solid', fgColor='E8F0F7')
        ws1['A1'].alignment = centro
        ws1.merge_cells('A2:E2')
        ws1['A2'].value = f'Período: {data_ini} a {data_fim}   |   Gerado em: {agora().strftime("%d/%m/%Y %H:%M")}'
        ws1['A2'].font = Font(italic=True, size=10, color='666666')
        ws1['A2'].alignment = centro
        for col, h in enumerate(['#', 'Funcionário', 'Total Retiradas', 'Itens Distintos', 'Participação (%)'], 1):
            c = ws1.cell(row=4, column=col, value=h)
            c.font = h_font; c.fill = h_fill; c.alignment = centro; c.border = borda
        for i, (nome, lista) in enumerate(sorted(por_pessoa.items(), key=lambda x: len(x[1]), reverse=True), 1):
            pct = round(len(lista) / total_geral * 100, 1) if total_geral else 0
            row = i + 4
            for col, val in enumerate([i, nome, len(lista), len(set(m.item_id for m in lista)), f'{pct}%'], 1):
                c = ws1.cell(row=row, column=col, value=val)
                c.border = borda
                c.alignment = esq if col == 2 else centro
                if row % 2 == 0: c.fill = z_fill
        r = len(por_pessoa) + 5
        for col, val in enumerate(['', f'{len(por_pessoa)} pessoa(s)', total_geral, '', '100%'], 1):
            c = ws1.cell(row=r, column=col, value=val)
            c.font = Font(bold=True)
            c.fill = PatternFill('solid', fgColor='D0E4F7')
            c.border = borda; c.alignment = centro
        ws2 = wb.create_sheet('Detalhes')
        for col, w in enumerate([35, 18, 14, 45, 14, 30, 25], 1):
            ws2.column_dimensions[get_column_letter(col)].width = w
        ws2.merge_cells('A1:G1')
        ws2['A1'].value = f'Detalhes por Funcionário — {alm_nome}'
        ws2['A1'].font = Font(bold=True, size=13, color='1A3A5C')
        ws2['A1'].fill = PatternFill('solid', fgColor='E8F0F7')
        ws2['A1'].alignment = centro
        ws2.merge_cells('A2:G2')
        ws2['A2'].value = f'Período: {data_ini} a {data_fim}   |   Gerado em: {agora().strftime("%d/%m/%Y %H:%M")}'
        ws2['A2'].font = Font(italic=True, size=10, color='666666')
        ws2['A2'].alignment = centro
        for col, h in enumerate(['Funcionário', 'Data', 'Código', 'Item', 'Quantidade', 'Almoxarifado', 'Liberado por'], 1):
            c = ws2.cell(row=4, column=col, value=h)
            c.font = h_font; c.fill = h_fill; c.alignment = centro; c.border = borda
        row_num = 5
        for nome, lista in sorted(por_pessoa.items()):
            ws2.merge_cells(f'A{row_num}:G{row_num}')
            c = ws2.cell(row=row_num, column=1, value=f'👷 {nome}  ({len(lista)} retirada(s))')
            c.font = Font(bold=True, color='FFFFFF', size=11)
            c.fill = PatternFill('solid', fgColor='2E6DA4')
            c.alignment = esq; c.border = borda
            row_num += 1
            for mov in sorted(lista, key=lambda m: m.data):
                for col, val in enumerate([nome, mov.data.strftime('%d/%m/%Y %H:%M'),
                                            mov.item.codigo, mov.item.nome,
                                            f'{mov.quantidade} {mov.item.unidade}',
                                            mov.item.almoxarifado.nome,
                                            mov.responsavel or '—'], 1):
                    c = ws2.cell(row=row_num, column=col, value=val)
                    c.font = Font(size=9); c.border = borda
                    c.alignment = esq if col in [1,3,4] else centro
                    if row_num % 2 == 0: c.fill = z_fill
                row_num += 1
            row_num += 1
        buf = io.BytesIO()
        wb.save(buf); buf.seek(0)
        return send_file(buf, as_attachment=True,
                         download_name=f'consumo_por_pessoa_{data_ini}_a_{data_fim}.xlsx',
                         mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

    @app.route('/relatorios/ficha-epi')
    @login_required
    def ficha_epi():
        u = usuario_atual()
        query = Movimentacao.query.join(Item).filter(
            Movimentacao.tipo == 'saida', Item.categoria == 'epi')
        if u.perfil != 'admin':
            query = query.filter(Item.almoxarifado_id.in_(u.almoxarifados_permitidos()))
        funcionarios = set()
        movs = query.all()
        for mov in movs:
            nome = extrair_colaborador(mov)
            if nome:
                funcionarios.add(nome)
        return render_template('ficha_epi.html',
                               funcionarios=sorted(funcionarios),
                               data_ini='2020-01-01',
                               data_fim=str(date.today()))

    @app.route('/relatorios/ficha-epi/exportar')
    @login_required
    def exportar_ficha_epi():
        funcionario = request.args.get('funcionario', '').strip()
        data_ini = request.args.get('data_ini', '2020-01-01')
        data_fim = request.args.get('data_fim', str(date.today()))
        u = usuario_atual()
        if not funcionario:
            flash('Selecione um funcionário.', 'warning')
            return redirect(url_for('ficha_epi'))
        query = (Movimentacao.query.join(Item)
                     .filter(Movimentacao.tipo == 'saida',
                             Item.categoria == 'epi',
                             Movimentacao.data >= data_ini,
                             Movimentacao.data <= data_fim + ' 23:59:59'))
        if u.perfil != 'admin':
            query = query.filter(Item.almoxarifado_id.in_(u.almoxarifados_permitidos()))
        movs_todas = query.order_by(Movimentacao.data.asc()).all()
        lista = [m for m in movs_todas if extrair_colaborador(m).lower() == funcionario.lower()]
        if not lista:
            flash(f'Nenhuma retirada de EPI encontrada para "{funcionario}" no período.', 'warning')
            return redirect(url_for('ficha_epi'))
        borda = Border(left=Side(style='thin'), right=Side(style='thin'),
                       top=Side(style='thin'), bottom=Side(style='thin'))
        borda_med = Border(left=Side(style='medium'), right=Side(style='medium'),
                           top=Side(style='medium'), bottom=Side(style='medium'))
        centro = Alignment(horizontal='center', vertical='center', wrap_text=True)
        esq = Alignment(horizontal='left', vertical='center', wrap_text=True)
        azul_esc = PatternFill('solid', fgColor='1F3864')
        azul_cla = PatternFill('solid', fgColor='BDD7EE')
        cinza = PatternFill('solid', fgColor='F2F2F2')
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = funcionario[:28]
        for col, w in zip('ABCDEFGH', [8, 40, 10, 14, 26, 14, 26, 18]):
            ws.column_dimensions[col].width = w
        def celula(ws, ref, val='', font=None, fill=None, aln=None, brd=None, height=None):
            c = ws[ref]
            c.value = val
            if font: c.font = font
            if fill: c.fill = fill
            if aln: c.alignment = aln
            if brd: c.border = brd
            return c
        def merge_row(ws, row, col_ini, col_fim, val='', font=None, fill=None, aln=None, height=None):
            ws.merge_cells(f'{col_ini}{row}:{col_fim}{row}')
            c = ws[f'{col_ini}{row}']
            c.value = val
            if font: c.font = font
            if fill: c.fill = fill
            if aln: c.alignment = aln
            for col in range(ord(col_ini)-64, ord(col_fim)-64+1):
                ws.cell(row=row, column=col).border = borda
            return c
        ws.row_dimensions[1].height = 36
        ws.merge_cells('A1:C1')
        ws['A1'].value = 'STANZA'
        ws['A1'].font = Font(bold=True, size=22, color='808080')
        ws['A1'].alignment = centro
        for c in range(1,4): ws.cell(1,c).border = borda_med
        ws.merge_cells('D1:F1')
        ws['D1'].value = 'FICHA DE CONTROLE DE EPI\'S E UNIFORMES'
        ws['D1'].font = Font(bold=True, size=13, color='1F3864')
        ws['D1'].alignment = centro
        ws['D1'].fill = azul_cla
        for c in range(4,7): ws.cell(1,c).border = borda_med
        ws.merge_cells('G1:H1')
        ws['G1'].value = 'FORM.SEG.014'
        ws['G1'].font = Font(bold=True, size=9, color='1F3864')
        ws['G1'].alignment = centro
        ws['G1'].fill = azul_cla
        for c in range(7,9): ws.cell(1,c).border = borda_med
        ws.row_dimensions[2].height = 14
        ws.merge_cells('D2:F2')
        ws['D2'].value = 'Data Elaboração/Revisão: 20/10/2024'
        ws['D2'].font = Font(size=8, italic=True, color='595959')
        ws['D2'].alignment = centro
        ws.merge_cells('G2:H2')
        ws['G2'].value = 'Revisão: 00'
        ws['G2'].font = Font(size=8, italic=True, color='595959')
        ws['G2'].alignment = centro
        for c in range(1,9): ws.cell(2,c).border = borda
        ws.row_dimensions[3].height = 22
        ws.merge_cells('A3:B3')
        ws['A3'].value = f'NOME: {funcionario.upper()}'
        ws['A3'].font = Font(bold=True, size=10)
        ws['A3'].alignment = esq
        ws['A3'].fill = cinza
        ws['C3'].value = f'MATRÍCULA:'
        ws['C3'].font = Font(size=9)
        ws['C3'].alignment = centro
        ws['C3'].fill = cinza
        ws.merge_cells('D3:E3')
        ws['D3'].value = 'FUNÇÃO:'
        ws['D3'].font = Font(size=9)
        ws['D3'].alignment = esq
        ws['D3'].fill = cinza
        ws.merge_cells('F3:G3')
        ws['F3'].value = 'ADMISSÃO:'
        ws['F3'].font = Font(size=9)
        ws['F3'].alignment = esq
        ws['F3'].fill = cinza
        ws['H3'].value = ''
        ws['H3'].fill = cinza
        for c in range(1,9): ws.cell(3,c).border = borda
        ws.row_dimensions[4].height = 20
        for ref, val in [('A4','QUANT'), ('B4','DESCRIÇÃO'), ('C4','C.A.')]:
            ws[ref].value = val
            ws[ref].font = Font(bold=True, color='FFFFFF', size=9)
            ws[ref].fill = azul_esc
            ws[ref].alignment = centro
            ws[ref].border = borda
        ws.merge_cells('D4:E4')
        ws['D4'].value = 'ENTREGA'
        ws['D4'].font = Font(bold=True, color='FFFFFF', size=9)
        ws['D4'].fill = azul_esc
        ws['D4'].alignment = centro
        for c in range(4,6): ws.cell(4,c).border = borda
        ws.merge_cells('F4:G4')
        ws['F4'].value = 'DEVOLUÇÃO'
        ws['F4'].font = Font(bold=True, color='FFFFFF', size=9)
        ws['F4'].fill = azul_esc
        ws['F4'].alignment = centro
        for c in range(6,8): ws.cell(4,c).border = borda
        ws['H4'].value = 'MOTIVO'
        ws['H4'].font = Font(bold=True, color='FFFFFF', size=9)
        ws['H4'].fill = azul_esc
        ws['H4'].alignment = centro
        ws['H4'].border = borda
        ws.row_dimensions[5].height = 16
        for ref, val in [('A5',''), ('B5',''), ('C5',''),
                         ('D5','DATA'), ('E5','ASSINATURA'),
                         ('F5','DATA'), ('G5','ASSINATURA'), ('H5','')]:
            ws[ref].value = val
            ws[ref].font = Font(bold=True, color='FFFFFF', size=8)
            ws[ref].fill = azul_esc
            ws[ref].alignment = centro
            ws[ref].border = borda
        row = 6
        for mov in lista:
            ws.row_dimensions[row].height = 18
            fill_z = PatternFill('solid', fgColor='EBF3FB') if row % 2 == 0 else None
            for col, val in zip('ABCDEFGH', [
                f'{mov.quantidade} {mov.item.unidade}',
                mov.item.nome, mov.item.ca or '',
                mov.data.strftime('%d/%m/%Y'), '',
                '', '', ''
            ]):
                c = ws[f'{col}{row}']
                c.value = val
                c.font = Font(size=9)
                c.alignment = esq if col == 'B' else centro
                c.border = borda
                if fill_z: c.fill = fill_z
            row += 1
        total_linhas = max(14, len(lista) + 4)
        while row <= 5 + total_linhas:
            ws.row_dimensions[row].height = 18
            for col in 'ABCDEFGH':
                ws[f'{col}{row}'].border = borda
                ws[f'{col}{row}'].value = '/    /' if col in ('D','F') else ''
                ws[f'{col}{row}'].font = Font(size=9, color='BFBFBF')
                ws[f'{col}{row}'].alignment = centro
            row += 1
        row += 1
        ws.row_dimensions[row].height = 16
        ws.merge_cells(f'A{row}:H{row}')
        ws[f'A{row}'].value = 'TERMO DE RESPONSABILIDADE'
        ws[f'A{row}'].font = Font(bold=True, size=10, color='1F3864')
        ws[f'A{row}'].alignment = centro
        ws[f'A{row}'].fill = azul_cla
        for c in range(1,9): ws.cell(row,c).border = borda
        row += 1
        ws.row_dimensions[row].height = 70
        ws.merge_cells(f'A{row}:H{row}')
        ws[f'A{row}'].value = (
            'Pelo presente declaro que recebi da empresa STANZA INCORPORAÇÃO E CONSTRUÇÃO LTDA., os materiais '
            'relacionados nesta ficha, assumindo o compromisso nos termos das letras "a" e "b" do ítem 1.8 da NR 1 '
            'e letras "a","b"e "c" do ítem 6.7.1 da NR 6, de usá-los em atividades ligadas ao trabalho, zelar pela '
            'sua guarda, conservação e devolvê-lo ao setor competente quando se tornar impróprio para uso ou por '
            'motivo de demissão ou afastamento.\n'
            'Em caso de perda, extravio e inutilização proposital do material recebido, autorizo a empresa, na forma '
            'prevista no parágrafo primeiro do art. 462 da CLT - Consolidação das leis do trabalho. A descontar de '
            'meu salário, inclusive no que me couber a título de indenização por rescisão de contrato de trabalho, '
            'a importância correspondente ao valor do material.'
        )
        ws[f'A{row}'].font = Font(size=8)
        ws[f'A{row}'].alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
        for c in range(1,9): ws.cell(row,c).border = borda
        row += 2
        ws.row_dimensions[row].height = 28
        ws.merge_cells(f'A{row}:C{row}')
        ws[f'A{row}'].value = 'Data:        /        /'
        ws[f'A{row}'].font = Font(size=10)
        ws[f'A{row}'].alignment = esq
        for c in range(1,4): ws.cell(row,c).border = borda
        ws.merge_cells(f'D{row}:H{row}')
        ws[f'D{row}'].value = 'EMPREGADO'
        ws[f'D{row}'].font = Font(bold=True, size=10, color='1F3864')
        ws[f'D{row}'].alignment = centro
        ws[f'D{row}'].fill = azul_cla
        for c in range(4,9): ws.cell(row,c).border = borda
        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        nome_safe = funcionario.replace(' ', '_').replace('/', '-')
        return send_file(buf, as_attachment=True,
                         download_name=f'FORM-SEG-014_{nome_safe}_{data_fim}.xlsx',
                         mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

    @app.route('/movimentacoes/excluir', methods=['POST'])
    @admin_required
    def excluir_movimentacoes():
        ids = request.form.getlist('mov_ids')
        if not ids:
            flash('Nenhuma movimentação selecionada.', 'warning')
            return redirect(request.referrer or url_for('relatorio_consumo_pessoa'))
        excluidas = 0
        for mov_id in ids:
            mov = db.session.get(Movimentacao, mov_id)
            if mov:
                if mov.tipo == 'saida':
                    mov.item.quantidade += mov.quantidade
                elif mov.tipo == 'entrada':
                    mov.item.quantidade -= mov.quantidade
                db.session.delete(mov)
                excluidas += 1
        db.session.commit()
        flash(f'{excluidas} movimentação(ões) excluída(s) e estoque revertido.', 'success')
        return redirect(request.referrer or url_for('relatorio_consumo_pessoa'))

    @app.route('/relatorios/alertas')
    @login_required
    def relatorio_alertas():
        u = usuario_atual()
        if u.perfil == 'admin':
            itens = Item.query.filter(Item.quantidade <= Item.estoque_minimo, Item.ativo == True).order_by(
                Item.fixado.desc(), Item.quantidade.asc()
            ).all()
        else:
            ids = u.almoxarifados_permitidos()
            itens = Item.query.filter(
                Item.quantidade <= Item.estoque_minimo,
                Item.almoxarifado_id.in_(ids),
                Item.ativo == True
            ).order_by(Item.fixado.desc(), Item.quantidade.asc()).all() if ids else []
        return render_template('relatorio_alertas.html', itens=itens)

    @app.route('/item/<int:id>/status_compra', methods=['POST'])
    @login_required
    def atualizar_status_compra(id):
        it = Item.query.get_or_404(id)
        u = usuario_atual()
        if u.perfil != 'admin' and it.almoxarifado_id not in u.almoxarifados_permitidos():
            return ('', 403)
        it.status_compra = request.form.get('status_compra', 'pendente')
        db.session.commit()
        return ('', 204)

    @app.route('/item/<int:id>/fixar', methods=['POST'])
    @login_required
    def fixar_item(id):
        it = Item.query.get_or_404(id)
        u = usuario_atual()
        if u.perfil != 'admin' and it.almoxarifado_id not in u.almoxarifados_permitidos():
            return jsonify({'error': 'Acesso negado.'}), 403
        it.fixado = not it.fixado
        db.session.commit()
        return jsonify({'fixado': it.fixado})

    @app.route('/item/<int:id>/desativar', methods=['POST'])
    @login_required
    def desativar_item(id):
        it = Item.query.get_or_404(id)
        u = usuario_atual()
        if u.perfil != 'admin' and it.almoxarifado_id not in u.almoxarifados_permitidos():
            flash('Acesso negado.', 'danger')
            return redirect(url_for('item', id=id))
        if it.quantidade > 0:
            mov = Movimentacao(
                tipo='saida',
                quantidade=it.quantidade,
                responsavel=u.nome,
                observacao=f'Item desativado - saldo restante: {it.quantidade} {it.unidade}',
                item_id=id
            )
            db.session.add(mov)
            it.quantidade = 0
        it.ativo = False
        db.session.commit()
        flash(f'Item "{it.nome}" desativado com sucesso!', 'warning')
        return redirect(url_for('item', id=id))

    @app.route('/item/<int:id>/reativar', methods=['POST'])
    @login_required
    def reativar_item(id):
        it = Item.query.get_or_404(id)
        u = usuario_atual()
        if u.perfil != 'admin' and it.almoxarifado_id not in u.almoxarifados_permitidos():
            flash('Acesso negado.', 'danger')
            return redirect(url_for('item', id=id))
        it.ativo = True
        db.session.commit()
        flash(f'Item "{it.nome}" reativado com sucesso!', 'success')
        return redirect(url_for('item', id=id))

    @app.route('/almoxarifado/<int:id>/exportar')
    @login_required
    def exportar_almoxarifado(id):
        u = usuario_atual()
        if u.perfil != 'admin' and id not in u.almoxarifados_permitidos():
            flash('Acesso negado.', 'danger')
            return redirect(url_for('index'))
        alm = Almoxarifado.query.get_or_404(id)
        itens = Item.query.filter_by(almoxarifado_id=id).all()
        wb = openpyxl.Workbook()
        h_fill = PatternFill('solid', fgColor='1A3A5C')
        h_font = Font(bold=True, color='FFFFFF', size=11)
        ok_fill = PatternFill('solid', fgColor='D4EDDA')
        al_fill = PatternFill('solid', fgColor='FFF3CD')
        cr_fill = PatternFill('solid', fgColor='F8D7DA')
        en_fill = PatternFill('solid', fgColor='D4EDDA')
        sa_fill = PatternFill('solid', fgColor='F8D7DA')
        borda = Border(left=Side(style='thin'), right=Side(style='thin'),
                       top=Side(style='thin'), bottom=Side(style='thin'))
        def titulo(ws, texto, cols):
            ws.merge_cells(f'A1:{get_column_letter(cols)}1')
            ws['A1'] = texto
            ws['A1'].font = Font(bold=True, size=13, color='1A3A5C')
            ws['A1'].alignment = Alignment(horizontal='center')
            ws.merge_cells(f'A2:{get_column_letter(cols)}2')
            ws['A2'] = f'Gerado em: {datetime.now().strftime("%d/%m/%Y %H:%M")}'
            ws['A2'].font = Font(italic=True, color='888888')
            ws['A2'].alignment = Alignment(horizontal='center')
        def cabecalho(ws, row, headers):
            for col, h in enumerate(headers, 1):
                c = ws.cell(row=row, column=col, value=h)
                c.font = h_font; c.fill = h_fill
                c.alignment = Alignment(horizontal='center'); c.border = borda
        ws1 = wb.active
        ws1.title = 'Estoque Atual'
        titulo(ws1, f'Estoque Atual — {alm.nome}', 7)
        cabecalho(ws1, 4, ['Codigo', 'Item', 'Unidade', 'Qtd. Atual', 'Est. Minimo', 'Deficit', 'Status'])
        for r, it in enumerate(itens, 5):
            deficit = max(0, it.estoque_minimo - it.quantidade)
            status = 'ZERADO' if it.status == 'critico' else ('ABAIXO DO MINIMO' if it.status == 'alerta' else 'OK')
            fill = cr_fill if it.status == 'critico' else (al_fill if it.status == 'alerta' else ok_fill)
            for c, v in enumerate([it.codigo, it.nome, it.unidade, it.quantidade, it.estoque_minimo, deficit, status], 1):
                cell = ws1.cell(row=r, column=c, value=v)
                cell.fill = fill; cell.border = borda
                cell.alignment = Alignment(horizontal='left' if c == 2 else 'center')
        for i, w in enumerate([14, 35, 10, 12, 14, 12, 20], 1):
            ws1.column_dimensions[get_column_letter(i)].width = w
        ws2 = wb.create_sheet('Movimentacoes')
        titulo(ws2, f'Historico de Movimentacoes — {alm.nome}', 6)
        cabecalho(ws2, 4, ['Data', 'Item', 'Tipo', 'Quantidade', 'Responsavel', 'Observacao'])
        movs = (Movimentacao.query.join(Item)
                .filter(Item.almoxarifado_id == id)
                .order_by(Movimentacao.data.desc()).all())
        for r, mov in enumerate(movs, 5):
            fill = en_fill if mov.tipo == 'entrada' else sa_fill
            for c, v in enumerate([
                mov.data.strftime('%d/%m/%Y %H:%M'), mov.item.nome,
                'Entrada' if mov.tipo == 'entrada' else 'Saida',
                f'{mov.quantidade} {mov.item.unidade}',
                mov.responsavel or '', mov.observacao or ''
            ], 1):
                cell = ws2.cell(row=r, column=c, value=v)
                cell.fill = fill; cell.border = borda
                cell.alignment = Alignment(horizontal='left')
        for i, w in enumerate([18, 35, 10, 14, 22, 30], 1):
            ws2.column_dimensions[get_column_letter(i)].width = w
        ws3 = wb.create_sheet('Consumo por Item')
        titulo(ws3, f'Consumo por Item — {alm.nome}', 4)
        cabecalho(ws3, 4, ['Item', 'Total Entradas', 'Total Saidas', 'Saldo Atual'])
        for r, it in enumerate(itens, 5):
            entradas = sum(m.quantidade for m in it.movimentacoes if m.tipo == 'entrada')
            saidas = sum(m.quantidade for m in it.movimentacoes if m.tipo == 'saida')
            for c, v in enumerate([it.nome, entradas, saidas, it.quantidade], 1):
                cell = ws3.cell(row=r, column=c, value=v)
                cell.border = borda
                cell.alignment = Alignment(horizontal='left' if c == 1 else 'center')
        for i, w in enumerate([35, 16, 14, 14], 1):
            ws3.column_dimensions[get_column_letter(i)].width = w
        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        nome = f"estoque_{alm.nome.replace(' ', '_')}_{date.today()}.xlsx"
        return send_file(buf, as_attachment=True, download_name=nome,
                         mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

    @app.route('/api/alertas')
    @login_required
    def api_alertas():
        itens = Item.query.filter(Item.quantidade <= Item.estoque_minimo, Item.ativo == True).all()
        return jsonify([{
            'id': i.id, 'nome': i.nome, 'codigo': i.codigo,
            'quantidade': i.quantidade, 'estoque_minimo': i.estoque_minimo,
            'unidade': i.unidade, 'status': i.status,
            'almoxarifado': i.almoxarifado.nome
        } for i in itens])

    @app.route('/api/colaboradores')
    @login_required
    def api_colaboradores():
        q = request.args.get('q', '').strip()
        nomes = set()
        like = f"%{q}%"
        is_pg = 'postgresql' in str(db.engine.url)
        ilike = 'ILIKE' if is_pg else 'LIKE'
        from sqlalchemy import text
        rows = db.session.execute(
            text(f"SELECT nome FROM colaborador WHERE ativo = TRUE AND nome {ilike} :q ORDER BY nome LIMIT 10"),
            {'q': like}
        ).fetchall()
        for r in rows:
            nomes.add(r[0])
        rows = db.session.execute(
            text(f"SELECT DISTINCT colaborador FROM requisicao_mestre WHERE colaborador {ilike} :q ORDER BY colaborador LIMIT 10"),
            {'q': like}
        ).fetchall()
        for r in rows:
            nomes.add(r[0])
        rows = db.session.execute(
            text(f"SELECT DISTINCT colaborador FROM requisicao WHERE colaborador {ilike} :q ORDER BY colaborador LIMIT 10"),
            {'q': like}
        ).fetchall()
        for r in rows:
            nomes.add(r[0])
        resultado = sorted(nomes)[:10]
        return jsonify([{'nome': n} for n in resultado])

    @app.route('/colaboradores')
    @login_required
    def colaboradores():
        u = usuario_atual()
        if u.perfil not in ('admin', 'almoxarife'):
            flash('Acesso negado.', 'danger')
            return redirect(url_for('index'))
        cols = Colaborador.query.order_by(Colaborador.ativo.desc(), Colaborador.nome).all()
        return render_template('colaboradores.html', colaboradores=cols)

    @app.route('/colaboradores/novo', methods=['POST'])
    @login_required
    def novo_colaborador():
        u = usuario_atual()
        if u.perfil not in ('admin', 'almoxarife'):
            flash('Acesso negado.', 'danger')
            return redirect(url_for('index'))
        nome = request.form.get('nome', '').strip()
        funcao = request.form.get('funcao', '').strip()
        if not nome:
            flash('Informe o nome do colaborador.', 'warning')
            return redirect(url_for('colaboradores'))
        if Colaborador.query.filter(Colaborador.nome.ilike(nome)).first():
            flash(f'Colaborador "{nome}" já está cadastrado.', 'warning')
            return redirect(url_for('colaboradores'))
        db.session.add(Colaborador(nome=nome, funcao=funcao or None))
        db.session.commit()
        flash(f'✅ Colaborador "{nome}" cadastrado!', 'success')
        return redirect(url_for('colaboradores'))

    @app.route('/colaboradores/<int:id>/desativar', methods=['POST'])
    @login_required
    def desativar_colaborador(id):
        u = usuario_atual()
        if u.perfil not in ('admin', 'almoxarife'):
            flash('Acesso negado.', 'danger')
            return redirect(url_for('index'))
        c = Colaborador.query.get_or_404(id)
        c.ativo = False
        db.session.commit()
        flash(f'Colaborador "{c.nome}" desativado.', 'warning')
        return redirect(url_for('colaboradores'))

    @app.route('/colaboradores/<int:id>/reativar', methods=['POST'])
    @login_required
    def reativar_colaborador(id):
        u = usuario_atual()
        if u.perfil not in ('admin', 'almoxarife'):
            flash('Acesso negado.', 'danger')
            return redirect(url_for('index'))
        c = Colaborador.query.get_or_404(id)
        c.ativo = True
        db.session.commit()
        flash(f'Colaborador "{c.nome}" reativado.', 'success')
        return redirect(url_for('colaboradores'))

    @app.route('/usuarios')
    @admin_required
    def usuarios():
        return render_template('usuarios.html', usuarios=Usuario.query.all())

    @app.route('/usuarios/novo', methods=['GET', 'POST'])
    @admin_required
    def novo_usuario():
        almoxarifados = Almoxarifado.query.all()
        if request.method == 'POST':
            u = Usuario(
                nome=request.form['nome'],
                login=request.form['login'],
                perfil=request.form['perfil'],
                almoxarifado_id=request.form.get('almoxarifado_id') or None,
                email=request.form.get('email', '').strip() or None
            )
            u.set_senha(request.form['senha'])
            db.session.add(u)
            db.session.commit()
            flash(f'Usuário "{u.nome}" criado!', 'success')
            return redirect(url_for('usuarios'))
        return render_template('form_usuario.html', usuario=None, almoxarifados=almoxarifados)

    @app.route('/usuarios/<int:id>/editar', methods=['GET', 'POST'])
    @admin_required
    def editar_usuario(id):
        u = Usuario.query.get_or_404(id)
        atual = usuario_atual()
        almoxarifados = Almoxarifado.query.all()
        if request.method == 'POST':
            novo_perfil = request.form['perfil']
            if u.id == atual.id and novo_perfil != 'admin':
                flash('Você não pode remover seu próprio perfil de administrador.', 'danger')
                return redirect(url_for('editar_usuario', id=id))
            u.nome = request.form['nome']
            u.login = request.form['login']
            u.perfil = novo_perfil
            u.almoxarifado_id = request.form.get('almoxarifado_id') or None
            u.email = request.form.get('email', '').strip() or None
            u.ativo = 'ativo' in request.form
            if u.id == atual.id:
                u.ativo = True
            if request.form.get('senha'):
                u.set_senha(request.form['senha'])
            db.session.commit()
            flash('Usuário atualizado!', 'success')
            return redirect(url_for('usuarios'))
        return render_template('form_usuario.html', usuario=u, almoxarifados=almoxarifados)

    @app.route('/usuarios/<int:id>/deletar', methods=['POST'])
    @admin_required
    def deletar_usuario(id):
        u = Usuario.query.get_or_404(id)
        atual = usuario_atual()
        if u.id == atual.id:
            flash('Você não pode remover sua própria conta.', 'danger')
            return redirect(url_for('usuarios'))
        try:
            RequisicaoMestre.query.filter_by(mestre_id=u.id).update({'mestre_id': atual.id})
            RequisicaoMestre.query.filter_by(entregue_por_id=u.id).update({'entregue_por_id': None})
            db.session.flush()
            db.session.delete(u)
            db.session.commit()
            flash(f'Usuário "{u.nome}" removido!', 'warning')
        except Exception:
            db.session.rollback()
            flash('Não foi possível remover o usuário. Ele pode ter registros vinculados no sistema.', 'danger')
        return redirect(url_for('usuarios'))

    @app.route('/usuarios/<int:id>/acesso_extra', methods=['POST'])
    @admin_required
    def conceder_acesso_extra(id):
        u = Usuario.query.get_or_404(id)
        admin = usuario_atual()
        alm_id = request.form.get('almoxarifado_id', type=int)
        motivo = request.form.get('motivo', '')
        data_fim_str = request.form.get('data_fim', '')
        data_fim = datetime.strptime(data_fim_str, '%Y-%m-%dT%H:%M') if data_fim_str else None
        acesso = AcessoExtra(
            usuario_id=id,
            almoxarifado_id=alm_id,
            motivo=motivo,
            data_fim=data_fim,
            concedido_por=admin.nome
        )
        db.session.add(acesso)
        db.session.commit()
        flash(f'Acesso temporário concedido a {u.nome}!', 'success')
        return redirect(url_for('editar_usuario', id=id))

    @app.route('/acesso_extra/<int:id>/revogar', methods=['POST'])
    @admin_required
    def revogar_acesso_extra(id):
        a = AcessoExtra.query.get_or_404(id)
        uid = a.usuario_id
        db.session.delete(a)
        db.session.commit()
        flash('Acesso revogado!', 'warning')
        return redirect(url_for('editar_usuario', id=uid))

    @app.route('/mestre/requisicoes')
    @login_required
    def mestre_requisicoes():
        u = usuario_atual()
        if u.perfil in ('mestre', 'tecnico_seguranca'):
            reqs = RequisicaoMestre.query.filter_by(mestre_id=u.id).order_by(RequisicaoMestre.data_criacao.desc()).all()
        elif u.perfil in ('admin', 'almoxarife'):
            if u.perfil == 'almoxarife' and u.almoxarifado_id:
                reqs = RequisicaoMestre.query.filter_by(almoxarifado_id=u.almoxarifado_id).order_by(RequisicaoMestre.data_criacao.desc()).all()
            else:
                reqs = RequisicaoMestre.query.order_by(RequisicaoMestre.data_criacao.desc()).all()
        else:
            flash('Acesso negado.', 'danger')
            return redirect(url_for('index'))
        return render_template('mestre_requisicoes.html', requisicoes=reqs)

    @app.route('/mestre/requisicoes/nova', methods=['GET', 'POST'])
    @login_required
    def mestre_requisicao_nova():
        u = usuario_atual()
        if u.perfil not in ('mestre', 'tecnico_seguranca', 'admin'):
            flash('Apenas mestres e técnicos podem criar requisições.', 'danger')
            return redirect(url_for('index'))
        if u.perfil in ('mestre', 'tecnico_seguranca'):
            if not u.almoxarifado_id:
                flash('Você não está vinculado a nenhum almoxarifado. Contate o administrador.', 'warning')
                return redirect(url_for('mestre_requisicoes'))
            almoxarifados = [u.almoxarifado]
        else:
            almoxarifados = Almoxarifado.query.all()
        itens_json = {}
        for alm in almoxarifados:
            itens_json[str(alm.id)] = [
                {'id': it.id, 'nome': it.nome, 'quantidade': it.quantidade, 'unidade': it.unidade}
                for it in alm.itens if it.ativo
            ]
        if request.method == 'POST':
            colaborador = request.form.get('colaborador', '').strip()
            alm_id = int(request.form.get('almoxarifado_id', u.almoxarifado_id or 0))
            observacao = request.form.get('observacao', '')
            if not colaborador:
                flash('Informe o nome do colaborador que vai buscar os materiais.', 'warning')
                return redirect(url_for('mestre_requisicao_nova'))
            indices = set()
            for key in request.form.keys():
                if key.startswith('item_id_'):
                    try:
                        indices.add(int(key.split('_')[-1]))
                    except ValueError:
                        pass
            if not indices:
                flash('Adicione pelo menos um item à requisição.', 'warning')
                return redirect(url_for('mestre_requisicao_nova'))
            req = RequisicaoMestre(
                mestre_id=u.id,
                colaborador=colaborador,
                almoxarifado_id=alm_id,
                observacao=observacao,
                status='pendente',
                data_criacao=agora()
            )
            db.session.add(req)
            db.session.flush()
            for i in sorted(indices):
                item_id = request.form.get(f'item_id_{i}')
                qtd_str = request.form.get(f'quantidade_{i}')
                obs_item = request.form.get(f'observacao_{i}', '')
                if not item_id or not qtd_str:
                    continue
                try:
                    qtd = float(qtd_str)
                except ValueError:
                    continue
                if qtd <= 0:
                    continue
                db.session.add(RequisicaoMestreItem(
                    requisicao_id=req.id,
                    item_id=int(item_id),
                    quantidade=qtd,
                    observacao=obs_item
                ))
            db.session.commit()
            flash_html(f'✅ Requisição <strong>#{escape(req.id)}</strong> enviada ao almoxarifado! Aguarde a separação.', 'success')
            return redirect(url_for('mestre_requisicoes'))
        return render_template('mestre_requisicao_nova.html',
                               almoxarifados=almoxarifados,
                               itens_json=json.dumps(itens_json))

    @app.route('/mestre/requisicoes/<int:id>')
    @login_required
    def mestre_requisicao_detalhe(id):
        req = RequisicaoMestre.query.get_or_404(id)
        u = usuario_atual()
        if u.perfil == 'mestre' and req.mestre_id != u.id:
            flash('Acesso negado.', 'danger')
            return redirect(url_for('mestre_requisicoes'))
        if u.perfil == 'almoxarife' and req.almoxarifado_id != u.almoxarifado_id:
            flash('Acesso negado.', 'danger')
            return redirect(url_for('mestre_requisicoes'))
        return render_template('mestre_requisicao_detalhe.html', req=req)

    @app.route('/mestre/requisicoes/<int:id>/editar', methods=['GET', 'POST'])
    @login_required
    def mestre_requisicao_editar(id):
        req = RequisicaoMestre.query.get_or_404(id)
        u = usuario_atual()
        if u.perfil not in ('admin', 'almoxarife'):
            flash('Apenas almoxarife ou admin pode editar requisições.', 'danger')
            return redirect(url_for('mestre_requisicoes'))
        if req.status == 'entregue':
            flash('Não é possível editar uma requisição já entregue.', 'warning')
            return redirect(url_for('mestre_requisicao_detalhe', id=id))
        if request.method == 'POST':
            req.colaborador = request.form.get('colaborador', req.colaborador).strip()
            req.observacao = request.form.get('observacao', req.observacao)
            for ri in req.itens:
                qtd_str = request.form.get(f'qtd_{ri.id}')
                obs_str = request.form.get(f'obs_{ri.id}', ri.observacao)
                if qtd_str:
                    try:
                        ri.quantidade = float(qtd_str)
                    except ValueError:
                        pass
                ri.observacao = obs_str
            db.session.commit()
            flash('Requisição atualizada!', 'success')
            return redirect(url_for('mestre_requisicao_detalhe', id=id))
        return render_template('mestre_requisicao_editar.html', req=req)

    @app.route('/mestre/requisicoes/<int:id>/aprovar', methods=['POST'])
    @login_required
    def mestre_requisicao_aprovar(id):
        req = RequisicaoMestre.query.get_or_404(id)
        u = usuario_atual()
        if u.perfil not in ('admin', 'almoxarife'):
            flash('Acesso negado.', 'danger')
            return redirect(url_for('mestre_requisicoes'))
        if req.status != 'pendente':
            flash('Requisição não está pendente.', 'warning')
            return redirect(url_for('mestre_requisicao_detalhe', id=id))
        decisao = request.form.get('decisao', 'aprovada')
        req.status = decisao
        db.session.commit()
        if req.status == 'aprovada':
            flash(f'✅ Requisição #{req.id} aprovada! Separe os materiais e confirme a entrega.', 'success')
        else:
            flash(f'❌ Requisição #{req.id} recusada.', 'danger')
        return redirect(url_for('mestre_requisicao_detalhe', id=id))

    @app.route('/mestre/requisicoes/<int:id>/entregar', methods=['POST'])
    @login_required
    def mestre_requisicao_entregar(id):
        req = RequisicaoMestre.query.get_or_404(id)
        u = usuario_atual()
        if u.perfil not in ('admin', 'almoxarife'):
            flash('Acesso negado.', 'danger')
            return redirect(url_for('mestre_requisicoes'))
        if req.status not in ('pendente', 'aprovada'):
            flash('Requisição já foi entregue, recusada ou cancelada.', 'warning')
            return redirect(url_for('mestre_requisicao_detalhe', id=id))
        erros = []
        for ri in req.itens:
            if ri.quantidade > ri.item.quantidade:
                erros.append(f'"{ri.item.nome}": apenas {ri.item.quantidade} {ri.item.unidade} disponível')
        if erros:
            for e in erros:
                flash(f'⚠️ {e}', 'danger')
            return redirect(url_for('mestre_requisicao_detalhe', id=id))
        for ri in req.itens:
            ri.item.quantidade = round(ri.item.quantidade - ri.quantidade, 4)
            db.session.add(Movimentacao(
                tipo='saida',
                quantidade=ri.quantidade,
                responsavel=req.mestre.nome,
                observacao=f'Req. Mestre #{req.id} — Colaborador: {req.colaborador}',
                item_id=ri.item.id
            ))
        req.status = 'entregue'
        req.data_entrega = agora()
        req.entregue_por_id = u.id
        db.session.commit()
        flash(f'✅ Entrega confirmada! Estoque atualizado para {len(req.itens)} item(ns).', 'success')
        return redirect(url_for('mestre_requisicao_detalhe', id=id))

    @app.route('/mestre/requisicoes/<int:id>/cancelar', methods=['POST'])
    @login_required
    def mestre_requisicao_cancelar(id):
        req = RequisicaoMestre.query.get_or_404(id)
        u = usuario_atual()
        if u.perfil in ('mestre', 'tecnico_seguranca') and req.mestre_id != u.id:
            flash('Acesso negado.', 'danger')
            return redirect(url_for('mestre_requisicoes'))
        if req.status == 'entregue':
            flash('Não é possível cancelar uma requisição já entregue.', 'warning')
            return redirect(url_for('mestre_requisicao_detalhe', id=id))
        req.status = 'cancelada'
        db.session.commit()
        flash(f'Requisição #{req.id} cancelada.', 'warning')
        return redirect(url_for('mestre_requisicoes'))

    @app.route('/api/mestre/notificacoes')
    @login_required
    def api_mestre_notificacoes():
        u = usuario_atual()
        if u.perfil not in ('mestre', 'tecnico_seguranca'):
            return jsonify([])
        reqs = RequisicaoMestre.query.filter_by(mestre_id=u.id).filter(
            RequisicaoMestre.status.in_(['aprovada', 'parcial', 'recusada', 'entregue'])
        ).order_by(RequisicaoMestre.data_criacao.desc()).limit(10).all()
        result = []
        for r in reqs:
            if r.status == 'aprovada':
                msg = f'✅ Requisição #{r.id} aprovada! Envie o colaborador buscar.'
                tipo = 'success'
            elif r.status == 'parcial':
                msg = f'⚠️ Requisição #{r.id} aprovada parcialmente. Verifique os itens.'
                tipo = 'warning'
            elif r.status == 'recusada':
                msg = f'❌ Requisição #{r.id} foi recusada pelo almoxarifado.'
                tipo = 'danger'
            elif r.status == 'entregue':
                msg = f'📦 Requisição #{r.id} entregue ao colaborador {r.colaborador}.'
                tipo = 'info'
            else:
                continue
            result.append({'id': r.id, 'msg': msg, 'tipo': tipo, 'status': r.status})
        return jsonify(result)

    @app.route('/api/mestre/notificacoes/marcar-lidas', methods=['POST'])
    @login_required
    def marcar_notificacoes_lidas():
        return jsonify({'ok': True})

    @app.route('/admin/reativar-todos-itens', methods=['GET', 'POST'])
    @admin_required
    def reativar_todos_itens():
        if request.method == 'POST':
            try:
                itens_desativados = Item.query.filter_by(ativo=False).all()
                count = 0
                for item in itens_desativados:
                    item.ativo = True
                    count += 1
                from sqlalchemy import text
                with db.engine.connect() as conn:
                    conn.execute(text('UPDATE item SET ativo = 1 WHERE ativo IS NULL OR ativo = 0'))
                    conn.commit()
                db.session.commit()
                total_ativos = Item.query.filter_by(ativo=True).count()
                flash(f'✅ Sucesso! {count} itens reativados. Total de itens ativos: {total_ativos}', 'success')
            except Exception as e:
                flash(f'❌ Erro ao reativar itens: {str(e)}', 'danger')
            return redirect(url_for('reativar_todos_itens'))
        itens_desativados = Item.query.filter_by(ativo=False).count()
        total_itens = Item.query.count()
        return render_template('admin_reativar_itens.html',
                               itens_desativados=itens_desativados,
                               total_itens=total_itens)

    @app.route('/admin/classificar-epis', methods=['POST'])
    @admin_required
    def classificar_epis():
        palavras_epi = [
            'bota', 'capacete', 'carneira', 'cinto de segurança', 'capa de chuva',
            'calça brim', 'camisa brim', 'macacão', 'mascara', 'máscara',
            'luva vaqueta', 'luva flextactil', 'perneira', 'protetor auricular',
            'óculos de proteção', 'óculos de segurança', 'óculos de sobrepor',
            'talabarte', 'trava-quedas', 'mosquetão oval', 'cinto paraquedista',
            'epi', 'uniforme', 'colete', 'capacete'
        ]
        palavras_maq = [
            'broca diamantada', 'disco diamantado', 'disco de desbaste',
            'maçarico', 'perfuratriz', 'abrasiva'
        ]
        atualizados_epi = 0
        atualizados_maq = 0
        itens = Item.query.all()
        for it in itens:
            nome_lower = it.nome.lower()
            if any(p in nome_lower for p in palavras_epi):
                if it.categoria != 'epi':
                    it.categoria = 'epi'
                    atualizados_epi += 1
            elif any(p in nome_lower for p in palavras_maq):
                if it.categoria != 'maquinario':
                    it.categoria = 'maquinario'
                    atualizados_maq += 1
        db.session.commit()
        flash(f'✅ Classificação concluída: {atualizados_epi} EPIs e {atualizados_maq} Maquinários atualizados.', 'success')
        return redirect(url_for('index'))

    @app.route('/admin/debug-env')
    @admin_required
    def debug_env():
        variaveis = ['BACKUP_EMAIL_FROM', 'BACKUP_EMAIL_PASS', 'BACKUP_EMAIL_TO',
                     'RESEND_API_KEY', 'RESEND_FROM_EMAIL', 'SECRET_KEY', 'DATABASE_URL']
        status = {v: '✅ Definida' if os.environ.get(v) else '❌ Não definida' for v in variaveis}
        linhas = '\n'.join(f'  {k} = {v}' for k, v in status.items())
        return f'<pre style="font-family:monospace;padding:20px">\nVariáveis de ambiente:\n\n{linhas}\n</pre>'

    @app.route('/admin/backup', methods=['GET', 'POST'])
    @admin_required
    def backup_manual():
        if request.method == 'POST':
            acao = request.form.get('acao', 'download')
            if acao == 'email':
                try:
                    ok, erro_msg = enviar_backup_por_almoxarifado()
                    if ok:
                        flash('✅ Backup enviado por email com sucesso! Admins receberam o backup completo e cada almoxarife recebeu o seu.', 'success')
                    else:
                        detalhe = f' Detalhe: {erro_msg}' if erro_msg else ' Verifique as configurações BACKUP_EMAIL_FROM e BACKUP_EMAIL_PASS no Railway.'
                        flash(f'❌ Erro ao enviar email.{detalhe}', 'danger')
                except Exception as e:
                    flash(f'❌ Erro inesperado ao enviar email: {str(e)}', 'danger')
                return redirect(url_for('backup_manual'))
            try:
                buf = gerar_excel_backup()
                nome = f"backup_estoque_{date.today()}.xlsx"
                return send_file(buf, as_attachment=True, download_name=nome,
                                 mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            except Exception as e:
                flash(f'❌ Erro ao gerar backup: {str(e)}', 'danger')
                return redirect(url_for('backup_manual'))
        email_configurado = bool(os.environ.get('BACKUP_EMAIL_FROM', '').strip())
        return render_template('backup.html', email_configurado=email_configurado)

    @app.route('/api/backup-automatico', methods=['GET'])
    def api_backup_automatico():
        chave = request.args.get('key', '')
        chave_esperada = os.environ.get('BACKUP_CRON_KEY', 'backup2024').strip()
        if chave.strip() != chave_esperada:
            return jsonify({'error': 'Não autorizado'}), 401
        import threading
        def executar_backup_background():
            with current_app.app_context():
                try:
                    ok, erro_msg = enviar_backup_por_almoxarifado()
                    if ok:
                        current_app.logger.info(f'✅ BACKUP API: enviado com sucesso às {agora().strftime("%d/%m/%Y %H:%M:%S")}')
                    else:
                        current_app.logger.error(f'❌ BACKUP API: erro — {erro_msg}')
                except Exception as e:
                    current_app.logger.error(f'❌ BACKUP API: erro inesperado — {str(e)}')
        thread = threading.Thread(target=executar_backup_background)
        thread.daemon = True
        thread.start()
        return jsonify({
            'success': True,
            'message': 'Backup iniciado!',
            'timestamp': agora().strftime('%d/%m/%Y %H:%M:%S')
        }), 200
