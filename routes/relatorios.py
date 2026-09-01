
from flask import Blueprint, render_template, request, redirect, url_for, jsonify, flash, send_file, session
from markupsafe import Markup, escape
from datetime import datetime, date, timedelta
import io, os, json, re, logging
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from extensions import db
from models import (agora, Almoxarifado, Item, Movimentacao, Requisicao,
    RequisicaoMestre, RequisicaoMestreItem, Usuario, Colaborador,
    Ferramenta, HistoricoFerramenta, ItemEPI, HistoricoEPI,
    AcessoExtra, PermissaoExtra)
from core import (login_required, admin_required, almoxarife_required,
    usuario_atual, flash_html, usuario_tem_acesso_almoxarifado,
    usuario_tem_acesso_item, PERMISSOES_DISPONIVEIS,
    _check_rate_limit, _register_attempt, _clear_attempts, _check_api_rate,
    ggo_cidade, almoxarifados_do_ggo)

logger = logging.getLogger(__name__)
from utils import calcular_ruptura
relatorios_bp = Blueprint('relatorios_bp', __name__)

@relatorios_bp.route('/relatorios/consumo')
@login_required
def relatorio_consumo():
    u = usuario_atual()
    alm_id = request.args.get('almoxarifado_id', type=int)
    data_ini = request.args.get('data_ini', str(date.today().replace(day=1)))
    data_fim = request.args.get('data_fim', str(date.today()))
    aba = request.args.get('aba', 'saidas')  # saidas | entradas

    # Determina almoxarifados permitidos filtrados por cidade
    if u.perfil == 'admin':
        ids_perm = None
    elif u.perfil == 'ggo':
        ids_perm = {a.id for a in almoxarifados_do_ggo(u)}
    else:
        ids_perm = set(u.almoxarifados_permitidos())
        if u.perfil in ('tecnico_seguranca', 'analista') and u.almoxarifado_id:
            alm_ref = db.session.get(Almoxarifado, u.almoxarifado_id)
            if alm_ref and alm_ref.cidade:
                ids_perm = {a.id for a in Almoxarifado.query.filter_by(cidade=alm_ref.cidade).all()}

    if alm_id and ids_perm and alm_id not in ids_perm:
        flash('Acesso negado.', 'danger')
        return redirect(url_for('main_bp.index'))

    tipo_mov = 'saida' if aba == 'saidas' else 'entrada'

    query = Movimentacao.query.filter(
        Movimentacao.tipo == tipo_mov,
        Movimentacao.data >= data_ini,
        Movimentacao.data <= data_fim + ' 23:59:59'
    )
    if alm_id:
        query = query.join(Item).filter(Item.almoxarifado_id == alm_id)
    elif u.perfil != 'admin':
        if ids_perm:
            query = query.join(Item).filter(Item.almoxarifado_id.in_(ids_perm))
        else:
            query = query.filter(False)

    movs = query.order_by(Movimentacao.data.desc()).all()

    if u.perfil == 'admin':
        almoxarifados = Almoxarifado.query.all()
    elif ids_perm:
        almoxarifados = Almoxarifado.query.filter(Almoxarifado.id.in_(ids_perm)).all()
    else:
        almoxarifados = []

    return render_template('relatorio_consumo.html', movimentacoes=movs,
                           almoxarifados=almoxarifados, data_ini=data_ini,
                           data_fim=data_fim, alm_id=alm_id, aba=aba)

@relatorios_bp.route('/relatorios/consumo/exportar')
@login_required
def exportar_consumo():
    u = usuario_atual()
    alm_id = request.args.get('almoxarifado_id', type=int)
    data_ini = request.args.get('data_ini', str(date.today().replace(day=1)))
    data_fim = request.args.get('data_fim', str(date.today()))
    if alm_id and u.perfil != 'admin' and alm_id not in u.almoxarifados_permitidos():
        flash('Acesso negado.', 'danger')
        return redirect(url_for('main_bp.index'))
    query = Movimentacao.query.filter(
        Movimentacao.tipo == 'saida',
        Movimentacao.data >= data_ini,
        Movimentacao.data <= data_fim + ' 23:59:59'
    )
    if alm_id:
        query = query.join(Item).filter(Item.almoxarifado_id == alm_id)
    elif u.perfil != 'admin':
        query = query.join(Item).filter(Item.almoxarifado_id.in_(u.almoxarifados_permitidos()))
    movs = query.order_by(Movimentacao.data.desc()).all()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Consumo'
    h_fill = PatternFill('solid', fgColor='1A3A5C')
    h_font = Font(bold=True, color='FFFFFF', size=11)
    borda   = Border(left=Side(style='thin'), right=Side(style='thin'),
                     top=Side(style='thin'), bottom=Side(style='thin'))

    alm_nome = db.session.get(Almoxarifado, alm_id).nome if alm_id else 'Todos os Almoxarifados'
    ws.merge_cells('A1:H1')
    ws['A1'] = f'Relatório de Consumo — {alm_nome}'
    ws['A1'].font = Font(bold=True, size=13, color='1A3A5C')
    ws.merge_cells('A2:H2')
    ws['A2'] = f'Período: {data_ini} a {data_fim}   |   Gerado em: {agora().strftime("%d/%m/%Y %H:%M")}'
    ws['A2'].font = Font(italic=True, size=10, color='666666')

    headers = ['Data', 'Código', 'Item', 'Categoria', 'Almoxarifado', 'Quantidade', 'Responsável', 'Colaborador', 'Observação']
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=4, column=col, value=h)
        c.font = h_font; c.fill = h_fill
        c.alignment = Alignment(horizontal='center'); c.border = borda

    cat_label = {
        'epi': 'EPI', 'maquinario': 'Maquinário', 'eletrica': 'Elétrica',
        'hidraulica': 'Hidráulica', 'gas': 'Gás', 'geral': 'Geral'
    }

    for row_num, mov in enumerate(movs, 5):
        # Separar colaborador da observação
        obs = mov.observacao or ''
        if 'liberado P/' in obs or 'liberado p/' in obs:
            partes = re.split(r'liberado [Pp]/', obs, maxsplit=1)
            resto = partes[-1].split(' | ', 1)
            colab_nome = resto[0].strip()
            obs_limpa = resto[1].strip() if len(resto) > 1 else ''
        elif 'Colaborador:' in obs:
            partes = obs.split('Colaborador:', 1)[-1].split('|', 1)
            colab_nome = partes[0].strip()
            obs_limpa = partes[1].strip() if len(partes) > 1 else ''
        else:
            colab_nome = ''
            obs_limpa = obs

        cat = cat_label.get(mov.item.categoria or 'geral', mov.item.categoria or 'Geral')
        dados = [
            mov.data.strftime('%d/%m/%Y %H:%M'),
            mov.item.codigo, mov.item.nome,
            cat,
            mov.item.almoxarifado.nome,
            f'{mov.quantidade} {mov.item.unidade}',
            mov.responsavel or '',
            colab_nome,
            obs_limpa
        ]
        for col, val in enumerate(dados, 1):
            c = ws.cell(row=row_num, column=col, value=val)
            c.border = borda
            if row_num % 2 == 0:
                c.fill = PatternFill('solid', fgColor='F0F4F8')

    for i, w in enumerate([18, 14, 45, 14, 35, 14, 20, 30, 30], 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    buf = io.BytesIO()
    wb.save(buf); buf.seek(0)
    return send_file(buf, as_attachment=True,
                     download_name=f'consumo_{data_ini}_a_{data_fim}.xlsx',
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

@relatorios_bp.route('/relatorios/consumo-por-pessoa')
@login_required
def relatorio_consumo_pessoa():
    """Relatório de consumo agrupado por colaborador (extraído da observação)."""
    u = usuario_atual()
    alm_id = request.args.get('almoxarifado_id', type=int)
    data_ini = request.args.get('data_ini', str(date.today().replace(day=1)))
    data_fim = request.args.get('data_fim', str(date.today()))
    responsavel_filtro = request.args.get('responsavel', '').strip()

    # Técnico de segurança e analista com permissão ver_relatorios vêem
    # todos os almoxarifados da sua cidade — não de outras cidades
    tem_perm_relatorio = u.perfil in ('tecnico_seguranca', 'analista') or any(
        p.permissao == 'ver_relatorios' and p.ativo for p in u.permissoes_extras
    )

    # IDs de almoxarifados que o usuário pode ver (filtrado por cidade para não-admins)
    if u.perfil == 'admin':
        ids_permitidos = None  # sem restrição
    elif u.perfil == 'ggo':
        ids_permitidos = {a.id for a in almoxarifados_do_ggo(u)}
    else:
        ids_permitidos = u.almoxarifados_permitidos()
        # Para técnico/analista: expande para todos da mesma cidade
        if tem_perm_relatorio and u.almoxarifado_id:
            alm_ref = db.session.get(Almoxarifado, u.almoxarifado_id)
            if alm_ref and alm_ref.cidade:
                alms_cidade = Almoxarifado.query.filter_by(cidade=alm_ref.cidade).all()
                ids_permitidos = set(a.id for a in alms_cidade)

    if alm_id and u.perfil != 'admin' and ids_permitidos and alm_id not in ids_permitidos:
        flash('Acesso negado.', 'danger')
        return redirect(url_for('main_bp.index'))

    query = Movimentacao.query.filter(
        Movimentacao.tipo == 'saida',
        Movimentacao.data >= data_ini,
        Movimentacao.data <= data_fim + ' 23:59:59'
    )
    if alm_id:
        query = query.join(Item).filter(Item.almoxarifado_id == alm_id)
    elif u.perfil != 'admin':
        if ids_permitidos:
            query = query.join(Item).filter(Item.almoxarifado_id.in_(ids_permitidos))
        else:
            # sem almoxarifados permitidos — retorna vazio
            query = query.filter(False)

    movs = query.order_by(Movimentacao.data.desc()).all()

    import re

    # Nomes de usuários do sistema (mestres, técnicos, almoxarifes) para excluir
    usuarios_sistema = {u.nome.strip().lower() for u in Usuario.query.filter(
        Usuario.perfil.in_(['admin', 'almoxarife', 'mestre', 'tecnico_seguranca'])
    ).all()}

    def extrair_colaborador(mov):
        """Extrai o nome do colaborador da observação."""
        obs = mov.observacao or ''
        # Formato requisição mestre: "Req. Mestre #X — Colaborador: Nome"
        m = re.search(r'[Cc]olaborador[:\s]+([^|]+)', obs)
        if m:
            nome = m.group(1).strip()
            if nome:
                return nome
        # Formato movimentação avulsa: "liberado P/ Nome | ..."
        m = re.search(r'liberado\s+[Pp][/\s]+([^|]+)', obs, re.IGNORECASE)
        if m:
            nome = m.group(1).strip()
            if nome:
                return nome
        return None  # sem colaborador identificável

    def normalizar_nome(nome):
        """Remove sufixos de requisição do nome e normaliza capitalização."""
        if not nome:
            return None
        # Remove ' | req XXXX' ou ' | REQ XXXX' ou ' | ajuste ...' do final
        nome_limpo = re.sub(r'\s*[|·]\s*.+$', '', nome).strip()
        if not nome_limpo:
            return None
        # Normaliza: remove espaços duplos, converte para título (João Silva)
        nome_limpo = re.sub(r'\s+', ' ', nome_limpo).strip().upper()
        return nome_limpo if nome_limpo else None

    def e_nome_valido(nome):
        """Filtra nomes inválidos: ajustes, vazios, nomes do sistema."""
        if not nome or len(nome) < 2:
            return False
        nome_lower = nome.lower().strip()
        # Excluir termos de ajuste sistêmico
        termos_invalidos = [
            'ajuste', 'sistemico', 'sistêmico', 'reajuste',
            'req -', 'req-', '- req', 'sem responsável',
            'sem responsavel', 'ajuste de estoque'
        ]
        if any(t in nome_lower for t in termos_invalidos):
            return False
        # Excluir nomes de usuários do sistema (mestres, admins, etc.)
        if nome_lower in usuarios_sistema:
            return False
        return True

    # Filtrar por nome se informado
    from collections import defaultdict
    por_pessoa = defaultdict(list)
    for mov in movs:
        colab_raw = extrair_colaborador(mov)
        colaborador = normalizar_nome(colab_raw)
        if not e_nome_valido(colaborador):
            continue
        if responsavel_filtro and responsavel_filtro.lower() not in colaborador.lower():
            continue
        por_pessoa[colaborador].append(mov)

    # Calcular totais por pessoa
    resumo = []
    for nome, lista in sorted(por_pessoa.items()):
        total_movs = len(lista)
        itens_distintos = len(set(m.item_id for m in lista))
        # Coletar números de requisição únicos associados a esta pessoa
        reqs = []
        for mov in lista:
            obs = mov.observacao or ''
            m_req = re.search(r'[Rr]eq\.?\s*(?:[Mm]estre\s*)?#?(\d+)', obs)
            if m_req:
                r_num = m_req.group(1)
                if r_num not in reqs:
                    reqs.append(r_num)
        resumo.append({
            'nome': nome,
            'movimentacoes': lista,
            'total_movs': total_movs,
            'itens_distintos': itens_distintos,
            'reqs': reqs,
        })

    # Lista de almoxarifados para o filtro — respeita cidade do usuário
    if u.perfil == 'admin':
        almoxarifados = Almoxarifado.query.all()
    elif ids_permitidos:
        almoxarifados = Almoxarifado.query.filter(Almoxarifado.id.in_(ids_permitidos)).all()
    else:
        almoxarifados = []

    return render_template('relatorio_consumo_pessoa.html',
                           resumo=resumo,
                           almoxarifados=almoxarifados,
                           data_ini=data_ini,
                           data_fim=data_fim,
                           alm_id=alm_id,
                           responsavel_filtro=responsavel_filtro,
                           total_geral=sum(p['total_movs'] for p in resumo))

@relatorios_bp.route('/relatorios/consumo-por-pessoa/exportar')
@login_required
def exportar_consumo_pessoa():
    """Exporta relatório de consumo por pessoa em Excel com 2 abas."""
    import re
    from collections import defaultdict

    u = usuario_atual()
    alm_id           = request.args.get('almoxarifado_id', type=int)
    data_ini         = request.args.get('data_ini', str(date.today().replace(day=1)))
    data_fim         = request.args.get('data_fim', str(date.today()))
    responsavel_filtro = request.args.get('responsavel', '').strip()

    if alm_id and u.perfil != 'admin' and alm_id not in u.almoxarifados_permitidos():
        flash('Acesso negado.', 'danger')
        return redirect(url_for('main_bp.index'))

    query = Movimentacao.query.filter(
        Movimentacao.tipo == 'saida',
        Movimentacao.data >= data_ini,
        Movimentacao.data <= data_fim + ' 23:59:59'
    )
    if alm_id:
        query = query.join(Item).filter(Item.almoxarifado_id == alm_id)
    elif u.perfil != 'admin':
        query = query.join(Item).filter(Item.almoxarifado_id.in_(u.almoxarifados_permitidos()))

    movs = query.order_by(Movimentacao.data.asc()).all()

    def extrair_colaborador(mov):
        obs = mov.observacao or ''
        m = re.search(r'liberado\s+[Pp][/\s]+(.+)', obs, re.IGNORECASE)
        if m: return m.group(1).strip()
        m = re.search(r'[Cc]olaborador[:\s]+([^|]+)', obs)
        if m: return m.group(1).strip()
        return mov.responsavel or 'Sem responsável'

    def normalizar_nome(nome):
        return re.sub(r'\s*[|·]\s*[Rr][Ee][Qq]\.?\s*\d+.*$', '', nome).strip() or nome

    por_pessoa = defaultdict(list)
    for mov in movs:
        colab = normalizar_nome(extrair_colaborador(mov))
        if responsavel_filtro and responsavel_filtro.lower() not in colab.lower():
            continue
        por_pessoa[colab].append(mov)

    # Estilos
    h_fill  = PatternFill('solid', fgColor='1A3A5C')
    h_font  = Font(bold=True, color='FFFFFF', size=11)
    z_fill  = PatternFill('solid', fgColor='F0F4F8')
    borda   = Border(left=Side(style='thin'), right=Side(style='thin'),
                     top=Side(style='thin'), bottom=Side(style='thin'))
    centro  = Alignment(horizontal='center', vertical='center')
    esq     = Alignment(horizontal='left',   vertical='center')
    alm_nome = db.session.get(Almoxarifado, alm_id).nome if alm_id else 'Todos'
    total_geral = sum(len(v) for v in por_pessoa.values())

    wb = openpyxl.Workbook()

    # ── ABA 1: Resumo ─────────────────────────────────────────────────────────
    ws1 = wb.active
    ws1.title = 'Resumo'
    for col, w in enumerate([6, 40, 20, 18, 18], 1):
        ws1.column_dimensions[get_column_letter(col)].width = w

    ws1.merge_cells('A1:E1')
    ws1['A1'].value = f'Consumo por Pessoa — {alm_nome}'
    ws1['A1'].font = Font(bold=True, size=13, color='1A3A5C')
    ws1['A1'].fill = PatternFill('solid', fgColor='E8F0F7')
    ws1['A1'].alignment = centro

    ws1.merge_cells('A2:E2')
    ws1['A2'].value = f'Período: {data_ini} a {data_fim}   |   Gerado em: {agora().strftime("%d/%m/%Y %H:%M")}'
    ws1['A2'].font = Font(italic=True, size=10, color='666666')
    ws1['A2'].alignment = centro

    for col, h in enumerate(['#', 'Funcionário', 'Total Retiradas', 'Itens Distintos', 'Participação (%)'], 1):
        c = ws1.cell(row=4, column=col, value=h)
        c.font = h_font; c.fill = h_fill; c.alignment = centro; c.border = borda

    for i, (nome, lista) in enumerate(sorted(por_pessoa.items(), key=lambda x: len(x[1]), reverse=True), 1):
        pct = round(len(lista) / total_geral * 100, 1) if total_geral else 0
        row = i + 4
        for col, val in enumerate([i, nome, len(lista), len(set(m.item_id for m in lista)), f'{pct}%'], 1):
            c = ws1.cell(row=row, column=col, value=val)
            c.border = borda
            c.alignment = esq if col == 2 else centro
            if row % 2 == 0: c.fill = z_fill

    # Linha total
    r = len(por_pessoa) + 5
    for col, val in enumerate(['', f'{len(por_pessoa)} pessoa(s)', total_geral, '', '100%'], 1):
        c = ws1.cell(row=r, column=col, value=val)
        c.font = Font(bold=True)
        c.fill = PatternFill('solid', fgColor='D0E4F7')
        c.border = borda; c.alignment = centro

    # ── ABA 2: Detalhes ───────────────────────────────────────────────────────
    ws2 = wb.create_sheet('Detalhes')
    for col, w in enumerate([35, 18, 14, 45, 14, 30, 25], 1):
        ws2.column_dimensions[get_column_letter(col)].width = w

    ws2.merge_cells('A1:G1')
    ws2['A1'].value = f'Detalhes por Funcionário — {alm_nome}'
    ws2['A1'].font = Font(bold=True, size=13, color='1A3A5C')
    ws2['A1'].fill = PatternFill('solid', fgColor='E8F0F7')
    ws2['A1'].alignment = centro

    ws2.merge_cells('A2:G2')
    ws2['A2'].value = f'Período: {data_ini} a {data_fim}   |   Gerado em: {agora().strftime("%d/%m/%Y %H:%M")}'
    ws2['A2'].font = Font(italic=True, size=10, color='666666')
    ws2['A2'].alignment = centro

    for col, h in enumerate(['Funcionário', 'Data', 'Código', 'Item', 'Quantidade', 'Almoxarifado', 'Liberado por'], 1):
        c = ws2.cell(row=4, column=col, value=h)
        c.font = h_font; c.fill = h_fill; c.alignment = centro; c.border = borda

    row_num = 5
    for nome, lista in sorted(por_pessoa.items()):
        # Cabeçalho do funcionário
        ws2.merge_cells(f'A{row_num}:G{row_num}')
        c = ws2.cell(row=row_num, column=1, value=f'👷 {nome}  ({len(lista)} retirada(s))')
        c.font = Font(bold=True, color='FFFFFF', size=11)
        c.fill = PatternFill('solid', fgColor='2E6DA4')
        c.alignment = esq; c.border = borda
        row_num += 1

        for mov in sorted(lista, key=lambda m: m.data):
            for col, val in enumerate([nome, mov.data.strftime('%d/%m/%Y %H:%M'),
                                        mov.item.codigo, mov.item.nome,
                                        f'{mov.quantidade} {mov.item.unidade}',
                                        mov.item.almoxarifado.nome,
                                        mov.responsavel or '—'], 1):
                c = ws2.cell(row=row_num, column=col, value=val)
                c.font = Font(size=9); c.border = borda
                c.alignment = esq if col in [1,3,4] else centro
                if row_num % 2 == 0: c.fill = z_fill
            row_num += 1
        row_num += 1  # linha em branco entre funcionários

    buf = io.BytesIO()
    wb.save(buf); buf.seek(0)
    return send_file(buf, as_attachment=True,
                     download_name=f'consumo_por_pessoa_{data_ini}_a_{data_fim}.xlsx',
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')


@relatorios_bp.route('/relatorios/ficha-epi')
@login_required
def ficha_epi():
    """Página para gerar ficha de EPI individual por funcionário."""
    import re
    u = usuario_atual()
    # Técnico de segurança e analista vêem todos os almoxarifados
    tem_perm_relatorio = u.perfil in ('tecnico_seguranca', 'analista') or any(
        p.permissao == 'ver_relatorios' and p.ativo for p in u.permissoes_extras
    )
    query = Movimentacao.query.join(Item).filter(
        Movimentacao.tipo == 'saida', Item.categoria == 'epi')
    if u.perfil not in ('admin', 'analista') and not tem_perm_relatorio:
        query = query.filter(Item.almoxarifado_id.in_(u.almoxarifados_permitidos()))
    movs = query.order_by(Movimentacao.data.desc()).all()

    # Coleta funcionários únicos com sua requisição mais recente
    visto = {}  # nome_lower -> {nome, req}
    for mov in movs:
        obs = mov.observacao or ''
        nome = None
        req_num = None

        m = re.search(r'liberado\s+[Pp][/\s]+(.+)', obs, re.IGNORECASE)
        if m:
            nome = m.group(1).strip()
        elif 'Colaborador:' in obs:
            partes = obs.split('Colaborador:', 1)
            nome = partes[-1].split('|')[0].strip()

        # Tenta extrair número da requisição da observação
        m_req = re.search(r'[Rr]eq\.?\s*(?:Mestre\s*)?#?(\d+)', obs)
        if m_req:
            req_num = m_req.group(1)

        if nome and nome.lower() not in visto:
            visto[nome.lower()] = {'nome': nome, 'req': req_num or ''}

    funcionarios = sorted(visto.values(), key=lambda x: x['nome'])
    return render_template('ficha_epi.html',
                           funcionarios=funcionarios,
                           data_ini='2020-01-01',
                           data_fim=str(date.today()))


@relatorios_bp.route('/relatorios/ficha-epi/exportar')
@login_required
def exportar_ficha_epi():
    """Exporta FORM.SEG.014 — Ficha de Controle de EPIs e Uniformes."""
    import re
    u = usuario_atual()
    funcionario = request.args.get('funcionario', '').strip()
    data_ini    = request.args.get('data_ini', '2020-01-01')
    data_fim    = request.args.get('data_fim', str(date.today()))

    if not funcionario:
        flash('Selecione um funcionário.', 'warning')
        return redirect(url_for('relatorios_bp.ficha_epi'))

    query = (Movimentacao.query.join(Item)
                  .filter(Movimentacao.tipo == 'saida',
                          Item.categoria == 'epi',
                          Movimentacao.data >= data_ini,
                          Movimentacao.data <= data_fim + ' 23:59:59'))
    if u.perfil not in ('admin', 'analista'):
        tem_perm_relatorio = u.perfil in ('tecnico_seguranca', 'analista') or any(
            p.permissao == 'ver_relatorios' and p.ativo for p in u.permissoes_extras
        )
        if not tem_perm_relatorio:
            query = query.filter(Item.almoxarifado_id.in_(u.almoxarifados_permitidos()))
    movs_todas = query.order_by(Movimentacao.data.asc()).all()

    def extrair_colab(mov):
        obs = mov.observacao or ''
        m = re.search(r'liberado\s+[Pp][/\s]+(.+)', obs, re.IGNORECASE)
        if m: return m.group(1).strip()
        m = re.search(r'[Cc]olaborador[:\s]+([^|]+)', obs)
        if m: return m.group(1).strip()
        return mov.responsavel or ''

    lista = [m for m in movs_todas if extrair_colab(m).lower() == funcionario.lower()]

    if not lista:
        flash(f'Nenhuma retirada de EPI encontrada para "{funcionario}" no período.', 'warning')
        return redirect(url_for('relatorios_bp.ficha_epi'))

    # ── Estilos ──────────────────────────────────────────────────────────────
    borda = Border(left=Side(style='thin'), right=Side(style='thin'),
                   top=Side(style='thin'), bottom=Side(style='thin'))
    borda_med = Border(left=Side(style='medium'), right=Side(style='medium'),
                       top=Side(style='medium'), bottom=Side(style='medium'))
    centro = Alignment(horizontal='center', vertical='center', wrap_text=True)
    esq    = Alignment(horizontal='left',   vertical='center', wrap_text=True)
    azul_esc = PatternFill('solid', fgColor='1F3864')
    azul_cla = PatternFill('solid', fgColor='BDD7EE')
    cinza    = PatternFill('solid', fgColor='F2F2F2')

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = funcionario[:28]

    # Larguras: A=Qtd | B=Descrição | C=C.A. | D=Data Ent | E=Assin Ent | F=Data Dev | G=Assin Dev | H=Motivo
    for col, w in zip('ABCDEFGH', [8, 40, 10, 14, 26, 14, 26, 18]):
        ws.column_dimensions[col].width = w

    def celula(ws, ref, val='', font=None, fill=None, aln=None, brd=None, height=None):
        c = ws[ref]
        c.value = val
        if font:  c.font = font
        if fill:  c.fill = fill
        if aln:   c.alignment = aln
        if brd:   c.border = brd
        return c

    def merge_row(ws, row, col_ini, col_fim, val='', font=None, fill=None, aln=None, height=None):
        ws.merge_cells(f'{col_ini}{row}:{col_fim}{row}')
        c = ws[f'{col_ini}{row}']
        c.value = val
        if font:  c.font = font
        if fill:  c.fill = fill
        if aln:   c.alignment = aln
        for col in range(ord(col_ini)-64, ord(col_fim)-64+1):
            ws.cell(row=row, column=col).border = borda
        if height: ws.row_dimensions[row].height = height
        return c

    # ── LINHA 1: Cabeçalho principal ─────────────────────────────────────────
    ws.row_dimensions[1].height = 36
    ws.merge_cells('A1:C1')
    ws['A1'].value = 'STANZA'
    ws['A1'].font = Font(bold=True, size=22, color='808080')
    ws['A1'].alignment = centro
    for c in range(1,4): ws.cell(1,c).border = borda_med

    ws.merge_cells('D1:F1')
    ws['D1'].value = 'FICHA DE CONTROLE DE EPI\'S E UNIFORMES'
    ws['D1'].font = Font(bold=True, size=13, color='1F3864')
    ws['D1'].alignment = centro
    ws['D1'].fill = azul_cla
    for c in range(4,7): ws.cell(1,c).border = borda_med

    ws.merge_cells('G1:H1')
    ws['G1'].value = 'FORM.SEG.014'
    ws['G1'].font = Font(bold=True, size=9, color='1F3864')
    ws['G1'].alignment = centro
    ws['G1'].fill = azul_cla
    for c in range(7,9): ws.cell(1,c).border = borda_med

    # ── LINHA 2: Data elaboração ──────────────────────────────────────────────
    ws.row_dimensions[2].height = 14
    ws.merge_cells('D2:F2')
    ws['D2'].value = 'Data Elaboração/Revisão: 20/10/2024'
    ws['D2'].font = Font(size=8, italic=True, color='595959')
    ws['D2'].alignment = centro
    ws.merge_cells('G2:H2')
    ws['G2'].value = 'Revisão: 00'
    ws['G2'].font = Font(size=8, italic=True, color='595959')
    ws['G2'].alignment = centro
    for c in range(1,9): ws.cell(2,c).border = borda

    # ── LINHA 3: Dados do funcionário ────────────────────────────────────────
    ws.row_dimensions[3].height = 22
    ws.merge_cells('A3:B3')
    ws['A3'].value = f'NOME: {funcionario.upper()}'
    ws['A3'].font = Font(bold=True, size=10)
    ws['A3'].alignment = esq
    ws['A3'].fill = cinza

    ws['C3'].value = f'MATRÍCULA:'
    ws['C3'].font = Font(size=9)
    ws['C3'].alignment = centro
    ws['C3'].fill = cinza

    ws.merge_cells('D3:E3')
    ws['D3'].value = 'FUNÇÃO:'
    ws['D3'].font = Font(size=9)
    ws['D3'].alignment = esq
    ws['D3'].fill = cinza

    ws.merge_cells('F3:G3')
    ws['F3'].value = f'ADMISSÃO:'
    ws['F3'].font = Font(size=9)
    ws['F3'].alignment = esq
    ws['F3'].fill = cinza

    ws['H3'].value = ''
    ws['H3'].fill = cinza
    for c in range(1,9): ws.cell(3,c).border = borda

    # ── LINHA 4: Cabeçalho da tabela ─────────────────────────────────────────
    ws.row_dimensions[4].height = 20
    for ref, val in [('A4','QUANT'), ('B4','DESCRIÇÃO'), ('C4','C.A.')]:
        ws[ref].value = val
        ws[ref].font = Font(bold=True, color='FFFFFF', size=9)
        ws[ref].fill = azul_esc
        ws[ref].alignment = centro
        ws[ref].border = borda

    ws.merge_cells('D4:E4')
    ws['D4'].value = 'ENTREGA'
    ws['D4'].font = Font(bold=True, color='FFFFFF', size=9)
    ws['D4'].fill = azul_esc
    ws['D4'].alignment = centro
    for c in range(4,6): ws.cell(4,c).border = borda

    ws.merge_cells('F4:G4')
    ws['F4'].value = 'DEVOLUÇÃO'
    ws['F4'].font = Font(bold=True, color='FFFFFF', size=9)
    ws['F4'].fill = azul_esc
    ws['F4'].alignment = centro
    for c in range(6,8): ws.cell(4,c).border = borda

    ws['H4'].value = 'MOTIVO'
    ws['H4'].font = Font(bold=True, color='FFFFFF', size=9)
    ws['H4'].fill = azul_esc
    ws['H4'].alignment = centro
    ws['H4'].border = borda

    # ── LINHA 5: Sub-cabeçalho ───────────────────────────────────────────────
    ws.row_dimensions[5].height = 16
    for ref, val in [('A5',''), ('B5',''), ('C5',''),
                     ('D5','DATA'), ('E5','ASSINATURA'),
                     ('F5','DATA'), ('G5','ASSINATURA'), ('H5','')]:
        ws[ref].value = val
        ws[ref].font = Font(bold=True, color='FFFFFF', size=8)
        ws[ref].fill = azul_esc
        ws[ref].alignment = centro
        ws[ref].border = borda

    # ── LINHAS DE DADOS ──────────────────────────────────────────────────────
    row = 6
    movs_com_foto = []  # guarda movs que têm foto para aba de comprovantes
    for mov in lista:
        ws.row_dimensions[row].height = 18
        fill_z = PatternFill('solid', fgColor='EBF3FB') if row % 2 == 0 else None
        tem_foto = bool(mov.foto_url)
        for col, val in zip('ABCDEFGH', [
            f'{mov.quantidade} {mov.item.unidade}',
            mov.item.nome, mov.item.ca or '',
            mov.data.strftime('%d/%m/%Y'), '',
            '', '',
            '📸 Ver aba' if tem_foto else ''
        ]):
            c = ws[f'{col}{row}']
            c.value = val
            c.font = Font(size=9)
            c.alignment = esq if col == 'B' else centro
            c.border = borda
            if fill_z: c.fill = fill_z
        if tem_foto:
            movs_com_foto.append((row, mov))
        row += 1

    # Linhas em branco (mínimo 14 no total conforme formulário)
    total_linhas = max(14, len(lista) + 4)
    while row <= 5 + total_linhas:
        ws.row_dimensions[row].height = 18
        for col in 'ABCDEFGH':
            ws[f'{col}{row}'].border = borda
            ws[f'{col}{row}'].value = '/    /' if col in ('D','F') else ''
            ws[f'{col}{row}'].font = Font(size=9, color='BFBFBF')
            ws[f'{col}{row}'].alignment = centro
        row += 1

    # ── TERMO DE RESPONSABILIDADE ────────────────────────────────────────────
    row += 1
    ws.row_dimensions[row].height = 16
    ws.merge_cells(f'A{row}:H{row}')
    ws[f'A{row}'].value = 'TERMO DE RESPONSABILIDADE'
    ws[f'A{row}'].font = Font(bold=True, size=10, color='1F3864')
    ws[f'A{row}'].alignment = centro
    ws[f'A{row}'].fill = azul_cla
    for c in range(1,9): ws.cell(row,c).border = borda
    row += 1

    ws.row_dimensions[row].height = 70
    ws.merge_cells(f'A{row}:H{row}')
    ws[f'A{row}'].value = (
        'Pelo presente declaro que recebi da empresa STANZA INCORPORAÇÃO E CONSTRUÇÃO LTDA., os materiais '
        'relacionados nesta ficha, assumindo o compromisso nos termos das letras "a" e "b" do ítem 1.8 da NR 1 '
        'e letras "a","b"e "c" do ítem 6.7.1 da NR 6, de usá-los em atividades ligadas ao trabalho, zelar pela '
        'sua guarda, conservação e devolvê-lo ao setor competente quando se tornar impróprio para uso ou por '
        'motivo de demissão ou afastamento.\n'
        'Em caso de perda, extravio e inutilização proposital do material recebido, autorizo a empresa, na forma '
        'prevista no parágrafo primeiro do art. 462 da CLT - Consolidação das leis do trabalho. A descontar de '
        'meu salário, inclusive no que me couber a título de indenização por rescisão de contrato de trabalho, '
        'a importância correspondente ao valor do material.'
    )
    ws[f'A{row}'].font = Font(size=8)
    ws[f'A{row}'].alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
    for c in range(1,9): ws.cell(row,c).border = borda
    row += 2

    # ── DATA ─────────────────────────────────────────────────────────────────
    ws.row_dimensions[row].height = 22
    ws.merge_cells(f'A{row}:H{row}')
    ws[f'A{row}'].value = 'Data: _______ / _______ / _____________'
    ws[f'A{row}'].font = Font(size=10)
    ws[f'A{row}'].alignment = esq
    for c in range(1,9): ws.cell(row,c).border = borda
    row += 2

    # ── CABEÇALHO BLOCO ASSINATURAS ──────────────────────────────────────────
    ws.row_dimensions[row].height = 16
    for col_ini, col_fim, label in [('A','B','FUNCIONÁRIO'), ('C','E','RESPONSÁVEL'), ('F','H','TESTEMUNHA')]:
        ws.merge_cells(f'{col_ini}{row}:{col_fim}{row}')
        ws[f'{col_ini}{row}'].value = label
        ws[f'{col_ini}{row}'].font = Font(bold=True, size=9, color='FFFFFF')
        ws[f'{col_ini}{row}'].fill = azul_esc
        ws[f'{col_ini}{row}'].alignment = centro
        for c in range(ord(col_ini)-64, ord(col_fim)-64+1):
            ws.cell(row, c).border = borda
    row += 1

    # Linha: Nome por extenso
    ws.row_dimensions[row].height = 22
    for col_ini, col_fim, placeholder in [('A','B', funcionario.upper()), ('C','E',''), ('F','H','')]:
        ws.merge_cells(f'{col_ini}{row}:{col_fim}{row}')
        ws[f'{col_ini}{row}'].value = placeholder
        ws[f'{col_ini}{row}'].font = Font(bold=True, size=9)
        ws[f'{col_ini}{row}'].alignment = centro
        ws[f'{col_ini}{row}'].fill = cinza
        for c in range(ord(col_ini)-64, ord(col_fim)-64+1):
            ws.cell(row, c).border = borda
    row += 1

    # Linha: Assinatura (espaço em branco para assinar)
    ws.row_dimensions[row].height = 38
    for col_ini, col_fim in [('A','B'), ('C','E'), ('F','H')]:
        ws.merge_cells(f'{col_ini}{row}:{col_fim}{row}')
        ws[f'{col_ini}{row}'].value = ''
        for c in range(ord(col_ini)-64, ord(col_fim)-64+1):
            ws.cell(row, c).border = borda
    row += 1

    # Linha: rótulo "Assinatura"
    ws.row_dimensions[row].height = 14
    for col_ini, col_fim in [('A','B'), ('C','E'), ('F','H')]:
        ws.merge_cells(f'{col_ini}{row}:{col_fim}{row}')
        ws[f'{col_ini}{row}'].value = 'Assinatura'
        ws[f'{col_ini}{row}'].font = Font(size=8, italic=True, color='595959')
        ws[f'{col_ini}{row}'].alignment = centro
        for c in range(ord(col_ini)-64, ord(col_fim)-64+1):
            ws.cell(row, c).border = borda

    # ── ABA DE COMPROVANTES (fotos embutidas) ─────────────────────────────────
    if movs_com_foto:
        try:
            from openpyxl.drawing.image import Image as XLImage
            ws2 = wb.create_sheet('Comprovantes')
            ws2.column_dimensions['A'].width = 20
            ws2.column_dimensions['B'].width = 50
            ws2.column_dimensions['C'].width = 15

            ws2.merge_cells('A1:C1')
            ws2['A1'].value = f'Comprovantes de Entrega — {funcionario.upper()}'
            ws2['A1'].font = Font(bold=True, size=12, color='1F3864')
            ws2['A1'].alignment = Alignment(horizontal='center', vertical='center')
            ws2['A1'].fill = PatternFill('solid', fgColor='BDD7EE')
            ws2.row_dimensions[1].height = 22

            foto_row = 3
            for _, mov in movs_com_foto:
                try:
                    import base64 as b64
                    # Extrai os bytes da imagem base64
                    header, data_b64 = mov.foto_url.split(',', 1)
                    img_bytes = b64.b64decode(data_b64)
                    img_buf = io.BytesIO(img_bytes)
                    xl_img = XLImage(img_buf)
                    # Redimensiona para caber na célula (max 300px largura)
                    scale = min(1.0, 300 / (xl_img.width or 300))
                    xl_img.width  = int(xl_img.width  * scale)
                    xl_img.height = int(xl_img.height * scale)

                    altura_linhas = max(20, int(xl_img.height * 0.75) + 5)
                    ws2.row_dimensions[foto_row].height = altura_linhas

                    ws2.cell(foto_row, 1, f'{mov.data.strftime("%d/%m/%Y")}').font = Font(size=9, bold=True)
                    ws2.cell(foto_row, 2, mov.item.nome).font = Font(size=9)
                    ws2.cell(foto_row, 3, f'{mov.quantidade} {mov.item.unidade}').font = Font(size=9)

                    foto_row += 1
                    ws2.row_dimensions[foto_row].height = altura_linhas
                    ws2.add_image(xl_img, f'A{foto_row}')
                    foto_row += max(2, int(xl_img.height / 15)) + 1
                except Exception as e_foto:
                    logger.warning(f'Foto não inserida no Excel: {e_foto}')
                    foto_row += 1
        except ImportError:
            pass  # openpyxl.drawing não disponível — ignora silenciosamente

    buf = io.BytesIO()
    wb.save(buf); buf.seek(0)
    nome_safe = funcionario.replace(' ', '_').replace('/', '-')
    return send_file(buf, as_attachment=True,
                     download_name=f'FORM-SEG-014_{nome_safe}_{data_fim}.xlsx',
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')


@relatorios_bp.route('/movimentacoes/excluir', methods=['POST'])
@admin_required
def excluir_movimentacoes():
    ids = request.form.getlist('mov_ids')
    if not ids:
        flash('Nenhuma movimentação selecionada.', 'warning')
        return redirect(request.referrer or url_for('relatorios_bp.relatorio_consumo_pessoa'))
    
    excluidas = 0
    for mov_id in ids:
        mov = db.session.get(Movimentacao, mov_id)
        if mov:
            # Reverter o estoque ao excluir saída
            if mov.tipo == 'saida':
                mov.item.quantidade += mov.quantidade
            elif mov.tipo == 'entrada':
                mov.item.quantidade -= mov.quantidade
            db.session.delete(mov)
            excluidas += 1
    
    db.session.commit()
    flash(f'{excluidas} movimentação(ões) excluída(s) e estoque revertido.', 'success')
    return redirect(request.referrer or url_for('relatorios_bp.relatorio_consumo_pessoa'))

@relatorios_bp.route('/relatorios/alertas')
@login_required
def relatorio_alertas():
    u = usuario_atual()
    # Engenheiro (colaborador) só vê alertas se tiver permissão 'ver_alertas'
    if u.perfil == 'colaborador':
        tem_permissao = any(p.permissao == 'ver_alertas' for p in u.permissoes_extras)
        if not tem_permissao:
            flash('Sem permissão para ver alertas de estoque. Solicite ao administrador.', 'warning')
            return redirect(url_for('main_bp.index'))
    if u.perfil == 'admin':
        itens = Item.query.filter(Item.quantidade <= Item.estoque_minimo, Item.ativo == True).order_by(
            Item.fixado.desc(), Item.quantidade.asc()
        ).all()
        todos_ativos = Item.query.filter(Item.ativo == True).all()
    elif u.perfil == 'ggo':
        ids_ggo = {a.id for a in almoxarifados_do_ggo(u)}
        itens = Item.query.filter(
            Item.quantidade <= Item.estoque_minimo,
            Item.almoxarifado_id.in_(ids_ggo),
            Item.ativo == True
        ).order_by(Item.fixado.desc(), Item.quantidade.asc()).all() if ids_ggo else []
        todos_ativos = Item.query.filter(
            Item.ativo == True, Item.almoxarifado_id.in_(ids_ggo)
        ).all() if ids_ggo else []
    elif u.perfil == 'analista':
        # Analista vê apenas alertas da sua cidade
        ids_alm = set()
        if u.almoxarifado_id:
            alm_ref = db.session.get(Almoxarifado, u.almoxarifado_id)
            if alm_ref and alm_ref.cidade:
                ids_alm = {a.id for a in Almoxarifado.query.filter(
                    Almoxarifado.cidade.ilike(alm_ref.cidade)
                ).all()}
            elif alm_ref:
                ids_alm = {alm_ref.id}
        if ids_alm:
            itens = Item.query.filter(
                Item.quantidade <= Item.estoque_minimo,
                Item.almoxarifado_id.in_(ids_alm),
                Item.ativo == True
            ).order_by(Item.fixado.desc(), Item.quantidade.asc()).all()
            todos_ativos = Item.query.filter(Item.ativo == True, Item.almoxarifado_id.in_(ids_alm)).all()
        else:
            itens = []
            todos_ativos = []
    else:
        # tecnico_seguranca, almoxarife, colaborador com pode_ver_alertas
        # Para técnico: expande para todos da sua cidade
        ids = set(u.almoxarifados_permitidos())
        if u.perfil == 'tecnico_seguranca' and u.almoxarifado_id:
            alm_ref = db.session.get(Almoxarifado, u.almoxarifado_id)
            if alm_ref and alm_ref.cidade:
                ids = {a.id for a in Almoxarifado.query.filter_by(cidade=alm_ref.cidade).all()}
        itens = Item.query.filter(
            Item.quantidade <= Item.estoque_minimo,
            Item.almoxarifado_id.in_(ids),
            Item.ativo == True
        ).order_by(Item.fixado.desc(), Item.quantidade.asc()).all() if ids else []
        todos_ativos = Item.query.filter(
            Item.ativo == True, Item.almoxarifado_id.in_(ids)
        ).all() if ids else []

    ruptura = calcular_ruptura(todos_ativos, limite_dias=None)
    ruptura_por_item = {r['item'].id: r for r in ruptura}
    return render_template('relatorio_alertas.html', itens=itens, ruptura=ruptura,
                           ruptura_por_item=ruptura_por_item)
