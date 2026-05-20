import os
import io
import base64
import logging
from datetime import datetime, date
from flask import current_app
from . import db
from .models import Usuario, Almoxarifado, Item
from .utils import agora
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

logger = logging.getLogger(__name__)


def gerar_excel_backup_almoxarifado(alm):
    wb = openpyxl.Workbook()
    h_fill = PatternFill('solid', fgColor='1A3A5C')
    h_font = Font(bold=True, color='FFFFFF', size=11)
    ok_fill = PatternFill('solid', fgColor='D4EDDA')
    al_fill = PatternFill('solid', fgColor='FFF3CD')
    cr_fill = PatternFill('solid', fgColor='F8D7DA')
    borda = Border(left=Side(style='thin'), right=Side(style='thin'),
                   top=Side(style='thin'), bottom=Side(style='thin'))

    itens = Item.query.filter_by(almoxarifado_id=alm.id).all()
    ws = wb.active
    ws.title = alm.nome[:31]

    ws.merge_cells('A1:G1')
    ws['A1'] = f'Backup — {alm.nome}'
    ws['A1'].font = Font(bold=True, size=13, color='1A3A5C')
    ws['A1'].alignment = Alignment(horizontal='center')
    ws.merge_cells('A2:G2')
    ws['A2'] = f'Gerado em: {datetime.now().strftime("%d/%m/%Y %H:%M")}'
    ws['A2'].font = Font(italic=True, color='888888')
    ws['A2'].alignment = Alignment(horizontal='center')

    headers = ['Codigo', 'Item', 'Unidade', 'Qtd. Atual', 'Est. Minimo', 'Status', 'Ativo']
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=4, column=col, value=h)
        c.font = h_font; c.fill = h_fill
        c.alignment = Alignment(horizontal='center'); c.border = borda

    for r, it in enumerate(itens, 5):
        status = 'ZERADO' if it.status == 'critico' else ('ABAIXO DO MINIMO' if it.status == 'alerta' else 'OK')
        fill = cr_fill if it.status == 'critico' else (al_fill if it.status == 'alerta' else ok_fill)
        for c, v in enumerate([it.codigo, it.nome, it.unidade, it.quantidade,
                                it.estoque_minimo, status, 'Sim' if it.ativo else 'Não'], 1):
            cell = ws.cell(row=r, column=c, value=v)
            cell.fill = fill; cell.border = borda
            cell.alignment = Alignment(horizontal='left' if c == 2 else 'center')

    for i, w in enumerate([14, 40, 10, 12, 14, 20, 8], 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


def gerar_excel_backup():
    wb = openpyxl.Workbook()
    h_fill = PatternFill('solid', fgColor='1A3A5C')
    h_font = Font(bold=True, color='FFFFFF', size=11)
    ok_fill = PatternFill('solid', fgColor='D4EDDA')
    al_fill = PatternFill('solid', fgColor='FFF3CD')
    cr_fill = PatternFill('solid', fgColor='F8D7DA')
    borda = Border(left=Side(style='thin'), right=Side(style='thin'),
                   top=Side(style='thin'), bottom=Side(style='thin'))

    almoxarifados = Almoxarifado.query.all()
    primeira = True

    for alm in almoxarifados:
        itens = Item.query.filter_by(almoxarifado_id=alm.id).all()
        ws = wb.active if primeira else wb.create_sheet()
        primeira = False
        ws.title = alm.nome[:31]

        ws.merge_cells('A1:G1')
        ws['A1'] = f'Backup — {alm.nome}'
        ws['A1'].font = Font(bold=True, size=13, color='1A3A5C')
        ws['A1'].alignment = Alignment(horizontal='center')
        ws.merge_cells('A2:G2')
        ws['A2'] = f'Gerado em: {datetime.now().strftime("%d/%m/%Y %H:%M")}'
        ws['A2'].font = Font(italic=True, color='888888')
        ws['A2'].alignment = Alignment(horizontal='center')

        headers = ['Codigo', 'Item', 'Unidade', 'Qtd. Atual', 'Est. Minimo', 'Status', 'Ativo']
        for col, h in enumerate(headers, 1):
            c = ws.cell(row=4, column=col, value=h)
            c.font = h_font; c.fill = h_fill
            c.alignment = Alignment(horizontal='center'); c.border = borda

        for r, it in enumerate(itens, 5):
            status = 'ZERADO' if it.status == 'critico' else ('ABAIXO DO MINIMO' if it.status == 'alerta' else 'OK')
            fill = cr_fill if it.status == 'critico' else (al_fill if it.status == 'alerta' else ok_fill)
            for c, v in enumerate([it.codigo, it.nome, it.unidade, it.quantidade,
                                    it.estoque_minimo, status, 'Sim' if it.ativo else 'Não'], 1):
                cell = ws.cell(row=r, column=c, value=v)
                cell.fill = fill; cell.border = borda
                cell.alignment = Alignment(horizontal='left' if c == 2 else 'center')

        for i, w in enumerate([14, 40, 10, 12, 14, 20, 8], 1):
            ws.column_dimensions[get_column_letter(i)].width = w

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


def enviar_backup_email(buf):
    import smtplib
    from email.mime.base import MIMEBase
    from email.mime.multipart import MIMEMultipart
    from email.mime.text import MIMEText
    from email import encoders

    remetente = os.environ.get('BACKUP_EMAIL_FROM')
    senha = os.environ.get('BACKUP_EMAIL_PASS')
    destinatario = os.environ.get('BACKUP_EMAIL_TO', 'rickgouveia17@gmail.com')

    if not remetente or not senha:
        logger.info('BACKUP: variáveis BACKUP_EMAIL_FROM e BACKUP_EMAIL_PASS não configuradas.')
        return False

    msg = MIMEMultipart()
    msg['From'] = remetente
    msg['To'] = destinatario
    msg['Subject'] = f'Backup Estoque Obra Patamares — {date.today().strftime("%d/%m/%Y")}'

    corpo = f"""
    Backup automático do sistema de estoque.
    Data: {datetime.now().strftime('%d/%m/%Y %H:%M')}
    
    Este arquivo contém todos os dados de estoque de todos os almoxarifados.
    Guarde em local seguro.
    """
    msg.attach(MIMEText(corpo, 'plain'))

    part = MIMEBase('application', 'octet-stream')
    part.set_payload(buf.read())
    encoders.encode_base64(part)
    part.add_header('Content-Disposition',
                    f'attachment; filename="backup_estoque_{date.today()}.xlsx"')
    msg.attach(part)

    try:
        with smtplib.SMTP_SSL('smtp.gmail.com', 465) as smtp:
            smtp.login(remetente, senha)
            smtp.send_message(msg)
        logger.info(f'BACKUP: enviado para {destinatario}')
        return True
    except Exception as e:
        logger.info(f'BACKUP: erro ao enviar email — {e}')
        return False


def _smtp_connect():
    import smtplib
    remetente = os.environ.get('BACKUP_EMAIL_FROM')
    senha = os.environ.get('BACKUP_EMAIL_PASS')
    if not remetente or not senha:
        return None, None, None
    smtp = smtplib.SMTP_SSL('smtp.gmail.com', 465)
    smtp.login(remetente, senha)
    return smtp, remetente, senha


def enviar_backup_por_almoxarifado():
    import resend

    resend_api_key = os.environ.get('RESEND_API_KEY')
    destinatario_fixo = os.environ.get('BACKUP_EMAIL_TO', 'rickgouveia157@gmail.com')
    if not resend_api_key:
        logger.info('BACKUP: variável RESEND_API_KEY não configurada.')
        return False, 'Variável RESEND_API_KEY não configurada no Railway.'

    resend.api_key = resend_api_key
    remetente = os.environ.get('RESEND_FROM_EMAIL', 'onboarding@resend.dev')
    hoje = date.today().strftime('%d/%m/%Y')
    enviados = 0
    erros = 0

    try:
        admins_emails = [u.email for u in Usuario.query.filter_by(perfil='admin', ativo=True).all() if u.email]
        almoxarifes_emails = [u.email for u in Usuario.query.filter_by(perfil='almoxarife', ativo=True).all() if u.email]
        cc_completo = list(set(admins_emails + almoxarifes_emails) - {destinatario_fixo})

        buf_completo = gerar_excel_backup()
        buf_completo.seek(0)
        arquivo_completo_base64 = base64.b64encode(buf_completo.read()).decode('utf-8')

        payload = {
            'from': f'Logi-Prime Backup <{remetente}>',
            'to': [destinatario_fixo],
            'subject': f'Backup Completo Estoque — {hoje}',
            'text': f'Backup automático completo.\nData: {datetime.now().strftime("%d/%m/%Y %H:%M")}\n\nContém todos os almoxarifados.',
            'attachments': [{'filename': f'backup_completo_{date.today()}.xlsx', 'content': arquivo_completo_base64}]
        }
        if cc_completo:
            payload['cc'] = cc_completo
        try:
            resend.Emails.send(payload)
            logger.info(f'BACKUP: completo enviado para {[destinatario_fixo] + cc_completo}')
            enviados += 1
        except Exception as e:
            logger.info(f'BACKUP: erro backup completo — {e}')
            erros += 1

        for alm in Almoxarifado.query.all():
            cc_alm = [u.email for u in alm.usuarios if u.perfil == 'almoxarife' and u.ativo and u.email and u.email != destinatario_fixo]
            if not cc_alm:
                continue
            buf_alm = gerar_excel_backup_almoxarifado(alm)
            buf_alm.seek(0)
            payload_alm = {
                'from': f'Logi-Prime Backup <{remetente}>',
                'to': [destinatario_fixo],
                'cc': cc_alm,
                'subject': f'Backup {alm.nome} — {hoje}',
                'text': f'Backup do almoxarifado: {alm.nome}\nData: {datetime.now().strftime("%d/%m/%Y %H:%M")}',
                'attachments': [{'filename': f'backup_{alm.nome.replace(" ", "_")}_{date.today()}.xlsx', 'content': base64.b64encode(buf_alm.read()).decode('utf-8')}]
            }
            try:
                resend.Emails.send(payload_alm)
                logger.info(f'BACKUP: "{alm.nome}" enviado cc={cc_alm}')
                enviados += 1
            except Exception as e:
                logger.info(f'BACKUP: erro "{alm.nome}" — {e}')
                erros += 1

    except Exception as e:
        logger.info(f'BACKUP: erro geral — {e}')
        return False, str(e)

    return erros == 0, None


def job_backup_diario():
    agora_str = datetime.now().strftime("%d/%m/%Y %H:%M")
    logger.info(f'BACKUP AUTOMÁTICO: iniciando às {agora_str}')

    remetente = os.environ.get('BACKUP_EMAIL_FROM', '')
    senha = os.environ.get('BACKUP_EMAIL_PASS', '')
    destino_fixo = os.environ.get('BACKUP_EMAIL_TO', 'rickgouveia17@gmail.com')
    logger.info(f'BACKUP AUTOMÁTICO: remetente configurado = {"SIM (" + remetente + ")" if remetente else "NÃO — BACKUP_EMAIL_FROM não definido"}')
    logger.info(f'BACKUP AUTOMÁTICO: senha configurada = {"SIM" if senha else "NÃO — BACKUP_EMAIL_PASS não definido"}')
    logger.info(f'BACKUP AUTOMÁTICO: destino fixo = {destino_fixo}')

    with current_app.app_context():
        try:
            admins = [u for u in Usuario.query.filter_by(perfil='admin', ativo=True).all() if u.email]
            almoxarifes = [u for u in Usuario.query.filter_by(perfil='almoxarife', ativo=True).all() if u.email]
            logger.info(f'BACKUP AUTOMÁTICO: admins com email = {[u.email for u in admins]}')
            logger.info(f'BACKUP AUTOMÁTICO: almoxarifes com email = {[u.email for u in almoxarifes]}')

            ok, erro_msg = enviar_backup_por_almoxarifado()
            if ok:
                logger.info('BACKUP AUTOMÁTICO: ✅ enviado com sucesso!')
            else:
                logger.info(f'BACKUP AUTOMÁTICO: ❌ falha no envio. Detalhe: {erro_msg}')
        except Exception as e:
            logger.info(f'BACKUP AUTOMÁTICO: erro — {e}')


def schedule_backup(app):
    try:
        from apscheduler.schedulers.background import BackgroundScheduler
        from apscheduler.triggers.cron import CronTrigger

        _already_started = os.environ.get('_SCHEDULER_STARTED', '')
        if not _already_started:
            os.environ['_SCHEDULER_STARTED'] = '1'
            scheduler = BackgroundScheduler(timezone='America/Sao_Paulo')
            scheduler.add_job(
                job_backup_diario,
                CronTrigger(hour=20, minute=0, timezone='America/Sao_Paulo'),
                id='backup_diario',
                replace_existing=True
            )
            scheduler.start()
            logger.info('BACKUP AUTOMÁTICO: ✅ agendado para todo dia às 20:00 (Brasília)')
        else:
            logger.info('BACKUP AUTOMÁTICO: scheduler já iniciado, ignorando.')
    except Exception as e:
        logger.error(f'BACKUP AUTOMÁTICO: ❌ erro ao iniciar agendador — {e}')
