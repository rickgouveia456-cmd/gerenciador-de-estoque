admin_bp = Blueprint('admin_bp', __name__)

from flask import Blueprint, render_template, request, redirect, url_for, jsonify, flash, send_file, session
from markupsafe import Markup, escape
from datetime import datetime, date, timedelta
import io, os, json, re, logging
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from extensions import db
from models import (agora, Almoxarifado, Item, Movimentacao, Requisicao,
    RequisicaoMestre, RequisicaoMestreItem, Usuario, Colaborador,
    Ferramenta, HistoricoFerramenta, ItemEPI, HistoricoEPI,
    AcessoExtra, PermissaoExtra)
from core import (login_required, admin_required, almoxarife_required,
    usuario_atual, flash_html, usuario_tem_acesso_almoxarifado,
    usuario_tem_acesso_item, PERMISSOES_DISPONIVEIS,
    _check_rate_limit, _register_attempt, _clear_attempts, _check_api_rate)

logger = logging.getLogger(__name__)

@admin_bp.route('/admin/reativar-todos-itens', methods=['GET', 'POST'])
@admin_required
def reativar_todos_itens():
    if request.method == 'POST':
        try:
            # Reativar todos os itens
            itens_desativados = Item.query.filter_by(ativo=False).all()
            count = 0
            for item in itens_desativados:
                item.ativo = True
                count += 1
            
            # Também garantir que itens com ativo=None sejam ativados
            from sqlalchemy import text
            with db.engine.connect() as conn:
                result = conn.execute(text("UPDATE item SET ativo = 1 WHERE ativo IS NULL OR ativo = 0"))
                conn.commit()
            
            db.session.commit()
            
            total_ativos = Item.query.filter_by(ativo=True).count()
            flash(f'✅ Sucesso! {count} itens reativados. Total de itens ativos: {total_ativos}', 'success')
            
        except Exception as e:
            flash(f'❌ Erro ao reativar itens: {str(e)}', 'danger')
        
        return redirect(url_for('reativar_todos_itens'))
    
    # GET - mostrar página de confirmação
    itens_desativados = Item.query.filter_by(ativo=False).count()
    total_itens = Item.query.count()
    return render_template('admin_reativar_itens.html',
                           itens_desativados=itens_desativados,
                           total_itens=total_itens)

# ── BACKUP ───────────────────────────────────────────────────────────────────


def gerar_excel_backup_almoxarifado(alm):
    """Gera um Excel com apenas um almoxarifado (para envio ao almoxarife)."""
    wb = openpyxl.Workbook()
    h_fill = PatternFill('solid', fgColor='1A3A5C')
    h_font = Font(bold=True, color='FFFFFF', size=11)
    ok_fill  = PatternFill('solid', fgColor='D4EDDA')
    al_fill  = PatternFill('solid', fgColor='FFF3CD')
    cr_fill  = PatternFill('solid', fgColor='F8D7DA')
    borda = Border(left=Side(style='thin'), right=Side(style='thin'),
                   top=Side(style='thin'), bottom=Side(style='thin'))

    itens = Item.query.filter_by(almoxarifado_id=alm.id).all()
    ws = wb.active
    ws.title = alm.nome[:31]

    ws.merge_cells('A1:G1')
    ws['A1'] = f'Backup — {alm.nome}'
    ws['A1'].font = Font(bold=True, size=13, color='1A3A5C')
    ws['A1'].alignment = Alignment(horizontal='center')
    ws.merge_cells('A2:G2')
    ws['A2'] = f'Gerado em: {datetime.now().strftime("%d/%m/%Y %H:%M")}'
    ws['A2'].font = Font(italic=True, color='888888')
    ws['A2'].alignment = Alignment(horizontal='center')

    headers = ['Codigo', 'Item', 'Unidade', 'Qtd. Atual', 'Est. Minimo', 'Status', 'Ativo']
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=4, column=col, value=h)
        c.font = h_font; c.fill = h_fill
        c.alignment = Alignment(horizontal='center'); c.border = borda

    for r, it in enumerate(itens, 5):
        status = 'ZERADO' if it.status == 'critico' else ('ABAIXO DO MINIMO' if it.status == 'alerta' else 'OK')
        fill = cr_fill if it.status == 'critico' else (al_fill if it.status == 'alerta' else ok_fill)
        for c, v in enumerate([it.codigo, it.nome, it.unidade, it.quantidade,
                                it.estoque_minimo, status, 'Sim' if it.ativo else 'Não'], 1):
            cell = ws.cell(row=r, column=c, value=v)
            cell.fill = fill; cell.border = borda
            cell.alignment = Alignment(horizontal='left' if c == 2 else 'center')

    for i, w in enumerate([14, 40, 10, 12, 14, 20, 8], 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf

def gerar_excel_backup():
    """Gera um Excel completo com todos os almoxarifados como backup."""
    wb = openpyxl.Workbook()
    h_fill = PatternFill('solid', fgColor='1A3A5C')
    h_font = Font(bold=True, color='FFFFFF', size=11)
    ok_fill  = PatternFill('solid', fgColor='D4EDDA')
    al_fill  = PatternFill('solid', fgColor='FFF3CD')
    cr_fill  = PatternFill('solid', fgColor='F8D7DA')
    borda = Border(left=Side(style='thin'), right=Side(style='thin'),
                   top=Side(style='thin'), bottom=Side(style='thin'))

    almoxarifados = Almoxarifado.query.all()
    primeira = True

    for alm in almoxarifados:
        itens = Item.query.filter_by(almoxarifado_id=alm.id).all()
        ws = wb.active if primeira else wb.create_sheet()
        primeira = False
        ws.title = alm.nome[:31]  # Excel limita 31 chars no nome da aba

        # Título
        ws.merge_cells('A1:G1')
        ws['A1'] = f'Backup — {alm.nome}'
        ws['A1'].font = Font(bold=True, size=13, color='1A3A5C')
        ws['A1'].alignment = Alignment(horizontal='center')
        ws.merge_cells('A2:G2')
        ws['A2'] = f'Gerado em: {datetime.now().strftime("%d/%m/%Y %H:%M")}'
        ws['A2'].font = Font(italic=True, color='888888')
        ws['A2'].alignment = Alignment(horizontal='center')

        # Cabeçalho
        headers = ['Codigo', 'Item', 'Unidade', 'Qtd. Atual', 'Est. Minimo', 'Status', 'Ativo']
        for col, h in enumerate(headers, 1):
            c = ws.cell(row=4, column=col, value=h)
            c.font = h_font; c.fill = h_fill
            c.alignment = Alignment(horizontal='center'); c.border = borda

        for r, it in enumerate(itens, 5):
            status = 'ZERADO' if it.status == 'critico' else ('ABAIXO DO MINIMO' if it.status == 'alerta' else 'OK')
            fill = cr_fill if it.status == 'critico' else (al_fill if it.status == 'alerta' else ok_fill)
            for c, v in enumerate([it.codigo, it.nome, it.unidade, it.quantidade,
                                    it.estoque_minimo, status, 'Sim' if it.ativo else 'Não'], 1):
                cell = ws.cell(row=r, column=c, value=v)
                cell.fill = fill; cell.border = borda
                cell.alignment = Alignment(horizontal='left' if c == 2 else 'center')

        for i, w in enumerate([14, 40, 10, 12, 14, 20, 8], 1):
            ws.column_dimensions[get_column_letter(i)].width = w

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf




@admin_bp.route('/admin/seed-colaboradores-infra', methods=['POST'])
@admin_required
def seed_colaboradores_infra():
    """Cadastra os colaboradores de infraestrutura da QLP — ignora duplicatas."""
    _lista = [
        ("ADAILTON JOSE DOS SANTOS", "Ajudante"),
        ("ALAN SOUZA DE OLIVEIRA", "Ajudante"),
        ("ANDRE DE JESUS MENDES", "Ajudante"),
        ("ANTONIO LUCAS NASCIMENTO BISPO", "Assistente de Produção"),
        ("ARODOALDO PEREIRA DA ROCHA", "Encanador"),
        ("CARLOS ALBERTO LOPES SILVA", "Ajudante"),
        ("CLAUDEMIRO GALVÃO DOS SANTOS", "Auxiliar de Serviços Gerais"),
        ("DEYVID DE SANTANA LOPES", "Almoxarife"),
        ("EDINEILSON DOS SANTOS DE OLIVEIRA", "Eletricista"),
        ("EDNILSON ASSIS DOS SANTOS", "Ajudante"),
        ("EDSON ANTONIO SANTOS DE OLIVEIRA", "Ajudante"),
        ("GERSON SILVA", "Mestre de Obras"),
        ("HEBERT DA SILVA MEDRADO", "Ajudante"),
        ("JACKSON DA SILVA MENEZES DOS SANTOS", "Ajudante"),
        ("JEFFERSON SANTOS RIBEIRO", "Ajudante"),
        ("JOSE SEVERINO MENDES DA SILVA", "Carpinteiro"),
        ("LAURA DOS SANTOS ARAÚJO", "Coordenadora"),
        ("LEONARDO VIDAL DOS SANTOS SENA", "Técnico de Segurança"),
        ("MARCOS VINICIUS SAMPAIO ROSA", "Ajudante"),
        ("MATEUS DE JESUS SANTANA SANTOS", "Ajudante"),
        ("RAILAN NASCIMENTO SANTOS", "Ajudante"),
        ("RODRIGO NASCIMENTO SANTANA", "Ajudante"),
        ("TIAGO DA SILVA FERREIRA", "Auxiliar de Serviços Gerais"),
        ("JOÃO FRANCISCO C DE JESUS RODRIGUES", "Ajudante Prático de Elétrica"),
        ("JADSON DIAS DE OLIVEIRA SOUSA", "Ajudante Comum"),
        ("JOSÉ AUGUSTO DOS SANTOS BISPO", "Ajudante Comum"),
        ("WALTER BATISTA DOS SANTOS FILHO", "Carpinteiro"),
    ]
    inseridos = 0
    for nome, funcao in _lista:
        if not Colaborador.query.filter(Colaborador.nome.ilike(nome)).first():
            db.session.add(Colaborador(nome=nome, funcao=funcao, escopo='infraestrutura', ativo=True))
            inseridos += 1
    db.session.commit()
    flash(f'✅ {inseridos} colaboradores de infraestrutura cadastrados!', 'success')
    return redirect(url_for('colaboradores'))


@admin_bp.route('/admin/seed-colaboradores', methods=['POST'])
@admin_required
def seed_colaboradores():
    """Cadastra os colaboradores da estrutura em massa — ignora duplicatas."""
    _lista = [
        ("ARILSON DE JESUS SOUZA","Profissional"),("ADSON MUNIZ","Ajudante"),
        ("MARCEL OLIVEIRA DA CONCEIÇÃO","Profissional"),("ROBERT WILLIAM DA HORA DE JESUS","Profissional"),
        ("MATEUS SANTOS DE JESUS","Profissional"),("ANIBAL SANTOS DANTAS","Profissional"),
        ("ROBERTO FELIX GONÇALVES","Profissional"),("EDNALDO DOS SANTOS","Profissional"),
        ("ALEXSANDRO TELES DOS SANTOS","Profissional"),("RUAN UITALO","Ajudante"),
        ("FELIPE MESSIAS","Ajudante"),("ROQUE DOS SANTOS","Profissional"),
        ("VALDOMIRO GOMES DE JESUS FILHO","Profissional"),("VALMIR GOMES DE JESUS","Profissional"),
        ("TIAGO GOMES DOS SANTOS","Ajudante"),("ADRIANO SOUZA DOS SANTOS","Profissional"),
        ("RONALDO DA CUNHA SANTOS","Profissional"),("EMERSON DE SANTANA ARAUJO","Profissional"),
        ("LUAN DOS SANTOS CARDOSO","Profissional"),("FRANCISCO CARLOS DOS SANTOS FILHO","Profissional"),
        ("NILSON MATIAS DOS SANTOS","Profissional"),("VINICIUS DANTAS DA SILVA","Profissional"),
        ("MAURICIO RAMON PINHEIRO MATOS","Ajudante"),("CARLOS ALBERTO BISPO DOS SANTOS","Ajudante"),
        ("DIEGO LIMA SANTOS","Ajudante"),("AILTON DA SILVA","Profissional"),
        ("EIDSON SILVA ROCHA","Ajudante"),("MARLEI ASSIS DE SOUZA","Profissional"),
        ("RICARDO VASQUES LEMOS LEONI","Profissional"),("ROBISON SANTOS DA CONCEIÇÃO","Profissional"),
        ("LUIS SILVAN LOPES DOS SANTOS","Profissional"),("JAIR CESAR BRITO RODRIGUES JUNIOR","Ajudante"),
        ("ROBSON BISPO DOS SANTOS","Profissional"),("EDVAN MACHADO SANTOS","Profissional"),
        ("LUIS ALBERTO MOREIRA DA SILVA","Ajudante"),("ISAAC GONÇALVES DA SILVA","Ajudante"),
        ("ANTONIO MARCOS DA SILVA COSTA","Ajudante"),("DENAILTON LEITE DOS SANTOS","Ajudante"),
        ("MARCIO DE JESUS DOS SANTOS","Profissional"),("WEBER OLIVEIRA DA LUZ","Ajudante"),
        ("RAFAEL DA SILVA BOMFIM","Ajudante"),("ANDERSON RODRIGUES DOS SANTOS","Profissional"),
        ("JAILTON RIBEIRO TOSTA","Profissional"),("JOAO LUIS OLIVEIRA DA SILVA","Profissional"),
        ("ANDERSON SOUZA DE FRIAS","Profissional"),("JOILSON DOS SANTOS","Profissional"),
        ("JOANDERSON ALMEIDA BISPO","Profissional"),("ANTONIO CARLOS SANTOS SILVA","Profissional"),
        ("NAILTON CONCEIÇÃO DE SOUZA","Ajudante"),("LUCAS SILVA DOS REIS","Ajudante"),
        ("ROBSON LIMA MACIEL","Profissional"),("ATILA ALMEIDA SILVA SANTOS","Ajudante"),
        ("JONAS DE SENA BARRETO","Ajudante"),("SAMUEL BISPO DOS SANTOS","Profissional"),
        ("GUILHERME SANTOS SAMPAIO","Ajudante"),("DANIEL SÃO PEDRO DOS SANTOS","Ajudante"),
        ("DIVINO CARDOSO DOS SANTOS","Ajudante"),("ANDERSON CONCEIÇÃO DE JESUS","Profissional"),
        ("ALEX DE JESUS DA SILVA","Profissional"),("ALEX VITÓRIO SILVA","Ajudante"),
        ("CARLOS DANIEL DA SILVA MARQUES","Ajudante"),("THIEGO DE OLIVEIRA REIS","Profissional"),
        ("UBIRATTAN SNATOS SOUZA","Ajudante"),("WALISSON SILVA COSTA","Ajudante"),
        ("VALMIR GONÇALVES DE OLIVEIRA","Profissional"),("JUDICAEL LEITE DOS SANTOS","Profissional"),
        ("JOÃO PEDRO SILVA DOS SANTOS","Profissional"),("JORGE DOS SANTOS","Profissional"),
        ("JEAN AUGUSTO DOS SANTOS TAVARES","Profissional"),
    ]
    inseridos = 0
    for nome, funcao in _lista:
        if not Colaborador.query.filter(Colaborador.nome.ilike(nome)).first():
            db.session.add(Colaborador(nome=nome, funcao=funcao, escopo='estrutura', ativo=True))
            inseridos += 1
    db.session.commit()
    flash(f'✅ {inseridos} colaboradores cadastrados da estrutura!', 'success')
    return redirect(url_for('colaboradores'))

@admin_bp.route('/admin/seed-ferramentas-estrutura', methods=['POST'])
@admin_required
def seed_ferramentas_estrutura():
    """Cadastra as ferramentas da Estrutura Ventura Patamares — ignora duplicatas pelo ID."""
    # Busca o almoxarifado de Estrutura
    alm = Almoxarifado.query.filter(Almoxarifado.nome.ilike('%estrutura%')).first()
    if not alm:
        flash('Almoxarifado de Estrutura não encontrado.', 'danger')
        return redirect(url_for('index'))

    ferramentas_lista = [
        ("INGFH007",  "PISTOLA DE FIXACAO A BATERIA BX 3-L A22MA HILTI GF",   "HILTI"),
        ("IN450348",  "MARTELO SDS PLUS C/ PUNHO",                              ""),
        ("INMRM158",  "MARTELO ROMPEDOR MMR1700 45J 12KG MONO 220V 60HZ",      "MENEGOTTI"),
        ("IN580121",  "MISTURADOR DE ARGAMASSA MAV1600 220V",                   "MENEGOTTI"),
        ("IN580039",  "MISTURADOR DE ARGAMASSA MAV1600 220V",                   "MENEGOTTI"),
        ("INGFH003",  "PISTOLA DE FIXACAO A BATERIA BX 3-L A22MA HILTI GF",    "HILTI"),
        ("INEAG119",  "ESMERILHADEIRA ANGULAR 5\" C/ ACESSORIOS",               "MAKITA"),
        ("IN270013",  "FURADEIRA 3/4 FUR3/4P",                                  ""),
        ("IN270028",  "FURADEIRA 3/4 FUR3/4P",                                  ""),
        ("IN270004",  "FURADEIRA 3/4 FUR3/4P",                                  ""),
        ("IN270005",  "FURADEIRA 3/4 FUR3/4P",                                  ""),
        ("INFPB017",  "FURADEIRA E PARAFUSADEIRA A BATERIA MFI-20 127/220V",    "MENEGOTTI"),
        ("IN340056",  "LAVA JATO HD 585 PROFISSIONAL MODELO 585",               ""),
        ("INMSV108",  "MARTELETE PERF/ROMP MPV1500 5,5J 220V",                  "VONDER"),
        ("INHT930012","KIT ASPIRADOR UNIV. HILTI / POLIDORA DE BETAO HILTI",    "HILTI"),
        ("INMSV261",  "MARTELETE PERF/ROMP MPV1500 5,5J 220V",                  "VONDER"),
        ("INMSV269",  "MARTELETE PERF/ROMP MPV1500 5,5J 220V",                  "VONDER"),
        ("IN240093",  "ESMERILHADEIRA ANGULAR 7\" 220V",                         ""),
        ("INMSV256",  "MARTELETE PERF/ROMP MPV1500 5,5J 220V",                  "VONDER"),
        ("INMSV257",  "MARTELETE PERF/ROMP MPV1500 5,5J 220V",                  "VONDER"),
        ("INMSU042",  "MISTURADOR ELETRICO MEL1600 MONO 220V 60HZ 1600W",       "MENEGOTTI"),
        ("INSER830025","SERRA CIRCULAR 7\"",                                      ""),
        ("INEAG243",  "ESMERILHADEIRA ANGULAR 5\" C/ ACESSORIOS",               "MAKITA"),
        ("INEAG233",  "ESMERILHADEIRA ANGULAR 5\" C/ ACESSORIOS",               "MAKITA"),
        ("INEAG236",  "ESMERILHADEIRA ANGULAR 5\" C/ ACESSORIOS",               "MAKITA"),
        ("INMRM149",  "MARTELO ROMPEDOR MMR1700 45J 12KG MONO 220V 60HZ",      "MENEGOTTI"),
        ("INMSV069",  "MARTELETE PERF/ROMP MPV1500 5,5J 220V",                  "VONDER"),
        ("INMSV156",  "MARTELETE PERF/ROMP MPV1500 5,5J 220V",                  "VONDER"),
        ("IN620015",  "KIT NIVELADOR A LASER HILTI SKR200",                     "HILTI"),
        ("INMSV067",  "MARTELETE PERF/ROMP MPV1500 5,5J 220V",                  "VONDER"),
        ("INMSV001",  "MARTELETE PERF/ROMP MPV1500 5,5J 220V",                  "VONDER"),
        ("INMSV099",  "MARTELETE PERF/ROMP MPV1500 5,5J 220V",                  "VONDER"),
        ("INMSV273",  "MARTELETE PERF/ROMP MPV1500 5,5J 220V",                  "VONDER"),
        ("INHT930015","KIT ASPIRADOR UNIV. HILTI / POLIDORA DE BETAO HILTI",    "HILTI"),
        ("INMSV024",  "MARTELETE PERF/ROMP MPV1500 5,5J 220V",                  "VONDER"),
        ("INMSV113",  "MARTELETE PERF/ROMP MPV1500 5,5J 220V",                  "VONDER"),
        ("INMSV131",  "MARTELETE PERF/ROMP MPV1500 5,5J 220V",                  "VONDER"),
        ("INMSV278",  "MARTELETE PERF/ROMP MPV1500 5,5J 220V",                  "VONDER"),
        ("INMSV148",  "MARTELETE PERF/ROMP MPV1500 5,5J 220V",                  "VONDER"),
        ("INHT930010","KIT ASPIRADOR UNIV. HILTI / POLIDORA DE BETAO HILTI",    "HILTI"),
        ("INHT930011","KIT ASPIRADOR UNIV. HILTI / POLIDORA DE BETAO HILTI",    "HILTI"),
        ("INMSV090",  "MARTELETE PERF/ROMP MPV1500 5,5J 220V",                  "VONDER"),
        ("INMSV281",  "MARTELETE PERF/ROMP MPV1500 5,5J 220V",                  "VONDER"),
        ("INMSV260",  "MARTELETE PERF/ROMP MPV1500 5,5J 220V",                  "VONDER"),
        ("IN850474",  "SERRA MARMORE C/ CHAVE",                                  ""),
    ]

    inseridas = 0
    ignoradas = 0
    for idf, nome, empresa in ferramentas_lista:
        existe = Ferramenta.query.filter_by(identificacao=idf, ativo=True).first()
        if existe:
            ignoradas += 1
            continue
        db.session.add(Ferramenta(
            identificacao=idf,
            nome=nome,
            empresa=empresa or None,
            almoxarifado_id=alm.id,
            status='disponivel'
        ))
        inseridas += 1

    db.session.commit()
    flash(f'✅ {inseridas} ferramentas cadastradas no {alm.nome}! ({ignoradas} já existiam)', 'success')
    return redirect(url_for('ferramentas', alm_id=alm.id))

@admin_bp.route('/admin/classificar-epis', methods=['POST'])
@admin_required
def classificar_epis():
    """Classifica automaticamente os itens conhecidos como EPI/Maquinário no banco."""
    # Palavras-chave que identificam EPIs
    palavras_epi = [
        'bota', 'capacete', 'carneira', 'cinto de segurança', 'capa de chuva',
        'calça brim', 'camisa brim', 'macacão', 'mascara', 'máscara',
        'luva vaqueta', 'luva flextactil', 'perneira', 'protetor auricular',
        'óculos de proteção', 'óculos de segurança', 'óculos de sobrepor',
        'talabarte', 'trava-quedas', 'mosquetão oval', 'cinto paraquedista',
        'epi', 'uniforme', 'colete', 'capacete'
    ]
    # Palavras-chave que identificam Maquinário/Peça
    palavras_maq = [
        'broca diamantada', 'disco diamantado', 'disco de desbaste',
        'maçarico', 'perfuratriz', 'abrasiva'
    ]

    atualizados_epi = 0
    atualizados_maq = 0
    itens = Item.query.all()
    for it in itens:
        nome_lower = it.nome.lower()
        if any(p in nome_lower for p in palavras_epi):
            if it.categoria != 'epi':
                it.categoria = 'epi'
                atualizados_epi += 1
        elif any(p in nome_lower for p in palavras_maq):
            if it.categoria != 'maquinario':
                it.categoria = 'maquinario'
                atualizados_maq += 1

    db.session.commit()
    flash(f'✅ Classificação concluída: {atualizados_epi} EPIs e {atualizados_maq} Maquinários atualizados.', 'success')
    return redirect(url_for('index'))


@admin_bp.route('/admin/debug-env')
@admin_required
def debug_env():
    """Rota de diagnóstico — apenas mostra quais variáveis estão definidas, sem expor valores."""
    variaveis = ['BACKUP_EMAIL_FROM', 'BACKUP_EMAIL_PASS', 'BACKUP_EMAIL_TO',
                 'RESEND_API_KEY', 'RESEND_FROM_EMAIL', 'SECRET_KEY', 'DATABASE_URL']
    status = {v: '✅ Definida' if os.environ.get(v) else '❌ Não definida' for v in variaveis}
    linhas = '\n'.join(f'  {k} = {v}' for k, v in status.items())
    return f'<pre style="font-family:monospace;padding:20px">\nVariáveis de ambiente:\n\n{linhas}\n</pre>'

@admin_bp.route('/admin/backup', methods=['GET', 'POST'])
@admin_required
def backup_manual():
    """Admin faz backup manual — baixa Excel."""
    if request.method == 'POST':
        # Download direto
        try:
            buf = gerar_excel_backup()
            nome = f"backup_estoque_{date.today()}.xlsx"
            return send_file(buf, as_attachment=True, download_name=nome,
                             mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        except Exception as e:
            flash(f'❌ Erro ao gerar backup: {str(e)}', 'danger')
            return redirect(url_for('backup_manual'))

    # GET — mostra a página
    return render_template('backup.html')