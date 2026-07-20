from flask import Flask, render_template, request, redirect, url_for, jsonify, flash, send_file, session
from flask_sqlalchemy import SQLAlchemy
from flask_wtf.csrf import CSRFProtect, CSRFError
from werkzeug.security import generate_password_hash, check_password_hash
from markupsafe import Markup, escape
from datetime import datetime, date, timezone, timedelta
from functools import wraps
import re
import io
import os
import json
import secrets
import logging
import openpyxl

# ── LOGGING ───────────────────────────────────────────────────────────────────
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s [%(levelname)s] %(message)s'
)
logger = logging.getLogger(__name__)
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ── DIAGNÓSTICO DE VARIÁVEIS DE AMBIENTE ─────────────────────────────────────
logger.info('=' * 60)
logger.info('DIAGNÓSTICO DE VARIÁVEIS DE AMBIENTE:')
logger.info(f'  DATABASE_URL      = {"SIM" if os.environ.get("DATABASE_URL") else "NÃO DEFINIDO"}')
logger.info('=' * 60)

# Fuso horário de Brasília (UTC-3)
TZ_BRASILIA = timezone(timedelta(hours=-3))

def agora():
    """Retorna o datetime atual no horário de Brasília"""
    return datetime.now(TZ_BRASILIA).replace(tzinfo=None)

db = SQLAlchemy()
csrf = CSRFProtect()


def configure_app(app):
    # SECRET_KEY deve ser definida como variável de ambiente no Railway.
    # Em produção, a aplicação deve falhar rápido se a variável não estiver definida.
    _secret = os.environ.get('SECRET_KEY')
    if not _secret:
        if os.environ.get('DATABASE_URL') or os.environ.get('URI_DO_BANCO_DE_DADOS'):
            raise RuntimeError('SECRET_KEY não definida em produção. Configure SECRET_KEY no Railway.')
        # Desenvolvimento local: gera chave aleatória por sessão (não previsível)
        import secrets as _secrets
        _secret = _secrets.token_hex(32)
    app.secret_key = _secret

    # Railway fornece DATABASE_URL com prefixo "postgres://" (formato antigo),
    # mas o SQLAlchemy 1.4+ exige "postgresql://". Corrige automaticamente.
    _db_url = (
        os.environ.get('DATABASE_URL') or
        os.environ.get('URI_DO_BANCO_DE_DADOS') or
        f'sqlite:///{os.path.join(os.path.dirname(os.path.abspath(__file__)), "instance", "estoque.db")}'
    )
    if _db_url.startswith('postgres://'):
        _db_url = _db_url.replace('postgres://', 'postgresql://', 1)

    app.config['SQLALCHEMY_DATABASE_URI'] = _db_url
    app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False
    app.config['WTF_CSRF_TIME_LIMIT'] = 7200
    app.config['PREFERRED_URL_SCHEME'] = 'https'
    app.config['SESSION_COOKIE_HTTPONLY'] = True
    app.config['SESSION_COOKIE_SAMESITE'] = 'Lax'
    app.config['SESSION_COOKIE_SECURE'] = bool(os.environ.get('DATABASE_URL') or os.environ.get('URI_DO_BANCO_DE_DADOS'))
    app.config['PERMANENT_SESSION_LIFETIME'] = timedelta(minutes=60)  # sessão expira em 60 min de inatividade
    app.config['JSONIFY_PRETTYPRINT_REGULAR'] = False


def create_app():
    app = Flask(__name__, static_folder='static', static_url_path='/static')
    configure_app(app)
    db.init_app(app)
    csrf.init_app(app)
    return app


app = create_app()


@app.before_request
def enforce_https_in_production():
    if app.config['SESSION_COOKIE_SECURE']:
        proto = request.headers.get('X-Forwarded-Proto')
        if proto and proto != 'https':
            return redirect(request.url.replace('http://', 'https://', 1), code=301)
    # Renova a sessão a cada request para manter o timeout de inatividade
    if 'usuario_id' in session:
        session.modified = True


def flash_html(message, category='info'):
    flash(Markup(message), category)


def usuario_tem_acesso_almoxarifado(u, alm_id):
    return u.perfil == 'admin' or (alm_id in u.almoxarifados_permitidos())


def usuario_tem_acesso_item(u, it):
    return u.perfil == 'admin' or (it and it.almoxarifado_id in u.almoxarifados_permitidos())

# ── HEADERS DE SEGURANÇA ──────────────────────────────────────────────────────
@app.after_request
def set_security_headers(response):
    response.headers['X-Content-Type-Options'] = 'nosniff'
    response.headers['X-Frame-Options'] = 'SAMEORIGIN'
    response.headers['X-XSS-Protection'] = '1; mode=block'
    response.headers['Referrer-Policy'] = 'strict-origin-when-cross-origin'
    response.headers['Strict-Transport-Security'] = 'max-age=63072000; includeSubDomains; preload'
    response.headers['Permissions-Policy'] = (
        'geolocation=(), microphone=(), camera=(), payment=(), usb=()'
    )
    response.headers['Content-Security-Policy'] = (
        "default-src 'self'; "
        "script-src 'self' 'unsafe-inline' https://cdn.jsdelivr.net; "
        "style-src 'self' 'unsafe-inline' https://cdn.jsdelivr.net; "
        "font-src 'self' https://cdn.jsdelivr.net data:; "
        "img-src 'self' data:; "
        "connect-src 'self';"
    )
    # Nunca cachear o app.css
    if request.path == '/static/css/app.css':
        response.headers['Cache-Control'] = 'no-cache, no-store, must-revalidate'
        response.headers['Pragma'] = 'no-cache'
        response.headers['Expires'] = '0'
    return response

# ── TRATAMENTO DE ERRO CSRF ───────────────────────────────────────────────────
@app.errorhandler(CSRFError)
def handle_csrf_error(e):
    flash('Sessão expirada ou requisição inválida. Tente novamente.', 'warning')
    return redirect(request.referrer or url_for('index'))

# ── RATE LIMITING DE LOGIN ────────────────────────────────────────────────────
_login_attempts: dict = {}
_MAX_ATTEMPTS = 10
_LOCKOUT_SECONDS = 300

def _check_rate_limit(ip: str) -> bool:
    now = datetime.now().timestamp()
    tentativas = [t for t in _login_attempts.get(ip, []) if now - t < _LOCKOUT_SECONDS]
    _login_attempts[ip] = tentativas
    return len(tentativas) >= _MAX_ATTEMPTS

def _register_attempt(ip: str):
    _login_attempts.setdefault(ip, []).append(datetime.now().timestamp())

def _clear_attempts(ip: str):
    _login_attempts.pop(ip, None)

# ── RATE LIMITING DE API ──────────────────────────────────────────────────────
_api_calls: dict = {}
_API_MAX = 120        # máximo de chamadas por janela
_API_WINDOW = 60      # janela de 60 segundos

def _check_api_rate(ip: str) -> bool:
    """Retorna True se o IP excedeu o limite de chamadas de API."""
    now = datetime.now().timestamp()
    calls = [t for t in _api_calls.get(ip, []) if now - t < _API_WINDOW]
    _api_calls[ip] = calls
    if len(calls) >= _API_MAX:
        return True
    _api_calls[ip].append(now)
    return False

# ── FILTRO JINJA2 — formata quantidades sem ponto flutuante feio ─────────────
@app.template_filter('fmt_qtd')
def fmt_qtd(value):
    """Formata quantidade: inteiros sem decimal, decimais com até 4 casas limpas."""
    try:
        v = round(float(value), 4)
        if v == int(v):
            return str(int(v))
        # Remove zeros à direita
        return f"{v:.4f}".rstrip('0').rstrip('.')
    except (TypeError, ValueError):
        return value

# ── MODELOS ──────────────────────────────────────────────────────────────────

class Almoxarifado(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    nome = db.Column(db.String(100), nullable=False)
    descricao = db.Column(db.String(200))
    obra = db.Column(db.String(100), nullable=True)    # ex: "Ventura Patamares", "Porto Aruana"
    cidade = db.Column(db.String(100), nullable=True)  # ex: "Salvador", "Aracaju"
    itens = db.relationship('Item', backref='almoxarifado', lazy=True, cascade='all, delete-orphan')

class Item(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    nome = db.Column(db.String(300), nullable=False)
    codigo = db.Column(db.String(50), nullable=False)
    unidade = db.Column(db.String(20), nullable=False)
    quantidade = db.Column(db.Float, default=0)
    estoque_minimo = db.Column(db.Float, default=0)
    almoxarifado_id = db.Column(db.Integer, db.ForeignKey('almoxarifado.id'), nullable=False)
    status_compra = db.Column(db.String(30), default='pendente')
    fixado = db.Column(db.Boolean, default=False)
    ativo = db.Column(db.Boolean, default=True)
    categoria = db.Column(db.String(30), default='geral')  # 'epi', 'maquinario', 'geral'
    ca = db.Column(db.String(20))  # Certificado de Aprovação (só EPIs)
    movimentacoes = db.relationship('Movimentacao', backref='item', lazy=True, cascade='all, delete-orphan')

    @property
    def status(self):
        if self.quantidade <= 0:
            return 'critico'
        elif self.quantidade <= self.estoque_minimo:
            return 'alerta'
        return 'ok'

class Movimentacao(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    tipo = db.Column(db.String(10), nullable=False)
    quantidade = db.Column(db.Float, nullable=False)
    responsavel = db.Column(db.String(100))
    observacao = db.Column(db.String(200))
    data = db.Column(db.DateTime, default=agora)
    item_id = db.Column(db.Integer, db.ForeignKey('item.id'), nullable=False)
    devolvido = db.Column(db.Boolean, nullable=True)  # None=não aplicável, True=devolvido, False=não devolvido
    foto_url = db.Column(db.Text, nullable=True)       # foto base64 — prova de entrega de EPI

class Requisicao(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    colaborador = db.Column(db.String(100), nullable=False)
    observacao = db.Column(db.String(200))
    quantidade = db.Column(db.Float, nullable=False)
    status = db.Column(db.String(20), default='aberta')  # aberta | devolvida
    data_retirada = db.Column(db.DateTime, default=agora)
    data_devolucao = db.Column(db.DateTime, nullable=True)
    item_id = db.Column(db.Integer, db.ForeignKey('item.id'), nullable=False)
    item = db.relationship('Item', backref='requisicoes')

# ── REQUISIÇÃO DO MESTRE ──────────────────────────────────────────────────────

class RequisicaoMestre(db.Model):
    """Requisição feita pelo mestre de obra ao almoxarifado.
    Fluxo: pendente → aprovada (almoxarife separa) → entregue (baixa no estoque)
    """
    id = db.Column(db.Integer, primary_key=True)
    protocolo = db.Column(db.String(30), unique=True, nullable=True)  # ex: REQ-20250611-0042
    mestre_id = db.Column(db.Integer, db.ForeignKey('usuario.id'), nullable=False)
    mestre = db.relationship('Usuario', foreign_keys=[mestre_id])
    colaborador = db.Column(db.String(100), nullable=False)
    almoxarifado_id = db.Column(db.Integer, db.ForeignKey('almoxarifado.id'), nullable=False)
    almoxarifado = db.relationship('Almoxarifado')
    observacao = db.Column(db.String(300))
    # Status: pendente | aprovada | parcial | recusada | entregue | cancelada
    status = db.Column(db.String(20), default='pendente')
    data_criacao = db.Column(db.DateTime, default=agora)
    data_entrega = db.Column(db.DateTime, nullable=True)
    entregue_por_id = db.Column(db.Integer, db.ForeignKey('usuario.id'), nullable=True)
    entregue_por = db.relationship('Usuario', foreign_keys=[entregue_por_id])
    foto_url = db.Column(db.Text, nullable=True)  # foto base64 — comprovante de entrega
    # Itens da requisição
    itens = db.relationship('RequisicaoMestreItem', backref='requisicao', lazy=True, cascade='all, delete-orphan')

class RequisicaoMestreItem(db.Model):
    """Cada item dentro de uma RequisicaoMestre."""
    id = db.Column(db.Integer, primary_key=True)
    requisicao_id = db.Column(db.Integer, db.ForeignKey('requisicao_mestre.id'), nullable=False)
    item_id = db.Column(db.Integer, db.ForeignKey('item.id'), nullable=False)
    item = db.relationship('Item')
    quantidade = db.Column(db.Float, nullable=False)
    observacao = db.Column(db.String(200))
    # Aprovação por item: pendente | aprovado | recusado
    status_item = db.Column(db.String(20), default='pendente')
    motivo_recusa = db.Column(db.String(200))

class Usuario(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    nome = db.Column(db.String(100), nullable=False)
    login = db.Column(db.String(50), unique=True, nullable=False)
    senha_hash = db.Column(db.String(256), nullable=False)
    perfil = db.Column(db.String(20), default='colaborador')  # admin | colaborador | mestre | almoxarife | tecnico_seguranca | analista
    almoxarifado_id = db.Column(db.Integer, db.ForeignKey('almoxarifado.id'), nullable=True)
    escopo = db.Column(db.String(50), nullable=True)  # estrutura | infraestrutura | acabamento — usado pelo analista
    email = db.Column(db.String(120), nullable=True)
    ativo = db.Column(db.Boolean, default=True)
    totp_secret = db.Column(db.String(32), nullable=True)   # 2FA — None = desativado
    almoxarifado = db.relationship('Almoxarifado', backref='usuarios')
    acessos_extras = db.relationship('AcessoExtra', backref='usuario', lazy=True, cascade='all, delete-orphan')
    pode_requisitar = db.Column(db.Boolean, default=False)
    pode_ver_alertas = db.Column(db.Boolean, default=False)

    def set_senha(self, senha):
        self.senha_hash = generate_password_hash(senha)

    def check_senha(self, senha):
        # Compatibilidade: hashes antigos eram SHA-256 puro (64 chars hex)
        if len(self.senha_hash) == 64 and not self.senha_hash.startswith('pbkdf2:'):
            import hashlib
            if self.senha_hash == hashlib.sha256(senha.encode()).hexdigest():
                # Migra automaticamente para o novo hash seguro
                self.set_senha(senha)
                db.session.commit()
                return True
            return False
        return check_password_hash(self.senha_hash, senha)

    def almoxarifados_permitidos(self):
        ids = set()
        if self.almoxarifado_id:
            ids.add(self.almoxarifado_id)
        for a in self.acessos_extras:
            if a.ativo:
                ids.add(a.almoxarifado_id)
        return ids

class Colaborador(db.Model):
    """Banco de dados de colaboradores/peões que retiram material no almoxarifado."""
    id = db.Column(db.Integer, primary_key=True)
    nome = db.Column(db.String(100), nullable=False)
    funcao = db.Column(db.String(50))   # ex: pedreiro, servente, eletricista
    escopo = db.Column(db.String(50))   # ex: estrutura, acabamento, infraestrutura, forma
    obra = db.Column(db.String(100), nullable=True)   # ex: Ventura Patamares, Porto Aruana
    cidade = db.Column(db.String(100), nullable=True) # ex: Salvador, Aracaju
    tipo = db.Column(db.String(30), default='peao')  # peao | mestre | tecnico_seguranca | engenheiro | almoxarife
    ativo = db.Column(db.Boolean, default=True)
    data_cadastro = db.Column(db.DateTime, default=agora)

class Ferramenta(db.Model):
    """Frota de ferramentas e máquinas por almoxarifado."""
    id = db.Column(db.Integer, primary_key=True)
    identificacao = db.Column(db.String(50), nullable=False)  # ID/patrimônio
    nome = db.Column(db.String(200), nullable=False)
    empresa = db.Column(db.String(100))                       # empresa proprietária (vazio = própria)
    almoxarifado_id = db.Column(db.Integer, db.ForeignKey('almoxarifado.id'), nullable=False)
    almoxarifado = db.relationship('Almoxarifado', backref='ferramentas')
    status = db.Column(db.String(20), default='disponivel')   # disponivel | em_uso | atrasado | manutencao
    responsavel_atual = db.Column(db.String(100))
    data_saida = db.Column(db.DateTime, nullable=True)        # quando foi retirada
    observacao = db.Column(db.String(200))
    local = db.Column(db.String(100))                         # localização física
    ativo = db.Column(db.Boolean, default=True)
    data_cadastro = db.Column(db.DateTime, default=agora)
    historico = db.relationship('HistoricoFerramenta', backref='ferramenta', lazy=True, order_by='HistoricoFerramenta.data_saida.desc()')

class HistoricoFerramenta(db.Model):
    """Registro de cada saída/devolução/manutenção de ferramenta."""
    id = db.Column(db.Integer, primary_key=True)
    ferramenta_id = db.Column(db.Integer, db.ForeignKey('ferramenta.id'), nullable=False)
    colaborador = db.Column(db.String(100), nullable=False)
    data_saida = db.Column(db.DateTime, nullable=False)
    data_devolucao = db.Column(db.DateTime, nullable=True)
    registrado_por = db.Column(db.String(100))
    tipo_evento = db.Column(db.String(20), default='uso')     # uso | manutencao
    motivo_manutencao = db.Column(db.String(300), nullable=True)
    foto_url = db.Column(db.Text, nullable=True)       # foto base64 comprimida — prova de retirada

class ItemEPI(db.Model):
    """Controle de EPIs e uniformes por almoxarifado (espelho de Ferramenta, sem prazo)."""
    id = db.Column(db.Integer, primary_key=True)
    identificacao = db.Column(db.String(50), nullable=False)  # CA ou código interno
    nome = db.Column(db.String(200), nullable=False)
    tamanho = db.Column(db.String(30))                        # P, M, G, GG, 38, 40...
    almoxarifado_id = db.Column(db.Integer, db.ForeignKey('almoxarifado.id'), nullable=False)
    almoxarifado = db.relationship('Almoxarifado', backref='epis')
    status = db.Column(db.String(20), default='disponivel')   # disponivel | em_uso | manutencao
    responsavel_atual = db.Column(db.String(100))
    quantidade = db.Column(db.Integer, default=1)             # estoque disponível deste EPI
    local = db.Column(db.String(100))                         # localização física
    observacao = db.Column(db.String(200))
    ativo = db.Column(db.Boolean, default=True)
    data_cadastro = db.Column(db.DateTime, default=agora)
    historico = db.relationship('HistoricoEPI', backref='item_epi', lazy=True, order_by='HistoricoEPI.data_saida.desc()')

class HistoricoEPI(db.Model):
    """Registro de cada entrega/devolução de EPI."""
    id = db.Column(db.Integer, primary_key=True)
    item_epi_id = db.Column(db.Integer, db.ForeignKey('item_epi.id'), nullable=False)
    colaborador = db.Column(db.String(100), nullable=False)
    data_saida = db.Column(db.DateTime, nullable=False)
    data_devolucao = db.Column(db.DateTime, nullable=True)
    registrado_por = db.Column(db.String(100))
    tipo_evento = db.Column(db.String(20), default='uso')     # uso | manutencao
    motivo_manutencao = db.Column(db.String(300), nullable=True)
    foto_url = db.Column(db.Text, nullable=True)              # foto base64 — comprovante de entrega

class AcessoExtra(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    usuario_id = db.Column(db.Integer, db.ForeignKey('usuario.id'), nullable=False)
    almoxarifado_id = db.Column(db.Integer, db.ForeignKey('almoxarifado.id'), nullable=False)
    motivo = db.Column(db.String(200))
    data_inicio = db.Column(db.DateTime, default=agora)
    data_fim = db.Column(db.DateTime, nullable=True)
    concedido_por = db.Column(db.String(100))
    almoxarifado = db.relationship('Almoxarifado')

    @property
    def ativo(self):
        if self.data_fim and agora() > self.data_fim:
            return False
        return True

class PermissaoExtra(db.Model):
    """Permissões especiais concedidas a usuários pelo fundador/admin.
    Ex: dar ao engenheiro a permissão de fazer requisições.
    permissao: 'fazer_requisicao' | 'ver_relatorios' | 'gerenciar_colaboradores'
    """
    id = db.Column(db.Integer, primary_key=True)
    usuario_id = db.Column(db.Integer, db.ForeignKey('usuario.id'), nullable=False)
    permissao = db.Column(db.String(50), nullable=False)
    concedido_por = db.Column(db.String(100))
    data_concessao = db.Column(db.DateTime, default=agora)
    usuario = db.relationship('Usuario', backref='permissoes_extras')

# ── DECORATORS DE ACESSO ─────────────────────────────────────────────────────

def login_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        if 'usuario_id' not in session:
            flash('Faça login para continuar.', 'warning')
            return redirect(url_for('login'))
        u = db.session.get(Usuario, session['usuario_id'])
        if not u or not u.ativo:
            session.clear()
            flash('Sessão expirada. Faça login novamente.', 'warning')
            return redirect(url_for('login'))
        return f(*args, **kwargs)
    return decorated

def admin_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        if 'usuario_id' not in session:
            return redirect(url_for('login'))
        u = db.session.get(Usuario, session['usuario_id'])
        if not u or u.perfil != 'admin':
            flash('Acesso restrito ao administrador.', 'danger')
            return redirect(url_for('index'))
        return f(*args, **kwargs)
    return decorated

def almoxarife_required(f):
    """Permite acesso a admin e almoxarife."""
    @wraps(f)
    def decorated(*args, **kwargs):
        if 'usuario_id' not in session:
            return redirect(url_for('login'))
        u = db.session.get(Usuario, session['usuario_id'])
        if not u or u.perfil not in ('admin', 'almoxarife'):
            flash('Acesso restrito ao almoxarife.', 'danger')
            return redirect(url_for('index'))
        return f(*args, **kwargs)
    return decorated

def usuario_atual():
    if 'usuario_id' in session:
        return db.session.get(Usuario, session['usuario_id'])
    return None

def is_fundador():
    """Retorna True se o usuário logado é o fundador (login 'rick')."""
    u = usuario_atual()
    return u is not None and u.login == 'rick'

def extrair_colaborador(mov):
    """Extrai o nome do colaborador da observação da movimentação.
    Suporta: 'liberado P/ Nome', 'Colaborador: Nome'.
    Fallback para mov.responsavel.
    """
    obs = mov.observacao or ''
    m = re.search(r'liberado\s+[Pp][/\s]+(.+)', obs, re.IGNORECASE)
    if m:
        return m.group(1).strip()
    m = re.search(r'[Cc]olaborador[:\s]+([^|]+)', obs)
    if m:
        return m.group(1).strip()
    return mov.responsavel or 'Sem responsável'

# ── CONTEXT PROCESSOR ────────────────────────────────────────────────────────

@app.context_processor
def inject_sidebar():
    u = usuario_atual()
    if not u:
        return dict(sidebar_alms=[], usuario_atual=None, sidebar_contadores={})
    if u.perfil == 'admin':
        alms = Almoxarifado.query.all()
    elif u.perfil == 'analista':
        # Analista vê apenas almoxarifados da sua cidade (via almoxarifado vinculado)
        if u.almoxarifado_id:
            alm_ref = db.session.get(Almoxarifado, u.almoxarifado_id)
            if alm_ref and alm_ref.cidade:
                alms = Almoxarifado.query.filter(
                    Almoxarifado.cidade.ilike(alm_ref.cidade)
                ).all()
            else:
                alms = [alm_ref] if alm_ref else []
        else:
            alms = Almoxarifado.query.all()
    elif u.perfil in ('mestre', 'tecnico_seguranca'):
        if u.perfil == 'tecnico_seguranca':
            ids = u.almoxarifados_permitidos()
            alms = Almoxarifado.query.filter(Almoxarifado.id.in_(ids)).all() if ids else (
                [u.almoxarifado] if u.almoxarifado_id else []
            )
        else:
            alms = [u.almoxarifado] if u.almoxarifado_id else []
    else:
        ids = u.almoxarifados_permitidos()
        alms = Almoxarifado.query.filter(Almoxarifado.id.in_(ids)).all() if ids else []

    # Ordenar por cidade → obra → nome para agrupamento consistente na sidebar
    alms = sorted(alms, key=lambda a: (a.cidade or 'zzz', a.obra or 'zzz', a.nome))

    # Contadores de ferramentas e EPIs por almoxarifado para exibir na sidebar
    sidebar_contadores = {}
    for alm in alms:
        n_ferr  = Ferramenta.query.filter_by(almoxarifado_id=alm.id, ativo=True).count()
        n_epi   = ItemEPI.query.filter_by(almoxarifado_id=alm.id, ativo=True).count()
        n_itens = Item.query.filter_by(almoxarifado_id=alm.id, ativo=True).count()
        # % de saúde: itens com estoque > mínimo / total
        itens_ok = Item.query.filter_by(almoxarifado_id=alm.id, ativo=True).filter(
            Item.quantidade > Item.estoque_minimo
        ).count()
        pct_saude = round(itens_ok / n_itens * 100) if n_itens > 0 else 100
        sidebar_contadores[alm.id] = {'ferr': n_ferr, 'epi': n_epi, 'itens': n_itens, 'pct': pct_saude}

    return dict(sidebar_alms=alms, usuario_atual=u, sidebar_contadores=sidebar_contadores)

def run_migrations():
    """Executa migrações de schema de forma segura usando SAVEPOINT no PostgreSQL."""
    try:
        from sqlalchemy import text
        is_pg = 'postgresql' in str(db.engine.url)

        def safe_exec(conn, sql):
            """Executa um comando DDL isolado — usa SAVEPOINT no PG para não quebrar a transação."""
            try:
                if is_pg:
                    conn.execute(text("SAVEPOINT mig"))
                conn.execute(text(sql))
                if is_pg:
                    conn.execute(text("RELEASE SAVEPOINT mig"))
                conn.commit()
            except Exception:
                if is_pg:
                    try:
                        conn.execute(text("ROLLBACK TO SAVEPOINT mig"))
                        conn.execute(text("RELEASE SAVEPOINT mig"))
                    except Exception:
                        pass
                else:
                    try:
                        conn.rollback()
                    except Exception:
                        pass

        with db.engine.connect() as conn:
            pk_type = 'SERIAL PRIMARY KEY' if is_pg else 'INTEGER PRIMARY KEY AUTOINCREMENT'
            dt_type = 'TIMESTAMP' if is_pg else 'DATETIME'

            # ── Colunas em item ──────────────────────────────────────────────
            safe_exec(conn, "ALTER TABLE item ADD COLUMN status_compra VARCHAR(30) DEFAULT 'pendente'")
            safe_exec(conn, "ALTER TABLE item ADD COLUMN fixado BOOLEAN DEFAULT 0")
            safe_exec(conn, "ALTER TABLE item ADD COLUMN ativo BOOLEAN DEFAULT 1")
            safe_exec(conn, "ALTER TABLE item ADD COLUMN categoria VARCHAR(30) DEFAULT 'geral'")
            safe_exec(conn, "ALTER TABLE item ADD COLUMN ca VARCHAR(20)")
            if is_pg:
                safe_exec(conn, "ALTER TABLE item ALTER COLUMN nome TYPE VARCHAR(300)")

            # ── Coluna email em usuario (crítica — deve rodar cedo) ──────────
            safe_exec(conn, "ALTER TABLE usuario ADD COLUMN email VARCHAR(120)")
            safe_exec(conn, "ALTER TABLE usuario ADD COLUMN senha_hash_new VARCHAR(256)")

            # ── Tabela acesso_extra ──────────────────────────────────────────
            if is_pg:
                safe_exec(conn, """
                    CREATE TABLE IF NOT EXISTS acesso_extra (
                        id SERIAL PRIMARY KEY,
                        usuario_id INTEGER NOT NULL REFERENCES usuario(id),
                        almoxarifado_id INTEGER NOT NULL REFERENCES almoxarifado(id),
                        motivo VARCHAR(200),
                        data_inicio TIMESTAMP,
                        data_fim TIMESTAMP,
                        concedido_por VARCHAR(100)
                    )
                """)
            else:
                safe_exec(conn, """
                    CREATE TABLE IF NOT EXISTS acesso_extra (
                        id INTEGER PRIMARY KEY AUTOINCREMENT,
                        usuario_id INTEGER NOT NULL REFERENCES usuario(id),
                        almoxarifado_id INTEGER NOT NULL REFERENCES almoxarifado(id),
                        motivo VARCHAR(200),
                        data_inicio DATETIME,
                        data_fim DATETIME,
                        concedido_por VARCHAR(100)
                    )
                """)

            # ── Tabelas do mestre ────────────────────────────────────────────
            safe_exec(conn, f"""
                CREATE TABLE IF NOT EXISTS requisicao_mestre (
                    id {pk_type},
                    mestre_id INTEGER NOT NULL REFERENCES usuario(id),
                    colaborador VARCHAR(100) NOT NULL,
                    almoxarifado_id INTEGER NOT NULL REFERENCES almoxarifado(id),
                    observacao VARCHAR(300),
                    status VARCHAR(20) DEFAULT 'pendente',
                    data_criacao {dt_type},
                    data_entrega {dt_type},
                    entregue_por_id INTEGER REFERENCES usuario(id),
                    notificado BOOLEAN DEFAULT FALSE
                )
            """)
            safe_exec(conn, f"""
                CREATE TABLE IF NOT EXISTS requisicao_mestre_item (
                    id {pk_type},
                    requisicao_id INTEGER NOT NULL REFERENCES requisicao_mestre(id),
                    item_id INTEGER NOT NULL REFERENCES item(id),
                    quantidade FLOAT NOT NULL,
                    observacao VARCHAR(200),
                    status_item VARCHAR(20) DEFAULT 'pendente',
                    motivo_recusa VARCHAR(200)
                )
            """)
            safe_exec(conn, f"""
                CREATE TABLE IF NOT EXISTS colaborador (
                    id {pk_type},
                    nome VARCHAR(100) NOT NULL,
                    funcao VARCHAR(50),
                    ativo BOOLEAN DEFAULT TRUE,
                    data_cadastro {dt_type}
                )
            """)

            # ── Colunas adicionais em tabelas existentes ─────────────────────
            safe_exec(conn, "ALTER TABLE requisicao_mestre ADD COLUMN notificado BOOLEAN DEFAULT FALSE")
            safe_exec(conn, "ALTER TABLE requisicao_mestre_item ADD COLUMN status_item VARCHAR(20) DEFAULT 'pendente'")
            safe_exec(conn, "ALTER TABLE requisicao_mestre_item ADD COLUMN motivo_recusa VARCHAR(200)")
            # ── Coluna protocolo em requisicao_mestre ─────────────────────────
            safe_exec(conn, "ALTER TABLE requisicao_mestre ADD COLUMN protocolo VARCHAR(30)")

            # ── Valores padrão para colunas que podem estar NULL ─────────────
            safe_exec(conn, "UPDATE requisicao_mestre_item SET status_item = 'pendente' WHERE status_item IS NULL")
            safe_exec(conn, "UPDATE requisicao_mestre SET notificado = FALSE WHERE notificado IS NULL")

            # ── Campo escopo em colaborador ──────────────────────────────────
            safe_exec(conn, "ALTER TABLE colaborador ADD COLUMN escopo VARCHAR(50)")
            # ── Campos obra e cidade em colaborador ───────────────────────────
            safe_exec(conn, "ALTER TABLE colaborador ADD COLUMN obra VARCHAR(100)")
            safe_exec(conn, "ALTER TABLE colaborador ADD COLUMN cidade VARCHAR(100)")

            # ── Campo devolvido em movimentacao ──────────────────────────────
            safe_exec(conn, "ALTER TABLE movimentacao ADD COLUMN devolvido BOOLEAN")
            # ── Campo foto_url em movimentacao ────────────────────────────────
            safe_exec(conn, "ALTER TABLE movimentacao ADD COLUMN foto_url TEXT")

            # ── Campo foto_url em requisicao_mestre ───────────────────────────
            safe_exec(conn, "ALTER TABLE requisicao_mestre ADD COLUMN foto_url TEXT")

            # ── Campo escopo em usuario (para analista) ───────────────────────
            safe_exec(conn, "ALTER TABLE usuario ADD COLUMN escopo VARCHAR(50)")

            # ── Campo local em ferramenta ─────────────────────────────────────
            safe_exec(conn, "ALTER TABLE ferramenta ADD COLUMN local VARCHAR(100)")
            # ── Colunas adicionais em item_epi (para bancos criados sem elas) ─
            safe_exec(conn, "ALTER TABLE item_epi ADD COLUMN quantidade INTEGER DEFAULT 1")
            safe_exec(conn, "ALTER TABLE item_epi ADD COLUMN local VARCHAR(100)")

            # ── Campo obra em almoxarifado ────────────────────────────────────
            safe_exec(conn, "ALTER TABLE almoxarifado ADD COLUMN obra VARCHAR(100)")
            # Setar obra padrão para almoxarifados existentes sem obra definida
            safe_exec(conn, "UPDATE almoxarifado SET obra = 'Salvador' WHERE obra IS NULL OR obra = ''")
            # ── Campo cidade em almoxarifado ──────────────────────────────────
            safe_exec(conn, "ALTER TABLE almoxarifado ADD COLUMN cidade VARCHAR(100)")
            # Setar cidade padrão para almoxarifados existentes sem cidade definida
            safe_exec(conn, "UPDATE almoxarifado SET cidade = 'Salvador' WHERE cidade IS NULL OR cidade = ''")
            # Se obra contém nome de cidade (Aracaju/Salvador), mover para cidade e limpar obra
            safe_exec(conn, "UPDATE almoxarifado SET cidade = obra, obra = NULL WHERE obra IN ('Aracaju', 'Salvador', 'aracaju', 'salvador') AND (cidade IS NULL OR cidade = '')")
            # ── Campos obra e cidade em colaborador ───────────────────────────
            safe_exec(conn, "ALTER TABLE colaborador ADD COLUMN obra VARCHAR(100)")
            safe_exec(conn, "ALTER TABLE colaborador ADD COLUMN cidade VARCHAR(100)")

            # ── Tabela permissao_extra ────────────────────────────────────────
            if is_pg:
                safe_exec(conn, """
                    CREATE TABLE IF NOT EXISTS permissao_extra (
                        id SERIAL PRIMARY KEY,
                        usuario_id INTEGER NOT NULL REFERENCES usuario(id),
                        permissao VARCHAR(50) NOT NULL,
                        concedido_por VARCHAR(100),
                        data_concessao TIMESTAMP
                    )
                """)
            else:
                safe_exec(conn, """
                    CREATE TABLE IF NOT EXISTS permissao_extra (
                        id INTEGER PRIMARY KEY AUTOINCREMENT,
                        usuario_id INTEGER NOT NULL REFERENCES usuario(id),
                        permissao VARCHAR(50) NOT NULL,
                        concedido_por VARCHAR(100),
                        data_concessao DATETIME
                    )
                """)

            # ── Tabela item_epi ───────────────────────────────────────────────
            if is_pg:
                safe_exec(conn, """
                    CREATE TABLE IF NOT EXISTS item_epi (
                        id SERIAL PRIMARY KEY,
                        identificacao VARCHAR(50) NOT NULL,
                        nome VARCHAR(200) NOT NULL,
                        tamanho VARCHAR(30),
                        almoxarifado_id INTEGER NOT NULL REFERENCES almoxarifado(id),
                        status VARCHAR(20) DEFAULT 'disponivel',
                        responsavel_atual VARCHAR(100),
                        quantidade INTEGER DEFAULT 1,
                        local VARCHAR(100),
                        observacao VARCHAR(200),
                        ativo BOOLEAN DEFAULT TRUE,
                        data_cadastro TIMESTAMP
                    )
                """)
            else:
                safe_exec(conn, """
                    CREATE TABLE IF NOT EXISTS item_epi (
                        id INTEGER PRIMARY KEY AUTOINCREMENT,
                        identificacao VARCHAR(50) NOT NULL,
                        nome VARCHAR(200) NOT NULL,
                        tamanho VARCHAR(30),
                        almoxarifado_id INTEGER NOT NULL REFERENCES almoxarifado(id),
                        status VARCHAR(20) DEFAULT 'disponivel',
                        responsavel_atual VARCHAR(100),
                        quantidade INTEGER DEFAULT 1,
                        local VARCHAR(100),
                        observacao VARCHAR(200),
                        ativo BOOLEAN DEFAULT TRUE,
                        data_cadastro DATETIME
                    )
                """)

            # ── Tabela historico_epi ──────────────────────────────────────────
            if is_pg:
                safe_exec(conn, """
                    CREATE TABLE IF NOT EXISTS historico_epi (
                        id SERIAL PRIMARY KEY,
                        item_epi_id INTEGER NOT NULL REFERENCES item_epi(id),
                        colaborador VARCHAR(100) NOT NULL,
                        data_saida TIMESTAMP NOT NULL,
                        data_devolucao TIMESTAMP,
                        registrado_por VARCHAR(100),
                        tipo_evento VARCHAR(20) DEFAULT 'uso',
                        motivo_manutencao VARCHAR(300),
                        foto_url TEXT
                    )
                """)
            else:
                safe_exec(conn, """
                    CREATE TABLE IF NOT EXISTS historico_epi (
                        id INTEGER PRIMARY KEY AUTOINCREMENT,
                        item_epi_id INTEGER NOT NULL REFERENCES item_epi(id),
                        colaborador VARCHAR(100) NOT NULL,
                        data_saida DATETIME NOT NULL,
                        data_devolucao DATETIME,
                        registrado_por VARCHAR(100),
                        tipo_evento VARCHAR(20) DEFAULT 'uso',
                        motivo_manutencao VARCHAR(300),
                        foto_url TEXT
                    )
                """)

            # ── Tabela ferramentas ────────────────────────────────────────────
            if is_pg:
                safe_exec(conn, """
                    CREATE TABLE IF NOT EXISTS ferramenta (
                        id SERIAL PRIMARY KEY,
                        identificacao VARCHAR(50) NOT NULL,
                        nome VARCHAR(200) NOT NULL,
                        almoxarifado_id INTEGER NOT NULL REFERENCES almoxarifado(id),
                        status VARCHAR(20) DEFAULT 'disponivel',
                        responsavel_atual VARCHAR(100),
                        observacao VARCHAR(200),
                        ativo BOOLEAN DEFAULT TRUE,
                        data_cadastro TIMESTAMP
                    )
                """)
            else:
                safe_exec(conn, """
                    CREATE TABLE IF NOT EXISTS ferramenta (
                        id INTEGER PRIMARY KEY AUTOINCREMENT,
                        identificacao VARCHAR(50) NOT NULL,
                        nome VARCHAR(200) NOT NULL,
                        almoxarifado_id INTEGER NOT NULL REFERENCES almoxarifado(id),
                        status VARCHAR(20) DEFAULT 'disponivel',
                        responsavel_atual VARCHAR(100),
                        observacao VARCHAR(200),
                        ativo BOOLEAN DEFAULT TRUE,
                        data_cadastro DATETIME
                    )
                """)
            # ── Campo tipo em colaborador ─────────────────────────────────────
            safe_exec(conn, "ALTER TABLE colaborador ADD COLUMN tipo VARCHAR(30) DEFAULT 'peao'")
            # ── Campo data_saida em ferramenta ────────────────────────────────
            if is_pg:
                safe_exec(conn, "ALTER TABLE ferramenta ADD COLUMN data_saida TIMESTAMP")
            else:
                safe_exec(conn, "ALTER TABLE ferramenta ADD COLUMN data_saida DATETIME")
            # ── Campo empresa em ferramenta ───────────────────────────────────
            safe_exec(conn, "ALTER TABLE ferramenta ADD COLUMN empresa VARCHAR(100)")

            # ── Campo 2FA em usuario ──────────────────────────────────────────
            safe_exec(conn, "ALTER TABLE usuario ADD COLUMN totp_secret VARCHAR(32)")
            # ── Permissões de função em usuario ──────────────────────────────
            safe_exec(conn, "ALTER TABLE usuario ADD COLUMN pode_requisitar BOOLEAN DEFAULT FALSE")
            safe_exec(conn, "ALTER TABLE usuario ADD COLUMN pode_ver_alertas BOOLEAN DEFAULT FALSE")

            # ── Tabela historico_ferramenta ───────────────────────────────────
            if is_pg:
                safe_exec(conn, """
                    CREATE TABLE IF NOT EXISTS historico_ferramenta (
                        id SERIAL PRIMARY KEY,
                        ferramenta_id INTEGER NOT NULL REFERENCES ferramenta(id),
                        colaborador VARCHAR(100) NOT NULL,
                        data_saida TIMESTAMP NOT NULL,
                        data_devolucao TIMESTAMP,
                        registrado_por VARCHAR(100),
                        tipo_evento VARCHAR(20) DEFAULT 'uso',
                        motivo_manutencao VARCHAR(300)
                    )
                """)
            else:
                safe_exec(conn, """
                    CREATE TABLE IF NOT EXISTS historico_ferramenta (
                        id INTEGER PRIMARY KEY AUTOINCREMENT,
                        ferramenta_id INTEGER NOT NULL REFERENCES ferramenta(id),
                        colaborador VARCHAR(100) NOT NULL,
                        data_saida DATETIME NOT NULL,
                        data_devolucao DATETIME,
                        registrado_por VARCHAR(100),
                        tipo_evento VARCHAR(20) DEFAULT 'uso',
                        motivo_manutencao VARCHAR(300)
                    )
                """)
            # ── Colunas novas em historico_ferramenta (bancos existentes) ────
            safe_exec(conn, "ALTER TABLE historico_ferramenta ADD COLUMN tipo_evento VARCHAR(20) DEFAULT 'uso'")
            safe_exec(conn, "ALTER TABLE historico_ferramenta ADD COLUMN motivo_manutencao VARCHAR(300)")
            safe_exec(conn, "ALTER TABLE historico_ferramenta ADD COLUMN foto_url TEXT")
            # Converter foto_url de VARCHAR(500) para TEXT (bancos com coluna antiga)
            if is_pg:
                safe_exec(conn, "ALTER TABLE historico_ferramenta ALTER COLUMN foto_url TYPE TEXT USING foto_url::TEXT")
            # Preencher tipo_evento nos registros antigos
            safe_exec(conn, "UPDATE historico_ferramenta SET tipo_evento = 'uso' WHERE tipo_evento IS NULL")
            # Garantir que ferramenta aceita status 'manutencao' (sem restrição de CHECK no SQLite)
            safe_exec(conn, "UPDATE ferramenta SET status = 'disponivel' WHERE status NOT IN ('disponivel','em_uso','manutencao','atrasado')")

    except Exception as e:
        logger.error(f'Migração: {e}')

# ── LOGIN / LOGOUT ────────────────────────────────────────────────────────────

@app.route('/static/sw.js')
def service_worker():
    """Serve o Service Worker com headers corretos para PWA."""
    from flask import send_from_directory
    response = send_from_directory('static', 'sw.js', mimetype='application/javascript')
    response.headers['Service-Worker-Allowed'] = '/'
    response.headers['Cache-Control'] = 'no-cache'
    return response

@app.route('/login', methods=['GET', 'POST'])
def login():
    if request.method == 'POST':
        ip = request.remote_addr or '0.0.0.0'
        if _check_rate_limit(ip):
            flash('Muitas tentativas. Aguarde 5 minutos.', 'danger')
            return render_template('login.html'), 429
        login_val = request.form.get('login', '').strip()
        senha_val = request.form.get('senha', '')
        totp_val  = request.form.get('totp_code', '').strip().replace(' ', '')
        if not login_val or not senha_val:
            flash('Preencha login e senha.', 'warning')
            return render_template('login.html')
        u = Usuario.query.filter_by(login=login_val, ativo=True).first()
        if u and u.check_senha(senha_val):
            # Verificar 2FA se estiver ativado
            if u.totp_secret:
                try:
                    import pyotp
                    totp = pyotp.TOTP(u.totp_secret)
                    if not totp_val or not totp.verify(totp_val, valid_window=1):
                        _register_attempt(ip)
                        flash('Código 2FA inválido ou expirado.', 'danger')
                        return render_template('login.html', requer_2fa=True,
                                               login_val=login_val)
                except Exception:
                    flash('Erro ao verificar 2FA. Tente novamente.', 'danger')
                    return render_template('login.html', requer_2fa=True,
                                           login_val=login_val)
            _clear_attempts(ip)
            session.clear()
            session.permanent = True
            session['usuario_id'] = u.id
            flash(f'Bem-vindo, {u.nome}!', 'success')
            return redirect(url_for('index'))
        _register_attempt(ip)
        import time as _time
        _time.sleep(0.3)  # timing uniforme — dificulta enumeration de usuários
        flash('Login ou senha incorretos.', 'danger')
    return render_template('login.html')

@app.route('/logout', methods=['POST'])
@login_required
def logout():
    session.clear()
    return redirect(url_for('login'))

# ── 2FA — AUTENTICAÇÃO DE DOIS FATORES ───────────────────────────────────────
@app.route('/perfil/2fa/ativar', methods=['GET', 'POST'])
@login_required
def ativar_2fa():
    """Gera QR Code para o usuário configurar o 2FA no Google Authenticator."""
    import pyotp, io as _io, base64
    u = usuario_atual()
    if request.method == 'POST':
        codigo = request.form.get('codigo', '').strip().replace(' ', '')
        secret = session.get('totp_secret_pendente')
        if not secret:
            flash('Sessão expirada. Tente novamente.', 'danger')
            return redirect(url_for('ativar_2fa'))
        totp = pyotp.TOTP(secret)
        if totp.verify(codigo, valid_window=1):
            u.totp_secret = secret
            session.pop('totp_secret_pendente', None)
            db.session.commit()
            flash('✅ 2FA ativado com sucesso! Seu login agora exige o código do app.', 'success')
            return redirect(url_for('index'))
        flash('Código inválido. Tente novamente.', 'danger')
        return redirect(url_for('ativar_2fa'))

    # Gerar novo secret
    secret = pyotp.random_base32()
    session['totp_secret_pendente'] = secret
    uri = pyotp.totp.TOTP(secret).provisioning_uri(
        name=u.login, issuer_name='Logi-Prime Obra Patamares'
    )

    # Gerar QR Code como SVG (sem dependência de Pillow/PIL)
    try:
        import qrcode, qrcode.image.svg
        factory = qrcode.image.svg.SvgPathImage
        img = qrcode.make(uri, image_factory=factory, box_size=10)
        buf = _io.BytesIO()
        img.save(buf)
        qr_svg = buf.getvalue().decode('utf-8')
        qr_b64 = None
    except Exception:
        qr_svg = None
        qr_b64 = None

    return render_template('2fa_ativar.html', qr_svg=qr_svg, qr_b64=qr_b64,
                           secret=secret, usuario=u, uri=uri)

@app.route('/perfil/2fa/desativar', methods=['POST'])
@login_required
def desativar_2fa():
    u = usuario_atual()
    u.totp_secret = None
    db.session.commit()
    flash('2FA desativado.', 'warning')
    return redirect(url_for('index'))

@app.route('/admin/2fa/desativar/<int:uid>', methods=['POST'])
@admin_required
def admin_desativar_2fa(uid):
    """Admin pode desativar 2FA de qualquer usuário (ex: perdeu o celular)."""
    u = Usuario.query.get_or_404(uid)
    u.totp_secret = None
    db.session.commit()
    flash(f'2FA de {u.nome} desativado pelo admin.', 'warning')
    return redirect(url_for('usuarios'))

@app.route('/healthz')
def healthz():
    """Healthcheck — responde imediatamente sem tocar no banco."""
    return 'ok', 200

# ── ROTAS PRINCIPAIS ─────────────────────────────────────────────────────────

@app.route('/')
@login_required
def index():
    u = usuario_atual()
    # Mestre e técnico de segurança só acessam a tela de requisições
    # Engenheiro com pode_requisitar também é redirecionado para lá
    if u.perfil in ('mestre', 'tecnico_seguranca'):
        return redirect(url_for('mestre_requisicoes'))
    if u.perfil == 'colaborador' and (
        u.pode_requisitar or
        any(p.permissao == 'fazer_requisicao' for p in u.permissoes_extras)
    ):
        return redirect(url_for('mestre_requisicoes'))
    if u.perfil == 'admin':
        almoxarifados = Almoxarifado.query.all()
        alertas = Item.query.filter(Item.quantidade <= Item.estoque_minimo, Item.ativo == True).all()
    elif u.perfil == 'analista':
        # Analista vê apenas almoxarifados da sua cidade (via almoxarifado vinculado)
        if u.almoxarifado_id:
            alm_ref = db.session.get(Almoxarifado, u.almoxarifado_id)
            cidade_analista = (alm_ref.cidade or '').strip() if alm_ref else None
            if cidade_analista:
                almoxarifados = Almoxarifado.query.filter(
                    Almoxarifado.cidade.ilike(cidade_analista)
                ).all()
            else:
                almoxarifados = [alm_ref] if alm_ref else []
        else:
            almoxarifados = Almoxarifado.query.all()
        ids_analista = {a.id for a in almoxarifados}
        alertas = Item.query.filter(
            Item.quantidade <= Item.estoque_minimo,
            Item.almoxarifado_id.in_(ids_analista),
            Item.ativo == True
        ).all()
    else:
        ids = u.almoxarifados_permitidos()
        almoxarifados = Almoxarifado.query.filter(Almoxarifado.id.in_(ids)).all() if ids else []
        alertas = Item.query.filter(
            Item.quantidade <= Item.estoque_minimo,
            Item.almoxarifado_id.in_(ids),
            Item.ativo == True
        ).all() if ids else []
    stats = {
        'total_almoxarifados': len(almoxarifados),
        'total_itens': sum(len(a.itens) for a in almoxarifados),
        'itens_alerta': len([a for a in alertas if a.quantidade > 0]),
        'itens_criticos': len([a for a in alertas if a.quantidade <= 0]),
    }
    # Previsão de ruptura no dashboard — itens OK mas em risco nos próximos 15 dias
    ids_alm = {a.id for a in almoxarifados}
    todos_ativos_dash = Item.query.filter(
        Item.ativo == True,
        Item.almoxarifado_id.in_(ids_alm)
    ).all() if ids_alm else []
    ruptura_dash = calcular_ruptura(todos_ativos_dash, limite_dias=15)

    return render_template('index.html', almoxarifados=almoxarifados, alertas=alertas,
                           stats=stats, ruptura=ruptura_dash)

@app.route('/almoxarifado/<int:id>')
@login_required
def almoxarifado(id):
    u = usuario_atual()
    # Mestre e técnico de segurança não acessam almoxarifado diretamente
    # Engenheiro com pode_requisitar também não — usa tela de requisições
    if u.perfil in ('mestre', 'tecnico_seguranca'):
        flash('Acesso restrito. Use a tela de requisições.', 'warning')
        return redirect(url_for('mestre_requisicoes'))
    if u.perfil == 'colaborador' and (
        u.pode_requisitar or
        any(p.permissao == 'fazer_requisicao' for p in u.permissoes_extras)
    ):
        flash('Acesso restrito. Use a tela de requisições.', 'warning')
        return redirect(url_for('mestre_requisicoes'))
    alm = Almoxarifado.query.get_or_404(id)
    if u.perfil not in ('admin', 'analista') and id not in u.almoxarifados_permitidos():
        flash('Acesso negado.', 'danger')
        return redirect(url_for('index'))
    # Analista só acessa almoxarifados da sua cidade
    if u.perfil == 'analista' and u.almoxarifado_id:
        alm_ref = db.session.get(Almoxarifado, u.almoxarifado_id)
        if alm_ref and alm_ref.cidade:
            alm_alvo = Almoxarifado.query.get_or_404(id)
            if (alm_alvo.cidade or '').lower().strip() != alm_ref.cidade.lower().strip():
                flash('Acesso negado.', 'danger')
                return redirect(url_for('index'))
    # Mostrar todos os itens (ativos e desativados) para permitir reativação
    itens = Item.query.filter_by(almoxarifado_id=id).order_by(Item.ativo.desc(), Item.nome).all()
    return render_template('almoxarifado.html', almoxarifado=alm, itens=itens)

@app.route('/item/<int:id>')
@login_required
def item(id):
    it = Item.query.get_or_404(id)
    u = usuario_atual()
    if u.perfil != 'admin' and it.almoxarifado_id not in u.almoxarifados_permitidos():
        flash('Acesso negado.', 'danger')
        return redirect(url_for('index'))
    movs = Movimentacao.query.filter_by(item_id=id).order_by(Movimentacao.data.desc()).limit(50).all()
    return render_template('item.html', item=it, movimentacoes=movs)

# ── CRUD ALMOXARIFADO ────────────────────────────────────────────────────────

@app.route('/almoxarifado/novo', methods=['GET', 'POST'])
@admin_required
def novo_almoxarifado():
    if request.method == 'POST':
        alm = Almoxarifado(nome=request.form['nome'], descricao=request.form.get('descricao', ''))
        alm.obra = request.form.get('obra', '').strip() or None
        alm.cidade = request.form.get('cidade', '').strip() or None
        db.session.add(alm)
        db.session.commit()
        flash(f'Almoxarifado "{alm.nome}" criado!', 'success')
        return redirect(url_for('index'))
    return render_template('form_almoxarifado.html', almoxarifado=None)

@app.route('/almoxarifado/<int:id>/editar', methods=['GET', 'POST'])
@login_required
def editar_almoxarifado(id):
    alm = Almoxarifado.query.get_or_404(id)
    u = usuario_atual()
    # Apenas admin ou almoxarife do próprio almoxarifado pode editar
    if u.perfil != 'admin' and id not in u.almoxarifados_permitidos():
        flash('Acesso negado.', 'danger')
        return redirect(url_for('index'))
    if request.method == 'POST':
        alm.nome = request.form['nome']
        alm.descricao = request.form.get('descricao', '')
        alm.obra = request.form.get('obra', '').strip() or None
        alm.cidade = request.form.get('cidade', '').strip() or None
        db.session.commit()
        flash('Almoxarifado atualizado!', 'success')
        return redirect(url_for('index'))
    return render_template('form_almoxarifado.html', almoxarifado=alm)

@app.route('/almoxarifado/<int:id>/deletar', methods=['POST'])
@admin_required
def deletar_almoxarifado(id):
    alm = Almoxarifado.query.get_or_404(id)
    try:
        # Desvincular usuários que têm este almoxarifado como principal
        Usuario.query.filter_by(almoxarifado_id=id).update({'almoxarifado_id': None})

        # Desvincular acessos extras para este almoxarifado
        AcessoExtra.query.filter_by(almoxarifado_id=id).delete()

        # Deletar requisições do mestre vinculadas a este almoxarifado
        # (RequisicaoMestreItem é deletado via cascade no relacionamento)
        for req in RequisicaoMestre.query.filter_by(almoxarifado_id=id).all():
            db.session.delete(req)

        # O cascade 'all, delete-orphan' já cuida de itens (Item model)
        # Ferramentas e EPIs também têm relationship com backref, deletar explicitamente
        Ferramenta.query.filter_by(almoxarifado_id=id).delete()
        ItemEPI.query.filter_by(almoxarifado_id=id).delete()

        db.session.delete(alm)
        db.session.commit()
        flash('Almoxarifado removido com sucesso!', 'warning')
    except Exception as e:
        db.session.rollback()
        flash(f'Erro ao remover almoxarifado: {str(e)}', 'danger')
    return redirect(url_for('index'))

# ── CRUD ITEM ────────────────────────────────────────────────────────────────

@app.route('/item/novo', methods=['GET', 'POST'])
@almoxarife_required
def novo_item():
    u = usuario_atual()
    # Almoxarife só vê seus próprios almoxarifados no select
    if u.perfil == 'admin':
        almoxarifados = Almoxarifado.query.all()
    else:
        ids = u.almoxarifados_permitidos()
        almoxarifados = Almoxarifado.query.filter(Almoxarifado.id.in_(ids)).all() if ids else []
    if request.method == 'POST':
        codigo = request.form['codigo'].strip()
        # Verificar duplicata de código DENTRO do mesmo almoxarifado
        alm_id = int(request.form['almoxarifado_id'])
        existente = Item.query.filter(
            Item.codigo.ilike(codigo),
            Item.almoxarifado_id == alm_id
        ).first()
        if existente:
            flash(f'⚠️ Código "{codigo}" já está em uso pelo item "{existente.nome}" neste almoxarifado. Use um código diferente.', 'danger')
            return render_template('form_item.html', item=None, almoxarifados=almoxarifados)
        try:
            it = Item(
                nome=request.form['nome'],
                codigo=codigo,
                unidade=request.form['unidade'],
                quantidade=float(request.form.get('quantidade', 0)),
                estoque_minimo=float(request.form.get('estoque_minimo', 0)),
                almoxarifado_id=int(request.form['almoxarifado_id']),
                categoria=request.form.get('categoria', 'geral'),
                ca=request.form.get('ca', '').strip() or None
            )
            db.session.add(it)
            db.session.commit()
            flash(f'Item "{it.nome}" cadastrado!', 'success')
            return redirect(url_for('almoxarifado', id=it.almoxarifado_id))
        except Exception as e:
            db.session.rollback()
            flash(f'Erro ao cadastrar item: {str(e)}', 'danger')
            return render_template('form_item.html', item=None, almoxarifados=almoxarifados)
    return render_template('form_item.html', item=None, almoxarifados=almoxarifados)

@app.route('/item/<int:id>/editar', methods=['GET', 'POST'])
@login_required
def editar_item(id):
    it = Item.query.get_or_404(id)
    u = usuario_atual()
    if u.perfil in ('mestre', 'tecnico_seguranca', 'analista') or (u.perfil != 'admin' and it.almoxarifado_id not in u.almoxarifados_permitidos()):
        flash('Acesso negado.', 'danger')
        return redirect(url_for('item', id=id))
    # Almoxarife só vê seus próprios almoxarifados no select
    if u.perfil == 'admin':
        almoxarifados = Almoxarifado.query.all()
    else:
        ids = u.almoxarifados_permitidos()
        almoxarifados = Almoxarifado.query.filter(Almoxarifado.id.in_(ids)).all() if ids else []
    if request.method == 'POST':
        it.nome = request.form['nome']
        it.codigo = request.form['codigo']
        it.unidade = request.form['unidade']
        it.estoque_minimo = float(request.form.get('estoque_minimo', 0))
        it.almoxarifado_id = int(request.form['almoxarifado_id'])
        it.categoria = request.form.get('categoria', 'geral')
        it.ca = request.form.get('ca', '').strip() or None
        # Atualizar quantidade se informada (admin pode corrigir o valor)
        qtd_str = request.form.get('quantidade')
        if qtd_str is not None and qtd_str != '':
            try:
                nova_qtd = float(qtd_str)
                if nova_qtd != it.quantidade:
                    # Registrar ajuste como movimentação
                    diff = nova_qtd - it.quantidade
                    tipo = 'entrada' if diff > 0 else 'saida'
                    db.session.add(Movimentacao(
                        tipo=tipo,
                        quantidade=abs(diff),
                        responsavel=u.nome if u else 'Sistema',
                        observacao=f'Ajuste manual: {it.quantidade} → {nova_qtd} {it.unidade}',
                        item_id=it.id
                    ))
                    it.quantidade = nova_qtd
            except ValueError:
                pass
        db.session.commit()
        flash('Item atualizado!', 'success')
        return redirect(url_for('almoxarifado', id=it.almoxarifado_id))
    return render_template('form_item.html', item=it, almoxarifados=almoxarifados)

@app.route('/item/<int:id>/deletar', methods=['POST'])
@login_required
def deletar_item(id):
    it = Item.query.get_or_404(id)
    u = usuario_atual()
    # Admin pode deletar qualquer item; almoxarife pode deletar itens do próprio almoxarifado
    if u.perfil == 'admin':
        pass  # permitido
    elif u.perfil == 'almoxarife' and it.almoxarifado_id in u.almoxarifados_permitidos():
        pass  # almoxarife do almoxarifado pode deletar
    else:
        flash('Sem permissão para deletar este item.', 'danger')
        return redirect(url_for('item', id=id))
    alm_id = it.almoxarifado_id
    # Soft delete — preserva histórico e evita erro de constraint com requisições/movimentações
    it.ativo = False
    db.session.commit()
    flash('Item removido!', 'warning')
    return redirect(url_for('almoxarifado', id=alm_id))

# ── MOVIMENTAÇÃO EM LOTE ─────────────────────────────────────────────────────

@app.route('/movimentacao/lote', methods=['GET', 'POST'])
@login_required
def movimentacao_lote():
    u = usuario_atual()
    # Analista só pode visualizar — sem movimentação
    if u.perfil == 'analista':
        flash('Analistas não têm permissão para registrar movimentações.', 'danger')
        return redirect(url_for('index'))
    almoxarifados = Almoxarifado.query.all() if u.perfil == 'admin' else \
        Almoxarifado.query.filter(Almoxarifado.id.in_(u.almoxarifados_permitidos())).all()

    # Histórico das últimas 20 movimentações
    if u.perfil == 'admin':
        historico = Movimentacao.query.order_by(Movimentacao.data.desc()).limit(20).all()
        requisicoes_hist = Requisicao.query.order_by(Requisicao.data_retirada.desc()).limit(20).all()
    else:
        ids = u.almoxarifados_permitidos()
        historico = (Movimentacao.query.join(Item)
                     .filter(Item.almoxarifado_id.in_(ids))
                     .order_by(Movimentacao.data.desc()).limit(20).all())
        requisicoes_hist = (Requisicao.query.join(Item)
                            .filter(Item.almoxarifado_id.in_(ids))
                            .order_by(Requisicao.data_retirada.desc()).limit(20).all())

    # Montar JSON com itens por almoxarifado
    itens_json = {}
    for alm in almoxarifados:
        itens_json[str(alm.id)] = [
            {'id': it.id, 'nome': it.nome, 'quantidade': it.quantidade,
             'unidade': it.unidade, 'categoria': it.categoria or 'geral',
             'ca': it.ca or ''}
            for it in alm.itens
        ]

    if request.method == 'POST':
        alm_id      = int(request.form['almoxarifado_id'])
        if u.perfil != 'admin' and alm_id not in u.almoxarifados_permitidos():
            flash('Acesso negado.', 'danger')
            return redirect(url_for('movimentacao_lote'))

        tipo        = request.form['tipo']
        responsavel = request.form.get('responsavel', '')
        observacao  = request.form.get('observacao', '')

        erros = []
        movs  = []

        # Coletar índices sem depender de total_linhas
        indices = set()
        for key in request.form.keys():
            if key.startswith('item_id_'):
                try:
                    indices.add(int(key.split('_')[-1]))
                except ValueError:
                    pass

        for i in sorted(indices):
            item_id    = request.form.get(f'item_id_{i}')
            qtd_str    = request.form.get(f'quantidade_{i}')
            colab      = request.form.get(f'colaborador_{i}', '').strip()
            resp_linha = request.form.get(f'responsavel_{i}', '').strip() or responsavel
            ca_linha   = request.form.get(f'ca_{i}', '').strip()

            if not item_id or not qtd_str:
                continue

            it = db.session.get(Item, item_id)
            try:
                qtd = float(qtd_str)
            except ValueError:
                continue

            if not it or qtd <= 0:
                continue

            if tipo == 'saida' and qtd > it.quantidade:
                erros.append(f'"{it.nome}": estoque insuficiente ({it.quantidade} {it.unidade})')
                continue

            # Devolução = entrada no estoque com observação especial
            tipo_real = 'entrada' if tipo in ('devolucao_epi', 'devolucao_ferramenta') else tipo
            it.quantidade = round(it.quantidade + qtd if tipo_real == 'entrada' else it.quantidade - qtd, 4)

            if tipo == 'saida' and colab:
                obs_linha = f'liberado P/ {colab}'
                if observacao:
                    obs_linha += f' | {observacao}'
            elif tipo == 'devolucao_epi':
                obs_linha = f'Devolução EPI — {colab}' if colab else 'Devolução EPI'
                if observacao:
                    obs_linha += f' | {observacao}'
            elif tipo == 'devolucao_ferramenta':
                obs_linha = f'Devolução Ferramenta — {colab}' if colab else 'Devolução Ferramenta'
                if observacao:
                    obs_linha += f' | {observacao}'
            else:
                obs_linha = observacao

            movs.append(Movimentacao(
                tipo=tipo_real, quantidade=qtd,
                responsavel=resp_linha,
                observacao=obs_linha,
                item_id=it.id
            ))

        if movs:
            db.session.add_all(movs)
            db.session.commit()
            tipo_label = '📥 Entrada' if request.form['tipo'] == 'entrada' else '📤 Saída'
            alm = db.session.get(Almoxarifado, alm_id)

            # Se é saída de EPI via AJAX, retorna JSON com IDs para abrir câmera
            if (request.headers.get('X-Requested-With') == 'XMLHttpRequest'
                    and request.form.get('tipo') == 'saida'):
                epi_movs = [m for m in movs if db.session.get(Item, m.item_id).categoria == 'epi']
                return jsonify({
                    'ok': True,
                    'msg': f'{len(movs)} item(ns) registrado(s).',
                    'epi_mov_ids': [m.id for m in epi_movs],
                    'redirect': url_for('movimentacao_lote')
                })

            flash_html(
                f'<strong>{escape(tipo_label)} registrada!</strong> '
                f'{len(movs)} item(ns) movimentado(s) em <strong>{escape(alm.nome if alm else "")}</strong>. '
                f'<a href="/almoxarifado/{alm_id}" class="alert-link">Ver Almoxarifado</a>',
                'success'
            )
        elif not erros:
            flash('Adicione pelo menos um item antes de confirmar.', 'warning')

        for e in erros:
            flash_html(
                f'<strong>Estoque insuficiente:</strong> {escape(e)} '
                f'<a href="/almoxarifado/{alm_id}" class="alert-link">Consultar Estoque</a>',
                'danger'
            )

        return redirect(url_for('movimentacao_lote'))

    import json
    return render_template('movimentacao_lote.html',
                           almoxarifados=almoxarifados,
                           itens_json=json.dumps(itens_json),
                           historico=historico,
                           requisicoes=requisicoes_hist)

@app.route('/item/<int:id>/movimentar', methods=['POST'])
@login_required
def movimentar(id):
    it = Item.query.get_or_404(id)
    u = usuario_atual()
    # Analista não pode fazer movimentação de forma alguma
    if u.perfil == 'analista':
        flash('Analistas não têm permissão para registrar movimentações.', 'danger')
        return redirect(url_for('item', id=id))
    if u.perfil != 'admin' and it.almoxarifado_id not in u.almoxarifados_permitidos():
        flash('Acesso negado.', 'danger')
        return redirect(url_for('item', id=id))

    tipo = request.form['tipo']
    qtd = float(request.form['quantidade'])
    responsavel = request.form.get('responsavel', '').strip()
    observacao = request.form.get('observacao', '').strip()

    if tipo == 'saida' and qtd > it.quantidade:
        flash('Quantidade insuficiente em estoque!', 'danger')
        return redirect(url_for('item', id=id))

    # Na saída, a observação já vem montada pelo JS como "liberado P/ Nome | req X"
    obs_final = observacao

    it.quantidade = round(it.quantidade + qtd if tipo == 'entrada' else it.quantidade - qtd, 4)
    mov = Movimentacao(
        tipo=tipo, quantidade=qtd,
        responsavel=responsavel,
        observacao=obs_final,
        item_id=id
    )
    db.session.add(mov)
    db.session.commit()
    flash(f'{"Entrada" if tipo == "entrada" else "Saída"} de {qtd} {it.unidade} registrada!', 'success')
    return redirect(url_for('item', id=id))

# ── REQUISIÇÕES ──────────────────────────────────────────────────────────────

@app.route('/requisicoes')
@login_required
def requisicoes():
    u = usuario_atual()
    colaborador  = request.args.get('colaborador', '')
    status       = request.args.get('status', '')
    data_ini     = request.args.get('data_ini', '')
    data_fim     = request.args.get('data_fim', '')

    q = Requisicao.query
    if u.perfil != 'admin':
        ids = u.almoxarifados_permitidos()
        q = q.join(Item).filter(Item.almoxarifado_id.in_(ids)) if ids else q.filter(False)
    if colaborador:
        q = q.filter(Requisicao.colaborador.ilike(f'%{colaborador}%'))
    if status:
        q = q.filter(Requisicao.status == status)
    if data_ini:
        q = q.filter(Requisicao.data_retirada >= data_ini)
    if data_fim:
        q = q.filter(Requisicao.data_retirada <= data_fim + ' 23:59:59')

    reqs = q.order_by(Requisicao.data_retirada.desc()).all()

    return render_template('requisicoes.html',
        requisicoes=reqs,
        total=len(reqs),
        em_uso=sum(1 for r in reqs if r.status == 'aberta'),
        devolvidos=sum(1 for r in reqs if r.status == 'devolvida'),
        filtro_colaborador=colaborador,
        filtro_status=status,
        filtro_data_ini=data_ini,
        filtro_data_fim=data_fim
    )

@app.route('/requisicoes/nova', methods=['GET', 'POST'])
@login_required
def requisicao_nova():
    u = usuario_atual()
    almoxarifados = Almoxarifado.query.all() if u.perfil == 'admin' else \
        Almoxarifado.query.filter(Almoxarifado.id.in_(u.almoxarifados_permitidos())).all()
    itens_json = {}
    for alm in almoxarifados:
        itens_json[str(alm.id)] = [
            {'id': it.id, 'nome': it.nome, 'quantidade': it.quantidade, 'unidade': it.unidade}
            for it in alm.itens
        ]

    if request.method == 'POST':
        colaborador = request.form.get('colaborador', '')
        observacao  = request.form.get('observacao', '')
        criados = 0

        # Coletar todos os item_id_N do formulário sem depender de total_linhas
        indices = set()
        for key in request.form.keys():
            if key.startswith('item_id_'):
                try:
                    indices.add(int(key.split('_')[-1]))
                except ValueError:
                    pass

        for i in sorted(indices):
            item_id = request.form.get(f'item_id_{i}')
            qtd_str = request.form.get(f'quantidade_{i}')
            if not item_id or not qtd_str:
                continue
            it  = db.session.get(Item, item_id)
            try:
                qtd = float(qtd_str)
            except ValueError:
                continue
            if not it or qtd <= 0:
                continue
            if u.perfil != 'admin' and it.almoxarifado_id not in u.almoxarifados_permitidos():
                continue
            if qtd > it.quantidade:
                flash_html(
                    f'<strong>Estoque insuficiente:</strong> "{escape(it.nome)}" tem apenas '
                    f'<strong>{it.quantidade} {escape(it.unidade)}</strong> disponível. '
                    f'<a href="/almoxarifado/{it.almoxarifado_id}" class="alert-link">Consultar Estoque</a>',
                    'danger'
                )
                continue

            it.quantidade -= qtd
            db.session.add(Requisicao(
                colaborador=colaborador,
                observacao=observacao,
                quantidade=qtd,
                item_id=it.id
            ))
            db.session.add(Movimentacao(
                tipo='saida', quantidade=qtd,
                responsavel=colaborador,
                observacao=f'Requisicao — {observacao}',
                item_id=it.id
            ))
            criados += 1

        if criados:
            db.session.commit()
            flash_html(
                f'<strong>✅ Requisição registrada!</strong> '
                f'{criados} item(ns) retirado(s) com sucesso para <strong>{escape(colaborador)}</strong>. '
                f'<a href="/requisicoes" class="alert-link">Ver Requisições</a>',
                'success'
            )
        elif not any(True for key in request.form.keys() if key.startswith('item_id_')):
            flash('Adicione pelo menos um item antes de registrar.', 'warning')
        return redirect(url_for('requisicoes'))

    import json
    return render_template('requisicao_nova.html',
                           almoxarifados=almoxarifados,
                           itens_json=json.dumps(itens_json))

@app.route('/requisicoes/<int:id>/devolver', methods=['POST'])
@login_required
def devolver_requisicao(id):
    req = Requisicao.query.get_or_404(id)
    u = usuario_atual()
    if u.perfil != 'admin' and req.item.almoxarifado_id not in u.almoxarifados_permitidos():
        flash('Acesso negado.', 'danger')
        return redirect(url_for('requisicoes'))
    if req.status == 'aberta':
        req.status = 'devolvida'
        req.data_devolucao = agora()
        req.item.quantidade += req.quantidade
        db.session.add(Movimentacao(
            tipo='entrada', quantidade=req.quantidade,
            responsavel=req.colaborador,
            observacao=f'Devolução de requisição #{req.id}',
            item_id=req.item_id
        ))
        db.session.commit()
        flash(f'Devolução de "{req.item.nome}" registrada!', 'success')
    return redirect(url_for('requisicoes'))

@app.route('/almoxarifado/<int:id>/importar', methods=['GET', 'POST'])
@login_required
def importar_itens(id):
    u = usuario_atual()
    alm = Almoxarifado.query.get_or_404(id)
    if u.perfil != 'admin' and id not in u.almoxarifados_permitidos():
        flash('Acesso negado.', 'danger')
        return redirect(url_for('index'))

    if request.method == 'POST':
        arquivo = request.files.get('arquivo')
        modo = request.form.get('modo', 'adicionar')  # adicionar | atualizar | substituir

        if not arquivo or not arquivo.filename.endswith(('.xlsx', '.xls')):
            flash('Envie um arquivo Excel (.xlsx ou .xls).', 'danger')
            return redirect(url_for('importar_itens', id=id))

        try:
            wb = openpyxl.load_workbook(arquivo, data_only=True)
            ws = wb.active

            inseridos = 0
            atualizados = 0
            erros = []

            # Pular cabeçalho (linha 1)
            for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                # Ignorar linhas completamente vazias
                if not any(row):
                    continue

                # Colunas esperadas: codigo | nome | unidade | quantidade | estoque_minimo
                try:
                    codigo    = str(row[0]).strip() if row[0] else None
                    nome      = str(row[1]).strip() if row[1] else None
                    unidade   = str(row[2]).strip() if row[2] else 'un'
                    quantidade = float(row[3]) if row[3] is not None else 0
                    est_min   = float(row[4]) if row[4] is not None else 0
                except Exception:
                    erros.append(f'Linha {row_num}: formato inválido')
                    continue

                if not codigo or not nome:
                    erros.append(f'Linha {row_num}: código ou nome vazio')
                    continue

                item_existente = Item.query.filter_by(codigo=codigo).first()

                if item_existente:
                    if modo in ('atualizar', 'substituir'):
                        item_existente.nome = nome
                        item_existente.unidade = unidade
                        item_existente.estoque_minimo = est_min
                        item_existente.almoxarifado_id = id
                        if modo == 'substituir':
                            # Só sobrescreve a quantidade se o item não tiver movimentações
                            # Isso protege contra reimportação acidental de dados antigos
                            tem_movimentacoes = Movimentacao.query.filter_by(item_id=item_existente.id).count() > 0
                            if not tem_movimentacoes:
                                item_existente.quantidade = quantidade
                            # Se tiver movimentações, ignora a quantidade do Excel (protege o saldo real)
                        else:  # atualizar = somar
                            item_existente.quantidade += quantidade
                        atualizados += 1
                    # modo 'adicionar' ignora duplicatas
                else:
                    novo = Item(
                        codigo=codigo, nome=nome, unidade=unidade,
                        quantidade=quantidade, estoque_minimo=est_min,
                        almoxarifado_id=id
                    )
                    db.session.add(novo)
                    inseridos += 1

            db.session.commit()

            msg = f'Importação concluída: {inseridos} inseridos, {atualizados} atualizados.'
            if erros:
                msg += f' {len(erros)} linha(s) com erro.'
            flash(msg, 'success' if not erros else 'warning')
            for e in erros[:10]:  # mostrar até 10 erros
                flash(e, 'danger')

        except Exception as e:
            flash(f'Erro ao processar arquivo: {str(e)}', 'danger')

        return redirect(url_for('almoxarifado', id=id))

    return render_template('importar_itens.html', almoxarifado=alm)

@app.route('/almoxarifado/<int:id>/modelo_excel')
@login_required
def modelo_excel(id):
    alm = Almoxarifado.query.get_or_404(id)
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Itens'
    h_fill = PatternFill('solid', fgColor='1A3A5C')
    h_font = Font(bold=True, color='FFFFFF', size=11)
    borda  = Border(left=Side(style='thin'), right=Side(style='thin'),
                    top=Side(style='thin'), bottom=Side(style='thin'))
    for col, h in enumerate(['Codigo', 'Nome', 'Unidade', 'Quantidade', 'Estoque Minimo'], 1):
        c = ws.cell(row=1, column=col, value=h)
        c.font = h_font; c.fill = h_fill
        c.alignment = Alignment(horizontal='center'); c.border = borda
    for r, ex in enumerate([('CIM-001','Cimento CP-II','sc',500,100),
                             ('ARG-002','Areia Grossa','m3',30,5),
                             ('EPI-003','Capacete','un',20,5)], 2):
        for c, v in enumerate(ex, 1):
            ws.cell(row=r, column=c, value=v).border = borda
    for i, w in enumerate([14,40,10,14,16], 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    buf = io.BytesIO()
    wb.save(buf); buf.seek(0)
    return send_file(buf, as_attachment=True,
                     download_name=f'modelo_{alm.nome.replace(" ","_")}.xlsx',
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

@app.route('/relatorios/consumo')
@login_required
def relatorio_consumo():
    u = usuario_atual()
    alm_id = request.args.get('almoxarifado_id', type=int)
    data_ini = request.args.get('data_ini', str(date.today().replace(day=1)))
    data_fim = request.args.get('data_fim', str(date.today()))

    # Determina almoxarifados permitidos filtrados por cidade
    if u.perfil == 'admin':
        ids_perm = None
    else:
        ids_perm = set(u.almoxarifados_permitidos())
        # Técnico/analista: expande para todos da sua cidade
        if u.perfil in ('tecnico_seguranca', 'analista') and u.almoxarifado_id:
            alm_ref = db.session.get(Almoxarifado, u.almoxarifado_id)
            if alm_ref and alm_ref.cidade:
                ids_perm = {a.id for a in Almoxarifado.query.filter_by(cidade=alm_ref.cidade).all()}

    if alm_id and ids_perm and alm_id not in ids_perm:
        flash('Acesso negado.', 'danger')
        return redirect(url_for('index'))

    query = Movimentacao.query.filter(
        Movimentacao.tipo == 'saida',
        Movimentacao.data >= data_ini,
        Movimentacao.data <= data_fim + ' 23:59:59'
    )
    if alm_id:
        query = query.join(Item).filter(Item.almoxarifado_id == alm_id)
    elif u.perfil != 'admin':
        if ids_perm:
            query = query.join(Item).filter(Item.almoxarifado_id.in_(ids_perm))
        else:
            query = query.filter(False)

    movs = query.order_by(Movimentacao.data.desc()).all()

    if u.perfil == 'admin':
        almoxarifados = Almoxarifado.query.all()
    elif ids_perm:
        almoxarifados = Almoxarifado.query.filter(Almoxarifado.id.in_(ids_perm)).all()
    else:
        almoxarifados = []

    return render_template('relatorio_consumo.html', movimentacoes=movs,
                           almoxarifados=almoxarifados, data_ini=data_ini,
                           data_fim=data_fim, alm_id=alm_id)

@app.route('/relatorios/consumo/exportar')
@login_required
def exportar_consumo():
    u = usuario_atual()
    alm_id = request.args.get('almoxarifado_id', type=int)
    data_ini = request.args.get('data_ini', str(date.today().replace(day=1)))
    data_fim = request.args.get('data_fim', str(date.today()))
    if alm_id and u.perfil != 'admin' and alm_id not in u.almoxarifados_permitidos():
        flash('Acesso negado.', 'danger')
        return redirect(url_for('index'))
    query = Movimentacao.query.filter(
        Movimentacao.tipo == 'saida',
        Movimentacao.data >= data_ini,
        Movimentacao.data <= data_fim + ' 23:59:59'
    )
    if alm_id:
        query = query.join(Item).filter(Item.almoxarifado_id == alm_id)
    elif u.perfil != 'admin':
        query = query.join(Item).filter(Item.almoxarifado_id.in_(u.almoxarifados_permitidos()))
    movs = query.order_by(Movimentacao.data.desc()).all()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Consumo'
    h_fill = PatternFill('solid', fgColor='1A3A5C')
    h_font = Font(bold=True, color='FFFFFF', size=11)
    borda   = Border(left=Side(style='thin'), right=Side(style='thin'),
                     top=Side(style='thin'), bottom=Side(style='thin'))

    alm_nome = db.session.get(Almoxarifado, alm_id).nome if alm_id else 'Todos os Almoxarifados'
    ws.merge_cells('A1:H1')
    ws['A1'] = f'Relatório de Consumo — {alm_nome}'
    ws['A1'].font = Font(bold=True, size=13, color='1A3A5C')
    ws.merge_cells('A2:H2')
    ws['A2'] = f'Período: {data_ini} a {data_fim}   |   Gerado em: {agora().strftime("%d/%m/%Y %H:%M")}'
    ws['A2'].font = Font(italic=True, size=10, color='666666')

    headers = ['Data', 'Código', 'Item', 'Categoria', 'Almoxarifado', 'Quantidade', 'Responsável', 'Colaborador', 'Observação']
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=4, column=col, value=h)
        c.font = h_font; c.fill = h_fill
        c.alignment = Alignment(horizontal='center'); c.border = borda

    cat_label = {
        'epi': 'EPI', 'maquinario': 'Maquinário', 'eletrica': 'Elétrica',
        'hidraulica': 'Hidráulica', 'gas': 'Gás', 'geral': 'Geral'
    }

    for row_num, mov in enumerate(movs, 5):
        # Separar colaborador da observação
        obs = mov.observacao or ''
        if 'liberado P/' in obs or 'liberado p/' in obs:
            partes = re.split(r'liberado [Pp]/', obs, maxsplit=1)
            resto = partes[-1].split(' | ', 1)
            colab_nome = resto[0].strip()
            obs_limpa = resto[1].strip() if len(resto) > 1 else ''
        elif 'Colaborador:' in obs:
            partes = obs.split('Colaborador:', 1)[-1].split('|', 1)
            colab_nome = partes[0].strip()
            obs_limpa = partes[1].strip() if len(partes) > 1 else ''
        else:
            colab_nome = ''
            obs_limpa = obs

        cat = cat_label.get(mov.item.categoria or 'geral', mov.item.categoria or 'Geral')
        dados = [
            mov.data.strftime('%d/%m/%Y %H:%M'),
            mov.item.codigo, mov.item.nome,
            cat,
            mov.item.almoxarifado.nome,
            f'{mov.quantidade} {mov.item.unidade}',
            mov.responsavel or '',
            colab_nome,
            obs_limpa
        ]
        for col, val in enumerate(dados, 1):
            c = ws.cell(row=row_num, column=col, value=val)
            c.border = borda
            if row_num % 2 == 0:
                c.fill = PatternFill('solid', fgColor='F0F4F8')

    for i, w in enumerate([18, 14, 45, 14, 35, 14, 20, 30, 30], 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    buf = io.BytesIO()
    wb.save(buf); buf.seek(0)
    return send_file(buf, as_attachment=True,
                     download_name=f'consumo_{data_ini}_a_{data_fim}.xlsx',
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

@app.route('/relatorios/consumo-por-pessoa')
@login_required
def relatorio_consumo_pessoa():
    """Relatório de consumo agrupado por colaborador (extraído da observação)."""
    u = usuario_atual()
    alm_id = request.args.get('almoxarifado_id', type=int)
    data_ini = request.args.get('data_ini', str(date.today().replace(day=1)))
    data_fim = request.args.get('data_fim', str(date.today()))
    responsavel_filtro = request.args.get('responsavel', '').strip()

    # Técnico de segurança e analista com permissão ver_relatorios vêem
    # todos os almoxarifados da sua cidade — não de outras cidades
    tem_perm_relatorio = u.perfil in ('tecnico_seguranca', 'analista') or any(
        p.permissao == 'ver_relatorios' and p.ativo for p in u.permissoes_extras
    )

    # IDs de almoxarifados que o usuário pode ver (filtrado por cidade para não-admins)
    if u.perfil == 'admin':
        ids_permitidos = None  # sem restrição
    else:
        ids_permitidos = u.almoxarifados_permitidos()
        # Para técnico/analista: expande para todos da mesma cidade
        if tem_perm_relatorio and u.almoxarifado_id:
            alm_ref = db.session.get(Almoxarifado, u.almoxarifado_id)
            if alm_ref and alm_ref.cidade:
                alms_cidade = Almoxarifado.query.filter_by(cidade=alm_ref.cidade).all()
                ids_permitidos = set(a.id for a in alms_cidade)

    if alm_id and u.perfil != 'admin' and ids_permitidos and alm_id not in ids_permitidos:
        flash('Acesso negado.', 'danger')
        return redirect(url_for('index'))

    query = Movimentacao.query.filter(
        Movimentacao.tipo == 'saida',
        Movimentacao.data >= data_ini,
        Movimentacao.data <= data_fim + ' 23:59:59'
    )
    if alm_id:
        query = query.join(Item).filter(Item.almoxarifado_id == alm_id)
    elif u.perfil != 'admin':
        if ids_permitidos:
            query = query.join(Item).filter(Item.almoxarifado_id.in_(ids_permitidos))
        else:
            # sem almoxarifados permitidos — retorna vazio
            query = query.filter(False)

    movs = query.order_by(Movimentacao.data.desc()).all()

    import re

    # Nomes de usuários do sistema (mestres, técnicos, almoxarifes) para excluir
    usuarios_sistema = {u.nome.strip().lower() for u in Usuario.query.filter(
        Usuario.perfil.in_(['admin', 'almoxarife', 'mestre', 'tecnico_seguranca'])
    ).all()}

    def extrair_colaborador(mov):
        """Extrai o nome do colaborador da observação."""
        obs = mov.observacao or ''
        # Formato requisição mestre: "Req. Mestre #X — Colaborador: Nome"
        m = re.search(r'[Cc]olaborador[:\s]+([^|]+)', obs)
        if m:
            nome = m.group(1).strip()
            if nome:
                return nome
        # Formato movimentação avulsa: "liberado P/ Nome | ..."
        m = re.search(r'liberado\s+[Pp][/\s]+([^|]+)', obs, re.IGNORECASE)
        if m:
            nome = m.group(1).strip()
            if nome:
                return nome
        return None  # sem colaborador identificável

    def normalizar_nome(nome):
        """Remove sufixos de requisição do nome e normaliza capitalização."""
        if not nome:
            return None
        # Remove ' | req XXXX' ou ' | REQ XXXX' ou ' | ajuste ...' do final
        nome_limpo = re.sub(r'\s*[|·]\s*.+$', '', nome).strip()
        if not nome_limpo:
            return None
        # Normaliza: remove espaços duplos, converte para título (João Silva)
        nome_limpo = re.sub(r'\s+', ' ', nome_limpo).strip().upper()
        return nome_limpo if nome_limpo else None

    def e_nome_valido(nome):
        """Filtra nomes inválidos: ajustes, vazios, nomes do sistema."""
        if not nome or len(nome) < 2:
            return False
        nome_lower = nome.lower().strip()
        # Excluir termos de ajuste sistêmico
        termos_invalidos = [
            'ajuste', 'sistemico', 'sistêmico', 'reajuste',
            'req -', 'req-', '- req', 'sem responsável',
            'sem responsavel', 'ajuste de estoque'
        ]
        if any(t in nome_lower for t in termos_invalidos):
            return False
        # Excluir nomes de usuários do sistema (mestres, admins, etc.)
        if nome_lower in usuarios_sistema:
            return False
        return True

    # Filtrar por nome se informado
    from collections import defaultdict
    por_pessoa = defaultdict(list)
    for mov in movs:
        colab_raw = extrair_colaborador(mov)
        colaborador = normalizar_nome(colab_raw)
        if not e_nome_valido(colaborador):
            continue
        if responsavel_filtro and responsavel_filtro.lower() not in colaborador.lower():
            continue
        por_pessoa[colaborador].append(mov)

    # Calcular totais por pessoa
    resumo = []
    for nome, lista in sorted(por_pessoa.items()):
        total_movs = len(lista)
        itens_distintos = len(set(m.item_id for m in lista))
        # Coletar números de requisição únicos associados a esta pessoa
        reqs = []
        for mov in lista:
            obs = mov.observacao or ''
            m_req = re.search(r'[Rr]eq\.?\s*(?:[Mm]estre\s*)?#?(\d+)', obs)
            if m_req:
                r_num = m_req.group(1)
                if r_num not in reqs:
                    reqs.append(r_num)
        resumo.append({
            'nome': nome,
            'movimentacoes': lista,
            'total_movs': total_movs,
            'itens_distintos': itens_distintos,
            'reqs': reqs,
        })

    # Lista de almoxarifados para o filtro — respeita cidade do usuário
    if u.perfil == 'admin':
        almoxarifados = Almoxarifado.query.all()
    elif ids_permitidos:
        almoxarifados = Almoxarifado.query.filter(Almoxarifado.id.in_(ids_permitidos)).all()
    else:
        almoxarifados = []

    return render_template('relatorio_consumo_pessoa.html',
                           resumo=resumo,
                           almoxarifados=almoxarifados,
                           data_ini=data_ini,
                           data_fim=data_fim,
                           alm_id=alm_id,
                           responsavel_filtro=responsavel_filtro,
                           total_geral=sum(p['total_movs'] for p in resumo))

@app.route('/relatorios/consumo-por-pessoa/exportar')
@login_required
def exportar_consumo_pessoa():
    """Exporta relatório de consumo por pessoa em Excel com 2 abas."""
    import re
    from collections import defaultdict

    u = usuario_atual()
    alm_id           = request.args.get('almoxarifado_id', type=int)
    data_ini         = request.args.get('data_ini', str(date.today().replace(day=1)))
    data_fim         = request.args.get('data_fim', str(date.today()))
    responsavel_filtro = request.args.get('responsavel', '').strip()

    if alm_id and u.perfil != 'admin' and alm_id not in u.almoxarifados_permitidos():
        flash('Acesso negado.', 'danger')
        return redirect(url_for('index'))

    query = Movimentacao.query.filter(
        Movimentacao.tipo == 'saida',
        Movimentacao.data >= data_ini,
        Movimentacao.data <= data_fim + ' 23:59:59'
    )
    if alm_id:
        query = query.join(Item).filter(Item.almoxarifado_id == alm_id)
    elif u.perfil != 'admin':
        query = query.join(Item).filter(Item.almoxarifado_id.in_(u.almoxarifados_permitidos()))

    movs = query.order_by(Movimentacao.data.asc()).all()

    def extrair_colaborador(mov):
        obs = mov.observacao or ''
        m = re.search(r'liberado\s+[Pp][/\s]+(.+)', obs, re.IGNORECASE)
        if m: return m.group(1).strip()
        m = re.search(r'[Cc]olaborador[:\s]+([^|]+)', obs)
        if m: return m.group(1).strip()
        return mov.responsavel or 'Sem responsável'

    def normalizar_nome(nome):
        return re.sub(r'\s*[|·]\s*[Rr][Ee][Qq]\.?\s*\d+.*$', '', nome).strip() or nome

    por_pessoa = defaultdict(list)
    for mov in movs:
        colab = normalizar_nome(extrair_colaborador(mov))
        if responsavel_filtro and responsavel_filtro.lower() not in colab.lower():
            continue
        por_pessoa[colab].append(mov)

    # Estilos
    h_fill  = PatternFill('solid', fgColor='1A3A5C')
    h_font  = Font(bold=True, color='FFFFFF', size=11)
    z_fill  = PatternFill('solid', fgColor='F0F4F8')
    borda   = Border(left=Side(style='thin'), right=Side(style='thin'),
                     top=Side(style='thin'), bottom=Side(style='thin'))
    centro  = Alignment(horizontal='center', vertical='center')
    esq     = Alignment(horizontal='left',   vertical='center')
    alm_nome = db.session.get(Almoxarifado, alm_id).nome if alm_id else 'Todos'
    total_geral = sum(len(v) for v in por_pessoa.values())

    wb = openpyxl.Workbook()

    # ── ABA 1: Resumo ─────────────────────────────────────────────────────────
    ws1 = wb.active
    ws1.title = 'Resumo'
    for col, w in enumerate([6, 40, 20, 18, 18], 1):
        ws1.column_dimensions[get_column_letter(col)].width = w

    ws1.merge_cells('A1:E1')
    ws1['A1'].value = f'Consumo por Pessoa — {alm_nome}'
    ws1['A1'].font = Font(bold=True, size=13, color='1A3A5C')
    ws1['A1'].fill = PatternFill('solid', fgColor='E8F0F7')
    ws1['A1'].alignment = centro

    ws1.merge_cells('A2:E2')
    ws1['A2'].value = f'Período: {data_ini} a {data_fim}   |   Gerado em: {agora().strftime("%d/%m/%Y %H:%M")}'
    ws1['A2'].font = Font(italic=True, size=10, color='666666')
    ws1['A2'].alignment = centro

    for col, h in enumerate(['#', 'Funcionário', 'Total Retiradas', 'Itens Distintos', 'Participação (%)'], 1):
        c = ws1.cell(row=4, column=col, value=h)
        c.font = h_font; c.fill = h_fill; c.alignment = centro; c.border = borda

    for i, (nome, lista) in enumerate(sorted(por_pessoa.items(), key=lambda x: len(x[1]), reverse=True), 1):
        pct = round(len(lista) / total_geral * 100, 1) if total_geral else 0
        row = i + 4
        for col, val in enumerate([i, nome, len(lista), len(set(m.item_id for m in lista)), f'{pct}%'], 1):
            c = ws1.cell(row=row, column=col, value=val)
            c.border = borda
            c.alignment = esq if col == 2 else centro
            if row % 2 == 0: c.fill = z_fill

    # Linha total
    r = len(por_pessoa) + 5
    for col, val in enumerate(['', f'{len(por_pessoa)} pessoa(s)', total_geral, '', '100%'], 1):
        c = ws1.cell(row=r, column=col, value=val)
        c.font = Font(bold=True)
        c.fill = PatternFill('solid', fgColor='D0E4F7')
        c.border = borda; c.alignment = centro

    # ── ABA 2: Detalhes ───────────────────────────────────────────────────────
    ws2 = wb.create_sheet('Detalhes')
    for col, w in enumerate([35, 18, 14, 45, 14, 30, 25], 1):
        ws2.column_dimensions[get_column_letter(col)].width = w

    ws2.merge_cells('A1:G1')
    ws2['A1'].value = f'Detalhes por Funcionário — {alm_nome}'
    ws2['A1'].font = Font(bold=True, size=13, color='1A3A5C')
    ws2['A1'].fill = PatternFill('solid', fgColor='E8F0F7')
    ws2['A1'].alignment = centro

    ws2.merge_cells('A2:G2')
    ws2['A2'].value = f'Período: {data_ini} a {data_fim}   |   Gerado em: {agora().strftime("%d/%m/%Y %H:%M")}'
    ws2['A2'].font = Font(italic=True, size=10, color='666666')
    ws2['A2'].alignment = centro

    for col, h in enumerate(['Funcionário', 'Data', 'Código', 'Item', 'Quantidade', 'Almoxarifado', 'Liberado por'], 1):
        c = ws2.cell(row=4, column=col, value=h)
        c.font = h_font; c.fill = h_fill; c.alignment = centro; c.border = borda

    row_num = 5
    for nome, lista in sorted(por_pessoa.items()):
        # Cabeçalho do funcionário
        ws2.merge_cells(f'A{row_num}:G{row_num}')
        c = ws2.cell(row=row_num, column=1, value=f'👷 {nome}  ({len(lista)} retirada(s))')
        c.font = Font(bold=True, color='FFFFFF', size=11)
        c.fill = PatternFill('solid', fgColor='2E6DA4')
        c.alignment = esq; c.border = borda
        row_num += 1

        for mov in sorted(lista, key=lambda m: m.data):
            for col, val in enumerate([nome, mov.data.strftime('%d/%m/%Y %H:%M'),
                                        mov.item.codigo, mov.item.nome,
                                        f'{mov.quantidade} {mov.item.unidade}',
                                        mov.item.almoxarifado.nome,
                                        mov.responsavel or '—'], 1):
                c = ws2.cell(row=row_num, column=col, value=val)
                c.font = Font(size=9); c.border = borda
                c.alignment = esq if col in [1,3,4] else centro
                if row_num % 2 == 0: c.fill = z_fill
            row_num += 1
        row_num += 1  # linha em branco entre funcionários

    buf = io.BytesIO()
    wb.save(buf); buf.seek(0)
    return send_file(buf, as_attachment=True,
                     download_name=f'consumo_por_pessoa_{data_ini}_a_{data_fim}.xlsx',
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')


@app.route('/relatorios/ficha-epi')
@login_required
def ficha_epi():
    """Página para gerar ficha de EPI individual por funcionário."""
    import re
    u = usuario_atual()
    # Técnico de segurança e analista vêem todos os almoxarifados
    tem_perm_relatorio = u.perfil in ('tecnico_seguranca', 'analista') or any(
        p.permissao == 'ver_relatorios' and p.ativo for p in u.permissoes_extras
    )
    query = Movimentacao.query.join(Item).filter(
        Movimentacao.tipo == 'saida', Item.categoria == 'epi')
    if u.perfil not in ('admin', 'analista') and not tem_perm_relatorio:
        query = query.filter(Item.almoxarifado_id.in_(u.almoxarifados_permitidos()))
    movs = query.order_by(Movimentacao.data.desc()).all()

    # Coleta funcionários únicos com sua requisição mais recente
    visto = {}  # nome_lower -> {nome, req}
    for mov in movs:
        obs = mov.observacao or ''
        nome = None
        req_num = None

        m = re.search(r'liberado\s+[Pp][/\s]+(.+)', obs, re.IGNORECASE)
        if m:
            nome = m.group(1).strip()
        elif 'Colaborador:' in obs:
            partes = obs.split('Colaborador:', 1)
            nome = partes[-1].split('|')[0].strip()

        # Tenta extrair número da requisição da observação
        m_req = re.search(r'[Rr]eq\.?\s*(?:Mestre\s*)?#?(\d+)', obs)
        if m_req:
            req_num = m_req.group(1)

        if nome and nome.lower() not in visto:
            visto[nome.lower()] = {'nome': nome, 'req': req_num or ''}

    funcionarios = sorted(visto.values(), key=lambda x: x['nome'])
    return render_template('ficha_epi.html',
                           funcionarios=funcionarios,
                           data_ini='2020-01-01',
                           data_fim=str(date.today()))


@app.route('/relatorios/ficha-epi/exportar')
@login_required
def exportar_ficha_epi():
    """Exporta FORM.SEG.014 — Ficha de Controle de EPIs e Uniformes."""
    import re
    u = usuario_atual()
    funcionario = request.args.get('funcionario', '').strip()
    data_ini    = request.args.get('data_ini', '2020-01-01')
    data_fim    = request.args.get('data_fim', str(date.today()))

    if not funcionario:
        flash('Selecione um funcionário.', 'warning')
        return redirect(url_for('ficha_epi'))

    query = (Movimentacao.query.join(Item)
                  .filter(Movimentacao.tipo == 'saida',
                          Item.categoria == 'epi',
                          Movimentacao.data >= data_ini,
                          Movimentacao.data <= data_fim + ' 23:59:59'))
    if u.perfil not in ('admin', 'analista'):
        tem_perm_relatorio = u.perfil in ('tecnico_seguranca', 'analista') or any(
            p.permissao == 'ver_relatorios' and p.ativo for p in u.permissoes_extras
        )
        if not tem_perm_relatorio:
            query = query.filter(Item.almoxarifado_id.in_(u.almoxarifados_permitidos()))
    movs_todas = query.order_by(Movimentacao.data.asc()).all()

    def extrair_colab(mov):
        obs = mov.observacao or ''
        m = re.search(r'liberado\s+[Pp][/\s]+(.+)', obs, re.IGNORECASE)
        if m: return m.group(1).strip()
        m = re.search(r'[Cc]olaborador[:\s]+([^|]+)', obs)
        if m: return m.group(1).strip()
        return mov.responsavel or ''

    lista = [m for m in movs_todas if extrair_colab(m).lower() == funcionario.lower()]

    if not lista:
        flash(f'Nenhuma retirada de EPI encontrada para "{funcionario}" no período.', 'warning')
        return redirect(url_for('ficha_epi'))

    # ── Estilos ──────────────────────────────────────────────────────────────
    borda = Border(left=Side(style='thin'), right=Side(style='thin'),
                   top=Side(style='thin'), bottom=Side(style='thin'))
    borda_med = Border(left=Side(style='medium'), right=Side(style='medium'),
                       top=Side(style='medium'), bottom=Side(style='medium'))
    centro = Alignment(horizontal='center', vertical='center', wrap_text=True)
    esq    = Alignment(horizontal='left',   vertical='center', wrap_text=True)
    azul_esc = PatternFill('solid', fgColor='1F3864')
    azul_cla = PatternFill('solid', fgColor='BDD7EE')
    cinza    = PatternFill('solid', fgColor='F2F2F2')

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = funcionario[:28]

    # Larguras: A=Qtd | B=Descrição | C=C.A. | D=Data Ent | E=Assin Ent | F=Data Dev | G=Assin Dev | H=Motivo
    for col, w in zip('ABCDEFGH', [8, 40, 10, 14, 26, 14, 26, 18]):
        ws.column_dimensions[col].width = w

    def celula(ws, ref, val='', font=None, fill=None, aln=None, brd=None, height=None):
        c = ws[ref]
        c.value = val
        if font:  c.font = font
        if fill:  c.fill = fill
        if aln:   c.alignment = aln
        if brd:   c.border = brd
        return c

    def merge_row(ws, row, col_ini, col_fim, val='', font=None, fill=None, aln=None, height=None):
        ws.merge_cells(f'{col_ini}{row}:{col_fim}{row}')
        c = ws[f'{col_ini}{row}']
        c.value = val
        if font:  c.font = font
        if fill:  c.fill = fill
        if aln:   c.alignment = aln
        for col in range(ord(col_ini)-64, ord(col_fim)-64+1):
            ws.cell(row=row, column=col).border = borda
        if height: ws.row_dimensions[row].height = height
        return c

    # ── LINHA 1: Cabeçalho principal ─────────────────────────────────────────
    ws.row_dimensions[1].height = 36
    ws.merge_cells('A1:C1')
    ws['A1'].value = 'STANZA'
    ws['A1'].font = Font(bold=True, size=22, color='808080')
    ws['A1'].alignment = centro
    for c in range(1,4): ws.cell(1,c).border = borda_med

    ws.merge_cells('D1:F1')
    ws['D1'].value = 'FICHA DE CONTROLE DE EPI\'S E UNIFORMES'
    ws['D1'].font = Font(bold=True, size=13, color='1F3864')
    ws['D1'].alignment = centro
    ws['D1'].fill = azul_cla
    for c in range(4,7): ws.cell(1,c).border = borda_med

    ws.merge_cells('G1:H1')
    ws['G1'].value = 'FORM.SEG.014'
    ws['G1'].font = Font(bold=True, size=9, color='1F3864')
    ws['G1'].alignment = centro
    ws['G1'].fill = azul_cla
    for c in range(7,9): ws.cell(1,c).border = borda_med

    # ── LINHA 2: Data elaboração ──────────────────────────────────────────────
    ws.row_dimensions[2].height = 14
    ws.merge_cells('D2:F2')
    ws['D2'].value = 'Data Elaboração/Revisão: 20/10/2024'
    ws['D2'].font = Font(size=8, italic=True, color='595959')
    ws['D2'].alignment = centro
    ws.merge_cells('G2:H2')
    ws['G2'].value = 'Revisão: 00'
    ws['G2'].font = Font(size=8, italic=True, color='595959')
    ws['G2'].alignment = centro
    for c in range(1,9): ws.cell(2,c).border = borda

    # ── LINHA 3: Dados do funcionário ────────────────────────────────────────
    ws.row_dimensions[3].height = 22
    ws.merge_cells('A3:B3')
    ws['A3'].value = f'NOME: {funcionario.upper()}'
    ws['A3'].font = Font(bold=True, size=10)
    ws['A3'].alignment = esq
    ws['A3'].fill = cinza

    ws['C3'].value = f'MATRÍCULA:'
    ws['C3'].font = Font(size=9)
    ws['C3'].alignment = centro
    ws['C3'].fill = cinza

    ws.merge_cells('D3:E3')
    ws['D3'].value = 'FUNÇÃO:'
    ws['D3'].font = Font(size=9)
    ws['D3'].alignment = esq
    ws['D3'].fill = cinza

    ws.merge_cells('F3:G3')
    ws['F3'].value = f'ADMISSÃO:'
    ws['F3'].font = Font(size=9)
    ws['F3'].alignment = esq
    ws['F3'].fill = cinza

    ws['H3'].value = ''
    ws['H3'].fill = cinza
    for c in range(1,9): ws.cell(3,c).border = borda

    # ── LINHA 4: Cabeçalho da tabela ─────────────────────────────────────────
    ws.row_dimensions[4].height = 20
    for ref, val in [('A4','QUANT'), ('B4','DESCRIÇÃO'), ('C4','C.A.')]:
        ws[ref].value = val
        ws[ref].font = Font(bold=True, color='FFFFFF', size=9)
        ws[ref].fill = azul_esc
        ws[ref].alignment = centro
        ws[ref].border = borda

    ws.merge_cells('D4:E4')
    ws['D4'].value = 'ENTREGA'
    ws['D4'].font = Font(bold=True, color='FFFFFF', size=9)
    ws['D4'].fill = azul_esc
    ws['D4'].alignment = centro
    for c in range(4,6): ws.cell(4,c).border = borda

    ws.merge_cells('F4:G4')
    ws['F4'].value = 'DEVOLUÇÃO'
    ws['F4'].font = Font(bold=True, color='FFFFFF', size=9)
    ws['F4'].fill = azul_esc
    ws['F4'].alignment = centro
    for c in range(6,8): ws.cell(4,c).border = borda

    ws['H4'].value = 'MOTIVO'
    ws['H4'].font = Font(bold=True, color='FFFFFF', size=9)
    ws['H4'].fill = azul_esc
    ws['H4'].alignment = centro
    ws['H4'].border = borda

    # ── LINHA 5: Sub-cabeçalho ───────────────────────────────────────────────
    ws.row_dimensions[5].height = 16
    for ref, val in [('A5',''), ('B5',''), ('C5',''),
                     ('D5','DATA'), ('E5','ASSINATURA'),
                     ('F5','DATA'), ('G5','ASSINATURA'), ('H5','')]:
        ws[ref].value = val
        ws[ref].font = Font(bold=True, color='FFFFFF', size=8)
        ws[ref].fill = azul_esc
        ws[ref].alignment = centro
        ws[ref].border = borda

    # ── LINHAS DE DADOS ──────────────────────────────────────────────────────
    row = 6
    movs_com_foto = []  # guarda movs que têm foto para aba de comprovantes
    for mov in lista:
        ws.row_dimensions[row].height = 18
        fill_z = PatternFill('solid', fgColor='EBF3FB') if row % 2 == 0 else None
        tem_foto = bool(mov.foto_url)
        for col, val in zip('ABCDEFGH', [
            f'{mov.quantidade} {mov.item.unidade}',
            mov.item.nome, mov.item.ca or '',
            mov.data.strftime('%d/%m/%Y'), '',
            '', '',
            '📸 Ver aba' if tem_foto else ''
        ]):
            c = ws[f'{col}{row}']
            c.value = val
            c.font = Font(size=9)
            c.alignment = esq if col == 'B' else centro
            c.border = borda
            if fill_z: c.fill = fill_z
        if tem_foto:
            movs_com_foto.append((row, mov))
        row += 1

    # Linhas em branco (mínimo 14 no total conforme formulário)
    total_linhas = max(14, len(lista) + 4)
    while row <= 5 + total_linhas:
        ws.row_dimensions[row].height = 18
        for col in 'ABCDEFGH':
            ws[f'{col}{row}'].border = borda
            ws[f'{col}{row}'].value = '/    /' if col in ('D','F') else ''
            ws[f'{col}{row}'].font = Font(size=9, color='BFBFBF')
            ws[f'{col}{row}'].alignment = centro
        row += 1

    # ── TERMO DE RESPONSABILIDADE ────────────────────────────────────────────
    row += 1
    ws.row_dimensions[row].height = 16
    ws.merge_cells(f'A{row}:H{row}')
    ws[f'A{row}'].value = 'TERMO DE RESPONSABILIDADE'
    ws[f'A{row}'].font = Font(bold=True, size=10, color='1F3864')
    ws[f'A{row}'].alignment = centro
    ws[f'A{row}'].fill = azul_cla
    for c in range(1,9): ws.cell(row,c).border = borda
    row += 1

    ws.row_dimensions[row].height = 70
    ws.merge_cells(f'A{row}:H{row}')
    ws[f'A{row}'].value = (
        'Pelo presente declaro que recebi da empresa STANZA INCORPORAÇÃO E CONSTRUÇÃO LTDA., os materiais '
        'relacionados nesta ficha, assumindo o compromisso nos termos das letras "a" e "b" do ítem 1.8 da NR 1 '
        'e letras "a","b"e "c" do ítem 6.7.1 da NR 6, de usá-los em atividades ligadas ao trabalho, zelar pela '
        'sua guarda, conservação e devolvê-lo ao setor competente quando se tornar impróprio para uso ou por '
        'motivo de demissão ou afastamento.\n'
        'Em caso de perda, extravio e inutilização proposital do material recebido, autorizo a empresa, na forma '
        'prevista no parágrafo primeiro do art. 462 da CLT - Consolidação das leis do trabalho. A descontar de '
        'meu salário, inclusive no que me couber a título de indenização por rescisão de contrato de trabalho, '
        'a importância correspondente ao valor do material.'
    )
    ws[f'A{row}'].font = Font(size=8)
    ws[f'A{row}'].alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
    for c in range(1,9): ws.cell(row,c).border = borda
    row += 2

    # ── DATA ─────────────────────────────────────────────────────────────────
    ws.row_dimensions[row].height = 22
    ws.merge_cells(f'A{row}:H{row}')
    ws[f'A{row}'].value = 'Data: _______ / _______ / _____________'
    ws[f'A{row}'].font = Font(size=10)
    ws[f'A{row}'].alignment = esq
    for c in range(1,9): ws.cell(row,c).border = borda
    row += 2

    # ── CABEÇALHO BLOCO ASSINATURAS ──────────────────────────────────────────
    ws.row_dimensions[row].height = 16
    for col_ini, col_fim, label in [('A','B','FUNCIONÁRIO'), ('C','E','RESPONSÁVEL'), ('F','H','TESTEMUNHA')]:
        ws.merge_cells(f'{col_ini}{row}:{col_fim}{row}')
        ws[f'{col_ini}{row}'].value = label
        ws[f'{col_ini}{row}'].font = Font(bold=True, size=9, color='FFFFFF')
        ws[f'{col_ini}{row}'].fill = azul_esc
        ws[f'{col_ini}{row}'].alignment = centro
        for c in range(ord(col_ini)-64, ord(col_fim)-64+1):
            ws.cell(row, c).border = borda
    row += 1

    # Linha: Nome por extenso
    ws.row_dimensions[row].height = 22
    for col_ini, col_fim, placeholder in [('A','B', funcionario.upper()), ('C','E',''), ('F','H','')]:
        ws.merge_cells(f'{col_ini}{row}:{col_fim}{row}')
        ws[f'{col_ini}{row}'].value = placeholder
        ws[f'{col_ini}{row}'].font = Font(bold=True, size=9)
        ws[f'{col_ini}{row}'].alignment = centro
        ws[f'{col_ini}{row}'].fill = cinza
        for c in range(ord(col_ini)-64, ord(col_fim)-64+1):
            ws.cell(row, c).border = borda
    row += 1

    # Linha: Assinatura (espaço em branco para assinar)
    ws.row_dimensions[row].height = 38
    for col_ini, col_fim in [('A','B'), ('C','E'), ('F','H')]:
        ws.merge_cells(f'{col_ini}{row}:{col_fim}{row}')
        ws[f'{col_ini}{row}'].value = ''
        for c in range(ord(col_ini)-64, ord(col_fim)-64+1):
            ws.cell(row, c).border = borda
    row += 1

    # Linha: rótulo "Assinatura"
    ws.row_dimensions[row].height = 14
    for col_ini, col_fim in [('A','B'), ('C','E'), ('F','H')]:
        ws.merge_cells(f'{col_ini}{row}:{col_fim}{row}')
        ws[f'{col_ini}{row}'].value = 'Assinatura'
        ws[f'{col_ini}{row}'].font = Font(size=8, italic=True, color='595959')
        ws[f'{col_ini}{row}'].alignment = centro
        for c in range(ord(col_ini)-64, ord(col_fim)-64+1):
            ws.cell(row, c).border = borda

    # ── ABA DE COMPROVANTES (fotos embutidas) ─────────────────────────────────
    if movs_com_foto:
        try:
            from openpyxl.drawing.image import Image as XLImage
            ws2 = wb.create_sheet('Comprovantes')
            ws2.column_dimensions['A'].width = 20
            ws2.column_dimensions['B'].width = 50
            ws2.column_dimensions['C'].width = 15

            ws2.merge_cells('A1:C1')
            ws2['A1'].value = f'Comprovantes de Entrega — {funcionario.upper()}'
            ws2['A1'].font = Font(bold=True, size=12, color='1F3864')
            ws2['A1'].alignment = Alignment(horizontal='center', vertical='center')
            ws2['A1'].fill = PatternFill('solid', fgColor='BDD7EE')
            ws2.row_dimensions[1].height = 22

            foto_row = 3
            for _, mov in movs_com_foto:
                try:
                    import base64 as b64
                    # Extrai os bytes da imagem base64
                    header, data_b64 = mov.foto_url.split(',', 1)
                    img_bytes = b64.b64decode(data_b64)
                    img_buf = io.BytesIO(img_bytes)
                    xl_img = XLImage(img_buf)
                    # Redimensiona para caber na célula (max 300px largura)
                    scale = min(1.0, 300 / (xl_img.width or 300))
                    xl_img.width  = int(xl_img.width  * scale)
                    xl_img.height = int(xl_img.height * scale)

                    altura_linhas = max(20, int(xl_img.height * 0.75) + 5)
                    ws2.row_dimensions[foto_row].height = altura_linhas

                    ws2.cell(foto_row, 1, f'{mov.data.strftime("%d/%m/%Y")}').font = Font(size=9, bold=True)
                    ws2.cell(foto_row, 2, mov.item.nome).font = Font(size=9)
                    ws2.cell(foto_row, 3, f'{mov.quantidade} {mov.item.unidade}').font = Font(size=9)

                    foto_row += 1
                    ws2.row_dimensions[foto_row].height = altura_linhas
                    ws2.add_image(xl_img, f'A{foto_row}')
                    foto_row += max(2, int(xl_img.height / 15)) + 1
                except Exception as e_foto:
                    logger.warning(f'Foto não inserida no Excel: {e_foto}')
                    foto_row += 1
        except ImportError:
            pass  # openpyxl.drawing não disponível — ignora silenciosamente

    buf = io.BytesIO()
    wb.save(buf); buf.seek(0)
    nome_safe = funcionario.replace(' ', '_').replace('/', '-')
    return send_file(buf, as_attachment=True,
                     download_name=f'FORM-SEG-014_{nome_safe}_{data_fim}.xlsx',
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')


@app.route('/movimentacoes/excluir', methods=['POST'])
@admin_required
def excluir_movimentacoes():
    ids = request.form.getlist('mov_ids')
    if not ids:
        flash('Nenhuma movimentação selecionada.', 'warning')
        return redirect(request.referrer or url_for('relatorio_consumo_pessoa'))
    
    excluidas = 0
    for mov_id in ids:
        mov = db.session.get(Movimentacao, mov_id)
        if mov:
            # Reverter o estoque ao excluir saída
            if mov.tipo == 'saida':
                mov.item.quantidade += mov.quantidade
            elif mov.tipo == 'entrada':
                mov.item.quantidade -= mov.quantidade
            db.session.delete(mov)
            excluidas += 1
    
    db.session.commit()
    flash(f'{excluidas} movimentação(ões) excluída(s) e estoque revertido.', 'success')
    return redirect(request.referrer or url_for('relatorio_consumo_pessoa'))

def calcular_ruptura(itens_ativos, limite_dias=None):
    """Calcula previsão de ruptura para uma lista de itens ativos.

    Usa média ponderada para maior precisão:
    - Últimos 7 dias: peso 3  (tendência recente)
    - Dias 8-30: peso 1       (tendência histórica)

    Se limite_dias=None, retorna previsão para TODOS os itens com consumo.
    Retorna lista ordenada por urgência (menos dias primeiro).
    """
    from datetime import timedelta
    agora_dt = datetime.utcnow()
    corte_30 = agora_dt - timedelta(days=30)
    corte_7  = agora_dt - timedelta(days=7)

    ruptura = []
    for it in itens_ativos:
        # Já zerado ou abaixo do mínimo — aparece nos alertas normais
        if it.quantidade <= it.estoque_minimo:
            continue

        movs = [m for m in it.movimentacoes if m.tipo == 'saida' and m.data >= corte_30]
        if not movs:
            continue  # sem consumo nos últimos 30 dias — sem previsão

        saidas_7  = sum(m.quantidade for m in movs if m.data >= corte_7)
        saidas_30 = sum(m.quantidade for m in movs)

        consumo_diario_7  = saidas_7  / 7  if saidas_7  > 0 else 0
        consumo_diario_30 = saidas_30 / 30 if saidas_30 > 0 else 0

        # Média ponderada: tendência recente (7 dias) tem peso 3x
        if consumo_diario_7 > 0 and consumo_diario_30 > 0:
            consumo_diario = (consumo_diario_7 * 3 + consumo_diario_30 * 1) / 4
        elif consumo_diario_7 > 0:
            consumo_diario = consumo_diario_7
        else:
            consumo_diario = consumo_diario_30

        if consumo_diario <= 0:
            continue

        # Dias até atingir o estoque mínimo
        estoque_disponivel = it.quantidade - it.estoque_minimo
        dias_ate_minimo = estoque_disponivel / consumo_diario

        # Dias até zerar completamente
        dias_ate_zero = it.quantidade / consumo_diario

        # Filtra por limite se definido
        if limite_dias is not None and dias_ate_minimo > limite_dias:
            continue

        ruptura.append({
            'item': it,
            'dias': int(round(dias_ate_minimo)),
            'dias_zero': int(round(dias_ate_zero)),
            'consumo_diario': round(consumo_diario, 2),
            'urgencia': (
                'critico' if dias_ate_minimo <= 3 else
                'alerta'  if dias_ate_minimo <= 7 else
                'aviso'   if dias_ate_minimo <= 15 else
                'normal'
            ),
        })

    ruptura.sort(key=lambda x: x['dias'])
    return ruptura


@app.route('/relatorios/alertas')
@login_required
def relatorio_alertas():
    u = usuario_atual()
    # Engenheiro (colaborador) só vê alertas se tiver permissão 'ver_alertas'
    if u.perfil == 'colaborador':
        tem_permissao = any(p.permissao == 'ver_alertas' for p in u.permissoes_extras)
        if not tem_permissao:
            flash('Sem permissão para ver alertas de estoque. Solicite ao administrador.', 'warning')
            return redirect(url_for('index'))
    if u.perfil == 'admin':
        itens = Item.query.filter(Item.quantidade <= Item.estoque_minimo, Item.ativo == True).order_by(
            Item.fixado.desc(), Item.quantidade.asc()
        ).all()
        todos_ativos = Item.query.filter(Item.ativo == True).all()
    elif u.perfil == 'analista':
        # Analista vê apenas alertas da sua cidade
        ids_alm = set()
        if u.almoxarifado_id:
            alm_ref = db.session.get(Almoxarifado, u.almoxarifado_id)
            if alm_ref and alm_ref.cidade:
                ids_alm = {a.id for a in Almoxarifado.query.filter(
                    Almoxarifado.cidade.ilike(alm_ref.cidade)
                ).all()}
            elif alm_ref:
                ids_alm = {alm_ref.id}
        if ids_alm:
            itens = Item.query.filter(
                Item.quantidade <= Item.estoque_minimo,
                Item.almoxarifado_id.in_(ids_alm),
                Item.ativo == True
            ).order_by(Item.fixado.desc(), Item.quantidade.asc()).all()
            todos_ativos = Item.query.filter(Item.ativo == True, Item.almoxarifado_id.in_(ids_alm)).all()
        else:
            itens = []
            todos_ativos = []
    else:
        # tecnico_seguranca, almoxarife, colaborador com pode_ver_alertas
        # Para técnico: expande para todos da sua cidade
        ids = set(u.almoxarifados_permitidos())
        if u.perfil == 'tecnico_seguranca' and u.almoxarifado_id:
            alm_ref = db.session.get(Almoxarifado, u.almoxarifado_id)
            if alm_ref and alm_ref.cidade:
                ids = {a.id for a in Almoxarifado.query.filter_by(cidade=alm_ref.cidade).all()}
        itens = Item.query.filter(
            Item.quantidade <= Item.estoque_minimo,
            Item.almoxarifado_id.in_(ids),
            Item.ativo == True
        ).order_by(Item.fixado.desc(), Item.quantidade.asc()).all() if ids else []
        todos_ativos = Item.query.filter(
            Item.ativo == True, Item.almoxarifado_id.in_(ids)
        ).all() if ids else []

    ruptura = calcular_ruptura(todos_ativos, limite_dias=None)
    # Dicionário item_id → previsão para uso inline nos templates
    ruptura_por_item = {r['item'].id: r for r in ruptura}
    return render_template('relatorio_alertas.html', itens=itens, ruptura=ruptura,
                           ruptura_por_item=ruptura_por_item)

@app.route('/item/<int:id>/status_compra', methods=['POST'])
@login_required
def atualizar_status_compra(id):
    it = Item.query.get_or_404(id)
    u = usuario_atual()
    if u.perfil != 'admin' and it.almoxarifado_id not in u.almoxarifados_permitidos():
        return ('', 403)
    it.status_compra = request.form.get('status_compra', 'pendente')
    db.session.commit()
    return ('', 204)

@app.route('/item/<int:id>/fixar', methods=['POST'])
@login_required
def fixar_item(id):
    it = Item.query.get_or_404(id)
    u = usuario_atual()
    if u.perfil != 'admin' and it.almoxarifado_id not in u.almoxarifados_permitidos():
        return jsonify({'error': 'Acesso negado.'}), 403
    it.fixado = not it.fixado
    db.session.commit()
    return jsonify({'fixado': it.fixado})

@app.route('/movimentacao/<int:id>/devolvido', methods=['POST'])
@login_required
def marcar_devolvido(id):
    """Marca/desmarca uma movimentação de saída como devolvida."""
    mov = Movimentacao.query.get_or_404(id)
    u = usuario_atual()
    if u.perfil not in ('admin', 'almoxarife'):
        return jsonify({'error': 'Acesso negado.'}), 403
    # Toggle: None/False → True, True → False
    mov.devolvido = not bool(mov.devolvido)
    db.session.commit()
    return jsonify({'devolvido': mov.devolvido})

@app.route('/item/<int:id>/desativar', methods=['POST'])
@login_required
def desativar_item(id):
    it = Item.query.get_or_404(id)
    u = usuario_atual()
    if u.perfil != 'admin' and it.almoxarifado_id not in u.almoxarifados_permitidos():
        flash('Acesso negado.', 'danger')
        return redirect(url_for('item', id=id))
    
    # Registrar saída do saldo restante se houver
    if it.quantidade > 0:
        mov = Movimentacao(
            tipo='saida',
            quantidade=it.quantidade,
            responsavel=u.nome,
            observacao=f'Item desativado - saldo restante: {it.quantidade} {it.unidade}',
            item_id=id
        )
        db.session.add(mov)
        it.quantidade = 0
    
    it.ativo = False
    db.session.commit()
    flash(f'Item "{it.nome}" desativado com sucesso!', 'warning')
    return redirect(url_for('item', id=id))

@app.route('/item/<int:id>/reativar', methods=['POST'])
@login_required
def reativar_item(id):
    it = Item.query.get_or_404(id)
    u = usuario_atual()
    if u.perfil != 'admin' and it.almoxarifado_id not in u.almoxarifados_permitidos():
        flash('Acesso negado.', 'danger')
        return redirect(url_for('item', id=id))
    
    it.ativo = True
    db.session.commit()
    flash(f'Item "{it.nome}" reativado com sucesso!', 'success')
    return redirect(url_for('item', id=id))

# ── EXPORTAR EXCEL ───────────────────────────────────────────────────────────

@app.route('/almoxarifado/<int:id>/exportar')
@login_required
def exportar_almoxarifado(id):
    u = usuario_atual()
    if u.perfil != 'admin' and id not in u.almoxarifados_permitidos():
        flash('Acesso negado.', 'danger')
        return redirect(url_for('index'))
    alm = Almoxarifado.query.get_or_404(id)
    itens = Item.query.filter_by(almoxarifado_id=id).all()

    wb = openpyxl.Workbook()
    h_fill = PatternFill('solid', fgColor='1A3A5C')
    h_font = Font(bold=True, color='FFFFFF', size=11)
    ok_fill  = PatternFill('solid', fgColor='D4EDDA')
    al_fill  = PatternFill('solid', fgColor='FFF3CD')
    cr_fill  = PatternFill('solid', fgColor='F8D7DA')
    en_fill  = PatternFill('solid', fgColor='D4EDDA')
    sa_fill  = PatternFill('solid', fgColor='F8D7DA')
    borda = Border(left=Side(style='thin'), right=Side(style='thin'),
                   top=Side(style='thin'), bottom=Side(style='thin'))

    def titulo(ws, texto, cols):
        ws.merge_cells(f'A1:{get_column_letter(cols)}1')
        ws['A1'] = texto
        ws['A1'].font = Font(bold=True, size=13, color='1A3A5C')
        ws['A1'].alignment = Alignment(horizontal='center')
        ws.merge_cells(f'A2:{get_column_letter(cols)}2')
        ws['A2'] = f'Gerado em: {datetime.now().strftime("%d/%m/%Y %H:%M")}'
        ws['A2'].font = Font(italic=True, color='888888')
        ws['A2'].alignment = Alignment(horizontal='center')

    def cabecalho(ws, row, headers):
        for col, h in enumerate(headers, 1):
            c = ws.cell(row=row, column=col, value=h)
            c.font = h_font; c.fill = h_fill
            c.alignment = Alignment(horizontal='center'); c.border = borda

    categoria_label = {
        'epi': 'EPI', 'maquinario': 'Maquinário', 'eletrica': 'Elétrica',
        'hidraulica': 'Hidráulica', 'gas': 'Gás', 'geral': 'Geral'
    }

    # Aba 1 — Estoque Atual
    ws1 = wb.active
    ws1.title = 'Estoque Atual'
    titulo(ws1, f'Estoque Atual — {alm.nome}', 8)
    cabecalho(ws1, 4, ['Codigo', 'Item', 'Categoria', 'Unidade', 'Qtd. Atual', 'Est. Minimo', 'Deficit', 'Status'])
    for r, it in enumerate(itens, 5):
        deficit = max(0, it.estoque_minimo - it.quantidade)
        status = 'ZERADO' if it.status == 'critico' else ('ABAIXO DO MINIMO' if it.status == 'alerta' else 'OK')
        fill = cr_fill if it.status == 'critico' else (al_fill if it.status == 'alerta' else ok_fill)
        cat = categoria_label.get(it.categoria or 'geral', it.categoria or 'Geral')
        for c, v in enumerate([it.codigo, it.nome, cat, it.unidade, it.quantidade, it.estoque_minimo, deficit, status], 1):
            cell = ws1.cell(row=r, column=c, value=v)
            cell.fill = fill; cell.border = borda
            cell.alignment = Alignment(horizontal='left' if c == 2 else 'center')
    for i, w in enumerate([14, 35, 14, 10, 12, 14, 12, 20], 1):
        ws1.column_dimensions[get_column_letter(i)].width = w

    # Aba 2 — Movimentações
    ws2 = wb.create_sheet('Movimentacoes')
    titulo(ws2, f'Historico de Movimentacoes — {alm.nome}', 6)
    cabecalho(ws2, 4, ['Data', 'Item', 'Tipo', 'Quantidade', 'Responsavel', 'Observacao'])
    movs = (Movimentacao.query.join(Item)
            .filter(Item.almoxarifado_id == id)
            .order_by(Movimentacao.data.desc()).all())
    for r, mov in enumerate(movs, 5):
        fill = en_fill if mov.tipo == 'entrada' else sa_fill
        for c, v in enumerate([
            mov.data.strftime('%d/%m/%Y %H:%M'), mov.item.nome,
            'Entrada' if mov.tipo == 'entrada' else 'Saida',
            f'{mov.quantidade} {mov.item.unidade}',
            mov.responsavel or '', mov.observacao or ''
        ], 1):
            cell = ws2.cell(row=r, column=c, value=v)
            cell.fill = fill; cell.border = borda
            cell.alignment = Alignment(horizontal='left')
    for i, w in enumerate([18, 35, 10, 14, 22, 30], 1):
        ws2.column_dimensions[get_column_letter(i)].width = w

    # Aba 3 — Consumo por Item
    ws3 = wb.create_sheet('Consumo por Item')
    titulo(ws3, f'Consumo por Item — {alm.nome}', 4)
    cabecalho(ws3, 4, ['Item', 'Total Entradas', 'Total Saidas', 'Saldo Atual'])
    for r, it in enumerate(itens, 5):
        entradas = sum(m.quantidade for m in it.movimentacoes if m.tipo == 'entrada')
        saidas   = sum(m.quantidade for m in it.movimentacoes if m.tipo == 'saida')
        for c, v in enumerate([it.nome, entradas, saidas, it.quantidade], 1):
            cell = ws3.cell(row=r, column=c, value=v)
            cell.border = borda
            cell.alignment = Alignment(horizontal='left' if c == 1 else 'center')
    for i, w in enumerate([35, 16, 14, 14], 1):
        ws3.column_dimensions[get_column_letter(i)].width = w

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    nome = f"estoque_{alm.nome.replace(' ', '_')}_{date.today()}.xlsx"
    return send_file(buf, as_attachment=True, download_name=nome,
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

# ── API ──────────────────────────────────────────────────────────────────────

@app.route('/api/alertas')
@login_required
def api_alertas():
    if _check_api_rate(request.remote_addr or '0.0.0.0'):
        return jsonify({'error': 'Too many requests'}), 429
    u = usuario_atual()
    if u.perfil == 'admin':
        itens = Item.query.filter(Item.quantidade <= Item.estoque_minimo, Item.ativo == True).all()
    else:
        ids = u.almoxarifados_permitidos()
        itens = Item.query.filter(
            Item.quantidade <= Item.estoque_minimo,
            Item.almoxarifado_id.in_(ids),
            Item.ativo == True
        ).all() if ids else []
    return jsonify([{
        'id': i.id, 'nome': i.nome, 'codigo': i.codigo,
        'quantidade': i.quantidade, 'estoque_minimo': i.estoque_minimo,
        'unidade': i.unidade, 'status': i.status,
        'almoxarifado': i.almoxarifado.nome
    } for i in itens])

@app.route('/api/almoxarife/notificacoes')
@login_required
def api_almoxarife_notificacoes():
    """Retorna requisições pendentes para o almoxarife logado — usado para popup de alerta."""
    u = usuario_atual()
    if u.perfil not in ('admin', 'almoxarife'):
        return jsonify([])
    # Busca requisições pendentes do almoxarifado do almoxarife
    if u.perfil == 'almoxarife' and u.almoxarifado_id:
        reqs = RequisicaoMestre.query.filter_by(
            almoxarifado_id=u.almoxarifado_id,
            status='pendente'
        ).order_by(RequisicaoMestre.data_criacao.desc()).limit(5).all()
    else:
        reqs = RequisicaoMestre.query.filter_by(
            status='pendente'
        ).order_by(RequisicaoMestre.data_criacao.desc()).limit(5).all()
    return jsonify([{
        'id': r.id,
        'protocolo': r.protocolo or f'#{r.id}',
        'mestre': r.mestre.nome,
        'colaborador': r.colaborador,
        'almoxarifado': r.almoxarifado.nome,
        'itens': len(r.itens),
        'data': r.data_criacao.strftime('%d/%m/%Y %H:%M')
    } for r in reqs])

@app.route('/api/colaboradores')
@login_required
def api_colaboradores():
    if _check_api_rate(request.remote_addr or '0.0.0.0'):
        return jsonify([]), 429
    """Autocomplete — busca colaboradores por nome, filtrado por cidade do usuário."""
    q = request.args.get('q', '').strip()
    u = usuario_atual()
    nomes = []
    like = f"%{q}%"
    is_pg = 'postgresql' in str(db.engine.url)
    ilike = 'ILIKE' if is_pg else 'LIKE'

    from sqlalchemy import text

    # Determinar cidade e escopo do usuário logado (via almoxarifado vinculado)
    escopo_usuario = None
    cidade_usuario = None
    if u and u.almoxarifado_id:
        alm = db.session.get(Almoxarifado, u.almoxarifado_id)
        if alm:
            cidade_usuario = (alm.cidade or '').strip() or None
            nome_lower = alm.nome.lower()
            for esc in ['estrutura', 'acabamento', 'infraestrutura', 'forma', 'acampamento']:
                if esc in nome_lower:
                    escopo_usuario = esc
                    break

    # Filtro de cidade — admin vê todos, demais só da sua cidade
    cidade_filtro = cidade_usuario if (u and u.perfil != 'admin') else None

    def _query_colabs(extra_where='', extra_params=None):
        params = {"q": like}
        cidade_clause = ''
        if cidade_filtro:
            cidade_clause = f" AND (cidade {ilike} :cidade OR cidade IS NULL OR cidade = '')" if not extra_where else \
                            f" AND (cidade {ilike} :cidade OR cidade IS NULL OR cidade = '')"
            params["cidade"] = cidade_filtro
        if extra_params:
            params.update(extra_params)
        sql = f"SELECT nome, funcao, escopo, tipo FROM colaborador WHERE ativo = TRUE AND nome {ilike} :q{cidade_clause}{extra_where} ORDER BY nome LIMIT 15"
        return db.session.execute(text(sql), params).fetchall()

    # 1. Colaboradores do mesmo escopo primeiro
    if escopo_usuario:
        rows = _query_colabs(
            extra_where=f" AND escopo {ilike} :esc",
            extra_params={"esc": f"%{escopo_usuario}%"}
        )
        for r in rows:
            nomes.append({'nome': r[0], 'funcao': r[1] or '', 'escopo': r[2] or '', 'tipo': r[3] or 'peao'})

    # 2. Demais colaboradores da mesma cidade
    nomes_ja = {n['nome'] for n in nomes}
    rows = _query_colabs()
    for r in rows:
        if r[0] not in nomes_ja:
            nomes.append({'nome': r[0], 'funcao': r[1] or '', 'escopo': r[2] or '', 'tipo': r[3] or 'peao'})

    # 3. Histórico de requisições dos almoxarifados da cidade (fallback)
    nomes_ja = {n['nome'] for n in nomes}
    if cidade_filtro:
        # Só requisições de almoxarifados da mesma cidade
        ids_cidade = [a.id for a in Almoxarifado.query.filter(
            Almoxarifado.cidade.ilike(cidade_filtro)
        ).all()]
        if ids_cidade:
            placeholders = ','.join(str(i) for i in ids_cidade)
            rows = db.session.execute(
                text(f"SELECT DISTINCT colaborador FROM requisicao_mestre WHERE colaborador {ilike} :q AND almoxarifado_id IN ({placeholders}) ORDER BY colaborador LIMIT 5"),
                {"q": like}
            ).fetchall()
            for r in rows:
                if r[0] not in nomes_ja:
                    nomes.append({'nome': r[0], 'funcao': '', 'escopo': '', 'tipo': ''})
    else:
        rows = db.session.execute(
            text(f"SELECT DISTINCT colaborador FROM requisicao_mestre WHERE colaborador {ilike} :q ORDER BY colaborador LIMIT 5"),
            {"q": like}
        ).fetchall()
        for r in rows:
            if r[0] not in nomes_ja:
                nomes.append({'nome': r[0], 'funcao': '', 'escopo': '', 'tipo': ''})

    return jsonify(nomes[:12])

# ── FROTA DE FERRAMENTAS ─────────────────────────────────────────────────────

@app.route('/almoxarifado/<int:alm_id>/ferramentas')
@login_required
def ferramentas(alm_id):
    u = usuario_atual()
    alm = Almoxarifado.query.get_or_404(alm_id)
    if u.perfil not in ('admin', 'almoxarife') and alm_id not in u.almoxarifados_permitidos():
        flash('Acesso negado.', 'danger')
        return redirect(url_for('index'))
    lista = Ferramenta.query.filter_by(almoxarifado_id=alm_id, ativo=True).order_by(Ferramenta.nome).all()
    return render_template('ferramentas.html', almoxarifado=alm, ferramentas=lista)

@app.route('/almoxarifado/<int:alm_id>/ferramentas/nova', methods=['GET', 'POST'])
@login_required
def nova_ferramenta(alm_id):
    u = usuario_atual()
    alm = Almoxarifado.query.get_or_404(alm_id)
    if u.perfil not in ('admin', 'almoxarife'):
        flash('Acesso negado.', 'danger')
        return redirect(url_for('ferramentas', alm_id=alm_id))
    if request.method == 'POST':
        identificacao = request.form['identificacao'].strip()
        # Verificar duplicidade de ID em qualquer almoxarifado ativo
        existente = Ferramenta.query.filter_by(identificacao=identificacao, ativo=True).first()
        if existente:
            alm_existente = existente.almoxarifado
            flash(
                f'⚠️ ID "{identificacao}" já está cadastrado: '
                f'<strong>{existente.nome}</strong> — '
                f'Almoxarifado: <strong>{alm_existente.nome}</strong> — '
                f'Status: <strong>{existente.status.replace("_", " ").title()}</strong>',
                'danger'
            )
            return render_template('ferramenta_form.html', almoxarifado=alm, ferramenta=None,
                                   form_data=request.form)
        f = Ferramenta(
            identificacao=identificacao,
            nome=request.form['nome'].strip(),
            empresa=request.form.get('empresa', '').strip() or None,
            almoxarifado_id=alm_id,
            local=request.form.get('local', '').strip() or None,
            observacao=request.form.get('observacao', '').strip() or None
        )
        db.session.add(f)
        db.session.commit()
        flash(f'Ferramenta "{f.nome}" cadastrada!', 'success')
        return redirect(url_for('ferramentas', alm_id=alm_id))
    return render_template('ferramenta_form.html', almoxarifado=alm, ferramenta=None, form_data={})

@app.route('/ferramenta/<int:id>/status', methods=['POST'])
@login_required
def atualizar_status_ferramenta(id):
    f = Ferramenta.query.get_or_404(id)
    u = usuario_atual()
    if u.perfil not in ('admin', 'almoxarife'):
        return jsonify({'error': 'Acesso negado'}), 403
    novo_status = request.form.get('status', 'disponivel')
    responsavel = request.form.get('responsavel', '').strip()
    motivo = request.form.get('motivo', '').strip()

    if novo_status == 'em_uso':
        f.status = 'em_uso'
        f.responsavel_atual = responsavel
        f.data_saida = agora()
        db.session.add(HistoricoFerramenta(
            ferramenta_id=f.id,
            colaborador=responsavel,
            data_saida=f.data_saida,
            registrado_por=u.nome,
            tipo_evento='uso'
        ))
    elif novo_status == 'manutencao':
        # Fechar registro aberto se houver
        hist_aberto = HistoricoFerramenta.query.filter_by(
            ferramenta_id=f.id, data_devolucao=None
        ).order_by(HistoricoFerramenta.data_saida.desc()).first()
        if hist_aberto:
            hist_aberto.data_devolucao = agora()
        f.status = 'manutencao'
        f.responsavel_atual = motivo or 'Em manutenção'
        f.data_saida = agora()
        db.session.add(HistoricoFerramenta(
            ferramenta_id=f.id,
            colaborador=u.nome,
            data_saida=f.data_saida,
            registrado_por=u.nome,
            tipo_evento='manutencao',
            motivo_manutencao=motivo or None
        ))
    else:
        # Devolver / disponivel
        hist = HistoricoFerramenta.query.filter_by(
            ferramenta_id=f.id, data_devolucao=None
        ).order_by(HistoricoFerramenta.data_saida.desc()).first()
        if hist:
            hist.data_devolucao = agora()
        f.status = 'disponivel'
        f.responsavel_atual = None
        f.data_saida = None

    db.session.commit()
    data_saida_iso = f.data_saida.isoformat() if f.data_saida else None
    # Retorna hist_id para o frontend abrir câmera de foto
    hist_novo = HistoricoFerramenta.query.filter_by(
        ferramenta_id=f.id, data_devolucao=None
    ).order_by(HistoricoFerramenta.data_saida.desc()).first()
    hist_id = hist_novo.id if hist_novo else None
    return jsonify({
        'status': f.status,
        'responsavel': f.responsavel_atual or '',
        'data_saida': data_saida_iso,
        'hist_id': hist_id
    })

@app.route('/ferramenta/<int:id>/historico')
@login_required
def historico_ferramenta(id):
    """Retorna o histórico de uso da ferramenta em JSON."""
    f = Ferramenta.query.get_or_404(id)
    hist = HistoricoFerramenta.query.filter_by(ferramenta_id=id).order_by(
        HistoricoFerramenta.data_saida.desc()
    ).limit(20).all()
    return jsonify({
        'ferramenta': f.nome,
        'id': f.identificacao,
        'empresa': f.empresa or '',
        'historico': [{
            'colaborador': h.colaborador,
            'data_saida': h.data_saida.strftime('%d/%m/%Y %H:%M'),
            'data_devolucao': h.data_devolucao.strftime('%d/%m/%Y %H:%M') if h.data_devolucao else None,
            'registrado_por': h.registrado_por or '—',
            'tipo_evento': h.tipo_evento or 'uso',
            'motivo_manutencao': h.motivo_manutencao or '',
            'foto_url': h.foto_url or ''
        } for h in hist]
    })

@app.route('/ferramenta/<int:id>/deletar', methods=['POST'])
@admin_required
def deletar_ferramenta(id):
    f = Ferramenta.query.get_or_404(id)
    alm_id = f.almoxarifado_id
    f.ativo = False
    db.session.commit()
    flash(f'Ferramenta "{f.nome}" removida.', 'warning')
    return redirect(url_for('ferramentas', alm_id=alm_id))

@app.route('/api/ferramenta/verificar-id')
@login_required
def verificar_id_ferramenta():
    if _check_api_rate(request.remote_addr or '0.0.0.0'):
        return jsonify({'disponivel': True}), 429
    """Verifica se um ID/patrimônio já está cadastrado em qualquer almoxarifado."""
    identificacao = request.args.get('id', '').strip()
    excluir_id = request.args.get('excluir', type=int)  # para edição futura
    if not identificacao or identificacao == '__noop__':
        return jsonify({'disponivel': True})
    q = Ferramenta.query.filter_by(identificacao=identificacao, ativo=True)
    if excluir_id:
        q = q.filter(Ferramenta.id != excluir_id)
    existente = q.first()
    if existente:
        return jsonify({
            'disponivel': False,
            'nome': existente.nome,
            'almoxarifado': existente.almoxarifado.nome,
            'status': existente.status,
            'empresa': existente.empresa or 'Própria'
        })
    return jsonify({'disponivel': True})

@app.route('/api/ferramentas/empresas')
@login_required
def api_empresas_ferramentas():
    """Retorna lista de empresas já cadastradas nas ferramentas para autocomplete."""
    q = request.args.get('q', '').strip()
    empresas = db.session.query(Ferramenta.empresa)\
        .filter(Ferramenta.ativo == True, Ferramenta.empresa != None)\
        .distinct().all()
    nomes = [e[0] for e in empresas if e[0] and q.lower() in e[0].lower()]
    return jsonify(sorted(nomes)[:10])

# ── FOTO DE RETIRADA — SALVA DIRETO NO BANCO ─────────────────────────────────
@app.route('/ferramenta/historico/<int:hist_id>/foto', methods=['POST'])
@login_required
def upload_foto_retirada(hist_id):
    """Recebe foto em base64 do frontend e salva direto no PostgreSQL."""
    hist = HistoricoFerramenta.query.get_or_404(hist_id)
    u = usuario_atual()
    ferr = db.session.get(Ferramenta, hist.ferramenta_id)
    if u.perfil not in ('admin', 'almoxarife') and (
            ferr and ferr.almoxarifado_id not in u.almoxarifados_permitidos()):
        return jsonify({'error': 'Acesso negado.'}), 403
    data = request.get_json()
    if not data or 'foto' not in data:
        return jsonify({'error': 'Nenhuma foto recebida.'}), 400

    foto = data['foto']
    # Aceita apenas imagem base64
    if not foto.startswith('data:image'):
        return jsonify({'error': 'Formato inválido.'}), 400
    # Limite de 5MB por foto
    if len(foto) > 5 * 1024 * 1024:
        return jsonify({'error': 'Foto muito grande. Máximo 5MB.'}), 400

    try:
        hist.foto_url = foto
        db.session.commit()
        logger.info(f'FOTO: salva no banco — hist_id={hist_id} tamanho={len(foto)} bytes')
        return jsonify({'ok': True})
    except Exception as e:
        logger.error(f'FOTO: erro ao salvar — {e}')
        return jsonify({'error': str(e)}), 500

# ── FOTO DE EPI — SALVA DIRETO NO BANCO ──────────────────────────────────────
@app.route('/movimentacao/<int:mov_id>/foto', methods=['POST'])
@login_required
def upload_foto_epi(mov_id):
    """Recebe foto em base64 do frontend e salva na movimentação de EPI."""
    mov = Movimentacao.query.get_or_404(mov_id)
    u = usuario_atual()
    it = db.session.get(Item, mov.item_id)
    if u.perfil not in ('admin', 'almoxarife') and (
            it and it.almoxarifado_id not in u.almoxarifados_permitidos()):
        return jsonify({'error': 'Acesso negado.'}), 403
    data = request.get_json()
    if not data or 'foto' not in data:
        return jsonify({'error': 'Nenhuma foto recebida.'}), 400

    foto = data['foto']
    if not foto.startswith('data:image'):
        return jsonify({'error': 'Formato inválido.'}), 400
    if len(foto) > 5 * 1024 * 1024:
        return jsonify({'error': 'Foto muito grande. Máximo 5MB.'}), 400

    try:
        mov.foto_url = foto
        db.session.commit()
        logger.info(f'FOTO EPI: salva — mov_id={mov_id} tamanho={len(foto)} bytes')
        return jsonify({'ok': True})
    except Exception as e:
        logger.error(f'FOTO EPI: erro — {e}')
        return jsonify({'error': str(e)}), 500

# ── FROTA DE EPIs / UNIFORMES ─────────────────────────────────────────────────

@app.route('/almoxarifado/<int:alm_id>/epis')
@login_required
def epis(alm_id):
    u = usuario_atual()
    alm = Almoxarifado.query.get_or_404(alm_id)
    if u.perfil not in ('admin', 'almoxarife') and alm_id not in u.almoxarifados_permitidos():
        flash('Acesso negado.', 'danger')
        return redirect(url_for('index'))
    lista = ItemEPI.query.filter_by(almoxarifado_id=alm_id, ativo=True).order_by(ItemEPI.nome).all()
    return render_template('epis.html', almoxarifado=alm, epis=lista)

@app.route('/almoxarifado/<int:alm_id>/epis/novo', methods=['GET', 'POST'])
@login_required
def novo_epi(alm_id):
    u = usuario_atual()
    alm = Almoxarifado.query.get_or_404(alm_id)
    if u.perfil not in ('admin', 'almoxarife'):
        flash('Acesso negado.', 'danger')
        return redirect(url_for('epis', alm_id=alm_id))
    if request.method == 'POST':
        e = ItemEPI(
            identificacao=request.form['identificacao'].strip(),
            nome=request.form['nome'].strip(),
            tamanho=request.form.get('tamanho', '').strip() or None,
            almoxarifado_id=alm_id,
            quantidade=int(request.form.get('quantidade', 1) or 1),
            local=request.form.get('local', '').strip() or None,
            observacao=request.form.get('observacao', '').strip() or None
        )
        db.session.add(e)
        db.session.commit()
        flash(f'EPI "{e.nome}" cadastrado!', 'success')
        return redirect(url_for('epis', alm_id=alm_id))
    return render_template('epi_form.html', almoxarifado=alm, epi=None, form_data={})

@app.route('/epi/<int:id>/status', methods=['POST'])
@login_required
def atualizar_status_epi(id):
    e = ItemEPI.query.get_or_404(id)
    u = usuario_atual()
    if u.perfil not in ('admin', 'almoxarife'):
        return jsonify({'error': 'Acesso negado'}), 403
    novo_status = request.form.get('status', 'disponivel')
    responsavel = request.form.get('responsavel', '').strip()
    motivo = request.form.get('motivo', '').strip()

    if novo_status == 'em_uso':
        e.status = 'em_uso'
        e.responsavel_atual = responsavel
        db.session.add(HistoricoEPI(
            item_epi_id=e.id,
            colaborador=responsavel,
            data_saida=agora(),
            registrado_por=u.nome,
            tipo_evento='uso'
        ))
    elif novo_status == 'manutencao':
        hist_aberto = HistoricoEPI.query.filter_by(
            item_epi_id=e.id, data_devolucao=None
        ).order_by(HistoricoEPI.data_saida.desc()).first()
        if hist_aberto:
            hist_aberto.data_devolucao = agora()
        e.status = 'manutencao'
        e.responsavel_atual = motivo or 'Em manutenção'
        db.session.add(HistoricoEPI(
            item_epi_id=e.id,
            colaborador=u.nome,
            data_saida=agora(),
            registrado_por=u.nome,
            tipo_evento='manutencao',
            motivo_manutencao=motivo or None
        ))
    else:
        hist = HistoricoEPI.query.filter_by(
            item_epi_id=e.id, data_devolucao=None
        ).order_by(HistoricoEPI.data_saida.desc()).first()
        if hist:
            hist.data_devolucao = agora()
        e.status = 'disponivel'
        e.responsavel_atual = None

    db.session.commit()
    hist_novo = HistoricoEPI.query.filter_by(
        item_epi_id=e.id, data_devolucao=None
    ).order_by(HistoricoEPI.data_saida.desc()).first()
    hist_id = hist_novo.id if hist_novo else None
    return jsonify({
        'status': e.status,
        'responsavel': e.responsavel_atual or '',
        'hist_id': hist_id
    })

@app.route('/epi/<int:id>/historico')
@login_required
def historico_epi(id):
    e = ItemEPI.query.get_or_404(id)
    hist = HistoricoEPI.query.filter_by(item_epi_id=id).order_by(
        HistoricoEPI.data_saida.desc()
    ).limit(20).all()
    return jsonify({
        'nome': e.nome,
        'id': e.identificacao,
        'historico': [{
            'colaborador': h.colaborador,
            'data_saida': h.data_saida.strftime('%d/%m/%Y'),
            'data_devolucao': h.data_devolucao.strftime('%d/%m/%Y') if h.data_devolucao else None,
            'registrado_por': h.registrado_por or '—',
            'tipo_evento': h.tipo_evento or 'uso',
            'motivo_manutencao': h.motivo_manutencao or '',
            'foto_url': h.foto_url or ''
        } for h in hist]
    })

@app.route('/epi/<int:id>/deletar', methods=['POST'])
@admin_required
def deletar_epi(id):
    e = ItemEPI.query.get_or_404(id)
    alm_id = e.almoxarifado_id
    e.ativo = False
    db.session.commit()
    flash(f'EPI "{e.nome}" removido.', 'warning')
    return redirect(url_for('epis', alm_id=alm_id))

@app.route('/epi/historico/<int:hist_id>/foto', methods=['POST'])
@login_required
def upload_foto_historico_epi(hist_id):
    hist = HistoricoEPI.query.get_or_404(hist_id)
    data = request.get_json()
    if not data or 'foto' not in data:
        return jsonify({'error': 'Nenhuma foto recebida.'}), 400
    foto = data['foto']
    if not foto.startswith('data:image'):
        return jsonify({'error': 'Formato inválido.'}), 400
    if len(foto) > 5 * 1024 * 1024:
        return jsonify({'error': 'Foto muito grande. Máximo 5MB.'}), 400
    try:
        hist.foto_url = foto
        db.session.commit()
        return jsonify({'ok': True})
    except Exception as exc:
        return jsonify({'error': str(exc)}), 500

# ── GERENCIAR COLABORADORES ──────────────────────────────────────────────────

@app.route('/colaboradores')
@login_required
def colaboradores():
    u = usuario_atual()
    if u.perfil not in ('admin', 'almoxarife', 'analista'):
        flash('Acesso negado.', 'danger')
        return redirect(url_for('index'))
    cols = Colaborador.query.order_by(Colaborador.ativo.desc(), Colaborador.nome).all()

    # Determinar escopo, obra e cidade do almoxarife pelo seu almoxarifado vinculado
    escopo_almoxarife = None
    obra_almoxarife = None
    cidade_almoxarife = None
    if u.perfil == 'almoxarife' and u.almoxarifado_id:
        alm = db.session.get(Almoxarifado, u.almoxarifado_id)
        if alm:
            obra_almoxarife = (alm.obra or '').lower().strip() or None
            cidade_almoxarife = (alm.cidade or '').lower().strip() or None
            nome_lower = alm.nome.lower()
            for esc in ['estrutura', 'acabamento', 'infraestrutura', 'forma', 'acampamento']:
                if esc in nome_lower:
                    escopo_almoxarife = esc
                    break

    # Almoxarife vê apenas colaboradores da sua obra E frente
    if u.perfil == 'almoxarife':
        def colab_pertence(c):
            # Filtro por obra — se almoxarife tem obra definida, só vê colaboradores
            # com a mesma obra. Colaboradores sem obra passam apenas se cidade bater.
            if obra_almoxarife:
                obra_colab = (c.obra or '').lower().strip()
                if obra_colab and obra_colab != obra_almoxarife:
                    return False
            # Filtro por cidade — se almoxarife tem cidade, colaboradores sem cidade
            # ou de outra cidade ficam fora
            if cidade_almoxarife:
                cidade_colab = (c.cidade or '').lower().strip()
                if cidade_colab != cidade_almoxarife:
                    return False
            # Filtro por escopo (frente de obra)
            if escopo_almoxarife and c.escopo:
                if c.escopo.lower().strip() != escopo_almoxarife:
                    return False
            return True
        cols = [c for c in cols if colab_pertence(c)]

    # Analista vê apenas colaboradores da sua cidade E escopo
    if u.perfil == 'analista':
        # Determina cidade do analista pelo almoxarifado vinculado
        cidade_analista = None
        if u.almoxarifado_id:
            alm_analista = db.session.get(Almoxarifado, u.almoxarifado_id)
            if alm_analista:
                cidade_analista = (alm_analista.cidade or '').lower().strip() or None

        def analista_pertence(c):
            # Filtro por cidade — se analista tem cidade definida, só vê colaboradores
            # da mesma cidade. Colaboradores sem cidade definida ficam visíveis apenas
            # para admin.
            if cidade_analista:
                cidade_colab = (c.cidade or '').lower().strip()
                if cidade_colab != cidade_analista:
                    return False
            # Filtro por escopo
            if u.escopo and c.escopo:
                if c.escopo.lower().strip() != u.escopo.lower().strip():
                    return False
            return True

        cols = [c for c in cols if analista_pertence(c)]
    from collections import OrderedDict
    grupos = OrderedDict([
        ('estrutura',      {'label': '🏗️ Estrutura',      'cor': '#f0a500', 'colaboradores': []}),
        ('infraestrutura', {'label': '🔧 Infraestrutura',  'cor': '#0ea5e9', 'colaboradores': []}),
        ('acabamento',     {'label': '🏕️ Acabamento',      'cor': '#22c55e', 'colaboradores': []}),
        ('sem_escopo',     {'label': '📋 Sem Escopo',       'cor': '#94a3b8', 'colaboradores': []}),
    ])
    for c in cols:
        escopo = (c.escopo or '').lower().strip()
        if escopo in grupos:
            grupos[escopo]['colaboradores'].append(c)
        else:
            grupos['sem_escopo']['colaboradores'].append(c)
    return render_template('colaboradores.html', colaboradores=cols, grupos=grupos)

@app.route('/colaboradores/novo', methods=['POST'])
@login_required
def novo_colaborador():
    u = usuario_atual()
    if u.perfil not in ('admin', 'almoxarife', 'analista'):
        flash('Acesso negado.', 'danger')
        return redirect(url_for('index'))
    nome = request.form.get('nome', '').strip()
    funcao = request.form.get('funcao', '').strip()
    escopo = request.form.get('escopo', '').strip()
    obra = request.form.get('obra', '').strip()
    cidade = request.form.get('cidade', '').strip()
    tipo = request.form.get('tipo', 'peao').strip()

    # Se almoxarife não preencheu obra/cidade, usa os do almoxarifado dele
    if not obra and u.perfil == 'almoxarife' and u.almoxarifado_id:
        alm = db.session.get(Almoxarifado, u.almoxarifado_id)
        if alm:
            obra = alm.obra or ''
            cidade = alm.cidade or ''
    if not nome:
        flash('Informe o nome do colaborador.', 'warning')
        return redirect(url_for('colaboradores'))
    # Evita duplicata
    if Colaborador.query.filter(Colaborador.nome.ilike(nome)).first():
        flash(f'Colaborador "{nome}" já está cadastrado.', 'warning')
        return redirect(url_for('colaboradores'))
    db.session.add(Colaborador(nome=nome, funcao=funcao or None, escopo=escopo or None,
                               obra=obra or None, cidade=cidade or None, tipo=tipo or 'peao'))
    db.session.commit()
    flash(f'✅ Colaborador "{nome}" cadastrado!', 'success')
    return redirect(url_for('colaboradores'))

@app.route('/colaboradores/<int:id>/desativar', methods=['POST'])
@login_required
def desativar_colaborador(id):
    u = usuario_atual()
    if u.perfil not in ('admin', 'almoxarife', 'analista'):
        flash('Acesso negado.', 'danger')
        return redirect(url_for('index'))
    c = Colaborador.query.get_or_404(id)
    c.ativo = False
    db.session.commit()
    flash(f'Colaborador "{c.nome}" desativado.', 'warning')
    return redirect(url_for('colaboradores'))

@app.route('/colaboradores/<int:id>/reativar', methods=['POST'])
@login_required
def reativar_colaborador(id):
    u = usuario_atual()
    if u.perfil not in ('admin', 'almoxarife', 'analista'):
        flash('Acesso negado.', 'danger')
        return redirect(url_for('index'))
    c = Colaborador.query.get_or_404(id)
    c.ativo = True
    db.session.commit()
    flash(f'Colaborador "{c.nome}" reativado.', 'success')
    return redirect(url_for('colaboradores'))

# ── GERENCIAR USUÁRIOS (só admin) ────────────────────────────────────────────

@app.route('/usuarios')
@admin_required
def usuarios():
    from collections import OrderedDict
    todos = Usuario.query.order_by(Usuario.nome).all()
    # Agrupar por perfil
    grupos = OrderedDict([
        ('admin',              {'label': '👑 Admin / Fundador',     'cor': '#7c3aed', 'usuarios': []}),
        ('almoxarife',         {'label': '📦 Almoxarife',           'cor': '#0ea5e9', 'usuarios': []}),
        ('mestre',             {'label': '🦺 Mestre de Obra',       'cor': '#f0a500', 'usuarios': []}),
        ('tecnico_seguranca',  {'label': '🔒 Técnico de Segurança', 'cor': '#3b82f6', 'usuarios': []}),
        ('analista',           {'label': '📊 Analista',             'cor': '#10b981', 'usuarios': []}),
        ('colaborador',        {'label': '👔 Engenheiro',           'cor': '#64748b', 'usuarios': []}),
    ])
    for u in todos:
        perfil = u.perfil if u.perfil in grupos else 'colaborador'
        grupos[perfil]['usuarios'].append(u)
    return render_template('usuarios.html', grupos=grupos,
                           usuarios=todos,
                           permissoes_disponiveis=PERMISSOES_DISPONIVEIS)

@app.route('/usuarios/novo', methods=['GET', 'POST'])
@admin_required
def novo_usuario():
    almoxarifados = Almoxarifado.query.all()
    if request.method == 'POST':
        u = Usuario(
            nome=request.form['nome'],
            login=request.form['login'],
            perfil=request.form['perfil'],
            almoxarifado_id=request.form.get('almoxarifado_id') or None,
            email=request.form.get('email', '').strip() or None
        )
        senha_nova = request.form.get('senha', '')
        if len(senha_nova) < 8:
            flash('A senha deve ter pelo menos 8 caracteres.', 'danger')
            return render_template('form_usuario.html', usuario=None, almoxarifados=almoxarifados)
        u.set_senha(senha_nova)
        db.session.add(u)
        db.session.commit()
        flash(f'Usuário "{u.nome}" criado!', 'success')
        return redirect(url_for('usuarios'))
    return render_template('form_usuario.html', usuario=None, almoxarifados=almoxarifados)

@app.route('/usuarios/<int:id>/editar', methods=['GET', 'POST'])
@admin_required
def editar_usuario(id):
    u = Usuario.query.get_or_404(id)
    atual = usuario_atual()

    # Admin não pode editar a si mesmo via esta rota (evita auto-lockout acidental)
    # e não pode editar outro admin de nível igual, exceto a si mesmo
    if u.id != atual.id and u.perfil == 'admin' and atual.perfil == 'admin':
        # Permite apenas se o usuário atual for o mesmo sendo editado
        # ou se o alvo não for admin — proteção contra escalada de privilégio
        pass  # admins podem editar outros admins (necessário para gestão)

    almoxarifados = Almoxarifado.query.all()
    if request.method == 'POST':
        # Impede que um admin remova o próprio perfil de admin acidentalmente
        novo_perfil = request.form['perfil']
        if u.id == atual.id and novo_perfil != 'admin':
            flash('Você não pode remover seu próprio perfil de administrador.', 'danger')
            return redirect(url_for('editar_usuario', id=id))
        u.nome = request.form['nome']
        u.login = request.form['login']
        u.perfil = novo_perfil
        u.almoxarifado_id = request.form.get('almoxarifado_id') or None
        u.email = request.form.get('email', '').strip() or None
        u.ativo = 'ativo' in request.form
        # Impede que admin desative a si mesmo
        if u.id == atual.id:
            u.ativo = True
        if request.form.get('senha'):
            if len(request.form['senha']) < 8:
                flash('A senha deve ter pelo menos 8 caracteres.', 'danger')
                return redirect(url_for('editar_usuario', id=id))
            u.set_senha(request.form['senha'])
        u.pode_requisitar = 'pode_requisitar' in request.form
        u.pode_ver_alertas = 'pode_ver_alertas' in request.form
        db.session.commit()
        flash('Usuário atualizado!', 'success')
        return redirect(url_for('usuarios'))
    return render_template('form_usuario.html', usuario=u, almoxarifados=almoxarifados,
                           permissoes_disponiveis=PERMISSOES_DISPONIVEIS)

@app.route('/usuarios/<int:id>/deletar', methods=['POST'])
@admin_required
def deletar_usuario(id):
    u = Usuario.query.get_or_404(id)
    atual = usuario_atual()

    # Fundador (rick) pode deletar qualquer conta exceto a própria
    # Admin comum não pode deletar a si mesmo
    if u.id == atual.id:
        flash('Você não pode remover sua própria conta.', 'danger')
        return redirect(url_for('usuarios'))

    try:
        # Desvincular requisições vinculadas antes de deletar (evita erro de FK)
        RequisicaoMestre.query.filter_by(mestre_id=u.id).update({'mestre_id': atual.id})
        RequisicaoMestre.query.filter_by(entregue_por_id=u.id).update({'entregue_por_id': None})
        db.session.flush()

        db.session.delete(u)
        db.session.commit()
        flash(f'Usuário "{u.nome}" removido!', 'warning')
    except Exception as e:
        db.session.rollback()
        flash('Não foi possível remover o usuário. Ele pode ter registros vinculados no sistema.', 'danger')

    return redirect(url_for('usuarios'))

# ── ACESSO EXTRA (substituto temporário) ─────────────────────────────────────

@app.route('/usuarios/<int:id>/acesso_extra', methods=['POST'])
@admin_required
def conceder_acesso_extra(id):
    u = Usuario.query.get_or_404(id)
    admin = usuario_atual()
    alm_id = request.form.get('almoxarifado_id', type=int)
    motivo = request.form.get('motivo', '')
    data_fim_str = request.form.get('data_fim', '')
    data_fim = datetime.strptime(data_fim_str, '%Y-%m-%dT%H:%M') if data_fim_str else None

    acesso = AcessoExtra(
        usuario_id=id,
        almoxarifado_id=alm_id,
        motivo=motivo,
        data_fim=data_fim,
        concedido_por=admin.nome
    )
    db.session.add(acesso)
    db.session.commit()
    flash(f'Acesso temporário concedido a {u.nome}!', 'success')
    return redirect(url_for('editar_usuario', id=id))

@app.route('/acesso_extra/<int:id>/revogar', methods=['POST'])
@admin_required
def revogar_acesso_extra(id):
    a = AcessoExtra.query.get_or_404(id)
    uid = a.usuario_id
    db.session.delete(a)
    db.session.commit()
    flash('Acesso revogado!', 'warning')
    return redirect(url_for('editar_usuario', id=uid))

# ── PERMISSÕES EXTRAS DE FUNÇÃO ───────────────────────────────────────────────

PERMISSOES_DISPONIVEIS = {
    'fazer_requisicao': 'Fazer Requisições ao Almoxarifado',
    'ver_relatorios':   'Ver Relatórios (Consumo, Ficha EPI)',
    'ver_alertas':      'Ver Alertas de Estoque',
}

@app.route('/usuarios/<int:id>/permissao', methods=['POST'])
@admin_required
def conceder_permissao(id):
    u = Usuario.query.get_or_404(id)
    admin = usuario_atual()
    permissao = request.form.get('permissao', '').strip()
    if permissao not in PERMISSOES_DISPONIVEIS:
        flash('Permissão inválida.', 'danger')
        return redirect(url_for('editar_usuario', id=id))
    # Evita duplicata
    ja_existe = PermissaoExtra.query.filter_by(usuario_id=id, permissao=permissao).first()
    if ja_existe:
        flash(f'Usuário já tem a permissão "{PERMISSOES_DISPONIVEIS[permissao]}".', 'info')
        return redirect(url_for('editar_usuario', id=id))
    db.session.add(PermissaoExtra(
        usuario_id=id,
        permissao=permissao,
        concedido_por=admin.nome,
        data_concessao=agora()
    ))
    db.session.commit()
    flash(f'Permissão "{PERMISSOES_DISPONIVEIS[permissao]}" concedida a {u.nome}!', 'success')
    return redirect(url_for('editar_usuario', id=id))

@app.route('/permissao_extra/<int:pid>/revogar', methods=['POST'])
@admin_required
def revogar_permissao(pid):
    p = PermissaoExtra.query.get_or_404(pid)
    uid = p.usuario_id
    nome_permissao = PERMISSOES_DISPONIVEIS.get(p.permissao, p.permissao)
    db.session.delete(p)
    db.session.commit()
    flash(f'Permissão "{nome_permissao}" revogada!', 'warning')
    return redirect(url_for('editar_usuario', id=uid))

# ── REQUISIÇÕES DO MESTRE DE OBRA ────────────────────────────────────────────

@app.route('/mestre/requisicoes')
@login_required
def mestre_requisicoes():
    """Lista de requisições do mestre logado."""
    u = usuario_atual()

    pode_fazer = (
        u.perfil in ('mestre', 'tecnico_seguranca', 'admin', 'almoxarife') or
        u.pode_requisitar or
        any(p.permissao == 'fazer_requisicao' for p in u.permissoes_extras)
    )
    if not pode_fazer:
        flash('Acesso negado.', 'danger')
        return redirect(url_for('index'))

    if u.perfil == 'admin':
        reqs = RequisicaoMestre.query.order_by(RequisicaoMestre.data_criacao.desc()).all()
    elif u.perfil == 'almoxarife' and u.almoxarifado_id:
        reqs = RequisicaoMestre.query.filter_by(almoxarifado_id=u.almoxarifado_id).order_by(RequisicaoMestre.data_criacao.desc()).all()
    else:
        # mestre, tecnico, engenheiro com pode_requisitar — vê só as suas
        reqs = RequisicaoMestre.query.filter_by(mestre_id=u.id).order_by(RequisicaoMestre.data_criacao.desc()).all()

    return render_template('mestre_requisicoes.html', requisicoes=reqs)

@app.route('/mestre/requisicoes/nova', methods=['GET', 'POST'])
@login_required
def mestre_requisicao_nova():
    """Mestre, técnico, engenheiro (pode_requisitar) e admin criam requisição."""
    u = usuario_atual()

    # ── Verificação de acesso ────────────────────────────────────────────────
    pode_fazer = (
        u.perfil in ('mestre', 'tecnico_seguranca', 'admin') or
        u.pode_requisitar or
        any(p.permissao == 'fazer_requisicao' for p in u.permissoes_extras)
    )
    if not pode_fazer:
        flash('Você não tem permissão para criar requisições.', 'danger')
        return redirect(url_for('index'))

    # ── Almoxarifados que o usuário pode requisitar ──────────────────────────
    if u.perfil == 'admin':
        almoxarifados = Almoxarifado.query.all()
    elif u.perfil == 'tecnico_seguranca':
        ids = u.almoxarifados_permitidos()
        almoxarifados = Almoxarifado.query.filter(Almoxarifado.id.in_(ids)).all() if ids else (
            [u.almoxarifado] if u.almoxarifado_id else []
        )
    else:
        # mestre, colaborador com pode_requisitar, engenheiro — usa almoxarifado vinculado
        if not u.almoxarifado_id:
            flash('Você não está vinculado a nenhum almoxarifado. Contate o administrador.', 'warning')
            return redirect(url_for('mestre_requisicoes'))
        almoxarifados = [u.almoxarifado]

    if not almoxarifados:
        flash('Nenhum almoxarifado disponível para requisição.', 'warning')
        return redirect(url_for('mestre_requisicoes'))

    itens_json = {}
    for alm in almoxarifados:
        # Mestre NÃO pode requisitar EPIs — filtra categoria epi
        # Demais perfis podem requisitar tudo
        if u.perfil == 'mestre':
            itens_filtrados = [it for it in alm.itens if it.ativo and it.categoria != 'epi']
        else:
            itens_filtrados = [it for it in alm.itens if it.ativo]

        itens_json[str(alm.id)] = [
            {'id': it.id, 'nome': it.nome, 'quantidade': it.quantidade,
             'unidade': it.unidade, 'categoria': it.categoria or 'geral'}
            for it in itens_filtrados
        ]

    if request.method == 'POST':
        colaborador = request.form.get('colaborador', '').strip()
        alm_id = int(request.form.get('almoxarifado_id', u.almoxarifado_id or 0))
        observacao = request.form.get('observacao', '')

        if not colaborador:
            flash('Informe o nome do colaborador que vai buscar os materiais.', 'warning')
            return redirect(url_for('mestre_requisicao_nova'))

        # Coletar itens
        indices = set()
        for key in request.form.keys():
            if key.startswith('item_id_'):
                try:
                    indices.add(int(key.split('_')[-1]))
                except ValueError:
                    pass

        if not indices:
            flash('Adicione pelo menos um item à requisição.', 'warning')
            return redirect(url_for('mestre_requisicao_nova'))

        req = RequisicaoMestre(
            mestre_id=u.id,
            colaborador=colaborador,
            almoxarifado_id=alm_id,
            observacao=observacao,
            status='pendente',
            data_criacao=agora()
        )
        db.session.add(req)
        db.session.flush()  # gera o id

        # Gera número de protocolo único: REQ-AAAAMMDD-XXXX
        req.protocolo = f"REQ-{req.data_criacao.strftime('%Y%m%d')}-{req.id:04d}"

        for i in sorted(indices):
            item_id = request.form.get(f'item_id_{i}')
            qtd_str = request.form.get(f'quantidade_{i}')
            obs_item = request.form.get(f'observacao_{i}', '')
            if not item_id or not qtd_str:
                continue
            try:
                qtd = float(qtd_str)
            except ValueError:
                continue
            if qtd <= 0:
                continue
            db.session.add(RequisicaoMestreItem(
                requisicao_id=req.id,
                item_id=int(item_id),
                quantidade=qtd,
                observacao=obs_item
            ))

        db.session.commit()
        flash_html(f'✅ Requisição <strong>{escape(req.protocolo or "#" + str(req.id))}</strong> enviada ao almoxarifado! Aguarde a separação.', 'success')
        return redirect(url_for('mestre_requisicoes'))

    return render_template('mestre_requisicao_nova.html',
                           almoxarifados=almoxarifados,
                           itens_json=json.dumps(itens_json))

@app.route('/mestre/requisicoes/<int:id>')
@login_required
def mestre_requisicao_detalhe(id):
    """Detalhe de uma requisição do mestre."""
    req = RequisicaoMestre.query.get_or_404(id)
    u = usuario_atual()
    # Mestre e técnico de segurança só veem as suas próprias
    if u.perfil in ('mestre', 'tecnico_seguranca') and req.mestre_id != u.id:
        flash('Acesso negado.', 'danger')
        return redirect(url_for('mestre_requisicoes'))
    # Colaborador com pode_requisitar só vê as suas próprias
    if u.perfil == 'colaborador' and req.mestre_id != u.id:
        flash('Acesso negado.', 'danger')
        return redirect(url_for('mestre_requisicoes'))
    # Almoxarife só vê do seu almoxarifado
    if u.perfil == 'almoxarife' and req.almoxarifado_id != u.almoxarifado_id:
        flash('Acesso negado.', 'danger')
        return redirect(url_for('mestre_requisicoes'))
    return render_template('mestre_requisicao_detalhe.html', req=req)

@app.route('/mestre/requisicoes/<int:id>/editar', methods=['GET', 'POST'])
@login_required
def mestre_requisicao_editar(id):
    """Almoxarife ou admin edita uma requisição pendente."""
    req = RequisicaoMestre.query.get_or_404(id)
    u = usuario_atual()
    if u.perfil not in ('admin', 'almoxarife'):
        flash('Apenas almoxarife ou admin pode editar requisições.', 'danger')
        return redirect(url_for('mestre_requisicoes'))
    if req.status == 'entregue':
        flash('Não é possível editar uma requisição já entregue.', 'warning')
        return redirect(url_for('mestre_requisicao_detalhe', id=id))

    if request.method == 'POST':
        req.colaborador = request.form.get('colaborador', req.colaborador).strip()
        req.observacao = request.form.get('observacao', req.observacao)
        # Atualizar quantidades dos itens
        for ri in req.itens:
            qtd_str = request.form.get(f'qtd_{ri.id}')
            obs_str = request.form.get(f'obs_{ri.id}', ri.observacao)
            if qtd_str:
                try:
                    ri.quantidade = float(qtd_str)
                except ValueError:
                    pass
            ri.observacao = obs_str
        db.session.commit()
        flash('Requisição atualizada!', 'success')
        return redirect(url_for('mestre_requisicao_detalhe', id=id))

    return render_template('mestre_requisicao_editar.html', req=req)

@app.route('/mestre/requisicoes/<int:id>/aprovar', methods=['POST'])
@login_required
def mestre_requisicao_aprovar(id):
    """Almoxarife aprova, recusa ou aprova parcialmente a requisição.

    Modos:
    - decisao=aprovada  → aprova todos os itens de uma vez
    - decisao=recusada  → recusa todos os itens
    - decisao=parcial   → avalia item a item via campos item_status_<ri_id>
                          e item_motivo_<ri_id> no formulário
    """
    req = RequisicaoMestre.query.get_or_404(id)
    u = usuario_atual()
    if u.perfil not in ('admin', 'almoxarife'):
        flash('Acesso negado.', 'danger')
        return redirect(url_for('mestre_requisicoes'))
    if req.status != 'pendente':
        flash('Requisição não está pendente.', 'warning')
        return redirect(url_for('mestre_requisicao_detalhe', id=id))

    decisao = request.form.get('decisao', 'aprovada')

    if decisao == 'parcial':
        # Aprovação item a item
        aprovados = 0
        recusados = 0
        for ri in req.itens:
            st = request.form.get(f'item_status_{ri.id}', 'aprovado')
            motivo = request.form.get(f'item_motivo_{ri.id}', '').strip()
            ri.status_item = st  # 'aprovado' ou 'recusado'
            ri.motivo_recusa = motivo if st == 'recusado' else None
            if st == 'aprovado':
                aprovados += 1
            else:
                recusados += 1

        # Define status geral da requisição
        if aprovados == 0:
            req.status = 'recusada'
            flash(f'❌ Requisição #{req.id} recusada — todos os itens foram recusados.', 'danger')
        elif recusados == 0:
            req.status = 'aprovada'
            flash(f'✅ Requisição #{req.id} aprovada! Separe os materiais e confirme a entrega.', 'success')
        else:
            req.status = 'parcial'
            flash(
                f'⚠️ Requisição #{req.id} aprovada parcialmente: '
                f'{aprovados} item(ns) aprovado(s), {recusados} recusado(s).',
                'warning'
            )
    elif decisao == 'recusada':
        req.status = 'recusada'
        for ri in req.itens:
            ri.status_item = 'recusado'
            ri.motivo_recusa = request.form.get('motivo_geral', '').strip() or None
        flash(f'❌ Requisição #{req.id} recusada.', 'danger')
    else:
        # aprovada inteira
        req.status = 'aprovada'
        for ri in req.itens:
            ri.status_item = 'aprovado'
            ri.motivo_recusa = None
        flash(f'✅ Requisição #{req.id} aprovada! Separe os materiais e confirme a entrega.', 'success')

    db.session.commit()
    return redirect(url_for('mestre_requisicao_detalhe', id=id))

@app.route('/mestre/requisicoes/<int:id>/entregar', methods=['POST'])
@login_required
def mestre_requisicao_entregar(id):
    """Almoxarife confirma entrega — baixa o estoque apenas dos itens aprovados."""
    req = RequisicaoMestre.query.get_or_404(id)
    u = usuario_atual()
    if u.perfil not in ('admin', 'almoxarife'):
        if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            return jsonify({'ok': False, 'msg': 'Acesso negado.'}), 403
        flash('Acesso negado.', 'danger')
        return redirect(url_for('mestre_requisicoes'))
    if req.status not in ('pendente', 'aprovada', 'parcial'):
        if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            return jsonify({'ok': False, 'msg': 'Requisição já processada.'})
        flash('Requisição já foi entregue, recusada ou cancelada.', 'warning')
        return redirect(url_for('mestre_requisicao_detalhe', id=id))

    # Itens a entregar: aprovados (ou todos se não houve aprovação por item)
    itens_a_entregar = [
        ri for ri in req.itens
        if ri.status_item in ('aprovado', 'pendente')
    ]

    erros = []
    for ri in itens_a_entregar:
        if ri.quantidade > ri.item.quantidade:
            erros.append(f'"{ri.item.nome}": apenas {ri.item.quantidade} {ri.item.unidade} disponível')

    if erros:
        if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            return jsonify({'ok': False, 'msg': ' | '.join(erros)})
        for e in erros:
            flash(f'⚠️ {e}', 'danger')
        return redirect(url_for('mestre_requisicao_detalhe', id=id))

    for ri in itens_a_entregar:
        ri.item.quantidade = round(ri.item.quantidade - ri.quantidade, 4)
        ri.status_item = 'aprovado'
        db.session.add(Movimentacao(
            tipo='saida',
            quantidade=ri.quantidade,
            responsavel=req.mestre.nome,
            observacao=f'Req. Mestre #{req.id} — Colaborador: {req.colaborador}',
            item_id=ri.item.id
        ))

    req.status = 'entregue'
    req.data_entrega = agora()
    req.entregue_por_id = u.id
    db.session.commit()

    if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
        return jsonify({'ok': True, 'req_id': req.id})

    flash(f'✅ Entrega confirmada! Estoque atualizado para {len(itens_a_entregar)} item(ns).', 'success')
    return redirect(url_for('mestre_requisicao_detalhe', id=id))


@app.route('/mestre/requisicoes/<int:id>/foto', methods=['POST'])
@login_required
def mestre_requisicao_foto(id):
    """Salva foto de comprovante de entrega na requisição."""
    req = RequisicaoMestre.query.get_or_404(id)
    u = usuario_atual()
    if u.perfil not in ('admin', 'almoxarife'):
        return jsonify({'ok': False, 'error': 'Acesso negado.'}), 403
    data = request.get_json(silent=True) or {}
    foto = data.get('foto', '')
    if not foto or not foto.startswith('data:image'):
        return jsonify({'ok': False, 'error': 'Foto inválida.'})
    req.foto_url = foto
    db.session.commit()
    return jsonify({'ok': True})

@app.route('/mestre/requisicoes/<int:id>/cancelar', methods=['POST'])
@login_required
def mestre_requisicao_cancelar(id):
    """Cancela uma requisição pendente ou aprovada."""
    req = RequisicaoMestre.query.get_or_404(id)
    u = usuario_atual()
    # Mestre, técnico de segurança e colaborador com pode_requisitar só cancelam as suas
    if u.perfil in ('mestre', 'tecnico_seguranca') and req.mestre_id != u.id:
        flash('Acesso negado.', 'danger')
        return redirect(url_for('mestre_requisicoes'))
    if u.perfil == 'colaborador' and req.mestre_id != u.id:
        flash('Acesso negado.', 'danger')
        return redirect(url_for('mestre_requisicoes'))
    if req.status == 'entregue':
        flash('Não é possível cancelar uma requisição já entregue.', 'warning')
        return redirect(url_for('mestre_requisicao_detalhe', id=id))

    req.status = 'cancelada'
    db.session.commit()
    flash(f'Requisição #{req.id} cancelada.', 'warning')
    return redirect(url_for('mestre_requisicoes'))

@app.route('/api/mestre/notificacoes')
@login_required
def api_mestre_notificacoes():
    """Retorna notificações baseadas no status das requisições do mestre."""
    u = usuario_atual()
    if u.perfil not in ('mestre', 'tecnico_seguranca'):
        return jsonify([])
    # Busca requisições não-pendentes e não-canceladas do mestre
    # A notificação é baseada no status atual — o frontend controla o que já foi visto via localStorage
    reqs = RequisicaoMestre.query.filter_by(mestre_id=u.id).filter(
        RequisicaoMestre.status.in_(['aprovada', 'parcial', 'recusada', 'entregue'])
    ).order_by(RequisicaoMestre.data_criacao.desc()).limit(10).all()

    result = []
    for r in reqs:
        if r.status == 'aprovada':
            msg = f'✅ Requisição #{r.id} aprovada! Envie o colaborador buscar.'
            tipo = 'success'
        elif r.status == 'parcial':
            msg = f'⚠️ Requisição #{r.id} aprovada parcialmente. Verifique os itens.'
            tipo = 'warning'
        elif r.status == 'recusada':
            msg = f'❌ Requisição #{r.id} foi recusada pelo almoxarifado.'
            tipo = 'danger'
        elif r.status == 'entregue':
            msg = f'📦 Requisição #{r.id} entregue ao colaborador {r.colaborador}.'
            tipo = 'info'
        else:
            continue
        result.append({'id': r.id, 'msg': msg, 'tipo': tipo, 'status': r.status})
    return jsonify(result)

@app.route('/api/mestre/notificacoes/marcar-lidas', methods=['POST'])
@login_required
def marcar_notificacoes_lidas():
    """Endpoint mantido para compatibilidade — controle feito no frontend."""
    return jsonify({'ok': True})

# ── ROTA ESPECIAL PARA REATIVAR TODOS OS ITENS ──────────────────────────────

@app.route('/admin/reativar-todos-itens', methods=['GET', 'POST'])
@admin_required
def reativar_todos_itens():
    if request.method == 'POST':
        try:
            # Reativar todos os itens
            itens_desativados = Item.query.filter_by(ativo=False).all()
            count = 0
            for item in itens_desativados:
                item.ativo = True
                count += 1
            
            # Também garantir que itens com ativo=None sejam ativados
            from sqlalchemy import text
            with db.engine.connect() as conn:
                result = conn.execute(text("UPDATE item SET ativo = 1 WHERE ativo IS NULL OR ativo = 0"))
                conn.commit()
            
            db.session.commit()
            
            total_ativos = Item.query.filter_by(ativo=True).count()
            flash(f'✅ Sucesso! {count} itens reativados. Total de itens ativos: {total_ativos}', 'success')
            
        except Exception as e:
            flash(f'❌ Erro ao reativar itens: {str(e)}', 'danger')
        
        return redirect(url_for('reativar_todos_itens'))
    
    # GET - mostrar página de confirmação
    itens_desativados = Item.query.filter_by(ativo=False).count()
    total_itens = Item.query.count()
    return render_template('admin_reativar_itens.html',
                           itens_desativados=itens_desativados,
                           total_itens=total_itens)

# ── BACKUP ───────────────────────────────────────────────────────────────────

def gerar_excel_backup_almoxarifado(alm):
    """Gera um Excel com apenas um almoxarifado (para envio ao almoxarife)."""
    wb = openpyxl.Workbook()
    h_fill = PatternFill('solid', fgColor='1A3A5C')
    h_font = Font(bold=True, color='FFFFFF', size=11)
    ok_fill  = PatternFill('solid', fgColor='D4EDDA')
    al_fill  = PatternFill('solid', fgColor='FFF3CD')
    cr_fill  = PatternFill('solid', fgColor='F8D7DA')
    borda = Border(left=Side(style='thin'), right=Side(style='thin'),
                   top=Side(style='thin'), bottom=Side(style='thin'))

    itens = Item.query.filter_by(almoxarifado_id=alm.id).all()
    ws = wb.active
    ws.title = alm.nome[:31]

    ws.merge_cells('A1:G1')
    ws['A1'] = f'Backup — {alm.nome}'
    ws['A1'].font = Font(bold=True, size=13, color='1A3A5C')
    ws['A1'].alignment = Alignment(horizontal='center')
    ws.merge_cells('A2:G2')
    ws['A2'] = f'Gerado em: {datetime.now().strftime("%d/%m/%Y %H:%M")}'
    ws['A2'].font = Font(italic=True, color='888888')
    ws['A2'].alignment = Alignment(horizontal='center')

    headers = ['Codigo', 'Item', 'Unidade', 'Qtd. Atual', 'Est. Minimo', 'Status', 'Ativo']
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=4, column=col, value=h)
        c.font = h_font; c.fill = h_fill
        c.alignment = Alignment(horizontal='center'); c.border = borda

    for r, it in enumerate(itens, 5):
        status = 'ZERADO' if it.status == 'critico' else ('ABAIXO DO MINIMO' if it.status == 'alerta' else 'OK')
        fill = cr_fill if it.status == 'critico' else (al_fill if it.status == 'alerta' else ok_fill)
        for c, v in enumerate([it.codigo, it.nome, it.unidade, it.quantidade,
                                it.estoque_minimo, status, 'Sim' if it.ativo else 'Não'], 1):
            cell = ws.cell(row=r, column=c, value=v)
            cell.fill = fill; cell.border = borda
            cell.alignment = Alignment(horizontal='left' if c == 2 else 'center')

    for i, w in enumerate([14, 40, 10, 12, 14, 20, 8], 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf

def gerar_excel_backup():
    """Gera um Excel completo com todos os almoxarifados como backup."""
    wb = openpyxl.Workbook()
    h_fill = PatternFill('solid', fgColor='1A3A5C')
    h_font = Font(bold=True, color='FFFFFF', size=11)
    ok_fill  = PatternFill('solid', fgColor='D4EDDA')
    al_fill  = PatternFill('solid', fgColor='FFF3CD')
    cr_fill  = PatternFill('solid', fgColor='F8D7DA')
    borda = Border(left=Side(style='thin'), right=Side(style='thin'),
                   top=Side(style='thin'), bottom=Side(style='thin'))

    almoxarifados = Almoxarifado.query.all()
    primeira = True

    for alm in almoxarifados:
        itens = Item.query.filter_by(almoxarifado_id=alm.id).all()
        ws = wb.active if primeira else wb.create_sheet()
        primeira = False
        ws.title = alm.nome[:31]  # Excel limita 31 chars no nome da aba

        # Título
        ws.merge_cells('A1:G1')
        ws['A1'] = f'Backup — {alm.nome}'
        ws['A1'].font = Font(bold=True, size=13, color='1A3A5C')
        ws['A1'].alignment = Alignment(horizontal='center')
        ws.merge_cells('A2:G2')
        ws['A2'] = f'Gerado em: {datetime.now().strftime("%d/%m/%Y %H:%M")}'
        ws['A2'].font = Font(italic=True, color='888888')
        ws['A2'].alignment = Alignment(horizontal='center')

        # Cabeçalho
        headers = ['Codigo', 'Item', 'Unidade', 'Qtd. Atual', 'Est. Minimo', 'Status', 'Ativo']
        for col, h in enumerate(headers, 1):
            c = ws.cell(row=4, column=col, value=h)
            c.font = h_font; c.fill = h_fill
            c.alignment = Alignment(horizontal='center'); c.border = borda

        for r, it in enumerate(itens, 5):
            status = 'ZERADO' if it.status == 'critico' else ('ABAIXO DO MINIMO' if it.status == 'alerta' else 'OK')
            fill = cr_fill if it.status == 'critico' else (al_fill if it.status == 'alerta' else ok_fill)
            for c, v in enumerate([it.codigo, it.nome, it.unidade, it.quantidade,
                                    it.estoque_minimo, status, 'Sim' if it.ativo else 'Não'], 1):
                cell = ws.cell(row=r, column=c, value=v)
                cell.fill = fill; cell.border = borda
                cell.alignment = Alignment(horizontal='left' if c == 2 else 'center')

        for i, w in enumerate([14, 40, 10, 12, 14, 20, 8], 1):
            ws.column_dimensions[get_column_letter(i)].width = w

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf



@app.route('/admin/seed-colaboradores-infra', methods=['POST'])
@admin_required
def seed_colaboradores_infra():
    """Cadastra os colaboradores de infraestrutura da QLP — ignora duplicatas."""
    _lista = [
        ("ADAILTON JOSE DOS SANTOS", "Ajudante"),
        ("ALAN SOUZA DE OLIVEIRA", "Ajudante"),
        ("ANDRE DE JESUS MENDES", "Ajudante"),
        ("ANTONIO LUCAS NASCIMENTO BISPO", "Assistente de Produção"),
        ("ARODOALDO PEREIRA DA ROCHA", "Encanador"),
        ("CARLOS ALBERTO LOPES SILVA", "Ajudante"),
        ("CLAUDEMIRO GALVÃO DOS SANTOS", "Auxiliar de Serviços Gerais"),
        ("DEYVID DE SANTANA LOPES", "Almoxarife"),
        ("EDINEILSON DOS SANTOS DE OLIVEIRA", "Eletricista"),
        ("EDNILSON ASSIS DOS SANTOS", "Ajudante"),
        ("EDSON ANTONIO SANTOS DE OLIVEIRA", "Ajudante"),
        ("GERSON SILVA", "Mestre de Obras"),
        ("HEBERT DA SILVA MEDRADO", "Ajudante"),
        ("JACKSON DA SILVA MENEZES DOS SANTOS", "Ajudante"),
        ("JEFFERSON SANTOS RIBEIRO", "Ajudante"),
        ("JOSE SEVERINO MENDES DA SILVA", "Carpinteiro"),
        ("LAURA DOS SANTOS ARAÚJO", "Coordenadora"),
        ("LEONARDO VIDAL DOS SANTOS SENA", "Técnico de Segurança"),
        ("MARCOS VINICIUS SAMPAIO ROSA", "Ajudante"),
        ("MATEUS DE JESUS SANTANA SANTOS", "Ajudante"),
        ("RAILAN NASCIMENTO SANTOS", "Ajudante"),
        ("RODRIGO NASCIMENTO SANTANA", "Ajudante"),
        ("TIAGO DA SILVA FERREIRA", "Auxiliar de Serviços Gerais"),
        ("JOÃO FRANCISCO C DE JESUS RODRIGUES", "Ajudante Prático de Elétrica"),
        ("JADSON DIAS DE OLIVEIRA SOUSA", "Ajudante Comum"),
        ("JOSÉ AUGUSTO DOS SANTOS BISPO", "Ajudante Comum"),
        ("WALTER BATISTA DOS SANTOS FILHO", "Carpinteiro"),
    ]
    inseridos = 0
    for nome, funcao in _lista:
        if not Colaborador.query.filter(Colaborador.nome.ilike(nome)).first():
            db.session.add(Colaborador(nome=nome, funcao=funcao, escopo='infraestrutura', ativo=True))
            inseridos += 1
    db.session.commit()
    flash(f'✅ {inseridos} colaboradores de infraestrutura cadastrados!', 'success')
    return redirect(url_for('colaboradores'))


@app.route('/admin/seed-colaboradores', methods=['POST'])
@admin_required
def seed_colaboradores():
    """Cadastra os colaboradores da estrutura em massa — ignora duplicatas."""
    _lista = [
        ("ARILSON DE JESUS SOUZA","Profissional"),("ADSON MUNIZ","Ajudante"),
        ("MARCEL OLIVEIRA DA CONCEIÇÃO","Profissional"),("ROBERT WILLIAM DA HORA DE JESUS","Profissional"),
        ("MATEUS SANTOS DE JESUS","Profissional"),("ANIBAL SANTOS DANTAS","Profissional"),
        ("ROBERTO FELIX GONÇALVES","Profissional"),("EDNALDO DOS SANTOS","Profissional"),
        ("ALEXSANDRO TELES DOS SANTOS","Profissional"),("RUAN UITALO","Ajudante"),
        ("FELIPE MESSIAS","Ajudante"),("ROQUE DOS SANTOS","Profissional"),
        ("VALDOMIRO GOMES DE JESUS FILHO","Profissional"),("VALMIR GOMES DE JESUS","Profissional"),
        ("TIAGO GOMES DOS SANTOS","Ajudante"),("ADRIANO SOUZA DOS SANTOS","Profissional"),
        ("RONALDO DA CUNHA SANTOS","Profissional"),("EMERSON DE SANTANA ARAUJO","Profissional"),
        ("LUAN DOS SANTOS CARDOSO","Profissional"),("FRANCISCO CARLOS DOS SANTOS FILHO","Profissional"),
        ("NILSON MATIAS DOS SANTOS","Profissional"),("VINICIUS DANTAS DA SILVA","Profissional"),
        ("MAURICIO RAMON PINHEIRO MATOS","Ajudante"),("CARLOS ALBERTO BISPO DOS SANTOS","Ajudante"),
        ("DIEGO LIMA SANTOS","Ajudante"),("AILTON DA SILVA","Profissional"),
        ("EIDSON SILVA ROCHA","Ajudante"),("MARLEI ASSIS DE SOUZA","Profissional"),
        ("RICARDO VASQUES LEMOS LEONI","Profissional"),("ROBISON SANTOS DA CONCEIÇÃO","Profissional"),
        ("LUIS SILVAN LOPES DOS SANTOS","Profissional"),("JAIR CESAR BRITO RODRIGUES JUNIOR","Ajudante"),
        ("ROBSON BISPO DOS SANTOS","Profissional"),("EDVAN MACHADO SANTOS","Profissional"),
        ("LUIS ALBERTO MOREIRA DA SILVA","Ajudante"),("ISAAC GONÇALVES DA SILVA","Ajudante"),
        ("ANTONIO MARCOS DA SILVA COSTA","Ajudante"),("DENAILTON LEITE DOS SANTOS","Ajudante"),
        ("MARCIO DE JESUS DOS SANTOS","Profissional"),("WEBER OLIVEIRA DA LUZ","Ajudante"),
        ("RAFAEL DA SILVA BOMFIM","Ajudante"),("ANDERSON RODRIGUES DOS SANTOS","Profissional"),
        ("JAILTON RIBEIRO TOSTA","Profissional"),("JOAO LUIS OLIVEIRA DA SILVA","Profissional"),
        ("ANDERSON SOUZA DE FRIAS","Profissional"),("JOILSON DOS SANTOS","Profissional"),
        ("JOANDERSON ALMEIDA BISPO","Profissional"),("ANTONIO CARLOS SANTOS SILVA","Profissional"),
        ("NAILTON CONCEIÇÃO DE SOUZA","Ajudante"),("LUCAS SILVA DOS REIS","Ajudante"),
        ("ROBSON LIMA MACIEL","Profissional"),("ATILA ALMEIDA SILVA SANTOS","Ajudante"),
        ("JONAS DE SENA BARRETO","Ajudante"),("SAMUEL BISPO DOS SANTOS","Profissional"),
        ("GUILHERME SANTOS SAMPAIO","Ajudante"),("DANIEL SÃO PEDRO DOS SANTOS","Ajudante"),
        ("DIVINO CARDOSO DOS SANTOS","Ajudante"),("ANDERSON CONCEIÇÃO DE JESUS","Profissional"),
        ("ALEX DE JESUS DA SILVA","Profissional"),("ALEX VITÓRIO SILVA","Ajudante"),
        ("CARLOS DANIEL DA SILVA MARQUES","Ajudante"),("THIEGO DE OLIVEIRA REIS","Profissional"),
        ("UBIRATTAN SNATOS SOUZA","Ajudante"),("WALISSON SILVA COSTA","Ajudante"),
        ("VALMIR GONÇALVES DE OLIVEIRA","Profissional"),("JUDICAEL LEITE DOS SANTOS","Profissional"),
        ("JOÃO PEDRO SILVA DOS SANTOS","Profissional"),("JORGE DOS SANTOS","Profissional"),
        ("JEAN AUGUSTO DOS SANTOS TAVARES","Profissional"),
    ]
    inseridos = 0
    for nome, funcao in _lista:
        if not Colaborador.query.filter(Colaborador.nome.ilike(nome)).first():
            db.session.add(Colaborador(nome=nome, funcao=funcao, escopo='estrutura', ativo=True))
            inseridos += 1
    db.session.commit()
    flash(f'✅ {inseridos} colaboradores cadastrados da estrutura!', 'success')
    return redirect(url_for('colaboradores'))

@app.route('/admin/seed-ferramentas-estrutura', methods=['POST'])
@admin_required
def seed_ferramentas_estrutura():
    """Cadastra as ferramentas da Estrutura Ventura Patamares — ignora duplicatas pelo ID."""
    # Busca o almoxarifado de Estrutura
    alm = Almoxarifado.query.filter(Almoxarifado.nome.ilike('%estrutura%')).first()
    if not alm:
        flash('Almoxarifado de Estrutura não encontrado.', 'danger')
        return redirect(url_for('index'))

    ferramentas_lista = [
        ("INGFH007",  "PISTOLA DE FIXACAO A BATERIA BX 3-L A22MA HILTI GF",   "HILTI"),
        ("IN450348",  "MARTELO SDS PLUS C/ PUNHO",                              ""),
        ("INMRM158",  "MARTELO ROMPEDOR MMR1700 45J 12KG MONO 220V 60HZ",      "MENEGOTTI"),
        ("IN580121",  "MISTURADOR DE ARGAMASSA MAV1600 220V",                   "MENEGOTTI"),
        ("IN580039",  "MISTURADOR DE ARGAMASSA MAV1600 220V",                   "MENEGOTTI"),
        ("INGFH003",  "PISTOLA DE FIXACAO A BATERIA BX 3-L A22MA HILTI GF",    "HILTI"),
        ("INEAG119",  "ESMERILHADEIRA ANGULAR 5\" C/ ACESSORIOS",               "MAKITA"),
        ("IN270013",  "FURADEIRA 3/4 FUR3/4P",                                  ""),
        ("IN270028",  "FURADEIRA 3/4 FUR3/4P",                                  ""),
        ("IN270004",  "FURADEIRA 3/4 FUR3/4P",                                  ""),
        ("IN270005",  "FURADEIRA 3/4 FUR3/4P",                                  ""),
        ("INFPB017",  "FURADEIRA E PARAFUSADEIRA A BATERIA MFI-20 127/220V",    "MENEGOTTI"),
        ("IN340056",  "LAVA JATO HD 585 PROFISSIONAL MODELO 585",               ""),
        ("INMSV108",  "MARTELETE PERF/ROMP MPV1500 5,5J 220V",                  "VONDER"),
        ("INHT930012","KIT ASPIRADOR UNIV. HILTI / POLIDORA DE BETAO HILTI",    "HILTI"),
        ("INMSV261",  "MARTELETE PERF/ROMP MPV1500 5,5J 220V",                  "VONDER"),
        ("INMSV269",  "MARTELETE PERF/ROMP MPV1500 5,5J 220V",                  "VONDER"),
        ("IN240093",  "ESMERILHADEIRA ANGULAR 7\" 220V",                         ""),
        ("INMSV256",  "MARTELETE PERF/ROMP MPV1500 5,5J 220V",                  "VONDER"),
        ("INMSV257",  "MARTELETE PERF/ROMP MPV1500 5,5J 220V",                  "VONDER"),
        ("INMSU042",  "MISTURADOR ELETRICO MEL1600 MONO 220V 60HZ 1600W",       "MENEGOTTI"),
        ("INSER830025","SERRA CIRCULAR 7\"",                                      ""),
        ("INEAG243",  "ESMERILHADEIRA ANGULAR 5\" C/ ACESSORIOS",               "MAKITA"),
        ("INEAG233",  "ESMERILHADEIRA ANGULAR 5\" C/ ACESSORIOS",               "MAKITA"),
        ("INEAG236",  "ESMERILHADEIRA ANGULAR 5\" C/ ACESSORIOS",               "MAKITA"),
        ("INMRM149",  "MARTELO ROMPEDOR MMR1700 45J 12KG MONO 220V 60HZ",      "MENEGOTTI"),
        ("INMSV069",  "MARTELETE PERF/ROMP MPV1500 5,5J 220V",                  "VONDER"),
        ("INMSV156",  "MARTELETE PERF/ROMP MPV1500 5,5J 220V",                  "VONDER"),
        ("IN620015",  "KIT NIVELADOR A LASER HILTI SKR200",                     "HILTI"),
        ("INMSV067",  "MARTELETE PERF/ROMP MPV1500 5,5J 220V",                  "VONDER"),
        ("INMSV001",  "MARTELETE PERF/ROMP MPV1500 5,5J 220V",                  "VONDER"),
        ("INMSV099",  "MARTELETE PERF/ROMP MPV1500 5,5J 220V",                  "VONDER"),
        ("INMSV273",  "MARTELETE PERF/ROMP MPV1500 5,5J 220V",                  "VONDER"),
        ("INHT930015","KIT ASPIRADOR UNIV. HILTI / POLIDORA DE BETAO HILTI",    "HILTI"),
        ("INMSV024",  "MARTELETE PERF/ROMP MPV1500 5,5J 220V",                  "VONDER"),
        ("INMSV113",  "MARTELETE PERF/ROMP MPV1500 5,5J 220V",                  "VONDER"),
        ("INMSV131",  "MARTELETE PERF/ROMP MPV1500 5,5J 220V",                  "VONDER"),
        ("INMSV278",  "MARTELETE PERF/ROMP MPV1500 5,5J 220V",                  "VONDER"),
        ("INMSV148",  "MARTELETE PERF/ROMP MPV1500 5,5J 220V",                  "VONDER"),
        ("INHT930010","KIT ASPIRADOR UNIV. HILTI / POLIDORA DE BETAO HILTI",    "HILTI"),
        ("INHT930011","KIT ASPIRADOR UNIV. HILTI / POLIDORA DE BETAO HILTI",    "HILTI"),
        ("INMSV090",  "MARTELETE PERF/ROMP MPV1500 5,5J 220V",                  "VONDER"),
        ("INMSV281",  "MARTELETE PERF/ROMP MPV1500 5,5J 220V",                  "VONDER"),
        ("INMSV260",  "MARTELETE PERF/ROMP MPV1500 5,5J 220V",                  "VONDER"),
        ("IN850474",  "SERRA MARMORE C/ CHAVE",                                  ""),
    ]

    inseridas = 0
    ignoradas = 0
    for idf, nome, empresa in ferramentas_lista:
        existe = Ferramenta.query.filter_by(identificacao=idf, ativo=True).first()
        if existe:
            ignoradas += 1
            continue
        db.session.add(Ferramenta(
            identificacao=idf,
            nome=nome,
            empresa=empresa or None,
            almoxarifado_id=alm.id,
            status='disponivel'
        ))
        inseridas += 1

    db.session.commit()
    flash(f'✅ {inseridas} ferramentas cadastradas no {alm.nome}! ({ignoradas} já existiam)', 'success')
    return redirect(url_for('ferramentas', alm_id=alm.id))

@app.route('/admin/classificar-epis', methods=['POST'])
@admin_required
def classificar_epis():
    """Classifica automaticamente os itens conhecidos como EPI/Maquinário no banco."""
    # Palavras-chave que identificam EPIs
    palavras_epi = [
        'bota', 'capacete', 'carneira', 'cinto de segurança', 'capa de chuva',
        'calça brim', 'camisa brim', 'macacão', 'mascara', 'máscara',
        'luva vaqueta', 'luva flextactil', 'perneira', 'protetor auricular',
        'óculos de proteção', 'óculos de segurança', 'óculos de sobrepor',
        'talabarte', 'trava-quedas', 'mosquetão oval', 'cinto paraquedista',
        'epi', 'uniforme', 'colete', 'capacete'
    ]
    # Palavras-chave que identificam Maquinário/Peça
    palavras_maq = [
        'broca diamantada', 'disco diamantado', 'disco de desbaste',
        'maçarico', 'perfuratriz', 'abrasiva'
    ]

    atualizados_epi = 0
    atualizados_maq = 0
    itens = Item.query.all()
    for it in itens:
        nome_lower = it.nome.lower()
        if any(p in nome_lower for p in palavras_epi):
            if it.categoria != 'epi':
                it.categoria = 'epi'
                atualizados_epi += 1
        elif any(p in nome_lower for p in palavras_maq):
            if it.categoria != 'maquinario':
                it.categoria = 'maquinario'
                atualizados_maq += 1

    db.session.commit()
    flash(f'✅ Classificação concluída: {atualizados_epi} EPIs e {atualizados_maq} Maquinários atualizados.', 'success')
    return redirect(url_for('index'))


@app.route('/admin/debug-env')
@admin_required
def debug_env():
    """Rota de diagnóstico — apenas mostra quais variáveis estão definidas, sem expor valores."""
    variaveis = ['BACKUP_EMAIL_FROM', 'BACKUP_EMAIL_PASS', 'BACKUP_EMAIL_TO',
                 'RESEND_API_KEY', 'RESEND_FROM_EMAIL', 'SECRET_KEY', 'DATABASE_URL']
    status = {v: '✅ Definida' if os.environ.get(v) else '❌ Não definida' for v in variaveis}
    linhas = '\n'.join(f'  {k} = {v}' for k, v in status.items())
    return f'<pre style="font-family:monospace;padding:20px">\nVariáveis de ambiente:\n\n{linhas}\n</pre>'

@app.route('/admin/backup', methods=['GET', 'POST'])
@admin_required
def backup_manual():
    """Admin faz backup manual — baixa Excel."""
    if request.method == 'POST':
        # Download direto
        try:
            buf = gerar_excel_backup()
            nome = f"backup_estoque_{date.today()}.xlsx"
            return send_file(buf, as_attachment=True, download_name=nome,
                             mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        except Exception as e:
            flash(f'❌ Erro ao gerar backup: {str(e)}', 'danger')
            return redirect(url_for('backup_manual'))

    # GET — mostra a página
    return render_template('backup.html')

def seed_data():
    if Almoxarifado.query.count() == 0:
        db.session.add_all([
            Almoxarifado(nome='Almoxarifado do Acampamento', descricao='Materiais de uso geral do acampamento'),
            Almoxarifado(nome='Almoxarifado de Infraestrutura', descricao='Materiais de construcao e manutencao'),
            Almoxarifado(nome='Almoxarifado de Forma', descricao='Formas, escoramentos e materiais de forma'),
        ])
        db.session.commit()
    if Usuario.query.count() == 0:
        # Gera senha aleatória segura — nunca usa senha padrão fixa
        senha_inicial = secrets.token_urlsafe(12)
        admin = Usuario(nome='Administrador', login='admin', perfil='admin')
        admin.set_senha(senha_inicial)
        db.session.add(admin)
        db.session.commit()
        logger.info('=' * 60)
        logger.warning('⚠️  PRIMEIRO ACESSO — GUARDE ESTAS CREDENCIAIS:')
        logger.info(f'   Login: admin')
        logger.info(f'   Senha: {senha_inicial}')
        logger.warning('   Altere a senha imediatamente após o primeiro login!')
        logger.info('=' * 60)

    # Seed de colaboradores da estrutura (roda uma vez se não houver nenhum)
    if Colaborador.query.count() == 0:
        _colaboradores_estrutura = [
            ("ARILSON DE JESUS SOUZA","Profissional"),("ADSON MUNIZ","Ajudante"),
            ("MARCEL OLIVEIRA DA CONCEIÇÃO","Profissional"),("ROBERT WILLIAM DA HORA DE JESUS","Profissional"),
            ("MATEUS SANTOS DE JESUS","Profissional"),("ANIBAL SANTOS DANTAS","Profissional"),
            ("ROBERTO FELIX GONÇALVES","Profissional"),("EDNALDO DOS SANTOS","Profissional"),
            ("ALEXSANDRO TELES DOS SANTOS","Profissional"),("RUAN UITALO","Ajudante"),
            ("FELIPE MESSIAS","Ajudante"),("ROQUE DOS SANTOS","Profissional"),
            ("VALDOMIRO GOMES DE JESUS FILHO","Profissional"),("VALMIR GOMES DE JESUS","Profissional"),
            ("TIAGO GOMES DOS SANTOS","Ajudante"),("ADRIANO SOUZA DOS SANTOS","Profissional"),
            ("RONALDO DA CUNHA SANTOS","Profissional"),("EMERSON DE SANTANA ARAUJO","Profissional"),
            ("LUAN DOS SANTOS CARDOSO","Profissional"),("FRANCISCO CARLOS DOS SANTOS FILHO","Profissional"),
            ("NILSON MATIAS DOS SANTOS","Profissional"),("VINICIUS DANTAS DA SILVA","Profissional"),
            ("MAURICIO RAMON PINHEIRO MATOS","Ajudante"),("CARLOS ALBERTO BISPO DOS SANTOS","Ajudante"),
            ("DIEGO LIMA SANTOS","Ajudante"),("AILTON DA SILVA","Profissional"),
            ("EIDSON SILVA ROCHA","Ajudante"),("MARLEI ASSIS DE SOUZA","Profissional"),
            ("RICARDO VASQUES LEMOS LEONI","Profissional"),("ROBISON SANTOS DA CONCEIÇÃO","Profissional"),
            ("LUIS SILVAN LOPES DOS SANTOS","Profissional"),("JAIR CESAR BRITO RODRIGUES JUNIOR","Ajudante"),
            ("ROBSON BISPO DOS SANTOS","Profissional"),("EDVAN MACHADO SANTOS","Profissional"),
            ("LUIS ALBERTO MOREIRA DA SILVA","Ajudante"),("ISAAC GONÇALVES DA SILVA","Ajudante"),
            ("ANTONIO MARCOS DA SILVA COSTA","Ajudante"),("DENAILTON LEITE DOS SANTOS","Ajudante"),
            ("MARCIO DE JESUS DOS SANTOS","Profissional"),("WEBER OLIVEIRA DA LUZ","Ajudante"),
            ("RAFAEL DA SILVA BOMFIM","Ajudante"),("ANDERSON RODRIGUES DOS SANTOS","Profissional"),
            ("JAILTON RIBEIRO TOSTA","Profissional"),("JOAO LUIS OLIVEIRA DA SILVA","Profissional"),
            ("ANDERSON SOUZA DE FRIAS","Profissional"),("JOILSON DOS SANTOS","Profissional"),
            ("JOANDERSON ALMEIDA BISPO","Profissional"),("ANTONIO CARLOS SANTOS SILVA","Profissional"),
            ("NAILTON CONCEIÇÃO DE SOUZA","Ajudante"),("LUCAS SILVA DOS REIS","Ajudante"),
            ("ROBSON LIMA MACIEL","Profissional"),("ATILA ALMEIDA SILVA SANTOS","Ajudante"),
            ("JONAS DE SENA BARRETO","Ajudante"),("SAMUEL BISPO DOS SANTOS","Profissional"),
            ("GUILHERME SANTOS SAMPAIO","Ajudante"),("DANIEL SÃO PEDRO DOS SANTOS","Ajudante"),
            ("DIVINO CARDOSO DOS SANTOS","Ajudante"),("ANDERSON CONCEIÇÃO DE JESUS","Profissional"),
            ("ALEX DE JESUS DA SILVA","Profissional"),("ALEX VITÓRIO SILVA","Ajudante"),
            ("CARLOS DANIEL DA SILVA MARQUES","Ajudante"),("THIEGO DE OLIVEIRA REIS","Profissional"),
            ("UBIRATTAN SNATOS SOUZA","Ajudante"),("WALISSON SILVA COSTA","Ajudante"),
            ("VALMIR GONÇALVES DE OLIVEIRA","Profissional"),("JUDICAEL LEITE DOS SANTOS","Profissional"),
            ("JOÃO PEDRO SILVA DOS SANTOS","Profissional"),("JORGE DOS SANTOS","Profissional"),
            ("JEAN AUGUSTO DOS SANTOS TAVARES","Profissional"),
        ]
        for nome, funcao in _colaboradores_estrutura:
            db.session.add(Colaborador(nome=nome, funcao=funcao, escopo='estrutura', ativo=True))
        db.session.commit()
        logger.info(f'Seed: {len(_colaboradores_estrutura)} colaboradores da estrutura cadastrados.')

def classificar_categorias_itens():
    """Classifica automaticamente itens pelo nome: EPI, Maquinário, Elétrica, Hidráulica, Gás."""
    palavras_epi = [
        'bota', 'capacete', 'carneira', 'cinto de segurança', 'capa de chuva',
        'calça brim', 'camisa brim', 'macacão', 'mascara', 'máscara',
        'luva vaqueta', 'luva flextactil', 'perneira', 'protetor auricular',
        'óculos de proteção', 'óculos de segurança', 'óculos de sobrepor',
        'talabarte', 'trava-quedas', 'mosquetão oval', 'cinto paraquedista',
        'uniforme', 'colete refletivo'
    ]
    palavras_maq = [
        'broca diamantada', 'disco diamantado', 'disco de desbaste',
        'maçarico', 'perfuratriz'
    ]
    palavras_eletrica = [
        'fio', 'cabo eletrico', 'cabo pp', 'eletroduto', 'disjuntor', 'tomada',
        'interruptor', 'condulete', 'luminaria', 'lampada', 'quadro de distribuicao',
        'wago', 'conector eletrico', 'passe fio eletrico', 'cabo flexivel',
        'fita isolante', 'eletrica', 'eletrico', 'no-break', 'estabilizador',
        'rgnh', 'thhn', 'sintenax', 'cabo 1.5', 'cabo 2.5', 'cabo 4mm', 'cabo 6mm',
        'cabo 10mm', 'cabo 16mm', 'cabo 35mm', 'canaleta', 'dimer', 'sensor de presença'
    ]
    palavras_hidraulica = [
        'tubo pvc', 'joelho pvc', 'te pvc', 'registro', 'torneira', 'chuveiro',
        'vaso sanitario', 'caixa dagua', 'sifao', 'valvula', 'te soldavel',
        'te reducao', 'joelho soldavel', 'luva soldavel', 'cap pvc', 'adaptador pvc',
        'bucha de reducao', 'curva pvc', 'tê reducao', 'cola pvc', 'lixa pvc',
        'cano pvc', 'tubo soldavel', 'pvc soldavel', 'flange', 'niple',
        'boia de nivel', 'bomba dagua', 'aquecedor', 'caixa sifonada', 'ralo',
        'vedacao', 'veda rosca', 'teflon', 'mangueira', 'pressao agua'
    ]
    palavras_gas = [
        'gas', 'tubulacao gas', 'registro gas', 'botijao', 'mangueira gas',
        'cobre gas', 'tubo cobre', 'solda cobre', 'abracadeira gas',
        'valvula gas', 'regulador gas', 'flexivel gas', 'conexao gas'
    ]
    try:
        itens = Item.query.filter(Item.categoria.in_(['geral', None])).all()
        atualizados = 0
        for it in itens:
            nome_lower = it.nome.lower()
            if any(p in nome_lower for p in palavras_epi):
                it.categoria = 'epi'; atualizados += 1
            elif any(p in nome_lower for p in palavras_maq):
                it.categoria = 'maquinario'; atualizados += 1
            elif any(p in nome_lower for p in palavras_eletrica):
                it.categoria = 'eletrica'; atualizados += 1
            elif any(p in nome_lower for p in palavras_hidraulica):
                it.categoria = 'hidraulica'; atualizados += 1
            elif any(p in nome_lower for p in palavras_gas):
                it.categoria = 'gas'; atualizados += 1
        if atualizados:
            db.session.commit()
            logger.info(f'Categorias: {atualizados} itens classificados automaticamente.')
    except Exception as e:
        logger.info(f'Categorias: erro ao classificar — {e}')


def migrar_itens_para_epi():
    """Migra automaticamente itens com categoria='epi' do estoque para ItemEPI.
    Executa apenas se ItemEPI estiver vazia mas houver itens EPI no estoque.
    Usa o nome+almoxarifado como chave para evitar duplicatas.
    """
    try:
        total_epi_items = Item.query.filter_by(categoria='epi', ativo=True).count()
        if total_epi_items == 0:
            return
        total_item_epi = ItemEPI.query.count()

        itens_epi = Item.query.filter_by(categoria='epi', ativo=True).order_by(Item.nome).all()
        migrados = 0
        for it in itens_epi:
            # Verifica se já existe um ItemEPI com o mesmo nome e almoxarifado
            existe = ItemEPI.query.filter_by(
                nome=it.nome,
                almoxarifado_id=it.almoxarifado_id,
                ativo=True
            ).first()
            if existe:
                continue
            novo = ItemEPI(
                identificacao=it.ca or it.codigo or f'EPI-{it.id}',
                nome=it.nome,
                tamanho=None,
                almoxarifado_id=it.almoxarifado_id,
                quantidade=max(1, int(it.quantidade)) if it.quantidade > 0 else 1,
                status='disponivel',
                observacao=f'Migrado do estoque (cod: {it.codigo})',
                ativo=True,
                data_cadastro=agora()
            )
            db.session.add(novo)
            migrados += 1

        if migrados > 0:
            db.session.commit()
            logger.info(f'MIGRAÇÃO EPI: {migrados} item(s) migrados do estoque para ItemEPI.')
        else:
            logger.info(f'MIGRAÇÃO EPI: sem itens novos para migrar ({total_item_epi} ItemEPI já existem).')
    except Exception as e:
        logger.error(f'MIGRAÇÃO EPI: erro — {e}')
        db.session.rollback()


def seed_ferramentas_estrutura_auto():
    """Cadastra automaticamente as ferramentas da Estrutura no startup — ignora duplicatas."""
    try:
        alm = Almoxarifado.query.filter(Almoxarifado.nome.ilike('%estrutura%')).first()
        if not alm:
            return
        ferramentas_lista = [
            ("INGFH007",   "PISTOLA DE FIXACAO A BATERIA BX 3-L A22MA HILTI GF",  "HILTI"),
            ("IN450348",   "MARTELO SDS PLUS C/ PUNHO",                             ""),
            ("INMRM158",   "MARTELO ROMPEDOR MMR1700 45J 12KG MONO 220V 60HZ",     "MENEGOTTI"),
            ("IN580121",   "MISTURADOR DE ARGAMASSA MAV1600 220V",                  "MENEGOTTI"),
            ("IN580039",   "MISTURADOR DE ARGAMASSA MAV1600 220V",                  "MENEGOTTI"),
            ("INGFH003",   "PISTOLA DE FIXACAO A BATERIA BX 3-L A22MA HILTI GF",   "HILTI"),
            ("INEAG119",   'ESMERILHADEIRA ANGULAR 5" C/ ACESSORIOS',               "MAKITA"),
            ("IN270013",   "FURADEIRA 3/4 FUR3/4P",                                 ""),
            ("IN270028",   "FURADEIRA 3/4 FUR3/4P",                                 ""),
            ("IN270004",   "FURADEIRA 3/4 FUR3/4P",                                 ""),
            ("IN270005",   "FURADEIRA 3/4 FUR3/4P",                                 ""),
            ("INFPB017",   "FURADEIRA E PARAFUSADEIRA A BATERIA MFI-20 127/220V",   "MENEGOTTI"),
            ("IN340056",   "LAVA JATO HD 585 PROFISSIONAL MODELO 585",              ""),
            ("INMSV108",   "MARTELETE PERF/ROMP MPV1500 5,5J 220V",                 "VONDER"),
            ("INHT930012", "KIT ASPIRADOR UNIV. HILTI / POLIDORA DE BETAO HILTI",   "HILTI"),
            ("INMSV261",   "MARTELETE PERF/ROMP MPV1500 5,5J 220V",                 "VONDER"),
            ("INMSV269",   "MARTELETE PERF/ROMP MPV1500 5,5J 220V",                 "VONDER"),
            ("IN240093",   'ESMERILHADEIRA ANGULAR 7" 220V',                         ""),
            ("INMSV256",   "MARTELETE PERF/ROMP MPV1500 5,5J 220V",                 "VONDER"),
            ("INMSV257",   "MARTELETE PERF/ROMP MPV1500 5,5J 220V",                 "VONDER"),
            ("INMSU042",   "MISTURADOR ELETRICO MEL1600 MONO 220V 60HZ 1600W",      "MENEGOTTI"),
            ("INSER830025",'SERRA CIRCULAR 7"',                                      ""),
            ("INEAG243",   'ESMERILHADEIRA ANGULAR 5" C/ ACESSORIOS',               "MAKITA"),
            ("INEAG233",   'ESMERILHADEIRA ANGULAR 5" C/ ACESSORIOS',               "MAKITA"),
            ("INEAG236",   'ESMERILHADEIRA ANGULAR 5" C/ ACESSORIOS',               "MAKITA"),
            ("INMRM149",   "MARTELO ROMPEDOR MMR1700 45J 12KG MONO 220V 60HZ",     "MENEGOTTI"),
            ("INMSV069",   "MARTELETE PERF/ROMP MPV1500 5,5J 220V",                 "VONDER"),
            ("INMSV156",   "MARTELETE PERF/ROMP MPV1500 5,5J 220V",                 "VONDER"),
            ("IN620015",   "KIT NIVELADOR A LASER HILTI SKR200",                    "HILTI"),
            ("INMSV067",   "MARTELETE PERF/ROMP MPV1500 5,5J 220V",                 "VONDER"),
            ("INMSV001",   "MARTELETE PERF/ROMP MPV1500 5,5J 220V",                 "VONDER"),
            ("INMSV099",   "MARTELETE PERF/ROMP MPV1500 5,5J 220V",                 "VONDER"),
            ("INMSV273",   "MARTELETE PERF/ROMP MPV1500 5,5J 220V",                 "VONDER"),
            ("INHT930015", "KIT ASPIRADOR UNIV. HILTI / POLIDORA DE BETAO HILTI",   "HILTI"),
            ("INMSV024",   "MARTELETE PERF/ROMP MPV1500 5,5J 220V",                 "VONDER"),
            ("INMSV113",   "MARTELETE PERF/ROMP MPV1500 5,5J 220V",                 "VONDER"),
            ("INMSV131",   "MARTELETE PERF/ROMP MPV1500 5,5J 220V",                 "VONDER"),
            ("INMSV278",   "MARTELETE PERF/ROMP MPV1500 5,5J 220V",                 "VONDER"),
            ("INMSV148",   "MARTELETE PERF/ROMP MPV1500 5,5J 220V",                 "VONDER"),
            ("INHT930010", "KIT ASPIRADOR UNIV. HILTI / POLIDORA DE BETAO HILTI",   "HILTI"),
            ("INHT930011", "KIT ASPIRADOR UNIV. HILTI / POLIDORA DE BETAO HILTI",   "HILTI"),
            ("INMSV090",   "MARTELETE PERF/ROMP MPV1500 5,5J 220V",                 "VONDER"),
            ("INMSV281",   "MARTELETE PERF/ROMP MPV1500 5,5J 220V",                 "VONDER"),
            ("INMSV260",   "MARTELETE PERF/ROMP MPV1500 5,5J 220V",                 "VONDER"),
            ("IN850474",   "SERRA MARMORE C/ CHAVE",                                 ""),
        ]
        inseridas = 0
        for idf, nome, empresa in ferramentas_lista:
            if not Ferramenta.query.filter_by(identificacao=idf, ativo=True).first():
                db.session.add(Ferramenta(
                    identificacao=idf,
                    nome=nome,
                    empresa=empresa or None,
                    almoxarifado_id=alm.id,
                    status='disponivel'
                ))
                inseridas += 1
        if inseridas:
            db.session.commit()
            logger.info(f'SEED FERRAMENTAS ESTRUTURA: {inseridas} ferramentas cadastradas.')
    except Exception as e:
        logger.error(f'SEED FERRAMENTAS ESTRUTURA: erro — {e}')
        db.session.rollback()


def inicializar_banco():
    """Roda migrações, cria tabelas e seed — executado uma única vez."""
    try:
        db.create_all()
        run_migrations()
        seed_data()
        classificar_categorias_itens()
        migrar_itens_para_epi()
        seed_ferramentas_estrutura_auto()
    except Exception as e:
        logger.error(f'Inicialização do banco: {e}')

# Inicializa na primeira requisição para não bloquear o boot do gunicorn
_banco_inicializado = False

@app.before_request
def init_on_first_request():
    global _banco_inicializado
    if not _banco_inicializado:
        _banco_inicializado = True
        try:
            inicializar_banco()
        except Exception as e:
            logger.error(f'Erro na inicialização: {e}')

if __name__ == '__main__':
    port = int(os.environ.get('PORT', 5000))
    app.run(debug=False, host='0.0.0.0', port=port)