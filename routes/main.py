
from flask import Blueprint, render_template, request, redirect, url_for, jsonify, flash, send_file, session
from markupsafe import Markup, escape
from datetime import datetime, date, timedelta
import io, os, json, re, logging
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from extensions import db
from models import (agora, Almoxarifado, Item, Movimentacao, Requisicao,
    RequisicaoMestre, RequisicaoMestreItem, Usuario, Colaborador,
    Ferramenta, HistoricoFerramenta, ItemEPI, HistoricoEPI,
    AcessoExtra, PermissaoExtra)
from core import (login_required, admin_required, almoxarife_required,
    usuario_atual, flash_html, usuario_tem_acesso_almoxarifado,
    usuario_tem_acesso_item, PERMISSOES_DISPONIVEIS,
    _check_rate_limit, _register_attempt, _clear_attempts, _check_api_rate)

logger = logging.getLogger(__name__)
from utils import calcular_ruptura
main_bp = Blueprint('main_bp', __name__)

@main_bp.route('/')
@login_required
def index():
    u = usuario_atual()
    # Mestre e técnico de segurança só acessam a tela de requisições
    # Engenheiro com pode_requisitar também é redirecionado para lá
    if u.perfil in ('mestre', 'tecnico_seguranca'):
        if u.perfil == 'tecnico_seguranca':
            return redirect(url_for('main_bp.tecnico_dashboard'))
        return redirect(url_for('requisicoes_bp.mestre_requisicoes'))
    if u.perfil == 'colaborador' and (
        u.pode_requisitar or
        any(p.permissao == 'fazer_requisicao' for p in u.permissoes_extras)
    ):
        return redirect(url_for('requisicoes_bp.mestre_requisicoes'))
    if u.perfil == 'admin':
        almoxarifados = Almoxarifado.query.all()
        alertas = Item.query.filter(Item.quantidade <= Item.estoque_minimo, Item.ativo == True).all()
    elif u.perfil == 'analista':
        # Analista vê apenas o almoxarifado vinculado a ele
        if u.almoxarifado_id:
            alm_ref = db.session.get(Almoxarifado, u.almoxarifado_id)
            almoxarifados = [alm_ref] if alm_ref else []
        else:
            almoxarifados = []
        ids_analista = {a.id for a in almoxarifados}
        alertas = Item.query.filter(
            Item.quantidade <= Item.estoque_minimo,
            Item.almoxarifado_id.in_(ids_analista),
            Item.ativo == True
        ).all()
    else:
        ids = u.almoxarifados_permitidos()
        almoxarifados = Almoxarifado.query.filter(Almoxarifado.id.in_(ids)).all() if ids else []
        alertas = Item.query.filter(
            Item.quantidade <= Item.estoque_minimo,
            Item.almoxarifado_id.in_(ids),
            Item.ativo == True
        ).all() if ids else []
    stats = {
        'total_almoxarifados': len(almoxarifados),
        'total_itens': sum(len(a.itens) for a in almoxarifados),
        'itens_alerta': len([a for a in alertas if a.quantidade > 0]),
        'itens_criticos': len([a for a in alertas if a.quantidade <= 0]),
    }
    # Previsão de ruptura no dashboard — itens OK mas em risco nos próximos 15 dias
    ids_alm = {a.id for a in almoxarifados}
    todos_ativos_dash = Item.query.filter(
        Item.ativo == True,
        Item.almoxarifado_id.in_(ids_alm)
    ).all() if ids_alm else []
    ruptura_dash = calcular_ruptura(todos_ativos_dash, limite_dias=15)

    return render_template('index.html', almoxarifados=almoxarifados, alertas=alertas,
                           stats=stats, ruptura=ruptura_dash)

@main_bp.route('/tecnico/dashboard')
@login_required
def tecnico_dashboard():
    from models import FichaEPI, ItemFichaEPI, RequisicaoMestre
    from datetime import date
    u = usuario_atual()
    if u.perfil != 'tecnico_seguranca':
        return redirect(url_for('main_bp.index'))

    ids = u.almoxarifados_permitidos() or set()

    # Requisições do técnico
    reqs = RequisicaoMestre.query.filter_by(mestre_id=u.id).order_by(
        RequisicaoMestre.data_criacao.desc()).all()
    total_reqs   = len(reqs)
    pendentes    = sum(1 for r in reqs if r.status == 'pendente')
    aprovadas    = sum(1 for r in reqs if r.status in ('aprovada', 'parcial'))
    entregues    = sum(1 for r in reqs if r.status == 'entregue')
    recusadas    = sum(1 for r in reqs if r.status == 'recusada')
    ultimas_reqs = reqs[:8]

    # EPI — fichas ativas nos almoxarifados do técnico
    q_fichas = FichaEPI.query
    if ids:
        q_fichas = q_fichas.filter(FichaEPI.almoxarifado_id.in_(ids))
    fichas_ativas = q_fichas.filter_by(status='ativa').count()

    # Devoluções de EPI em aberto
    q_dev = ItemFichaEPI.query.join(FichaEPI).filter(
        ItemFichaEPI.data_entrega != None,
        ItemFichaEPI.data_devolucao == None
    )
    if ids:
        q_dev = q_dev.filter(FichaEPI.almoxarifado_id.in_(ids))
    devolucoes_abertas = q_dev.count()

    # Entregas de EPI no mês atual
    ini_mes = date.today().replace(day=1)
    q_ent = ItemFichaEPI.query.join(FichaEPI).filter(
        ItemFichaEPI.data_entrega >= ini_mes
    )
    if ids:
        q_ent = q_ent.filter(FichaEPI.almoxarifado_id.in_(ids))
    entregas_mes = q_ent.count()

    return render_template('tecnico_dashboard.html',
        total_reqs=total_reqs, pendentes=pendentes,
        aprovadas=aprovadas, entregues=entregues, recusadas=recusadas,
        ultimas_reqs=ultimas_reqs,
        fichas_ativas=fichas_ativas,
        devolucoes_abertas=devolucoes_abertas,
        entregas_mes=entregas_mes,
        now_date=date.today().strftime('%d/%m/%Y'),
    )


@main_bp.route('/almoxarifado/<int:id>')
@login_required
def almoxarifado(id):
    u = usuario_atual()
    # Mestre e técnico de segurança não acessam almoxarifado diretamente
    # Engenheiro com pode_requisitar também não — usa tela de requisições
    if u.perfil in ('mestre', 'tecnico_seguranca'):
        flash('Acesso restrito. Use a tela de requisições.', 'warning')
        return redirect(url_for('requisicoes_bp.mestre_requisicoes'))
    if u.perfil == 'colaborador' and (
        u.pode_requisitar or
        any(p.permissao == 'fazer_requisicao' for p in u.permissoes_extras)
    ):
        flash('Acesso restrito. Use a tela de requisições.', 'warning')
        return redirect(url_for('requisicoes_bp.mestre_requisicoes'))
    alm = Almoxarifado.query.get_or_404(id)
    if u.perfil not in ('admin', 'analista') and id not in u.almoxarifados_permitidos():
        flash('Acesso negado.', 'danger')
        return redirect(url_for('main_bp.index'))
    # Analista só acessa almoxarifados da sua cidade
    if u.perfil == 'analista' and u.almoxarifado_id:
        alm_ref = db.session.get(Almoxarifado, u.almoxarifado_id)
        if alm_ref and alm_ref.cidade:
            alm_alvo = Almoxarifado.query.get_or_404(id)
            if (alm_alvo.cidade or '').lower().strip() != alm_ref.cidade.lower().strip():
                flash('Acesso negado.', 'danger')
                return redirect(url_for('main_bp.index'))
    # Mostrar todos os itens (ativos e desativados) para permitir reativação
    itens = Item.query.filter_by(almoxarifado_id=id).order_by(Item.ativo.desc(), Item.nome).all()
    todos_almoxarifados = Almoxarifado.query.order_by(Almoxarifado.nome).all() if u.perfil == 'admin' else []
    return render_template('almoxarifado.html', almoxarifado=alm, itens=itens,
                           todos_almoxarifados=todos_almoxarifados)

@main_bp.route('/item/<int:id>')
@login_required
def item(id):
    it = Item.query.get_or_404(id)
    u = usuario_atual()
    if u.perfil != 'admin' and it.almoxarifado_id not in u.almoxarifados_permitidos():
        flash('Acesso negado.', 'danger')
        return redirect(url_for('main_bp.index'))
    movs = Movimentacao.query.filter_by(item_id=id).order_by(Movimentacao.data.desc()).limit(50).all()
    return render_template('item.html', item=it, movimentacoes=movs)

# ── CRUD ALMOXARIFADO ────────────────────────────────────────────────────────

@main_bp.route('/almoxarifado/novo', methods=['GET', 'POST'])
@admin_required
def novo_almoxarifado():
    if request.method == 'POST':
        alm = Almoxarifado(nome=request.form['nome'], descricao=request.form.get('descricao', ''))
        alm.obra = request.form.get('obra', '').strip() or None
        alm.cidade = request.form.get('cidade', '').strip() or None
        db.session.add(alm)
        db.session.commit()
        flash(f'Almoxarifado "{alm.nome}" criado!', 'success')
        return redirect(url_for('main_bp.index'))
    return render_template('form_almoxarifado.html', almoxarifado=None)

@main_bp.route('/almoxarifado/<int:id>/editar', methods=['GET', 'POST'])
@login_required
def editar_almoxarifado(id):
    alm = Almoxarifado.query.get_or_404(id)
    u = usuario_atual()
    # Apenas admin ou almoxarife do próprio almoxarifado pode editar
    if u.perfil != 'admin' and id not in u.almoxarifados_permitidos():
        flash('Acesso negado.', 'danger')
        return redirect(url_for('main_bp.index'))
    if request.method == 'POST':
        alm.nome = request.form['nome']
        alm.descricao = request.form.get('descricao', '')
        alm.obra = request.form.get('obra', '').strip() or None
        alm.cidade = request.form.get('cidade', '').strip() or None
        db.session.commit()
        flash('Almoxarifado atualizado!', 'success')
        return redirect(url_for('main_bp.index'))
    return render_template('form_almoxarifado.html', almoxarifado=alm)

@main_bp.route('/almoxarifado/<int:id>/deletar', methods=['POST'])
@admin_required
def deletar_almoxarifado(id):
    alm = Almoxarifado.query.get_or_404(id)
    try:
        # Desvincular usuários que têm este almoxarifado como principal
        Usuario.query.filter_by(almoxarifado_id=id).update({'almoxarifado_id': None})

        # Desvincular acessos extras para este almoxarifado
        AcessoExtra.query.filter_by(almoxarifado_id=id).delete()

        # Deletar requisições do mestre vinculadas a este almoxarifado
        # (RequisicaoMestreItem é deletado via cascade no relacionamento)
        for req in RequisicaoMestre.query.filter_by(almoxarifado_id=id).all():
            db.session.delete(req)

        # O cascade 'all, delete-orphan' já cuida de itens (Item model)
        # Ferramentas e EPIs também têm relationship com backref, deletar explicitamente
        Ferramenta.query.filter_by(almoxarifado_id=id).delete()
        ItemEPI.query.filter_by(almoxarifado_id=id).delete()

        db.session.delete(alm)
        db.session.commit()
        flash('Almoxarifado removido com sucesso!', 'warning')
    except Exception as e:
        db.session.rollback()
        flash(f'Erro ao remover almoxarifado: {str(e)}', 'danger')
    return redirect(url_for('main_bp.index'))

# ── CRUD ITEM ────────────────────────────────────────────────────────────────

@main_bp.route('/item/novo', methods=['GET', 'POST'])
@almoxarife_required
def novo_item():
    u = usuario_atual()
    # Almoxarife só vê seus próprios almoxarifados no select
    if u.perfil == 'admin':
        almoxarifados = Almoxarifado.query.all()
    else:
        ids = u.almoxarifados_permitidos()
        almoxarifados = Almoxarifado.query.filter(Almoxarifado.id.in_(ids)).all() if ids else []
    if request.method == 'POST':
        codigo = request.form['codigo'].strip()
        # Verificar duplicata de código DENTRO do mesmo almoxarifado
        alm_id = int(request.form['almoxarifado_id'])
        existente = Item.query.filter(
            Item.codigo.ilike(codigo),
            Item.almoxarifado_id == alm_id
        ).first()
        if existente:
            flash(f'⚠️ Código "{codigo}" já está em uso pelo item "{existente.nome}" neste almoxarifado. Use um código diferente.', 'danger')
            return render_template('form_item.html', item=None, almoxarifados=almoxarifados)
        try:
            it = Item(
                nome=request.form['nome'],
                codigo=codigo,
                unidade=request.form['unidade'],
                quantidade=float(request.form.get('quantidade', 0)),
                estoque_minimo=float(request.form.get('estoque_minimo', 0)),
                almoxarifado_id=int(request.form['almoxarifado_id']),
                categoria=request.form.get('categoria', 'geral'),
                ca=request.form.get('ca', '').strip() or None
            )
            db.session.add(it)
            db.session.commit()
            flash(f'Item "{it.nome}" cadastrado!', 'success')
            return redirect(url_for('main_bp.almoxarifado', id=it.almoxarifado_id))
        except Exception as e:
            db.session.rollback()
            flash(f'Erro ao cadastrar item: código já existe ou dados inválidos.', 'danger')
            return render_template('form_item.html', item=None, almoxarifados=almoxarifados)
    return render_template('form_item.html', item=None, almoxarifados=almoxarifados)

@main_bp.route('/item/<int:id>/editar', methods=['GET', 'POST'])
@login_required
def editar_item(id):
    it = Item.query.get_or_404(id)
    u = usuario_atual()
    if u.perfil in ('mestre', 'tecnico_seguranca', 'analista') or (u.perfil != 'admin' and it.almoxarifado_id not in u.almoxarifados_permitidos()):
        flash('Acesso negado.', 'danger')
        return redirect(url_for('main_bp.item', id=id))
    # Almoxarife só vê seus próprios almoxarifados no select
    if u.perfil == 'admin':
        almoxarifados = Almoxarifado.query.all()
    else:
        ids = u.almoxarifados_permitidos()
        almoxarifados = Almoxarifado.query.filter(Almoxarifado.id.in_(ids)).all() if ids else []
    if request.method == 'POST':
        it.nome = request.form['nome']
        it.codigo = request.form['codigo']
        it.unidade = request.form['unidade']
        it.estoque_minimo = float(request.form.get('estoque_minimo', 0))
        it.almoxarifado_id = int(request.form['almoxarifado_id'])
        it.categoria = request.form.get('categoria', 'geral')
        it.ca = request.form.get('ca', '').strip() or None
        # Atualizar quantidade se informada (admin pode corrigir o valor)
        qtd_str = request.form.get('quantidade')
        if qtd_str is not None and qtd_str != '':
            try:
                nova_qtd = float(qtd_str)
                if nova_qtd != it.quantidade:
                    # Registrar ajuste como movimentação
                    diff = nova_qtd - it.quantidade
                    tipo = 'entrada' if diff > 0 else 'saida'
                    db.session.add(Movimentacao(
                        tipo=tipo,
                        quantidade=abs(diff),
                        responsavel=u.nome if u else 'Sistema',
                        observacao=f'Ajuste manual: {it.quantidade} → {nova_qtd} {it.unidade}',
                        item_id=it.id
                    ))
                    it.quantidade = nova_qtd
            except ValueError:
                pass
        db.session.commit()
        flash('Item atualizado!', 'success')
        return redirect(url_for('main_bp.almoxarifado', id=it.almoxarifado_id))
    return render_template('form_item.html', item=it, almoxarifados=almoxarifados)

@main_bp.route('/item/<int:id>/deletar', methods=['POST'])
@login_required
def deletar_item(id):
    it = Item.query.get_or_404(id)
    u = usuario_atual()
    # Admin pode deletar qualquer item; almoxarife pode deletar itens do próprio almoxarifado
    if u.perfil == 'admin':
        pass  # permitido
    elif u.perfil == 'almoxarife' and it.almoxarifado_id in u.almoxarifados_permitidos():
        pass  # almoxarife do almoxarifado pode deletar
    else:
        flash('Sem permissão para deletar este item.', 'danger')
        return redirect(url_for('main_bp.item', id=id))
    alm_id = it.almoxarifado_id
    # Soft delete — preserva histórico e evita erro de constraint com requisições/movimentações
    it.ativo = False
    db.session.commit()
    flash('Item removido!', 'warning')
    return redirect(url_for('main_bp.almoxarifado', id=alm_id))

@main_bp.route('/almoxarifado/<int:alm_id>/deletar-lote', methods=['POST'])
@login_required
def deletar_itens_lote(alm_id):
    alm = Almoxarifado.query.get_or_404(alm_id)
    u = usuario_atual()
    # Verificar permissão: admin ou almoxarife do próprio almoxarifado
    if u.perfil == 'admin':
        pass
    elif u.perfil == 'almoxarife' and alm_id in u.almoxarifados_permitidos():
        pass
    else:
        flash('Sem permissão para excluir itens em lote.', 'danger')
        return redirect(url_for('main_bp.almoxarifado', id=alm_id))

    ids_str = request.form.getlist('item_ids')
    if not ids_str:
        flash('Nenhum item selecionado.', 'warning')
        return redirect(url_for('main_bp.almoxarifado', id=alm_id))

    try:
        ids = [int(i) for i in ids_str]
    except ValueError:
        flash('Seleção inválida.', 'danger')
        return redirect(url_for('main_bp.almoxarifado', id=alm_id))

    # Soft delete apenas nos itens que pertencem a este almoxarifado
    itens = Item.query.filter(
        Item.id.in_(ids),
        Item.almoxarifado_id == alm_id,
        Item.ativo == True
    ).all()

    if not itens:
        flash('Nenhum item ativo encontrado na seleção.', 'warning')
        return redirect(url_for('main_bp.almoxarifado', id=alm_id))

    for it in itens:
        it.ativo = False
    db.session.commit()

    flash(f'{len(itens)} item(ns) removido(s) com sucesso.', 'warning')
    return redirect(url_for('main_bp.almoxarifado', id=alm_id))


@main_bp.route('/almoxarifado/<int:alm_id>/transferir-lote', methods=['POST'])
@admin_required
def transferir_itens_lote(alm_id):
    """Admin transfere itens para outro almoxarifado OU exclui permanentemente."""
    ids_str = request.form.getlist('item_ids')
    acao    = request.form.get('acao', '')          # 'transferir' | 'excluir'
    dest_id = request.form.get('destino_id', type=int)

    if not ids_str:
        flash('Nenhum item selecionado.', 'warning')
        return redirect(url_for('main_bp.almoxarifado', id=alm_id))

    try:
        ids = [int(i) for i in ids_str]
    except ValueError:
        flash('Seleção inválida.', 'danger')
        return redirect(url_for('main_bp.almoxarifado', id=alm_id))

    itens = Item.query.filter(
        Item.id.in_(ids),
        Item.almoxarifado_id == alm_id
    ).all()

    if not itens:
        flash('Nenhum item encontrado na seleção.', 'warning')
        return redirect(url_for('main_bp.almoxarifado', id=alm_id))

    if acao == 'transferir':
        if not dest_id:
            flash('Selecione o almoxarifado de destino.', 'danger')
            return redirect(url_for('main_bp.almoxarifado', id=alm_id))
        dest = Almoxarifado.query.get_or_404(dest_id)
        if dest_id == alm_id:
            flash('O destino deve ser diferente do almoxarifado atual.', 'danger')
            return redirect(url_for('main_bp.almoxarifado', id=alm_id))

        u = usuario_atual()
        transferidos = 0
        for it in itens:
            # Verifica se já existe item com mesmo código no destino
            existente = Item.query.filter_by(
                codigo=it.codigo, almoxarifado_id=dest_id
            ).first()
            if existente:
                # Soma a quantidade ao item existente
                existente.quantidade = round(existente.quantidade + it.quantidade, 4)
                db.session.add(Movimentacao(
                    tipo='entrada',
                    quantidade=it.quantidade,
                    responsavel=u.nome,
                    observacao=f'Transferido de {it.almoxarifado.nome} (item id={it.id})',
                    item_id=existente.id
                ))
                # Remove o item original com cascade nas movimentações
                Movimentacao.query.filter_by(item_id=it.id).delete()
                db.session.delete(it)
            else:
                # Move o item para o destino
                it.almoxarifado_id = dest_id
                db.session.add(Movimentacao(
                    tipo='entrada',
                    quantidade=it.quantidade,
                    responsavel=u.nome,
                    observacao=f'Transferido de {it.almoxarifado.nome if it.almoxarifado else alm_id}',
                    item_id=it.id
                ))
            transferidos += 1

        db.session.commit()
        flash(f'✅ {transferidos} item(ns) transferido(s) para "{dest.nome}"!', 'success')

    elif acao == 'excluir':
        u = usuario_atual()
        excluidos = 0
        for it in itens:
            # Hard delete — remove movimentações primeiro (cascade pode não funcionar em todos os casos)
            Movimentacao.query.filter_by(item_id=it.id).delete()
            # Remove itens de requisições vinculadas
            from models import RequisicaoMestreItem, KitItem
            RequisicaoMestreItem.query.filter_by(item_id=it.id).delete()
            KitItem.query.filter_by(item_id=it.id).delete()
            db.session.delete(it)
            excluidos += 1

        db.session.commit()
        flash(f'🗑️ {excluidos} item(ns) excluído(s) permanentemente!', 'warning')

    else:
        flash('Ação inválida.', 'danger')

    return redirect(url_for('main_bp.almoxarifado', id=alm_id))

# ── MOVIMENTAÇÃO EM LOTE ─────────────────────────────────────────────────────

@main_bp.route('/movimentacao/lote', methods=['GET', 'POST'])
@login_required
def movimentacao_lote():
    u = usuario_atual()
    # Analista só pode visualizar — sem movimentação
    if u.perfil == 'analista':
        flash('Analistas não têm permissão para registrar movimentações.', 'danger')
        return redirect(url_for('main_bp.index'))
    almoxarifados = Almoxarifado.query.all() if u.perfil == 'admin' else \
        Almoxarifado.query.filter(Almoxarifado.id.in_(u.almoxarifados_permitidos())).all()

    # Histórico das últimas 20 movimentações
    if u.perfil == 'admin':
        historico = Movimentacao.query.order_by(Movimentacao.data.desc()).limit(20).all()
        requisicoes_hist = Requisicao.query.order_by(Requisicao.data_retirada.desc()).limit(20).all()
    else:
        ids = u.almoxarifados_permitidos()
        historico = (Movimentacao.query.join(Item)
                     .filter(Item.almoxarifado_id.in_(ids))
                     .order_by(Movimentacao.data.desc()).limit(20).all())
        requisicoes_hist = (Requisicao.query.join(Item)
                            .filter(Item.almoxarifado_id.in_(ids))
                            .order_by(Requisicao.data_retirada.desc()).limit(20).all())

    # Montar JSON com itens por almoxarifado
    itens_json = {}
    for alm in almoxarifados:
        itens_json[str(alm.id)] = [
            {'id': it.id, 'nome': it.nome, 'quantidade': it.quantidade,
             'unidade': it.unidade, 'categoria': it.categoria or 'geral',
             'ca': it.ca or ''}
            for it in alm.itens
        ]

    if request.method == 'POST':
        alm_id      = int(request.form['almoxarifado_id'])
        if u.perfil != 'admin' and alm_id not in u.almoxarifados_permitidos():
            flash('Acesso negado.', 'danger')
            return redirect(url_for('main_bp.movimentacao_lote'))

        tipo        = request.form['tipo']
        responsavel = request.form.get('responsavel', '')
        observacao  = request.form.get('observacao', '')

        erros = []
        movs  = []

        # Coletar índices sem depender de total_linhas
        indices = set()
        for key in request.form.keys():
            if key.startswith('item_id_'):
                try:
                    indices.add(int(key.split('_')[-1]))
                except ValueError:
                    pass

        for i in sorted(indices):
            item_id    = request.form.get(f'item_id_{i}')
            qtd_str    = request.form.get(f'quantidade_{i}')
            colab      = request.form.get(f'colaborador_{i}', '').strip()
            resp_linha = request.form.get(f'responsavel_{i}', '').strip() or responsavel
            ca_linha   = request.form.get(f'ca_{i}', '').strip()

            if not item_id or not qtd_str:
                continue

            it = db.session.get(Item, item_id)
            try:
                qtd = float(qtd_str)
            except ValueError:
                continue

            if not it or qtd <= 0:
                continue

            if tipo == 'saida' and qtd > it.quantidade:
                erros.append(f'"{it.nome}": estoque insuficiente ({it.quantidade} {it.unidade})')
                continue

            # Devolução = entrada no estoque com observação especial
            tipo_real = 'entrada' if tipo in ('devolucao_epi', 'devolucao_ferramenta') else tipo
            it.quantidade = round(it.quantidade + qtd if tipo_real == 'entrada' else it.quantidade - qtd, 4)

            if tipo == 'saida' and colab:
                obs_linha = f'liberado P/ {colab}'
                if observacao:
                    obs_linha += f' | {observacao}'
            elif tipo == 'devolucao_epi':
                obs_linha = f'Devolução EPI — {colab}' if colab else 'Devolução EPI'
                if observacao:
                    obs_linha += f' | {observacao}'
            elif tipo == 'devolucao_ferramenta':
                obs_linha = f'Devolução Ferramenta — {colab}' if colab else 'Devolução Ferramenta'
                if observacao:
                    obs_linha += f' | {observacao}'
            else:
                obs_linha = observacao

            movs.append(Movimentacao(
                tipo=tipo_real, quantidade=qtd,
                responsavel=resp_linha,
                observacao=obs_linha,
                item_id=it.id
            ))
            # Registrar automaticamente na FichaEPI se for saída de EPI com colaborador
            if tipo == 'saida' and it.categoria == 'epi' and colab:
                try:
                    from routes.epi_modulo import registrar_epi_na_ficha
                    registrar_epi_na_ficha(
                        colaborador=colab,
                        almoxarifado_id=it.almoxarifado_id,
                        descricao=it.nome,
                        ca=ca_linha or it.ca,
                        quantidade=qtd,
                        tamanho=None,
                        registrado_por=resp_linha or (u.nome if u else 'Sistema')
                    )
                except Exception as _e:
                    logger.error(f'registrar_epi_na_ficha (lote): {_e}')

        if movs:
            db.session.add_all(movs)
            db.session.commit()
            tipo_label = '📥 Entrada' if request.form['tipo'] == 'entrada' else '📤 Saída'
            alm = db.session.get(Almoxarifado, alm_id)

            # Se é saída de EPI via AJAX, retorna JSON com IDs para abrir câmera
            if (request.headers.get('X-Requested-With') == 'XMLHttpRequest'
                    and request.form.get('tipo') == 'saida'):
                epi_movs = [m for m in movs if db.session.get(Item, m.item_id).categoria == 'epi']
                return jsonify({
                    'ok': True,
                    'msg': f'{len(movs)} item(ns) registrado(s).',
                    'epi_mov_ids': [m.id for m in epi_movs],
                    'redirect': url_for('main_bp.movimentacao_lote')
                })

            flash_html(
                f'<strong>{escape(tipo_label)} registrada!</strong> '
                f'{len(movs)} item(ns) movimentado(s) em <strong>{escape(alm.nome if alm else "")}</strong>. '
                f'<a href="/almoxarifado/{alm_id}" class="alert-link">Ver Almoxarifado</a>',
                'success'
            )
        elif not erros:
            flash('Adicione pelo menos um item antes de confirmar.', 'warning')

        for e in erros:
            flash_html(
                f'<strong>Estoque insuficiente:</strong> {escape(e)} '
                f'<a href="/almoxarifado/{alm_id}" class="alert-link">Consultar Estoque</a>',
                'danger'
            )

        return redirect(url_for('main_bp.movimentacao_lote'))

    import json
    return render_template('movimentacao_lote.html',
                           almoxarifados=almoxarifados,
                           itens_json=json.dumps(itens_json),
                           historico=historico,
                           requisicoes=requisicoes_hist)

@main_bp.route('/item/<int:id>/movimentar', methods=['POST'])
@login_required
def movimentar(id):
    it = Item.query.get_or_404(id)
    u = usuario_atual()
    # Analista não pode fazer movimentação de forma alguma
    if u.perfil == 'analista':
        flash('Analistas não têm permissão para registrar movimentações.', 'danger')
        return redirect(url_for('main_bp.item', id=id))
    if u.perfil != 'admin' and it.almoxarifado_id not in u.almoxarifados_permitidos():
        flash('Acesso negado.', 'danger')
        return redirect(url_for('main_bp.item', id=id))

    tipo = request.form['tipo']
    qtd = float(request.form['quantidade'])
    responsavel = request.form.get('responsavel', '').strip()
    observacao = request.form.get('observacao', '').strip()

    if tipo == 'saida' and qtd > it.quantidade:
        flash('Quantidade insuficiente em estoque!', 'danger')
        return redirect(url_for('main_bp.item', id=id))

    # Na saída, a observação já vem montada pelo JS como "liberado P/ Nome | req X"
    obs_final = observacao

    it.quantidade = round(it.quantidade + qtd if tipo == 'entrada' else it.quantidade - qtd, 4)
    mov = Movimentacao(
        tipo=tipo, quantidade=qtd,
        responsavel=responsavel,
        observacao=obs_final,
        item_id=id
    )
    db.session.add(mov)
    db.session.commit()
    flash(f'{"Entrada" if tipo == "entrada" else "Saída"} de {qtd} {it.unidade} registrada!', 'success')
    return redirect(url_for('main_bp.item', id=id))

# ── REQUISIÇÕES ──────────────────────────────────────────────────────────────

@main_bp.route('/requisicoes')
@login_required
def requisicoes():
    u = usuario_atual()
    colaborador  = request.args.get('colaborador', '')
    status       = request.args.get('status', '')
    data_ini     = request.args.get('data_ini', '')
    data_fim     = request.args.get('data_fim', '')

    q = Requisicao.query
    if u.perfil != 'admin':
        ids = u.almoxarifados_permitidos()
        q = q.join(Item).filter(Item.almoxarifado_id.in_(ids)) if ids else q.filter(False)
    if colaborador:
        q = q.filter(Requisicao.colaborador.ilike(f'%{colaborador}%'))
    if status:
        q = q.filter(Requisicao.status == status)
    if data_ini:
        q = q.filter(Requisicao.data_retirada >= data_ini)
    if data_fim:
        q = q.filter(Requisicao.data_retirada <= data_fim + ' 23:59:59')

    reqs = q.order_by(Requisicao.data_retirada.desc()).all()

    return render_template('requisicoes.html',
        requisicoes=reqs,
        total=len(reqs),
        em_uso=sum(1 for r in reqs if r.status == 'aberta'),
        devolvidos=sum(1 for r in reqs if r.status == 'devolvida'),
        filtro_colaborador=colaborador,
        filtro_status=status,
        filtro_data_ini=data_ini,
        filtro_data_fim=data_fim
    )

@main_bp.route('/requisicoes/nova', methods=['GET', 'POST'])
@login_required
def requisicao_nova():
    u = usuario_atual()
    almoxarifados = Almoxarifado.query.all() if u.perfil == 'admin' else \
        Almoxarifado.query.filter(Almoxarifado.id.in_(u.almoxarifados_permitidos())).all()
    itens_json = {}
    for alm in almoxarifados:
        itens_json[str(alm.id)] = [
            {'id': it.id, 'nome': it.nome, 'quantidade': it.quantidade, 'unidade': it.unidade}
            for it in alm.itens
            if it.ativo and it.quantidade > 0
        ]

    if request.method == 'POST':
        colaborador = request.form.get('colaborador', '')
        observacao  = request.form.get('observacao', '')
        criados = 0

        # Coletar todos os item_id_N do formulário sem depender de total_linhas
        indices = set()
        for key in request.form.keys():
            if key.startswith('item_id_'):
                try:
                    indices.add(int(key.split('_')[-1]))
                except ValueError:
                    pass

        for i in sorted(indices):
            item_id = request.form.get(f'item_id_{i}')
            qtd_str = request.form.get(f'quantidade_{i}')
            if not item_id or not qtd_str:
                continue
            it  = db.session.get(Item, item_id)
            try:
                qtd = float(qtd_str)
            except ValueError:
                continue
            if not it or qtd <= 0:
                continue
            if u.perfil != 'admin' and it.almoxarifado_id not in u.almoxarifados_permitidos():
                continue
            if qtd > it.quantidade:
                flash_html(
                    f'<strong>Estoque insuficiente:</strong> "{escape(it.nome)}" tem apenas '
                    f'<strong>{it.quantidade} {escape(it.unidade)}</strong> disponível. '
                    f'<a href="/almoxarifado/{it.almoxarifado_id}" class="alert-link">Consultar Estoque</a>',
                    'danger'
                )
                continue

            it.quantidade -= qtd
            db.session.add(Requisicao(
                colaborador=colaborador,
                observacao=observacao,
                quantidade=qtd,
                item_id=it.id
            ))
            db.session.add(Movimentacao(
                tipo='saida', quantidade=qtd,
                responsavel=colaborador,
                observacao=f'Requisicao — {observacao}',
                item_id=it.id
            ))
            criados += 1

        if criados:
            db.session.commit()
            flash_html(
                f'<strong>✅ Requisição registrada!</strong> '
                f'{criados} item(ns) retirado(s) com sucesso para <strong>{escape(colaborador)}</strong>. '
                f'<a href="/requisicoes" class="alert-link">Ver Requisições</a>',
                'success'
            )
        elif not any(True for key in request.form.keys() if key.startswith('item_id_')):
            flash('Adicione pelo menos um item antes de registrar.', 'warning')
        return redirect(url_for('main_bp.requisicoes'))

    import json
    return render_template('requisicao_nova.html',
                           almoxarifados=almoxarifados,
                           itens_json=json.dumps(itens_json))

@main_bp.route('/requisicoes/<int:id>/devolver', methods=['POST'])
@login_required
def devolver_requisicao(id):
    req = Requisicao.query.get_or_404(id)
    u = usuario_atual()
    if u.perfil != 'admin' and req.item.almoxarifado_id not in u.almoxarifados_permitidos():
        flash('Acesso negado.', 'danger')
        return redirect(url_for('main_bp.requisicoes'))
    if req.status == 'aberta':
        req.status = 'devolvida'
        req.data_devolucao = agora()
        req.item.quantidade += req.quantidade
        db.session.add(Movimentacao(
            tipo='entrada', quantidade=req.quantidade,
            responsavel=req.colaborador,
            observacao=f'Devolução de requisição #{req.id}',
            item_id=req.item_id
        ))
        db.session.commit()
        flash(f'Devolução de "{req.item.nome}" registrada!', 'success')
    return redirect(url_for('main_bp.requisicoes'))

@main_bp.route('/almoxarifado/<int:id>/importar', methods=['GET', 'POST'])
@login_required
def importar_itens(id):
    u = usuario_atual()
    alm = Almoxarifado.query.get_or_404(id)
    if u.perfil != 'admin' and id not in u.almoxarifados_permitidos():
        flash('Acesso negado.', 'danger')
        return redirect(url_for('main_bp.index'))

    if request.method == 'POST':
        arquivo = request.files.get('arquivo')
        modo = request.form.get('modo', 'adicionar')  # adicionar | atualizar | substituir

        if not arquivo or not arquivo.filename.endswith(('.xlsx', '.xls')):
            flash('Envie um arquivo Excel (.xlsx ou .xls).', 'danger')
            return redirect(url_for('main_bp.importar_itens', id=id))

        try:
            wb = openpyxl.load_workbook(arquivo, data_only=True)
            ws = wb.active

            inseridos = 0
            atualizados = 0
            erros = []

            # Pular cabeçalho (linha 1)
            for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                # Ignorar linhas completamente vazias
                if not any(row):
                    continue

                # Colunas esperadas: codigo | nome | unidade | quantidade | estoque_minimo
                try:
                    codigo    = str(row[0]).strip() if row[0] else None
                    nome      = str(row[1]).strip() if row[1] else None
                    unidade   = str(row[2]).strip() if row[2] else 'un'
                    quantidade = float(row[3]) if row[3] is not None else 0
                    est_min   = float(row[4]) if row[4] is not None else 0
                except Exception:
                    erros.append(f'Linha {row_num}: formato inválido')
                    continue

                if not codigo or not nome:
                    erros.append(f'Linha {row_num}: código ou nome vazio')
                    continue

                item_existente = Item.query.filter_by(codigo=codigo).first()

                if item_existente:
                    if modo in ('atualizar', 'substituir'):
                        item_existente.nome = nome
                        item_existente.unidade = unidade
                        item_existente.estoque_minimo = est_min
                        item_existente.almoxarifado_id = id
                        if modo == 'substituir':
                            # Só sobrescreve a quantidade se o item não tiver movimentações
                            # Isso protege contra reimportação acidental de dados antigos
                            tem_movimentacoes = Movimentacao.query.filter_by(item_id=item_existente.id).count() > 0
                            if not tem_movimentacoes:
                                item_existente.quantidade = quantidade
                            # Se tiver movimentações, ignora a quantidade do Excel (protege o saldo real)
                        else:  # atualizar = somar
                            item_existente.quantidade += quantidade
                        atualizados += 1
                    # modo 'adicionar' ignora duplicatas
                else:
                    novo = Item(
                        codigo=codigo, nome=nome, unidade=unidade,
                        quantidade=quantidade, estoque_minimo=est_min,
                        almoxarifado_id=id
                    )
                    db.session.add(novo)
                    inseridos += 1

            db.session.commit()

            msg = f'Importação concluída: {inseridos} inseridos, {atualizados} atualizados.'
            if erros:
                msg += f' {len(erros)} linha(s) com erro.'
            flash(msg, 'success' if not erros else 'warning')
            for e in erros[:10]:  # mostrar até 10 erros
                flash(e, 'danger')

        except Exception as e:
            flash(f'Erro ao processar arquivo: {str(e)}', 'danger')

        return redirect(url_for('main_bp.almoxarifado', id=id))

    return render_template('importar_itens.html', almoxarifado=alm)

@main_bp.route('/almoxarifado/<int:id>/modelo_excel')
@login_required
def modelo_excel(id):
    alm = Almoxarifado.query.get_or_404(id)
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Itens'
    h_fill = PatternFill('solid', fgColor='1A3A5C')
    h_font = Font(bold=True, color='FFFFFF', size=11)
    borda  = Border(left=Side(style='thin'), right=Side(style='thin'),
                    top=Side(style='thin'), bottom=Side(style='thin'))
    for col, h in enumerate(['Codigo', 'Nome', 'Unidade', 'Quantidade', 'Estoque Minimo'], 1):
        c = ws.cell(row=1, column=col, value=h)
        c.font = h_font; c.fill = h_fill
        c.alignment = Alignment(horizontal='center'); c.border = borda
    for r, ex in enumerate([('CIM-001','Cimento CP-II','sc',500,100),
                             ('ARG-002','Areia Grossa','m3',30,5),
                             ('EPI-003','Capacete','un',20,5)], 2):
        for c, v in enumerate(ex, 1):
            ws.cell(row=r, column=c, value=v).border = borda
    for i, w in enumerate([14,40,10,14,16], 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    buf = io.BytesIO()
    wb.save(buf); buf.seek(0)
    return send_file(buf, as_attachment=True,
                     download_name=f'modelo_{alm.nome.replace(" ","_")}.xlsx',
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

@main_bp.route('/item/<int:id>/status_compra', methods=['POST'])
@login_required
def atualizar_status_compra(id):
    it = Item.query.get_or_404(id)
    u = usuario_atual()
    if u.perfil != 'admin' and it.almoxarifado_id not in u.almoxarifados_permitidos():
        return ('', 403)
    it.status_compra = request.form.get('status_compra', 'pendente')
    db.session.commit()
    return ('', 204)

@main_bp.route('/item/<int:id>/fixar', methods=['POST'])
@login_required
def fixar_item(id):
    it = Item.query.get_or_404(id)
    u = usuario_atual()
    if u.perfil != 'admin' and it.almoxarifado_id not in u.almoxarifados_permitidos():
        return jsonify({'error': 'Acesso negado.'}), 403
    it.fixado = not it.fixado
    db.session.commit()
    return jsonify({'fixado': it.fixado})

@main_bp.route('/movimentacao/<int:id>/devolvido', methods=['POST'])
@login_required
def marcar_devolvido(id):
    """Marca/desmarca uma movimentação de saída como devolvida."""
    mov = Movimentacao.query.get_or_404(id)
    u = usuario_atual()
    if u.perfil not in ('admin', 'almoxarife'):
        return jsonify({'error': 'Acesso negado.'}), 403
    # Toggle: None/False → True, True → False
    mov.devolvido = not bool(mov.devolvido)
    db.session.commit()
    return jsonify({'devolvido': mov.devolvido})

@main_bp.route('/item/<int:id>/desativar', methods=['POST'])
@login_required
def desativar_item(id):
    it = Item.query.get_or_404(id)
    u = usuario_atual()
    if u.perfil != 'admin' and it.almoxarifado_id not in u.almoxarifados_permitidos():
        flash('Acesso negado.', 'danger')
        return redirect(url_for('main_bp.item', id=id))
    
    # Registrar saída do saldo restante se houver
    if it.quantidade > 0:
        mov = Movimentacao(
            tipo='saida',
            quantidade=it.quantidade,
            responsavel=u.nome,
            observacao=f'Item desativado - saldo restante: {it.quantidade} {it.unidade}',
            item_id=id
        )
        db.session.add(mov)
        it.quantidade = 0
    
    it.ativo = False
    db.session.commit()
    flash(f'Item "{it.nome}" desativado com sucesso!', 'warning')
    return redirect(url_for('main_bp.item', id=id))

@main_bp.route('/item/<int:id>/reativar', methods=['POST'])
@login_required
def reativar_item(id):
    it = Item.query.get_or_404(id)
    u = usuario_atual()
    if u.perfil != 'admin' and it.almoxarifado_id not in u.almoxarifados_permitidos():
        flash('Acesso negado.', 'danger')
        return redirect(url_for('main_bp.item', id=id))
    
    it.ativo = True
    db.session.commit()
    flash(f'Item "{it.nome}" reativado com sucesso!', 'success')
    return redirect(url_for('main_bp.item', id=id))

# ── EXPORTAR EXCEL ───────────────────────────────────────────────────────────

@main_bp.route('/almoxarifado/<int:id>/exportar')
@login_required
def exportar_almoxarifado(id):
    u = usuario_atual()
    if u.perfil != 'admin' and id not in u.almoxarifados_permitidos():
        flash('Acesso negado.', 'danger')
        return redirect(url_for('main_bp.index'))
    alm = Almoxarifado.query.get_or_404(id)
    itens = Item.query.filter_by(almoxarifado_id=id).all()

    wb = openpyxl.Workbook()
    h_fill = PatternFill('solid', fgColor='1A3A5C')
    h_font = Font(bold=True, color='FFFFFF', size=11)
    ok_fill  = PatternFill('solid', fgColor='D4EDDA')
    al_fill  = PatternFill('solid', fgColor='FFF3CD')
    cr_fill  = PatternFill('solid', fgColor='F8D7DA')
    en_fill  = PatternFill('solid', fgColor='D4EDDA')
    sa_fill  = PatternFill('solid', fgColor='F8D7DA')
    borda = Border(left=Side(style='thin'), right=Side(style='thin'),
                   top=Side(style='thin'), bottom=Side(style='thin'))

    def titulo(ws, texto, cols):
        ws.merge_cells(f'A1:{get_column_letter(cols)}1')
        ws['A1'] = texto
        ws['A1'].font = Font(bold=True, size=13, color='1A3A5C')
        ws['A1'].alignment = Alignment(horizontal='center')
        ws.merge_cells(f'A2:{get_column_letter(cols)}2')
        ws['A2'] = f'Gerado em: {datetime.now().strftime("%d/%m/%Y %H:%M")}'
        ws['A2'].font = Font(italic=True, color='888888')
        ws['A2'].alignment = Alignment(horizontal='center')

    def cabecalho(ws, row, headers):
        for col, h in enumerate(headers, 1):
            c = ws.cell(row=row, column=col, value=h)
            c.font = h_font; c.fill = h_fill
            c.alignment = Alignment(horizontal='center'); c.border = borda

    categoria_label = {
        'epi': 'EPI', 'maquinario': 'Maquinário', 'eletrica': 'Elétrica',
        'hidraulica': 'Hidráulica', 'gas': 'Gás', 'geral': 'Geral'
    }

    # Aba 1 — Estoque Atual
    ws1 = wb.active
    ws1.title = 'Estoque Atual'
    titulo(ws1, f'Estoque Atual — {alm.nome}', 8)
    cabecalho(ws1, 4, ['Codigo', 'Item', 'Categoria', 'Unidade', 'Qtd. Atual', 'Est. Minimo', 'Deficit', 'Status'])
    for r, it in enumerate(itens, 5):
        deficit = max(0, it.estoque_minimo - it.quantidade)
        status = 'ZERADO' if it.status == 'critico' else ('ABAIXO DO MINIMO' if it.status == 'alerta' else 'OK')
        fill = cr_fill if it.status == 'critico' else (al_fill if it.status == 'alerta' else ok_fill)
        cat = categoria_label.get(it.categoria or 'geral', it.categoria or 'Geral')
        for c, v in enumerate([it.codigo, it.nome, cat, it.unidade, it.quantidade, it.estoque_minimo, deficit, status], 1):
            cell = ws1.cell(row=r, column=c, value=v)
            cell.fill = fill; cell.border = borda
            cell.alignment = Alignment(horizontal='left' if c == 2 else 'center')
    for i, w in enumerate([14, 35, 14, 10, 12, 14, 12, 20], 1):
        ws1.column_dimensions[get_column_letter(i)].width = w

    # Aba 2 — Movimentações
    ws2 = wb.create_sheet('Movimentacoes')
    titulo(ws2, f'Historico de Movimentacoes — {alm.nome}', 6)
    cabecalho(ws2, 4, ['Data', 'Item', 'Tipo', 'Quantidade', 'Responsavel', 'Observacao'])
    movs = (Movimentacao.query.join(Item)
            .filter(Item.almoxarifado_id == id)
            .order_by(Movimentacao.data.desc()).all())
    for r, mov in enumerate(movs, 5):
        fill = en_fill if mov.tipo == 'entrada' else sa_fill
        for c, v in enumerate([
            mov.data.strftime('%d/%m/%Y %H:%M'), mov.item.nome,
            'Entrada' if mov.tipo == 'entrada' else 'Saida',
            f'{mov.quantidade} {mov.item.unidade}',
            mov.responsavel or '', mov.observacao or ''
        ], 1):
            cell = ws2.cell(row=r, column=c, value=v)
            cell.fill = fill; cell.border = borda
            cell.alignment = Alignment(horizontal='left')
    for i, w in enumerate([18, 35, 10, 14, 22, 30], 1):
        ws2.column_dimensions[get_column_letter(i)].width = w

    # Aba 3 — Consumo por Item
    ws3 = wb.create_sheet('Consumo por Item')
    titulo(ws3, f'Consumo por Item — {alm.nome}', 4)
    cabecalho(ws3, 4, ['Item', 'Total Entradas', 'Total Saidas', 'Saldo Atual'])
    for r, it in enumerate(itens, 5):
        entradas = sum(m.quantidade for m in it.movimentacoes if m.tipo == 'entrada')
        saidas   = sum(m.quantidade for m in it.movimentacoes if m.tipo == 'saida')
        for c, v in enumerate([it.nome, entradas, saidas, it.quantidade], 1):
            cell = ws3.cell(row=r, column=c, value=v)
            cell.border = borda
            cell.alignment = Alignment(horizontal='left' if c == 1 else 'center')
    for i, w in enumerate([35, 16, 14, 14], 1):
        ws3.column_dimensions[get_column_letter(i)].width = w

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    nome = f"estoque_{alm.nome.replace(' ', '_')}_{date.today()}.xlsx"
    return send_file(buf, as_attachment=True, download_name=nome,
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

# ── API ──────────────────────────────────────────────────────────────────────
