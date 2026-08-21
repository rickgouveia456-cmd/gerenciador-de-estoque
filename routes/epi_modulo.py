"""Módulo EPI — Painel, Fichas de Entrega, Certificados CA, Matriz, Devoluções."""
import json
import io
import os
from datetime import datetime, date, timedelta
from flask import Blueprint, render_template, request, redirect, url_for, flash, jsonify, send_file
from extensions import db
from models import (agora, Almoxarifado, Colaborador, ItemEPI, HistoricoEPI,
                    FichaEPI, ItemFichaEPI, CertificadoCA, MatrizEPI,
                    Treinamento, TreinamentoParticipante, HabilitacaoFuncionario)
from core import login_required, usuario_atual
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

epi_modulo_bp = Blueprint('epi_modulo_bp', __name__)

# Caminho fixo onde o template PPTX é salvo
_TEMPLATE_PATH = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))),
                               'certificado_template.pptx')

_BLOQUEADOS = ('mestre',)

_BORDA = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin')
)
_BORDA_MED = Border(
    left=Side(style='medium'), right=Side(style='medium'),
    top=Side(style='medium'), bottom=Side(style='medium')
)
_CENTRO = Alignment(horizontal='center', vertical='center', wrap_text=True)
_ESQ    = Alignment(horizontal='left',   vertical='center', wrap_text=True)


def _checar(u):
    return u.perfil not in _BLOQUEADOS


def _template_existe_no_banco():
    """Verifica se há template PPTX salvo no banco."""
    try:
        from models import ConfiguracaoSistema
        cfg = ConfiguracaoSistema.query.filter_by(chave='certificado_template').first()
        return cfg is not None and cfg.binario is not None
    except Exception:
        return os.path.exists(_TEMPLATE_PATH)


def _alms_do_usuario(u):
    if u.perfil == 'admin':
        return Almoxarifado.query.all()
    ids = u.almoxarifados_permitidos()
    return Almoxarifado.query.filter(Almoxarifado.id.in_(ids)).all() if ids else []


def _alm_id_filtro(u):
    """Retorna set de alm_ids permitidos ou None (admin = sem filtro)."""
    if u.perfil == 'admin':
        return None
    return u.almoxarifados_permitidos()


# ══════════════════════════════════════════════════════════════════════════════
# ROTA PRINCIPAL — redireciona para o painel
# ══════════════════════════════════════════════════════════════════════════════
@epi_modulo_bp.route('/epi')
@login_required
def epi_index():
    return redirect(url_for('epi_modulo_bp.epi_painel'))


# ══════════════════════════════════════════════════════════════════════════════
# ABA 1 — PAINEL
# ══════════════════════════════════════════════════════════════════════════════
@epi_modulo_bp.route('/epi/painel')
@login_required
def epi_painel():
    u = usuario_atual()
    ids = _alm_id_filtro(u)

    # Fichas
    q_fichas = FichaEPI.query
    if ids is not None:
        q_fichas = q_fichas.filter(FichaEPI.almoxarifado_id.in_(ids))
    total_fichas   = q_fichas.count()
    fichas_ativas  = q_fichas.filter_by(status='ativa').count()
    fichas_enc     = q_fichas.filter_by(status='encerrada').count()

    # Itens de ficha — entregas do mês
    hoje = date.today()
    ini_mes = hoje.replace(day=1)
    q_itens = ItemFichaEPI.query.join(FichaEPI)
    if ids is not None:
        q_itens = q_itens.filter(FichaEPI.almoxarifado_id.in_(ids))
    entregas_mes = q_itens.filter(ItemFichaEPI.data_entrega >= ini_mes).count()

    # Devoluções em aberto (entregue mas sem devolução)
    devolucoes_abertas = q_itens.filter(
        ItemFichaEPI.data_entrega != None,
        ItemFichaEPI.data_devolucao == None
    ).count()

    # CAs — vencidos e a vencer em 90 dias
    q_ca = CertificadoCA.query.filter_by(ativo=True)
    if ids is not None:
        q_ca = q_ca.filter(CertificadoCA.almoxarifado_id.in_(ids))
    limite_90 = datetime.now() + timedelta(days=90)
    cas_vencidos  = q_ca.filter(CertificadoCA.data_validade < datetime.now()).count()
    cas_a_vencer  = q_ca.filter(
        CertificadoCA.data_validade >= datetime.now(),
        CertificadoCA.data_validade <= limite_90
    ).count()

    # Últimas 10 fichas abertas
    ultimas_fichas = (FichaEPI.query
        .filter_by(status='ativa')
        .order_by(FichaEPI.data_abertura.desc())
        .limit(10).all()
    ) if ids is None else (FichaEPI.query
        .filter(FichaEPI.almoxarifado_id.in_(ids), FichaEPI.status == 'ativa')
        .order_by(FichaEPI.data_abertura.desc())
        .limit(10).all()
    )

    return render_template('epi_modulo.html', aba='painel',
        total_fichas=total_fichas, fichas_ativas=fichas_ativas,
        fichas_enc=fichas_enc, entregas_mes=entregas_mes,
        devolucoes_abertas=devolucoes_abertas,
        cas_vencidos=cas_vencidos, cas_a_vencer=cas_a_vencer,
        ultimas_fichas=ultimas_fichas,
    )


# ══════════════════════════════════════════════════════════════════════════════
# ABA 2 — FICHAS DE ENTREGA
# ══════════════════════════════════════════════════════════════════════════════
@epi_modulo_bp.route('/epi/fichas')
@login_required
def epi_fichas():
    u = usuario_atual()
    ids = _alm_id_filtro(u)
    q = FichaEPI.query
    if ids is not None:
        q = q.filter(FichaEPI.almoxarifado_id.in_(ids))

    busca  = request.args.get('q', '').strip()
    status = request.args.get('status', '')
    if busca:
        q = q.filter(FichaEPI.colaborador.ilike(f'%{busca}%'))
    if status:
        q = q.filter_by(status=status)

    fichas = q.order_by(FichaEPI.data_abertura.desc()).all()
    almoxarifados = _alms_do_usuario(u)
    return render_template('epi_modulo.html', aba='fichas',
        fichas=fichas, busca=busca, status_filtro=status,
        almoxarifados=almoxarifados,
    )


@epi_modulo_bp.route('/epi/fichas/nova', methods=['GET', 'POST'])
@login_required
def epi_ficha_nova():
    u = usuario_atual()
    if not _checar(u):
        flash('Acesso negado.', 'danger')
        return redirect(url_for('epi_modulo_bp.epi_fichas'))

    almoxarifados = _alms_do_usuario(u)
    colaboradores = Colaborador.query.filter_by(ativo=True).order_by(Colaborador.nome).all()

    if request.method == 'POST':
        colaborador = request.form.get('colaborador', '').strip()
        funcao      = request.form.get('funcao', '').strip() or None
        obra        = request.form.get('obra', '').strip() or None
        alm_id      = request.form.get('almoxarifado_id', type=int)

        if not colaborador or not alm_id:
            flash('Colaborador e almoxarifado são obrigatórios.', 'danger')
            return render_template('epi_modulo.html', aba='ficha_nova',
                almoxarifados=almoxarifados, colaboradores=colaboradores,
                form_data=request.form)

        # Verificar se já existe ficha ativa para este colaborador neste almoxarifado
        ficha_existente = FichaEPI.query.filter_by(
            colaborador=colaborador,
            almoxarifado_id=alm_id,
            status='ativa'
        ).first()
        if ficha_existente:
            flash(f'{colaborador} já possui uma ficha ativa. Novos EPIs são adicionados à ficha existente.', 'info')
            return redirect(url_for('epi_modulo_bp.epi_ficha_detalhe', ficha_id=ficha_existente.id))

        ficha = FichaEPI(
            colaborador=colaborador, funcao=funcao, obra=obra,
            almoxarifado_id=alm_id, criado_por=u.nome,
            data_abertura=agora()
        )
        db.session.add(ficha)
        db.session.flush()

        # Salvar itens
        idx = 0
        while True:
            desc = request.form.get(f'desc_{idx}', '').strip()
            if not desc and idx > 0:
                break
            if desc:
                ca    = request.form.get(f'ca_{idx}', '').strip() or None
                qtd   = float(request.form.get(f'qtd_{idx}', 1) or 1)
                tam   = request.form.get(f'tam_{idx}', '').strip() or None
                db.session.add(ItemFichaEPI(
                    ficha_id=ficha.id, descricao=desc, ca=ca,
                    quantidade=qtd, tamanho=tam,
                    data_entrega=agora(), registrado_por=u.nome
                ))
            idx += 1
            if idx > 50:
                break

        db.session.commit()
        flash(f'Ficha de {colaborador} criada com sucesso!', 'success')
        return redirect(url_for('epi_modulo_bp.epi_fichas'))

    return render_template('epi_modulo.html', aba='ficha_nova',
        almoxarifados=almoxarifados, colaboradores=colaboradores, form_data={})


@epi_modulo_bp.route('/epi/fichas/<int:ficha_id>')
@login_required
def epi_ficha_detalhe(ficha_id):
    ficha = FichaEPI.query.get_or_404(ficha_id)
    return render_template('epi_modulo.html', aba='ficha_detalhe', ficha=ficha)


@epi_modulo_bp.route('/epi/fichas/<int:ficha_id>/encerrar', methods=['POST'])
@login_required
def epi_ficha_encerrar(ficha_id):
    u = usuario_atual()
    if not _checar(u):
        flash('Acesso negado.', 'danger')
        return redirect(url_for('epi_modulo_bp.epi_fichas'))
    ficha = FichaEPI.query.get_or_404(ficha_id)
    # Verifica se há EPIs ainda em uso (sem devolução)
    epis_em_uso = [it for it in ficha.itens if it.data_entrega and not it.data_devolucao]
    if epis_em_uso:
        flash(
            f'⚠️ Não é possível encerrar: {len(epis_em_uso)} EPI(s) ainda em uso sem devolução registrada. '
            'Registre a devolução antes de encerrar a ficha.',
            'warning'
        )
        return redirect(url_for('epi_modulo_bp.epi_ficha_detalhe', ficha_id=ficha_id))
    ficha.status = 'encerrada'
    ficha.data_encerramento = agora()
    db.session.commit()
    flash(f'Ficha de {ficha.colaborador} encerrada.', 'warning')
    return redirect(url_for('epi_modulo_bp.epi_fichas'))


@epi_modulo_bp.route('/epi/fichas/<int:ficha_id>/exportar')
@login_required
def epi_ficha_exportar(ficha_id):
    ficha = FichaEPI.query.get_or_404(ficha_id)
    wb = _gerar_form_seg014(ficha)
    buf = io.BytesIO()
    wb.save(buf); buf.seek(0)
    nome = f'FORM-SEG-014_{ficha.colaborador[:20]}_{date.today()}.xlsx'
    return send_file(buf, as_attachment=True, download_name=nome,
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')


@epi_modulo_bp.route('/epi/fichas/<int:ficha_id>/devolver/<int:item_id>', methods=['POST'])
@login_required
def epi_item_devolver(ficha_id, item_id):
    item = ItemFichaEPI.query.get_or_404(item_id)
    item.data_devolucao   = agora()
    item.motivo_devolucao = request.form.get('motivo', '').strip() or None
    db.session.commit()
    return jsonify({'ok': True})


# ══════════════════════════════════════════════════════════════════════════════
# ABA 3 — CERTIFICADOS CA
# ══════════════════════════════════════════════════════════════════════════════
@epi_modulo_bp.route('/epi/certificados')
@login_required
def epi_certificados():
    u = usuario_atual()
    almoxarifados = _alms_do_usuario(u)
    return render_template('epi_modulo.html', aba='certificados',
        almoxarifados=almoxarifados,
        busca='',
        cas=[],
    )


@epi_modulo_bp.route('/epi/certificados/novo', methods=['POST'])
@login_required
def epi_ca_novo():
    u = usuario_atual()
    if not _checar(u):
        return jsonify({'error': 'Acesso negado'}), 403

    alm_id = request.form.get('almoxarifado_id', type=int)
    val_str = request.form.get('data_validade', '').strip()
    emi_str = request.form.get('data_emissao', '').strip()

    ca = CertificadoCA(
        numero_ca  = request.form.get('numero_ca', '').strip(),
        nome_epi   = request.form.get('nome_epi', '').strip(),
        fabricante = request.form.get('fabricante', '').strip() or None,
        tipo       = request.form.get('tipo', '').strip() or None,
        data_validade = datetime.strptime(val_str, '%Y-%m-%d') if val_str else None,
        data_emissao  = datetime.strptime(emi_str, '%Y-%m-%d') if emi_str else None,
        almoxarifado_id = alm_id,
        criado_por = u.nome,
    )
    db.session.add(ca)
    db.session.commit()
    flash(f'CA {ca.numero_ca} cadastrado!', 'success')
    return redirect(url_for('epi_modulo_bp.epi_certificados'))


@epi_modulo_bp.route('/epi/certificados/<int:ca_id>/deletar', methods=['POST'])
@login_required
def epi_ca_deletar(ca_id):
    ca = CertificadoCA.query.get_or_404(ca_id)
    ca.ativo = False
    db.session.commit()
    flash(f'CA {ca.numero_ca} removido.', 'warning')
    return redirect(url_for('epi_modulo_bp.epi_certificados'))


@epi_modulo_bp.route('/epi/certificados/exportar-funcionario')
@login_required
def epi_ca_exportar_funcionario():
    """Gera Excel com CAs dos EPIs entregues a um colaborador."""
    colaborador = request.args.get('colaborador', '').strip()
    alm_id      = request.args.get('almoxarifado_id', type=int)
    if not colaborador:
        flash('Informe o colaborador.', 'warning')
        return redirect(url_for('epi_modulo_bp.epi_certificados'))

    # Busca fichas do colaborador
    q = FichaEPI.query.filter(FichaEPI.colaborador.ilike(f'%{colaborador}%'))
    if alm_id:
        q = q.filter_by(almoxarifado_id=alm_id)
    fichas = q.all()

    itens = []
    for ficha in fichas:
        for item in ficha.itens:
            itens.append(item)

    wb = _gerar_excel_certificados(colaborador, itens)
    buf = io.BytesIO()
    wb.save(buf); buf.seek(0)
    nome = f'Certificados_{colaborador[:20]}_{date.today()}.xlsx'
    return send_file(buf, as_attachment=True, download_name=nome,
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')


# ══════════════════════════════════════════════════════════════════════════════
# ABA 4 — MATRIZ DE EPI
# ══════════════════════════════════════════════════════════════════════════════
@epi_modulo_bp.route('/epi/matriz')
@login_required
def epi_matriz():
    u = usuario_atual()
    ids = _alm_id_filtro(u)
    q = MatrizEPI.query
    if ids is not None:
        q = q.filter(MatrizEPI.almoxarifado_id.in_(ids))
    matrizes = q.order_by(MatrizEPI.funcao).all()

    # Deserializa epis_obrigatorios
    for m in matrizes:
        try:
            m._epis_lista = json.loads(m.epis_obrigatorios or '[]')
        except Exception:
            m._epis_lista = []

    almoxarifados = _alms_do_usuario(u)
    return render_template('epi_modulo.html', aba='matriz',
        matrizes=matrizes, almoxarifados=almoxarifados,
    )


@epi_modulo_bp.route('/epi/matriz/nova', methods=['POST'])
@login_required
def epi_matriz_nova():
    u = usuario_atual()
    if not _checar(u):
        flash('Acesso negado.', 'danger')
        return redirect(url_for('epi_modulo_bp.epi_matriz'))

    funcao  = request.form.get('funcao', '').strip()
    obra    = request.form.get('obra', '').strip() or None
    alm_id  = request.form.get('almoxarifado_id', type=int)
    norma   = request.form.get('norma', '').strip() or None
    epis_raw = request.form.get('epis_obrigatorios', '').strip()

    # Processa lista de EPIs (separados por vírgula ou quebra de linha)
    import re
    epis_lista = [e.strip() for e in re.split(r'[,\n]+', epis_raw) if e.strip()]

    m = MatrizEPI(
        funcao=funcao, obra=obra, almoxarifado_id=alm_id,
        norma=norma, criado_por=u.nome,
        epis_obrigatorios=json.dumps(epis_lista, ensure_ascii=False),
    )
    db.session.add(m)
    db.session.commit()
    flash(f'Matriz para "{funcao}" criada!', 'success')
    return redirect(url_for('epi_modulo_bp.epi_matriz'))


@epi_modulo_bp.route('/epi/matriz/<int:mid>/deletar', methods=['POST'])
@login_required
def epi_matriz_deletar(mid):
    m = MatrizEPI.query.get_or_404(mid)
    db.session.delete(m)
    db.session.commit()
    flash('Matriz removida.', 'warning')
    return redirect(url_for('epi_modulo_bp.epi_matriz'))


# ══════════════════════════════════════════════════════════════════════════════
# ABA 5 — DEVOLUÇÕES
# ══════════════════════════════════════════════════════════════════════════════
@epi_modulo_bp.route('/epi/devolucoes')
@login_required
def epi_devolucoes():
    u = usuario_atual()
    ids = _alm_id_filtro(u)

    q = (ItemFichaEPI.query
         .join(FichaEPI)
         .filter(ItemFichaEPI.data_entrega != None,
                 ItemFichaEPI.data_devolucao == None))
    if ids is not None:
        q = q.filter(FichaEPI.almoxarifado_id.in_(ids))

    busca = request.args.get('q', '').strip()
    if busca:
        q = q.filter(FichaEPI.colaborador.ilike(f'%{busca}%'))

    itens = q.order_by(ItemFichaEPI.data_entrega.asc()).all()

    # Calcular dias em aberto
    hoje = datetime.now()
    for it in itens:
        it._dias_aberto = (hoje - it.data_entrega).days if it.data_entrega else 0

    return render_template('epi_modulo.html', aba='devolucoes',
        itens=itens, busca=busca,
    )


# ══════════════════════════════════════════════════════════════════════════════
# ABA 6 — TREINAMENTOS
# ══════════════════════════════════════════════════════════════════════════════
@epi_modulo_bp.route('/epi/treinamentos')
@login_required
def epi_treinamentos():
    u = usuario_atual()
    almoxarifados = _alms_do_usuario(u)
    return render_template('epi_modulo.html', aba='treinamentos',
        almoxarifados=almoxarifados,
        template_existe=_template_existe_no_banco(),
    )

@epi_modulo_bp.route('/epi/treinamentos/novo', methods=['POST'])
@login_required
def epi_treinamento_novo():
    u = usuario_atual()
    if not _checar(u):
        flash('Acesso negado.', 'danger')
        return redirect(url_for('epi_modulo_bp.epi_treinamentos'))

    tipo       = request.form.get('tipo', '').strip()
    descricao  = request.form.get('descricao', '').strip() or None
    data_str   = request.form.get('data_realizacao', '').strip()
    val_meses  = request.form.get('validade_meses', '').strip()
    responsavel      = request.form.get('responsavel', '').strip() or None
    cargo_responsavel = request.form.get('cargo_responsavel', '').strip() or None
    registro_mte     = request.form.get('registro_mte', '').strip() or None
    local_trein      = request.form.get('local', '').strip() or None
    carga_h          = request.form.get('carga_horaria', '').strip()
    alm_id     = request.form.get('almoxarifado_id', type=int)

    if not tipo or not data_str:
        flash('Tipo e data são obrigatórios.', 'danger')
        return redirect(url_for('epi_modulo_bp.epi_treinamentos'))

    try:
        data_real = datetime.strptime(data_str, '%Y-%m-%d')
    except ValueError:
        flash('Data inválida.', 'danger')
        return redirect(url_for('epi_modulo_bp.epi_treinamentos'))

    # Configurações padrão por tipo
    from certificado_pptx import NR_CONFIG
    cfg = NR_CONFIG.get(tipo, {})

    t = Treinamento(
        tipo=tipo, descricao=descricao,
        data_realizacao=data_real,
        validade_meses=int(val_meses) if val_meses else cfg.get('validade'),
        responsavel=responsavel,
        cargo_responsavel=cargo_responsavel,
        registro_mte=registro_mte,
        local=local_trein,
        carga_horaria=int(carga_h) if carga_h else cfg.get('carga'),
        nr_referencia=cfg.get('nr'),
        portaria=cfg.get('portaria'),
        almoxarifado_id=alm_id or None,
        criado_por=u.nome,
    )
    db.session.add(t)
    db.session.flush()

    # Participantes via campos dinâmicos: nome_0, cpf_0, funcao_0, ...
    idx = 0
    total_participantes = 0
    while True:
        nome = request.form.get(f'p_nome_{idx}', '').strip()
        if not nome and idx > 0:
            break
        if nome:
            db.session.add(TreinamentoParticipante(
                treinamento_id=t.id,
                colaborador=nome,
                cpf=request.form.get(f'p_cpf_{idx}', '').strip() or None,
                funcao=request.form.get(f'p_funcao_{idx}', '').strip() or None,
                concluiu=True
            ))
            total_participantes += 1
        idx += 1
        if idx > 200:
            break

    db.session.commit()
    flash(f'Treinamento "{tipo}" registrado com {total_participantes} participante(s)!', 'success')
    return redirect(url_for('epi_modulo_bp.epi_treinamentos'))


# ══════════════════════════════════════════════════════════════════════════════
# TEMPLATE PPTX — upload e status
# ══════════════════════════════════════════════════════════════════════════════
@epi_modulo_bp.route('/epi/template/upload', methods=['POST'])
@login_required
def epi_template_upload():
    """Admin faz upload do template PPTX do certificado — salvo no banco."""
    u = usuario_atual()
    if u.perfil not in ('admin', 'almoxarife'):
        flash('Acesso negado.', 'danger')
        return redirect(url_for('epi_modulo_bp.epi_treinamentos'))

    arquivo = request.files.get('template_pptx')
    if not arquivo or not arquivo.filename:
        flash('Nenhum arquivo enviado.', 'danger')
        return redirect(url_for('epi_modulo_bp.epi_treinamentos'))

    if not arquivo.filename.lower().endswith('.pptx'):
        flash('Apenas arquivos .pptx são aceitos.', 'danger')
        return redirect(url_for('epi_modulo_bp.epi_treinamentos'))

    try:
        from models import ConfiguracaoSistema
        conteudo = arquivo.read()
        cfg = ConfiguracaoSistema.query.filter_by(chave='certificado_template').first()
        if cfg:
            cfg.binario = conteudo
        else:
            cfg = ConfiguracaoSistema(chave='certificado_template', binario=conteudo)
            db.session.add(cfg)
        db.session.commit()
        # Salva também em disco como cache (para uso local)
        try:
            with open(_TEMPLATE_PATH, 'wb') as f:
                f.write(conteudo)
        except Exception:
            pass
        flash('Template de certificado atualizado com sucesso!', 'success')
    except Exception as e:
        flash(f'Erro ao salvar template: {e}', 'danger')

    return redirect(url_for('epi_modulo_bp.epi_treinamentos'))


@epi_modulo_bp.route('/epi/template/remover', methods=['POST'])
@login_required
def epi_template_remover():
    """Remove o template do banco e do disco."""
    u = usuario_atual()
    if u.perfil not in ('admin', 'almoxarife'):
        flash('Acesso negado.', 'danger')
        return redirect(url_for('epi_modulo_bp.epi_treinamentos'))
    try:
        from models import ConfiguracaoSistema
        cfg = ConfiguracaoSistema.query.filter_by(chave='certificado_template').first()
        if cfg:
            db.session.delete(cfg)
            db.session.commit()
        if os.path.exists(_TEMPLATE_PATH):
            os.remove(_TEMPLATE_PATH)
        flash('Template removido. O certificado padrão será usado.', 'warning')
    except Exception as e:
        flash(f'Erro ao remover: {e}', 'danger')
    return redirect(url_for('epi_modulo_bp.epi_treinamentos'))


@epi_modulo_bp.route('/epi/treinamentos/<int:tid>/certificado')
@login_required
def epi_treinamento_certificado(tid):
    from certificado_pptx import gerar_certificado_pptx
    from models import ConfiguracaoSistema
    t = Treinamento.query.get_or_404(tid)
    if not t.participantes:
        flash('Nenhum participante cadastrado neste treinamento.', 'warning')
        return redirect(url_for('epi_modulo_bp.epi_treinamentos'))

    # Tenta usar template do banco primeiro; fallback para disco
    template_path = None
    cfg = ConfiguracaoSistema.query.filter_by(chave='certificado_template').first()
    if cfg and cfg.binario:
        # Salva temporariamente em disco para o python-pptx ler
        try:
            with open(_TEMPLATE_PATH, 'wb') as f:
                f.write(cfg.binario)
            template_path = _TEMPLATE_PATH
        except Exception:
            template_path = _TEMPLATE_PATH if os.path.exists(_TEMPLATE_PATH) else None
    elif os.path.exists(_TEMPLATE_PATH):
        template_path = _TEMPLATE_PATH

    buf = gerar_certificado_pptx(t.participantes, t, template_path=template_path)
    if buf is None:
        flash('Template inválido. Usando layout padrão.', 'warning')
        buf = gerar_certificado_pptx(t.participantes, t)
    nome_arq = f'Certificado_{t.tipo.replace(" ","_")}_{t.data_realizacao.strftime("%Y%m%d")}.pptx'
    return send_file(buf, as_attachment=True, download_name=nome_arq,
                     mimetype='application/vnd.openxmlformats-officedocument.presentationml.presentation')


@epi_modulo_bp.route('/epi/treinamentos/<int:tid>/deletar', methods=['POST'])
@login_required
def epi_treinamento_deletar(tid):
    u = usuario_atual()
    if not _checar(u):
        flash('Acesso negado.', 'danger')
        return redirect(url_for('epi_modulo_bp.epi_treinamentos'))
    t = Treinamento.query.get_or_404(tid)
    db.session.delete(t)
    db.session.commit()
    flash('Treinamento removido.', 'warning')
    return redirect(url_for('epi_modulo_bp.epi_treinamentos'))


# ══════════════════════════════════════════════════════════════════════════════
# ABA 7 — HABILITAÇÕES
# ══════════════════════════════════════════════════════════════════════════════
@epi_modulo_bp.route('/epi/habilitacoes')
@login_required
def epi_habilitacoes():
    u = usuario_atual()
    almoxarifados = _alms_do_usuario(u)
    return render_template('epi_modulo.html', aba='habilitacoes',
        almoxarifados=almoxarifados,
    )


@epi_modulo_bp.route('/epi/habilitacoes/nova', methods=['POST'])
@login_required
def epi_habilitacao_nova():
    u = usuario_atual()
    if not _checar(u):
        flash('Acesso negado.', 'danger')
        return redirect(url_for('epi_modulo_bp.epi_habilitacoes'))

    colaborador = request.form.get('colaborador', '').strip()
    tipo        = request.form.get('tipo', '').strip()
    numero      = request.form.get('numero', '').strip() or None
    emissor     = request.form.get('emissor', '').strip() or None
    funcao_hab  = request.form.get('funcao_habilitada', '').strip() or None
    alm_id      = request.form.get('almoxarifado_id', type=int)
    emi_str     = request.form.get('data_emissao', '').strip()
    val_str     = request.form.get('data_validade', '').strip()

    if not colaborador or not tipo:
        flash('Colaborador e tipo são obrigatórios.', 'danger')
        return redirect(url_for('epi_modulo_bp.epi_habilitacoes'))

    h = HabilitacaoFuncionario(
        colaborador=colaborador, tipo=tipo, numero=numero,
        emissor=emissor, funcao_habilitada=funcao_hab,
        almoxarifado_id=alm_id or None,
        data_emissao=datetime.strptime(emi_str, '%Y-%m-%d') if emi_str else None,
        data_validade=datetime.strptime(val_str, '%Y-%m-%d') if val_str else None,
        criado_por=u.nome,
    )
    db.session.add(h)
    db.session.commit()
    flash(f'Habilitação "{tipo}" de {colaborador} cadastrada!', 'success')
    return redirect(url_for('epi_modulo_bp.epi_habilitacoes'))


@epi_modulo_bp.route('/epi/habilitacoes/<int:hid>/deletar', methods=['POST'])
@login_required
def epi_habilitacao_deletar(hid):
    u = usuario_atual()
    if not _checar(u):
        flash('Acesso negado.', 'danger')
        return redirect(url_for('epi_modulo_bp.epi_habilitacoes'))
    h = HabilitacaoFuncionario.query.get_or_404(hid)
    h.ativo = False
    db.session.commit()
    flash('Habilitação removida.', 'warning')
    return redirect(url_for('epi_modulo_bp.epi_habilitacoes'))


@epi_modulo_bp.route('/epi/habilitacoes/exportar')
@login_required
def epi_habilitacoes_exportar():
    """Exporta todas as habilitações em Excel."""
    u = usuario_atual()
    ids = _alm_id_filtro(u)
    q = HabilitacaoFuncionario.query.filter_by(ativo=True)
    if ids is not None:
        q = q.filter(HabilitacaoFuncionario.almoxarifado_id.in_(ids))
    habs = q.order_by(HabilitacaoFuncionario.colaborador).all()

    wb = _gerar_excel_habilitacoes(habs)
    buf = io.BytesIO()
    wb.save(buf); buf.seek(0)
    return send_file(buf, as_attachment=True,
                     download_name=f'Habilitacoes_{date.today()}.xlsx',
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')


# ══════════════════════════════════════════════════════════════════════════════
# HELPERS — Geração de Excel
# ══════════════════════════════════════════════════════════════════════════════
def _gerar_form_seg014(ficha: FichaEPI):
    """Gera planilha FORM.SEG.014 a partir de uma FichaEPI."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = ficha.colaborador[:28]

    azul_esc = PatternFill('solid', fgColor='1F3864')
    azul_cla = PatternFill('solid', fgColor='BDD7EE')
    cinza    = PatternFill('solid', fgColor='F2F2F2')

    for col, w in zip('ABCDEFGH', [8, 40, 10, 14, 26, 14, 26, 18]):
        ws.column_dimensions[col].width = w

    # Linha 1 — cabeçalho
    ws.row_dimensions[1].height = 36
    ws.merge_cells('A1:C1')
    ws['A1'].value = 'STANZA'
    ws['A1'].font = Font(bold=True, size=22, color='808080')
    ws['A1'].alignment = _CENTRO
    for c in range(1, 4): ws.cell(1, c).border = _BORDA_MED

    ws.merge_cells('D1:F1')
    ws['D1'].value = "FICHA DE CONTROLE DE EPI'S E UNIFORMES"
    ws['D1'].font = Font(bold=True, size=13, color='1F3864')
    ws['D1'].alignment = _CENTRO
    ws['D1'].fill = azul_cla
    for c in range(4, 7): ws.cell(1, c).border = _BORDA_MED

    ws.merge_cells('G1:H1')
    ws['G1'].value = 'FORM.SEG.014'
    ws['G1'].font = Font(bold=True, size=9, color='1F3864')
    ws['G1'].alignment = _CENTRO
    ws['G1'].fill = azul_cla
    for c in range(7, 9): ws.cell(1, c).border = _BORDA_MED

    # Linha 2 — data elaboração
    ws.row_dimensions[2].height = 14
    ws.merge_cells('D2:F2')
    ws['D2'].value = 'Data Elaboração/Revisão: 20/10/2024'
    ws['D2'].font = Font(size=8, italic=True, color='595959')
    ws['D2'].alignment = _CENTRO
    ws.merge_cells('G2:H2')
    ws['G2'].value = 'Revisão: 00'
    ws['G2'].font = Font(size=8, italic=True, color='595959')
    ws['G2'].alignment = _CENTRO
    for c in range(1, 9): ws.cell(2, c).border = _BORDA

    # Linha 3 — dados do funcionário
    ws.row_dimensions[3].height = 22
    ws.merge_cells('A3:B3')
    ws['A3'].value = f'NOME: {ficha.colaborador.upper()}'
    ws['A3'].font = Font(bold=True, size=10)
    ws['A3'].alignment = _ESQ
    ws['A3'].fill = cinza
    ws['C3'].value = 'MATRÍCULA:'
    ws['C3'].font = Font(size=9)
    ws['C3'].alignment = _CENTRO
    ws['C3'].fill = cinza
    ws.merge_cells('D3:E3')
    ws['D3'].value = f'FUNÇÃO: {ficha.funcao or ""}'
    ws['D3'].font = Font(size=9)
    ws['D3'].alignment = _ESQ
    ws['D3'].fill = cinza
    ws.merge_cells('F3:G3')
    ws['F3'].value = f'OBRA: {ficha.obra or ""}'
    ws['F3'].font = Font(size=9)
    ws['F3'].alignment = _ESQ
    ws['F3'].fill = cinza
    ws['H3'].fill = cinza
    for c in range(1, 9): ws.cell(3, c).border = _BORDA

    # Linha 4 — cabeçalho tabela
    ws.row_dimensions[4].height = 20
    for ref, val in [('A4', 'QUANT'), ('B4', 'DESCRIÇÃO'), ('C4', 'C.A.')]:
        ws[ref].value = val
        ws[ref].font = Font(bold=True, color='FFFFFF', size=9)
        ws[ref].fill = azul_esc
        ws[ref].alignment = _CENTRO
        ws[ref].border = _BORDA

    ws.merge_cells('D4:E4')
    ws['D4'].value = 'ENTREGA'
    ws['D4'].font = Font(bold=True, color='FFFFFF', size=9)
    ws['D4'].fill = azul_esc
    ws['D4'].alignment = _CENTRO
    for c in range(4, 6): ws.cell(4, c).border = _BORDA

    ws.merge_cells('F4:G4')
    ws['F4'].value = 'DEVOLUÇÃO'
    ws['F4'].font = Font(bold=True, color='FFFFFF', size=9)
    ws['F4'].fill = azul_esc
    ws['F4'].alignment = _CENTRO
    for c in range(6, 8): ws.cell(4, c).border = _BORDA

    ws['H4'].value = 'MOTIVO'
    ws['H4'].font = Font(bold=True, color='FFFFFF', size=9)
    ws['H4'].fill = azul_esc
    ws['H4'].alignment = _CENTRO
    ws['H4'].border = _BORDA

    # Linha 5 — sub-cabeçalho
    ws.row_dimensions[5].height = 16
    for ref, val in [('A5', ''), ('B5', ''), ('C5', ''),
                     ('D5', 'DATA'), ('E5', 'ASSINATURA'),
                     ('F5', 'DATA'), ('G5', 'ASSINATURA'), ('H5', '')]:
        ws[ref].value = val
        ws[ref].font = Font(bold=True, color='FFFFFF', size=8)
        ws[ref].fill = azul_esc
        ws[ref].alignment = _CENTRO
        ws[ref].border = _BORDA

    # Linhas de dados
    row = 6
    for item in ficha.itens:
        ws.row_dimensions[row].height = 18
        fill_z = PatternFill('solid', fgColor='EBF3FB') if row % 2 == 0 else None
        dev_data = item.data_devolucao.strftime('%d/%m/%Y') if item.data_devolucao else ''
        for col, val in zip('ABCDEFGH', [
            f'{item.quantidade} un',
            item.descricao,
            item.ca or '',
            item.data_entrega.strftime('%d/%m/%Y') if item.data_entrega else '',
            '',
            dev_data,
            '',
            item.motivo_devolucao or ''
        ]):
            c = ws[f'{col}{row}']
            c.value = val
            c.font = Font(size=9)
            c.alignment = _ESQ if col == 'B' else _CENTRO
            c.border = _BORDA
            if fill_z: c.fill = fill_z
        row += 1

    # Completar mínimo 14 linhas
    total_linhas = max(14, len(ficha.itens) + 4)
    while row <= 5 + total_linhas:
        ws.row_dimensions[row].height = 18
        for col in 'ABCDEFGH':
            ws[f'{col}{row}'].border = _BORDA
            ws[f'{col}{row}'].value = '/    /' if col in ('D', 'F') else ''
            ws[f'{col}{row}'].font = Font(size=9, color='BFBFBF')
            ws[f'{col}{row}'].alignment = _CENTRO
        row += 1

    # Termo de responsabilidade
    row += 1
    ws.row_dimensions[row].height = 16
    ws.merge_cells(f'A{row}:H{row}')
    ws[f'A{row}'].value = 'TERMO DE RESPONSABILIDADE'
    ws[f'A{row}'].font = Font(bold=True, size=10, color='1F3864')
    ws[f'A{row}'].alignment = _CENTRO
    ws[f'A{row}'].fill = azul_cla
    for c in range(1, 9): ws.cell(row, c).border = _BORDA
    row += 1

    ws.row_dimensions[row].height = 70
    ws.merge_cells(f'A{row}:H{row}')
    ws[f'A{row}'].value = (
        'Pelo presente declaro que recebi da empresa STANZA INCORPORAÇÃO E CONSTRUÇÃO LTDA., os materiais '
        'relacionados nesta ficha, assumindo o compromisso nos termos das letras "a" e "b" do ítem 1.8 da NR 1 '
        'e letras "a","b"e "c" do ítem 6.7.1 da NR 6, de usá-los em atividades ligadas ao trabalho, zelar pela '
        'sua guarda, conservação e devolvê-lo ao setor competente quando se tornar impróprio para uso ou por '
        'motivo de demissão ou afastamento.'
    )
    ws[f'A{row}'].font = Font(size=8)
    ws[f'A{row}'].alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
    for c in range(1, 9): ws.cell(row, c).border = _BORDA
    row += 2

    # Data e assinaturas
    ws.row_dimensions[row].height = 22
    ws.merge_cells(f'A{row}:H{row}')
    ws[f'A{row}'].value = 'Data: _______ / _______ / _____________'
    ws[f'A{row}'].font = Font(size=10)
    ws[f'A{row}'].alignment = _ESQ
    for c in range(1, 9): ws.cell(row, c).border = _BORDA
    row += 2

    ws.row_dimensions[row].height = 30
    ws.merge_cells(f'A{row}:C{row}')
    ws[f'A{row}'].value = 'Assinatura do Funcionário'
    ws[f'A{row}'].font = Font(size=9)
    ws[f'A{row}'].alignment = _CENTRO
    ws.merge_cells(f'D{row}:F{row}')
    ws[f'D{row}'].value = 'Responsável pela Entrega'
    ws[f'D{row}'].font = Font(size=9)
    ws[f'D{row}'].alignment = _CENTRO
    ws.merge_cells(f'G{row}:H{row}')
    ws[f'G{row}'].value = 'Testemunha'
    ws[f'G{row}'].font = Font(size=9)
    ws[f'G{row}'].alignment = _CENTRO
    for c in range(1, 9): ws.cell(row, c).border = _BORDA

    return wb


def _gerar_excel_certificados(colaborador: str, itens):
    """Gera Excel com CAs dos EPIs entregues a um colaborador."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Certificados'

    azul = PatternFill('solid', fgColor='1F3864')
    for col, w in zip('ABCDE', [35, 15, 20, 18, 18]):
        ws.column_dimensions[get_column_letter(ord(col) - 64)].width = w

    ws.merge_cells('A1:E1')
    ws['A1'].value = f'Certificados de Aprovação (CA) — {colaborador.upper()}'
    ws['A1'].font = Font(bold=True, size=13, color='1F3864')
    ws['A1'].fill = PatternFill('solid', fgColor='BDD7EE')
    ws['A1'].alignment = _CENTRO

    ws.merge_cells('A2:E2')
    ws['A2'].value = f'Gerado em: {date.today().strftime("%d/%m/%Y")}'
    ws['A2'].font = Font(italic=True, size=9, color='595959')
    ws['A2'].alignment = _CENTRO

    headers = ['EPI / Descrição', 'CA', 'Quantidade', 'Data Entrega', 'Data Devolução']
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=4, column=col, value=h)
        c.font = Font(bold=True, color='FFFFFF', size=10)
        c.fill = azul
        c.alignment = _CENTRO
        c.border = _BORDA

    for i, item in enumerate(itens, 5):
        fill_z = PatternFill('solid', fgColor='F0F4F8') if i % 2 == 0 else None
        vals = [
            item.descricao,
            item.ca or '—',
            item.quantidade,
            item.data_entrega.strftime('%d/%m/%Y') if item.data_entrega else '—',
            item.data_devolucao.strftime('%d/%m/%Y') if item.data_devolucao else 'Em uso',
        ]
        for col, val in enumerate(vals, 1):
            c = ws.cell(row=i, column=col, value=val)
            c.font = Font(size=10)
            c.border = _BORDA
            c.alignment = _CENTRO if col != 1 else _ESQ
            if fill_z: c.fill = fill_z

    if not itens:
        ws.merge_cells('A5:E5')
        ws['A5'].value = 'Nenhum EPI entregue encontrado para este colaborador.'
        ws['A5'].font = Font(italic=True, color='999999')
        ws['A5'].alignment = _CENTRO

    return wb


def _gerar_excel_habilitacoes(habs):
    """Gera Excel com todas as habilitações dos funcionários."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Habilitações'

    azul  = PatternFill('solid', fgColor='1F3864')
    verde = PatternFill('solid', fgColor='D9F2E4')
    amar  = PatternFill('solid', fgColor='FFF3CD')
    verm  = PatternFill('solid', fgColor='FDE8E8')

    cols = ['Colaborador', 'Tipo / Certificado', 'Nº Certificado', 'Emissor',
            'Função Habilitada', 'Emissão', 'Validade', 'Status']
    widths = [35, 28, 20, 25, 30, 14, 14, 14]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    ws.merge_cells('A1:H1')
    ws['A1'].value = f'Habilitações dos Funcionários — Gerado em {date.today().strftime("%d/%m/%Y")}'
    ws['A1'].font = Font(bold=True, size=13, color='1F3864')
    ws['A1'].fill = PatternFill('solid', fgColor='BDD7EE')
    ws['A1'].alignment = _CENTRO

    for col, h in enumerate(cols, 1):
        c = ws.cell(row=3, column=col, value=h)
        c.font = Font(bold=True, color='FFFFFF', size=10)
        c.fill = azul
        c.alignment = _CENTRO
        c.border = _BORDA

    status_label = {
        'valido': '🟢 Válido',
        'a_vencer': '🟡 A vencer',
        'vencido': '🔴 Vencido',
        'sem_vencimento': '— Sem venc.',
    }

    for i, h in enumerate(habs, 4):
        st = h.status
        fill_st = verde if st == 'valido' else (amar if st == 'a_vencer' else (verm if st == 'vencido' else None))
        vals = [
            h.colaborador, h.tipo, h.numero or '—', h.emissor or '—',
            h.funcao_habilitada or '—',
            h.data_emissao.strftime('%d/%m/%Y') if h.data_emissao else '—',
            h.data_validade.strftime('%d/%m/%Y') if h.data_validade else '—',
            status_label.get(st, '—'),
        ]
        for col, val in enumerate(vals, 1):
            c = ws.cell(row=i, column=col, value=val)
            c.font = Font(size=10)
            c.border = _BORDA
            c.alignment = _ESQ if col <= 2 else _CENTRO
            if fill_st:
                c.fill = fill_st

    return wb


# ══════════════════════════════════════════════════════════════════════════════
# FUNÇÃO AUXILIAR — chamada pelos 3 pontos de entrega de EPI
# ══════════════════════════════════════════════════════════════════════════════
def registrar_epi_na_ficha(
    colaborador: str,
    almoxarifado_id: int,
    descricao: str,
    ca=None,
    quantidade: float = 1.0,
    tamanho=None,
    registrado_por: str = 'Sistema'
):
    """
    Busca ou cria a FichaEPI ativa do colaborador no almoxarifado
    e adiciona um ItemFichaEPI registrando a entrega.
    NÃO faz commit — o chamador é responsável.
    Seguro chamar dentro de qualquer transação existente.
    """
    if not colaborador or not colaborador.strip():
        return  # sem colaborador identificado, não registra

    try:
        ficha = FichaEPI.query.filter_by(
            colaborador=colaborador.strip(),
            almoxarifado_id=almoxarifado_id,
            status='ativa'
        ).first()

        if not ficha:
            ficha = FichaEPI(
                colaborador=colaborador.strip(),
                almoxarifado_id=almoxarifado_id,
                status='ativa',
                criado_por=registrado_por,
                data_abertura=agora()
            )
            db.session.add(ficha)
            db.session.flush()  # gera ficha.id antes de referenciar

        db.session.add(ItemFichaEPI(
            ficha_id=ficha.id,
            descricao=descricao,
            ca=ca,
            quantidade=float(quantidade),
            tamanho=tamanho,
            data_entrega=agora(),
            registrado_por=registrado_por
        ))
    except Exception as exc:
        import logging as _log
        _log.getLogger(__name__).error(f'registrar_epi_na_ficha: {exc}')
